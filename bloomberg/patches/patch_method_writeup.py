"""Add a Method sheet (numbered build steps) and cut the long notes down to desk style."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
H1 = Font(name="Calibri", size=12, bold=True)
STEP = Font(name="Calibri", size=11, bold=True, color="1F3864")
BODY = Font(name="Calibri", size=10)
MONO = Font(name="Consolas", size=10)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
BAND = PatternFill("solid", fgColor="D9E1F2")

# --- terse replacements for the notes I wrote
TRIM = {
 ("SOFR_OIS_Quotes","N2"): "Paste export bid/ask in N:O. Q/R/S = BDP minus export, bp. Mixed signs of a few tenths = ticked apart. Consistent shift = wrong source/field.",
 ("SOFR_OIS_Quotes","J2"): "S490 screen mids, frozen 07/21/26. Used only when BDP fails.",
 ("SOFR_OIS_Quotes","A38"): "Maturity and Tenor(yrs) are parsed from the tenor label, business-day rolled. Display only; the curve reads col H.",
 ("SOFR_OIS_Quotes","M4"): "Dates modified-following, weekends only. Matches S490.",
 ("SOFR_Fixings","F6"): "O/N fixing. Informational only since Step 3 sets DF(settle)=1. Live BDP overrides this cell.",
 ("SOFR_Futures","C3"): "301 BDP/BDS calls live here and nothing reads them. Left on they fill the sheet with errors for no benefit. Yes re-enables; formulas unchanged.",
 ("Bootstrap","A5"): "Purely-OIS USD SOFR curve. Short single-payment OIS to 1Y, annual OIS swaps 18M-50Y. The production curve: Swap_Pricer, Curve_Interface and the CDS module all read it. See Method.",
 ("Bootstrap","A74"): "12Y+ pillars solved on Curve_Solver. Gap rows keep log-linear DF interpolation. See Method step 6.",
 ("Bootstrap","R3"): "12Y+ solved on Curve_Solver to break the gap-row annuity circularity.",
 ("Bootstrap","R5"): "Rebuilds each solved pillar from its own annuity and differences it against the solver. ~0 confirms the par identity.",
 ("Bootstrap","T80"): "Bootstrap vs S490. Only the 32 quoted pillars match; interpolated gap rows have no counterpart.",
 ("Bootstrap","AA6"): "Targets live on S490_Snapshot.",
 ("Curve_Interface","J8"): "Step-function forward reproduces S490 (confirmed on screen: Step Forward (Cont)). Piecewise Linear (Simple) is 48x worse here. Only switch if the screen says so.",
 ("Curve_Interface","I10"): "Method 1: DF=1/(1+r_s*t), r_s linear, ACT/360, flat outside the pillars.",
 ("Curve_Interface","I11"): "Method 3: flat forward between pillars = log-linear in DF.",
 ("Curve_Interface","I12"): "Identical at the pillars. They differ off-pillar, which is what CDS_Schedule looks up. ~0.07bp in a 2Y gap.",
 ("Swap_Pricer","N4"): "Do not write text into K or L. MATCH needs a strictly ascending date column; a stray label here flattened every DF past 17Y.",
 ("Swap_Pricer","C17"): "Type a date for an off-grid or sub-annual swap. Blank uses Tenor (years).",
 ("Swap_Pricer","A89"): "SWPM Fixed vs SOFR, 10mm, curve date 07/21/26, valuation 07/23/26. Reproduce with Effective 07/23/2026, Maturity override 07/30/2026, Coupon 3.63840.",
 ("Swap_Pricer","A96"): "Leg NPVs not compared: SWPM discounts curve date to a separate valuation date, so its leg PVs are on a different basis. See row 32.",
 ("Bloomberg_S490_Validation","A78"): "Col L = frozen S490 zeros, hard-coded 07/21/26. Col E compares our bootstrap to it. D (live BDS) and G (snapshot) are secondary.",
 ("Bloomberg_S490_Validation","H6"): "Verify I3/I4/I5 against the dump at J12. Defaults are a guess at the schema.",
 ("Bloomberg_S490_Validation","H8"): "Dump lives at J12. A second copy here would spill into it.",
 ("CDS_Quotes","A14"): "Manual spreads in col E are DEMO values, not market. Col H shows which source each tenor is using. Replace col E or connect a terminal before using any CDS output.",
 ("CDS_Quotes","A16"): "Help Desk H#1330731572. Step 1 BDP(ticker,\"CDS_SPREAD_TICKER_nY\"); step 2 BDP(returned&\" BEST Curncy\",\"PX_LAST\"). Tenors 1Y/3Y/5Y/7Y/10Y.",
 ("CDS_Quotes","A17"): "No Bloomberg export exists for zero-coupon CDS spreads or bootstrapped hazard rates (Help Desk). The credit curve has no external ground truth, unlike the SOFR curve.",
 ("CDS_Quotes","A18"): "Bloomberg suggest the credit triangle, hazard = S/(1-R). Exact only for a flat curve. Hazard_Solver bootstraps properly; CDS_Validation 13-18 compares the two.",
 ("CDS_Quotes","A2"): "Step 1 (col D) resolves each tenor's CDS ticker from CDS_Parameters!B19; step 2 (col F) quotes it. Col E used only when BDP fails.",
 ("CDS_Validation","A20"): "Internal consistency only: these confirm the curve reprices its own inputs, not that it agrees with Bloomberg. No hazard-rate export exists. Compare CDSW output for a known trade.",
 ("CDS_Pricer","A29"): "Settlement figures in CDSW order. S = par spread from the curve (B15), C = contractual coupon (B6).",
 ("CDS_Pricer","A39"): "CDSW reference (CINDBK 5Y, 07/21/26): pts upfront -1.97596265, price 101.97596265, principal -197,597, accrued -8,333 (30d), cash -205,930, def exp 6,197,596. Targets for a like-for-like trade; the demo spreads here are not that entity.",
 ("Hazard_Bootstrap","A2"): "Piecewise-constant hazard. Col D solved live by Hazard_Solver, no Goal Seek. CDS_Parameters!B17 switches to a flat hazard.",
 ("Hazard_Bootstrap","A4"): "Maturity by maturity: hold earlier hazards fixed, solve the current segment so model spread = market. Col J (repricing error) stays ~0 as inputs move.",
 ("CDS_Parameters","C17"): "Bootstrap fits the term structure to CDS_Quotes; Flat overrides every tenor with B18.",
 ("CDS_Parameters","C18"): "Continuously-compounded intensity. Credit-triangle guide: S/(1-R).",
 ("CDS_Parameters","C19"): "Drives the CDS_SPREAD_TICKER_nY lookups on CDS_Quotes.",
 ("S490_Snapshot","A3"): "Single place to update the comparison. Market Rate feeds SOFR_OIS_Quotes!J (input); Zero/Discount feed Bootstrap T:U and Bloomberg_S490_Validation (target).",
 ("S490_Snapshot","A4"): "Matched pair: Market Rate is the mid of the export's bid/ask. Valid while VAL_DATE = 07/21/2026.",
 ("Curve_Solver","A2"): "Solves each long-end pillar's DF by bisection with its interpolated gap points inside the annuity. Removes the circularity that stopped the curve past 10Y.",
 ("Curve_Solver","A3"): "DF_B = (DFspot - S_B*[A_anchor + SUM tau_g*DF_g]) / (1 + S_B*tau_B),  DF_g = DF_A*(DF_B/DF_A)^w_g",
 ("Curve_Solver","A4"): "Bracket [0,1.5], 40 steps, DF resolution ~1e-12. f is monotone decreasing in DF_B.",
 ("Hazard_Solver","A2"): "Solves each piecewise-constant hazard by bisection in cells. Replaces Goal Seek; refits live as spreads, recovery or the SOFR curve move.",
 ("Hazard_Solver","A3"): "Reads only the hazard-independent columns of CDS_Schedule (D alpha, E dt, G DF_end, H DF_mid), so no circular reference.",
}

STEPS = [
 ("STEP 1", "Pull the OIS strip", "SOFR_OIS_Quotes", [
   "32 tickers, USOSFR1Z to USOSFR50, source BGN (Instructions!B12).",
   "E/F = BDP PX_BID / PX_ASK.  H = (bid+ask)/2.  Curve Side = Mid, matching S490.",
   "H falls back to J (frozen S490 mids) when BDP is unavailable, so the curve always builds.",
   "N:S reconciles the BDP pull against a Bloomberg curve export."]),
 ("STEP 2", "Dates and settlement", "Bootstrap col B", [
   "Spot = VAL_DATE + 2 business days.  Maturities = spot anniversaries.",
   "Modified following, weekends only, no holiday calendar.",
   "Bloomberg rolls 07/23/2033 to 07/25/2033; plain EDATE does not.",
   "Settle Date = curve date on the S490 screen, so DF(settle) = 1. No spot-lag stub."]),
 ("STEP 3", "Short end, 1W to 1Y", "Bootstrap rows 8-22", [
   "Single-payment OIS.   DF = 1 / (1 + S * tau)",
   "tau = ACT/360 measured from VAL_DATE, not from spot."]),
 ("STEP 4", "Long end, 18M to 50Y", "Bootstrap rows 23-72", [
   "Annual coupons.   DF = (1 - S*A) / (1 + S*tau)",
   "A = sum of tau_i * DF_i over prior annual coupon dates (col G).",
   "18M is off the annual grid: uses the annuity through 1Y and contributes nothing",
   "onward, so I23 = 0 and the 2Y accrual runs from 1Y."]),
 ("STEP 5", "Gap years", "Bootstrap, interpolated rows", [
   "11Y, 13-14Y, 16-19Y, 21-24Y, 26-29Y, 31-39Y, 41-49Y are not quoted.",
   "Log-linear in DF between surrounding pillars = step-function forward.",
   "Confirmed on the S490 screen: Interpolation = Step Forward (Cont).",
   "Tested against the guide's method 1 (piecewise linear simple): 48x worse here."]),
 ("STEP 6", "Long pillar solver", "Curve_Solver (hidden)", [
   "Gap DFs feed the annuity of the pillar they interpolate from, which is circular.",
   "Each quoted pillar 12Y+ is bisection-solved so the par identity holds with its",
   "gap points inside the annuity.",
   "Par check: Bootstrap col R, residuals ~1e-13."]),
 ("STEP 7", "Zero rates", "Bootstrap col J", [
   "z = -ln(DF) / t,  ACT/365 from VAL_DATE, continuously compounded."]),
 ("STEP 8", "Validate against S490", "Bloomberg_S490_Validation, Bootstrap T:W", [
   "Col L = S490 zero curve, hard-coded from the 07/21/26 capture.",
   "Result: 32/32 pillars, max |dz| 0.397bp, max |dDF| 1.88e-05, 0.08bp from 2Y.",
   "Residual is simple interest vs Bloomberg's daily-compounded short OIS.",
   "It is not fitted out: every curve input is a pulled quote."]),
 ("STEP 9", "Curve interface", "Curve_Interface", [
   "Exposes D(0,t) for any date. Interpolation selector at J7.",
   "CDS_Schedule and Swap_Pricer each hold their own inline copy of the",
   "interpolation; all three read the J7 selector."]),
 ("STEP 10", "Swap pricing", "Swap_Pricer", [
   "Par = (DF_eff - DF_mat) / annuity.   PV01 = notional * annuity * 1bp.",
   "B17 maturity override prices sub-annual and off-grid deals.",
   "Reprices its own inputs to <=0.16bp across 1Y-50Y.",
   "SWPM cross-check at A88."]),
 ("STEP 11", "CDS hazard curve", "CDS_Quotes, Hazard_Bootstrap, Hazard_Solver", [
   "Two-step pull per Help Desk H#1330731572:",
   "   1.  BDP(ticker, \"CDS_SPREAD_TICKER_nY\")      -> CDS ticker",
   "   2.  BDP(ticker & \" BEST Curncy\", \"PX_LAST\")  -> spread",
   "Hazard_Solver bisection-solves each piecewise-constant hazard. No Goal Seek.",
   "CDS_Parameters!B17 switches between bootstrap and a flat hazard."]),
 ("STEP 12", "CDS pricing", "CDS_Schedule, CDS_Pricer", [
   "Quarterly IMM grid. DF from Curve_Interface, survival from the hazard curve.",
   "Protection leg  (1-R) * sum Z(t)*dQ(t)",
   "Premium leg     S * [ sum d_i Z_i Q_i  +  1/2 sum d_i Z_i (Q_i-1 - Q_i) ]",
   "CDS_Pricer rows 28+ give the CDSW settlement panel.",
   "No Bloomberg export of hazard rates exists, so the credit side has no external",
   "ground truth. Validate against CDSW output for a known trade."]),
]
GAPS = [
 "CDS spreads on CDS_Quotes are demo values, not market.",
 "Short end prices ~0.15bp rich to Bloomberg (simple vs daily-compounded OIS).",
 "VAL_DATE = TODAY(). The frozen S490 comparison only lines up on 07/21/2026.",
 "SOFR_Futures feeds nothing and its pulls are switched off at B3.",
]

wb = load_workbook(WB)
for (sh, cell), txt in TRIM.items():
    if sh not in wb.sheetnames:
        continue
    try:
        wb[sh][cell].value = txt
    except AttributeError:
        pass

if "Method" in wb.sheetnames:
    del wb["Method"]
ws = wb.create_sheet("Method", 1)
ws.sheet_properties.tabColor = "808080"
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 96
ws["A1"] = "USD SOFR curve and CDS module - build method"
ws["A1"].font = H1
ws["A2"] = "Order of construction. Each step names the sheet that does the work."
ws["A2"].font = NOTE
r = 4
for tag, title, where, lines in STEPS:
    for c in "ABC":
        ws[f"{c}{r}"].fill = BAND
    ws[f"A{r}"] = tag; ws[f"A{r}"].font = STEP
    ws[f"B{r}"] = title; ws[f"B{r}"].font = STEP
    ws[f"C{r}"] = where; ws[f"C{r}"].font = NOTE
    r += 1
    for ln in lines:
        ws[f"C{r}"] = ln
        ws[f"C{r}"].font = MONO if any(x in ln for x in ("=", "->", "sum", "DF ")) else BODY
        r += 1
    r += 1
ws[f"A{r}"] = "KNOWN GAPS"; ws[f"A{r}"].font = WARN
r += 1
for g in GAPS:
    ws[f"C{r}"] = g; ws[f"C{r}"].font = BODY; r += 1

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"Method sheet written ({len(STEPS)} steps); {len(TRIM)} notes trimmed")
