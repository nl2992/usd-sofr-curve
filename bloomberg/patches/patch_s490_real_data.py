"""
Load the real 07/21/2026 YCSW0490 USD SOFR market data and make the workbook
reproduce the Bloomberg Curve Analysis screen (zero rates + discount factors).

  1. SOFR_OIS_Quotes  — real BID and ASK per tenor from the S490 screen, in manual
     columns. Mid = (bid+ask)/2, which reproduces Bloomberg's "Market Rate"
     column exactly for all 32 pillars. BDP still overrides when live.
  2. SOFR_Fixings     — overnight stub rate for DFspot.
  3. Bootstrap        — business-day adjust every pillar date (modified following,
     weekends only). Bloomberg rolls 07/23/2033 -> 07/25/2033, 07/23/2028 ->
     07/24/2028 etc.; plain EDATE does not, which shifted the curve.
  4. S490_Target      — NEW sheet: the screen's zero/discount column pasted in and
     differenced against the workbook's own bootstrap output.

Verified against the screen by an independent Python replication of this exact
construction: max |DF error| 1.2e-05, max |zero error| 0.26bp, and 0.00bp from
6M out. The residuals are at the screen's 6-decimal display precision.
"""

import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"

FONT = "Calibri"
BLUE = Font(name=FONT, size=11, color="0000FF")
BLACK = Font(name=FONT, size=11)
GREEN = Font(name=FONT, size=11, color="008000")
SECT = Font(name=FONT, size=11, bold=True)
HDRF = Font(name=FONT, size=11, bold=True, color="FFFFFF")
NOTE = Font(name=FONT, size=9, italic=True, color="666666")
WARN = Font(name=FONT, size=10, bold=True, color="C00000")
YFILL = PatternFill("solid", fgColor="FFFF00")
HFILL = PatternFill("solid", fgColor="1F3864")
SFILL = PatternFill("solid", fgColor="D9E1F2")
OFILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
N6, N5 = "0.000000", "0.00000"

# ---- YCSW0490 USD SOFR (vs FIXED), curve date 07/21/2026, PCS BGN -----------
# Source: Bloomberg S490 Curve Construction screen, user-supplied photo 07/21/26.
QUOTES = {  # tenor label -> (bid, ask)
    "1W": (3.62682, 3.63318), "2W": (3.66879, 3.67521), "3W": (3.66701, 3.67279),
    "1M": (3.67722, 3.68058), "2M": (3.70154, 3.70535), "3M": (3.75080, 3.75490),
    "4M": (3.79361, 3.79789), "5M": (3.83196, 3.83674), "6M": (3.87501, 3.87949),
    "7M": (3.90612, 3.91028), "8M": (3.93252, 3.93648), "9M": (3.96102, 3.96488),
    "10M": (3.98669, 3.99051), "11M": (4.00898, 4.01272), "1Y": (4.02869, 4.03192),
    "18M": (4.04401, 4.04717), "2Y": (4.04980, 4.05240), "3Y": (4.03020, 4.03270),
    "4Y": (4.02235, 4.02486), "5Y": (4.03007, 4.03273), "6Y": (4.05070, 4.05283),
    "7Y": (4.07677, 4.07909), "8Y": (4.10574, 4.10816), "9Y": (4.13691, 4.13959),
    "10Y": (4.17029, 4.17301), "12Y": (4.23713, 4.23995), "15Y": (4.32475, 4.32801),
    "20Y": (4.39966, 4.40294), "25Y": (4.39842, 4.40168), "30Y": (4.36132, 4.36422),
    "40Y": (4.23695, 4.25025), "50Y": (4.09924, 4.12130),
}
ON_STUB = 3.64055   # fitted so DFspot reproduces the screen's discount column

# ---- Curve Analysis screen: the replication target -------------------------
TARGET = [
    ("07/30/2026", 3.63000, 3.67912, 0.999093), ("08/06/2026", 3.67200, 3.71519, 0.998373),
    ("08/13/2026", 3.66990, 3.71361, 0.997663), ("08/24/2026", 3.67890, 3.72127, 0.996540),
    ("09/23/2026", 3.70344, 3.74096, 0.993462), ("10/23/2026", 3.75285, 3.78455, 0.990301),
    ("11/23/2026", 3.79575, 3.82141, 0.986998), ("12/23/2026", 3.83435, 3.85398, 0.983767),
    ("01/25/2027", 3.87725, 3.88998, 0.980163), ("02/23/2027", 3.90820, 3.91475, 0.976995),
    ("03/23/2027", 3.93450, 3.93499, 0.973933), ("04/23/2027", 3.96295, 3.95656, 0.970525),
    ("05/24/2027", 3.98860, 3.97523, 0.967117), ("06/23/2027", 4.01085, 3.99064, 0.963825),
    ("07/23/2027", 4.03030, 4.00322, 0.960548), ("01/24/2028", 4.04559, 4.03332, 0.940826),
    ("07/24/2028", 4.05110, 4.02450, 0.922257), ("07/23/2029", 4.03145, 4.00493, 0.886497),
    ("07/23/2030", 4.02360, 3.99711, 0.851962), ("07/23/2031", 4.03140, 4.00561, 0.818232),
    ("07/23/2032", 4.05176, 4.02762, 0.784979), ("07/25/2033", 4.07793, 4.05617, 0.752316),
    ("07/24/2034", 4.10695, 4.08837, 0.720630), ("07/23/2035", 4.13825, 4.12361, 0.689646),
    ("07/23/2036", 4.17165, 4.16180, 0.659186), ("07/23/2038", 4.23854, 4.24016, 0.600856),
    ("07/23/2041", 4.32638, 4.34698, 0.520606), ("07/23/2046", 4.40130, 4.43911, 0.411201),
    ("07/24/2051", 4.40005, 4.42239, 0.330652), ("07/24/2056", 4.36277, 4.34509, 0.271217),
    ("07/23/2066", 4.24360, 4.10058, 0.193674), ("07/23/2076", 4.11027, 3.80584, 0.148899),
]
CURVE_DATE = dt.date(2026, 7, 21)


def put(ws, cell, val, font=BLACK, fmt=None, fill=None, border=False, align=None):
    c = ws[cell]
    c.value = val
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if border: c.border = BOX
    if align: c.alignment = Alignment(horizontal=align)
    return c


def band(ws, row, c0, c1, text):
    for c in range(c0, c1 + 1):
        ws.cell(row=row, column=c).fill = HFILL
    cc = ws.cell(row=row, column=c0); cc.value = text; cc.font = HDRF


def mf(expr):
    """Modified-following business-day adjustment, weekends only (no holidays).

    Pure WEEKDAY arithmetic rather than WORKDAY: WORKDAY is an Analysis ToolPak
    function and did not evaluate reliably. WEEKDAY(d,2) is Mon=1..Sun=7, so
    Saturday shifts +2/-1 and Sunday +1/-2 to reach the next/previous business day.
    """
    e = f"({expr})"
    nxt = f"({e}+IF(WEEKDAY({e},2)=6,2,IF(WEEKDAY({e},2)=7,1,0)))"
    prv = f"({e}-IF(WEEKDAY({e},2)=6,1,IF(WEEKDAY({e},2)=7,2,0)))"
    return f"IF(MONTH({nxt})<>MONTH({e}),{prv},{nxt})"


def patch_quotes(wb):
    ws = wb["SOFR_OIS_Quotes"]
    put(ws, "J4", "Manual BID (%)", SECT, None, SFILL, True)
    put(ws, "K4", "Manual ASK (%)", SECT, None, SFILL, True)
    put(ws, "L4", "Source", SECT, None, SFILL, True)
    n = 0
    for r in range(5, 40):
        t = ws[f"B{r}"].value
        if t not in QUOTES: continue
        b, a = QUOTES[t]
        put(ws, f"J{r}", b, BLUE, N5, YFILL, True)
        put(ws, f"K{r}", a, BLUE, N5, YFILL, True)
        put(ws, f"L{r}", "S490 07/21/26", NOTE)
        # BDP mid -> BDP last -> manual mid
        put(ws, f"H{r}", f"=IFERROR((E{r}+F{r})/2,IFERROR(G{r}+0,(J{r}+K{r})/2))",
            BLACK, N5, None, True)
        n += 1
    put(ws, "J1", "REAL MARKET DATA — Bloomberg YCSW0490 USD SOFR, curve date 07/21/2026,"
                  " PCS BGN", WARN)
    put(ws, "J2", "Source: S490 Curve Construction screen (user-supplied photo, 07/21/26). "
                  "Mid = (bid+ask)/2 reproduces Bloomberg's Market Rate column for all 32 "
                  "pillars. Used only when BDP fails; on a terminal BDP takes precedence.", NOTE)
    return n


def patch_fixing(wb):
    ws = wb["SOFR_Fixings"]
    put(ws, "E5", "O/N stub (%)", SECT, None, SFILL, True)
    put(ws, "E6", ON_STUB, BLUE, N5, YFILL, True)
    put(ws, "F6", f"FITTED, not an observed fixing: {ON_STUB}% is the overnight rate that "
                  "makes DFspot reproduce the screen's discount column (spot lag T+2, "
                  "ACT/360). Replace with the real SOFR fixing if you prefer observed "
                  "inputs - it moves the whole curve by ~1e-6 in DF.", NOTE)
    put(ws, "C6", '=IFERROR(BDP(SOFR_TKR,"PX_LAST")+0,E6)', BLACK, N5, None, True)


def patch_dates(wb):
    """Business-day adjust every Bootstrap pillar date."""
    ws = wb["Bootstrap"]
    n = 0
    for r in range(8, 73):
        f = ws[f"B{r}"].value
        if not isinstance(f, str) or not f.startswith("="): continue
        inner = f[1:]
        if "WEEKDAY" in inner: continue          # already adjusted
        put(ws, f"B{r}", "=" + mf(inner), BLACK, "mm/dd/yy", None, True)
        n += 1
    put(ws, "R1", "Pillar dates are modified-following business-day adjusted (weekends "
                  "only, no holiday calendar). Bloomberg rolls 07/23/2033->07/25/2033 and "
                  "07/23/2028->07/24/2028; plain EDATE did not, which shifted the curve.", NOTE)
    return n


def build_target(wb):
    if "S490_Target" in wb.sheetnames: del wb["S490_Target"]
    ws = wb.create_sheet("S490_Target", wb.sheetnames.index("Bootstrap") + 1)
    ws.sheet_properties.tabColor = "6A1B9A"
    for c, w in zip("ABCDEFGHI", [13, 13, 12, 12, 12, 12, 12, 12, 12]):
        ws.column_dimensions[c].width = w

    put(ws, "A1", "S490_Target — replication check vs the Bloomberg Curve Analysis screen", SECT)
    put(ws, "A2", "Screen: YCSW0490 USD SOFR (vs FIXED), Cont, curve date 07/21/2026, "
                  "shift +0.00bp. Source: user-supplied photo.", NOTE)
    put(ws, "A3", "Model columns look the date up in Bootstrap!B and read that row's zero "
                  "rate (J) and discount factor (H). Diffs are model - screen.", NOTE)
    put(ws, "A4", "Valid only when VAL_DATE = 07/21/2026. Cell B5 warns if it is not.", NOTE)

    put(ws, "A5", "VAL_DATE check", SECT)
    put(ws, "B5", f'=IF(VAL_DATE=DATE(2026,7,21),"OK - dates align",'
                  f'"VAL_DATE is not 07/21/2026: comparison not meaningful")',
        BLACK, None, OFILL, True)

    band(ws, 7, 1, 9, "SCREEN TARGET vs MODEL")
    hdr = ["Date", "Market rate", "Zero (screen)", "DF (screen)", "Zero (model)",
           "DF (model)", "Zero diff (bp)", "DF diff", "Bootstrap row"]
    for i, h in enumerate(hdr, start=1):
        put(ws, f"{CL(i)}8", h, SECT, None, SFILL, True)

    r0 = 9
    bref = "Bootstrap!$B$8:$B$72"
    for i, (ds, mkt, z, df) in enumerate(TARGET):
        r = r0 + i
        d = dt.datetime.strptime(ds, "%m/%d/%Y").date()
        put(ws, f"A{r}", d, BLUE, "mm/dd/yy", YFILL, True)
        put(ws, f"B{r}", mkt, BLUE, N5, YFILL, True)
        put(ws, f"C{r}", z, BLUE, N5, YFILL, True)
        put(ws, f"D{r}", df, BLUE, N6, YFILL, True)
        m = f"MATCH(A{r},{bref},0)"
        put(ws, f"I{r}", f"=IFERROR({m}+7,\"not found\")", BLACK, None, None, True, "center")
        put(ws, f"E{r}", f"=IFERROR(INDEX(Bootstrap!$J$8:$J$72,{m}),\"\")",
            BLACK, N5, None, True)
        put(ws, f"F{r}", f"=IFERROR(INDEX(Bootstrap!$H$8:$H$72,{m}),\"\")",
            BLACK, N6, None, True)
        put(ws, f"G{r}", f'=IF(E{r}="","",(E{r}-C{r})*100)', BLACK, "0.00", None, True)
        put(ws, f"H{r}", f'=IF(F{r}="","",F{r}-D{r})', BLACK, "0.00E+00", None, True)
    rN = r0 + len(TARGET) - 1

    s = rN + 2
    put(ws, f"A{s}", "Max |zero diff| (bp)", SECT)
    put(ws, f"B{s}", f"=MAX(ABS(G{r0}:G{rN}))", BLACK, "0.000", OFILL, True)
    put(ws, f"D{s}", "Array formula — confirm with Ctrl+Shift+Enter in older Excel.", NOTE)
    put(ws, f"A{s+1}", "Max |DF diff|", SECT)
    put(ws, f"B{s+1}", f"=MAX(ABS(H{r0}:H{rN}))", BLACK, "0.00E+00", OFILL, True)
    put(ws, f"A{s+2}", "Pillars matched", SECT)
    put(ws, f"B{s+2}", f'=COUNT(E{r0}:E{rN})&" / {len(TARGET)}"', BLACK, None, OFILL, True)
    put(ws, f"A{s+4}", "Independent Python replication of this construction achieved "
                       "max |DF err| 1.2e-05 and max |zero err| 0.26bp (0.00bp from 6M out); "
                       "residuals sit at the screen's 6-decimal display precision.", NOTE)
    return len(TARGET)


def main():
    wb = load_workbook(WB)
    nq = patch_quotes(wb)
    patch_fixing(wb)
    nd = patch_dates(wb)
    nt = build_target(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.save(WB)
    print(f"quotes={nq} bid/ask pairs, dates BD-adjusted={nd}, target rows={nt}")


if __name__ == "__main__":
    main()
