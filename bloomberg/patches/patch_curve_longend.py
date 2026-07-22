"""
Fix the circular reference in the Bootstrap long end, so the SOFR curve
dynamically bootstraps discount factors all the way to 50Y.

THE BUG
-------
Quoted pillars from 12Y out are separated by "gap" rows (11Y, 13Y, 14Y, 16-19Y,
21-24Y, 26-29Y, 31-39Y, 41-49Y) whose DF is log-linearly interpolated between the
surrounding quoted pillars. But those gap rows also contribute tau*DF to the
annuity used to bootstrap the very pillar they interpolate from:

    H33 (11Y) = f(H34)              gap interpolates off the 12Y pillar
    H34 (12Y) = f(G34), G34 = SUM(I8:I33)
    I33       = F33 * H33           gap feeds the 12Y annuity
    -> H33 -> H34 -> G34 -> I33 -> H33          CIRCULAR

Excel flags this; nothing past 10Y ever computes.

THE FIX
-------
For each quoted pillar B (anchor A = previous quoted pillar), solve DF_B so the
par-swap identity holds WITH the interpolated gap points inside its annuity:

    DF_B = ( DFspot - S_B * [ A_anchor + sum_g tau_g * DF_g(DF_B) ] )
           / ( 1 + S_B * tau_B )

    where DF_g(DF_B) = DF_A * (DF_B / DF_A) ^ w_g,
          w_g = (C_g - C_A) / (C_B - C_A)          the existing interpolation

That is a 1-D root find, monotone decreasing in DF_B, solved by in-cell bisection
on a new Curve_Solver sheet. Bootstrap!H<pillar> then reads the solver, which
breaks the dependency cycle: the gap rows keep their original formulas and the
annuity column G becomes a consistent display rather than an input to H.

A par-check column on Bootstrap proves the fixed point still holds.
"""

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"

FONT = "Calibri"
BLACK = Font(name=FONT, size=11)
GREEN = Font(name=FONT, size=11, color="008000")
SECT = Font(name=FONT, size=11, bold=True)
HDRF = Font(name=FONT, size=11, bold=True, color="FFFFFF")
NOTE = Font(name=FONT, size=9, italic=True, color="666666")
HFILL = PatternFill("solid", fgColor="1F3864")
SFILL = PatternFill("solid", fgColor="D9E1F2")
OFILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
N10, N6 = "0.0000000000", "0.000000"

# (label, pillar row B, anchor row A, [gap rows between them])
PILLARS = [
    ("12Y", 34, 32, [33]),
    ("15Y", 37, 34, [35, 36]),
    ("20Y", 42, 37, [38, 39, 40, 41]),
    ("25Y", 47, 42, [43, 44, 45, 46]),
    ("30Y", 52, 47, [48, 49, 50, 51]),
    ("40Y", 62, 52, list(range(53, 62))),
    ("50Y", 72, 62, list(range(63, 72))),
]
NGAP = 9
NITER = 40          # bracket 1.5 / 2^40 ~ 1.4e-12 on the discount factor
DFSPOT = "Bootstrap!$D$4"


def put(ws, cell, val, font=BLACK, fmt=None, fill=None, border=False, align=None):
    c = ws[cell]
    c.value = val
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = BOX
    if align:
        c.alignment = Alignment(horizontal=align)
    return c


def band(ws, row, c0, c1, text):
    for c in range(c0, c1 + 1):
        ws.cell(row=row, column=c).fill = HFILL
    cc = ws.cell(row=row, column=c0)
    cc.value = text
    cc.font = HDRF


def build(wb):
    if "Curve_Solver" in wb.sheetnames:
        del wb["Curve_Solver"]
    ws = wb.create_sheet("Curve_Solver", wb.sheetnames.index("Bootstrap") + 1)
    ws.sheet_properties.tabColor = "2E7D32"
    ws.column_dimensions["A"].width = 20
    for c in range(2, 17):
        ws.column_dimensions[CL(c)].width = 13

    put(ws, "A1", "Curve_Solver", SECT)
    put(ws, "A2", "Solves each long-end quoted pillar's discount factor by bisection, "
                  "with its interpolated gap points inside the annuity. Removes the "
                  "circular reference that stopped the curve computing past 10Y.", NOTE)
    put(ws, "A3", "DF_B solves:  DF_B = (DFspot - S_B*[A_anchor + SUM_g tau_g*DF_g]) "
                  "/ (1 + S_B*tau_B),   DF_g = DF_A*(DF_B/DF_A)^w_g", NOTE)
    put(ws, "A4", f"Bracket [0, 1.5], {NITER} bisection steps -> DF resolution ~1e-12. "
                  "f is monotone decreasing in DF_B.", NOTE)

    c_g0 = 5                       # E .. first gap slot
    c_gN = c_g0 + NGAP - 1         # M
    c_sum, c_f = c_gN + 1, c_gN + 2

    row = 6
    sol = {}
    for label, b, a, gaps in PILLARS:
        band(ws, row, 1, c_f, f"PILLAR {label}  —  Bootstrap row {b}  "
                              f"—  anchor row {a}  —  {len(gaps)} gap row(s)")
        h1 = row + 1
        put(ws, f"A{h1}", "S_B (%)", SECT)
        put(ws, f"B{h1}", f"=Bootstrap!$E${b}", GREEN, "0.0000", None, True)
        put(ws, f"C{h1}", "tau_B", SECT)
        put(ws, f"D{h1}", f"=Bootstrap!$F${b}", GREEN, N6, None, True)
        put(ws, f"E{h1}", "DFspot", SECT)
        put(ws, f"F{h1}", f"={DFSPOT}", GREEN, N10, None, True)
        put(ws, f"G{h1}", "C_B", SECT)
        put(ws, f"H{h1}", f"=Bootstrap!$C${b}", GREEN, N6, None, True)
        rS, rTau, rSpot, rCB = f"$B${h1}", f"$D${h1}", f"$F${h1}", f"$H${h1}"

        h2 = row + 2
        put(ws, f"A{h2}", "DF_A (anchor)", SECT)
        # anchor pillars <=10Y are plain Bootstrap cells; later ones are solved here
        put(ws, f"B{h2}", f"=Bootstrap!$H${a}", GREEN, N10, None, True)
        put(ws, f"C{h2}", "A_anchor", SECT)
        put(ws, f"D{h2}", f"=SUM(Bootstrap!$I$8:$I${a})", GREEN, N10, None, True)
        put(ws, f"E{h2}", "C_A", SECT)
        put(ws, f"F{h2}", f"=Bootstrap!$C${a}", GREEN, N6, None, True)
        rDFA, rAnn, rCA = f"$B${h2}", f"$D${h2}", f"$F${h2}"

        # ---- gap slot headers
        r_row, r_tau, r_w = row + 3, row + 4, row + 5
        for lr, lab in zip([r_row, r_tau, r_w],
                           ["gap row", "tau_g", "weight w_g"]):
            put(ws, f"A{lr}", lab, SECT)
        for s in range(NGAP):
            col = CL(c_g0 + s)
            if s < len(gaps):
                g = gaps[s]
                put(ws, f"{col}{r_row}", g, BLACK, None, None, True, "center")
                put(ws, f"{col}{r_tau}", f"=Bootstrap!$F${g}", GREEN, N6, None, True)
                put(ws, f"{col}{r_w}",
                    f"=(Bootstrap!$C${g}-{rCA})/({rCB}-{rCA})", BLACK, N6, None, True)
            else:
                put(ws, f"{col}{r_row}", "", BLACK, None, None, True, "center")
                put(ws, f"{col}{r_tau}", 0, BLACK, N6, None, True)
                put(ws, f"{col}{r_w}", 0, BLACK, N6, None, True)
        rng_tau = f"${CL(c_g0)}${r_tau}:${CL(c_gN)}${r_tau}"

        # ---- bisection
        ith = row + 6
        put(ws, f"A{ith}", "bisection", SECT)
        for c, lab in ((2, "lo"), (3, "hi"), (4, "DF_B trial"),
                       (c_g0, "DF_g ->"), (c_sum, "gap annuity"), (c_f, "f(DF_B)")):
            put(ws, f"{CL(c)}{ith}", lab, SECT, None, SFILL, True)

        it0 = ith + 1
        for it in range(NITER):
            r = it0 + it
            if it == 0:
                put(ws, f"B{r}", 0.0, BLACK, N10, None, True)
                put(ws, f"C{r}", 1.5, BLACK, N10, None, True)
            else:
                p = r - 1
                # f is DECREASING in DF_B: f>0 -> root is to the right
                put(ws, f"B{r}", f"=IF({CL(c_f)}{p}>0,D{p},B{p})", BLACK, N10, None, True)
                put(ws, f"C{r}", f"=IF({CL(c_f)}{p}>0,C{p},D{p})", BLACK, N10, None, True)
            put(ws, f"D{r}", f"=(B{r}+C{r})/2", BLACK, N10, None, True)
            for s in range(NGAP):
                col = CL(c_g0 + s)
                if s < len(gaps):
                    put(ws, f"{col}{r}",
                        f"=IF($D{r}<=0,0,{rDFA}*($D{r}/{rDFA})^{col}${r_w})",
                        BLACK, N10, None, True)
                else:
                    put(ws, f"{col}{r}", 0, BLACK, N10, None, True)
            rng_df = f"${CL(c_g0)}{r}:${CL(c_gN)}{r}"
            put(ws, f"{CL(c_sum)}{r}", f"=SUMPRODUCT({rng_tau},{rng_df})",
                BLACK, N10, None, True)
            put(ws, f"{CL(c_f)}{r}",
                f"=({rSpot}-({rS}/100)*({rAnn}+{CL(c_sum)}{r}))"
                f"/(1+({rS}/100)*{rTau})-$D{r}",
                BLACK, "0.00E+00", None, True)

        rl = it0 + NITER - 1
        rs = rl + 1
        put(ws, f"A{rs}", f"DF({label}) SOLVED", SECT)
        put(ws, f"B{rs}", f"=D{rl}", BLACK, N10, OFILL, True)
        put(ws, f"C{rs}", "resid", SECT)
        put(ws, f"D{rs}", f"={CL(c_f)}{rl}", BLACK, "0.00E+00", None, True)
        put(ws, f"E{rs}", "zero %", SECT)
        put(ws, f"F{rs}", f"=-LN(B{rs})/{rCB}*100", BLACK, "0.0000", None, True)
        sol[b] = rs
        row = rs + 3

    return sol


def rewire(wb, sol):
    ws = wb["Bootstrap"]
    put(ws, "R7", "Par check (~0)", SECT, None, SFILL, True)
    for label, b, a, gaps in PILLARS:
        rs = sol[b]
        put(ws, f"H{b}", f"=Curve_Solver!$B${rs}", BLACK, N10, OFILL, True)
        # the fixed point: rebuilding H from the (now consistent) annuity must agree
        put(ws, f"R{b}",
            f"=($D$4-(E{b}/100)*G{b})/(1+(E{b}/100)*F{b})-H{b}",
            BLACK, "0.00E+00", None, True)
    put(ws, "R5", "Column R rebuilds each solved pillar from its own annuity column G "
                  "and differences it against the solver answer. ~0 confirms the par "
                  "identity still holds with the gap rows included.", NOTE)
    put(ws, "R3", "Long-end quoted pillars (12Y+) are solved on Curve_Solver: their "
                  "interpolated gap rows feed the annuity, which used to make column H "
                  "circular. Gap rows keep their original interpolation formulas.", NOTE)


def main():
    wb = load_workbook(WB)
    sol = build(wb)
    rewire(wb, sol)
    wb.calculation.fullCalcOnLoad = True
    wb.save(WB)
    print("solved pillar rows ->", sol)


if __name__ == "__main__":
    main()
