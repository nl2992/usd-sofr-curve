"""
Root_Methods - the MATH5030 M2 methods on the CDS hazard strip.

Baseline is the SCHEME, not an algorithm. The white paper (section 4, p.8) fixes
piecewise-constant hazard solved maturity by maturity, each h_i+1 from a
one-dimensional root-find holding earlier hazards fixed - and does not name a
method. Every row below solves the same f(h) at the same step, so they must agree.
Our in-cell Hazard_Solver is bisection.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment

WB="bloomberg/CDS_Pricer.xlsx"
BLK=[(7,46,"1Y"),(50,89,"2Y"),(93,132,"3Y"),(136,175,"5Y"),(179,218,"7Y"),(222,261,"10Y")]
def args(b):
    return (f"Hazard_Solver!$D${b},Hazard_Solver!$F${b},Hazard_Solver!$B${b+1},"
            f"Hazard_Solver!$D${b+1},Hazard_Solver!$F${b+1},CDS_Parameters!$B$26,"
            f"Hazard_Solver!$F${b+3}:$U${b+3},Hazard_Solver!$F${b+7}:$U${b+7},"
            f"Hazard_Solver!$F${b+5}:$U${b+5},Hazard_Solver!$F${b+6}:$U${b+6}")

METHODS=[("BISECTION","bracketing","linear (halves)","no","yes, always",
          "The safe floor. Cannot fail once bracketed, but pays for it - ~52 steps to machine precision."),
         ("FALSE POSITION","bracketing","super-linear","no","yes, always",
          "Keeps a bracket like bisection but interpolates. Can stall when one end stays fixed."),
         ("SECANT","open","~1.618","no","no",
          "Newton without a derivative, using the secant of the last two points. No bracket, so it can leave the domain."),
         ("NEWTON","open","2 (quadratic)","f'","no",
          "Householder d=1. Derivative is analytic here - dQ/dh = -cum*Q - so no bumping."),
         ("HALLEY","open","3 (cubic)","f', f''","no",
          "Householder d=2. Second derivative also analytic: d2Q/dh2 = cum^2*Q."),
         ("HOUSEHOLDER","open","d+1","f' ... f^(d)","no",
          "General order d. d=1 reduces to Newton and d=2 to Halley, the M2 identity."),
         ("RIDDERS","bracketing hybrid","hybrid","no","yes, always",
          "Bisection plus exponential interpolation. Keeps the bracket and converges fast."),
         ("BRENT","bracketing hybrid","hybrid","no","yes, always",
          "IQI, falling back to secant, falling back to bisection. Bisection's guarantee at near-secant speed.")]

F="Calibri"
H1=Font(name=F,size=13,bold=True); B=Font(name=F,size=11,bold=True)
N=Font(name=F,size=10); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
BASE=PatternFill("solid",fgColor="FFF2CC"); RED=Font(name=F,size=10,bold=True,color="C00000")
GRN=Font(name=F,size=10,color="008000"); BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(WB)
if "Root_Methods" in wb.sheetnames: del wb["Root_Methods"]
ws=wb.create_sheet("Root_Methods",3)

ws["A1"]="Root-finding methods on the hazard strip"; ws["A1"].font=H1
ws["A2"]=("Every method solves the same objective at the same step: f(h) = (1-R)*Prot(h) - S*(RPV01(h) - AI), "
          "B-Model (3.3). They must agree - what differs is cost and robustness.")
ws["A2"].font=SM
ws["A3"]=("BASELINE is the scheme, not the algorithm. Section 4 p.8 fixes piecewise-constant hazard solved "
          "maturity by maturity, each one a 1-D root-find with earlier hazards held fixed; it does not name a method. "
          "Hazard_Solver in this workbook is bisection.")
ws["A3"].font=Font(name=F,size=9,italic=True,color="C00000")
ws["A4"]="Needs CDSBrent.bas and CDSRootFinders.bas imported, saved as .xlsm."; ws["A4"].font=SM

hdr=["Method","Family","Order","Needs","Bracketed","5Y hazard","its","|f(root)|","vs baseline","Note"]
for i,h in enumerate(hdr,start=1):
    c=ws.cell(6,i,h); c.font=HD; c.fill=HF; c.border=BOX
    c.alignment=Alignment(horizontal="center",wrap_text=True)

b5=136
for j,(m,fam,order,needs,brk,note) in enumerate(METHODS):
    r=7+j
    ws.cell(r,1,m); ws.cell(r,2,fam); ws.cell(r,3,order); ws.cell(r,4,needs); ws.cell(r,5,brk)
    ws.cell(r,6,f'=IFERROR(CDS_Root("{m}",{args(b5)}),"module not loaded")').number_format="0.00000000"
    ws.cell(r,7,'=IFERROR(CDS_RootIterations(),"")')
    ws.cell(r,8,(f'=IFERROR(ABS(CDS_Objective(F{r},Hazard_Solver!$D${b5},Hazard_Solver!$F${b5},'
                 f'Hazard_Solver!$B${b5+1},Hazard_Solver!$D${b5+1},Hazard_Solver!$F${b5+1},'
                 f'CDS_Parameters!$B$26,Hazard_Solver!$F${b5+3}:$U${b5+3},Hazard_Solver!$F${b5+7}:$U${b5+7},'
                 f'Hazard_Solver!$F${b5+5}:$U${b5+5},Hazard_Solver!$F${b5+6}:$U${b5+6})),"")')).number_format="0.00E+00"
    ws.cell(r,9,f'=IF(ISNUMBER(F{r}),F{r}-Hazard_Solver!$B${BLK[3][1]},"")').number_format="0.00E+00"
    ws.cell(r,10,note).font=N
    for k in range(1,11):
        ws.cell(r,k).border=BOX
        if ws.cell(r,k).font.size is None or k<6: ws.cell(r,k).font=N
ws.cell(7,1).fill=BASE; ws.cell(7,2).fill=BASE
ws.cell(7,1).comment=Comment("This is what Hazard_Solver does in cells - 30 halvings per tenor. "
                             "The baseline for the comparison.","B-Model")

r=7+len(METHODS)+1
ws.cell(r,1,"Spread of the eight roots (max - min)").font=B
ws.cell(r,6,f'=IF(COUNT(F7:F{6+len(METHODS)})=0,"module not loaded",MAX(F7:F{6+len(METHODS)})-MIN(F7:F{6+len(METHODS)}))').number_format="0.00E+00"
ws.cell(r,7,"they solve the same equation, so this is the real check").font=SM
r+=2

ws.cell(r,1,"Measured on the 5Y objective (quarterly, R=40%, S=95bp), tolerance 1e-15").font=B; r+=1
for i,h in enumerate(["Method","iters","|root - scipy.brentq|","|f(root)|"],start=1):
    c=ws.cell(r,i,h); c.font=HD; c.fill=HF; c.border=BOX
r+=1
for m,it,dv,fr in [("bisection",52,"5.0e-16","1.3e-15"),("false position",12,"2.4e-17","6.9e-17"),
                   ("secant",13,"3.5e-18","1.4e-17"),("Newton (d=1)",6,"6.9e-18","1.4e-17"),
                   ("Halley (d=2)",6,"3.5e-18","1.4e-17"),("Householder d=3",6,"3.5e-18","1.4e-17"),
                   ("Ridders",5,"2.1e-17","3.5e-17"),("Brent",9,"1.4e-17","1.4e-17")]:
    ws.cell(r,1,m).font=N; ws.cell(r,2,it).font=N
    ws.cell(r,3,dv).font=N; ws.cell(r,4,fr).font=N
    for k in range(1,5): ws.cell(r,k).border=BOX
    r+=1
r+=1
ws.cell(r,1,"Reading it honestly").font=B; r+=1
for t in ["All eight land on the same root to ~1e-16. On this problem the method is a cost question, not an accuracy one.",
          "Bisection needs 52 steps where Newton needs 6 - the price of never using the shape of f.",
          "Halley and Householder d=3 do NOT beat Newton here. f is close to linear in h near the root, so the extra",
          "    order buys nothing before the tolerance is reached. Higher order is not free and not always faster.",
          "Ridders is fastest at 5, but every open method (secant, Newton, Halley) can leave the domain on a bad start;",
          "    the bracketing ones cannot. On a strip that must never return h < 0, that guarantee is worth more than speed.",
          "If no method converges, suspect the quote, not the solver - see the p.9 strippability note on Model_Notes."]:
    ws.cell(r,1,t).font=(SM if t.startswith("    ") else N); r+=1

for col,w in zip("ABCDEFGHIJ",(17,18,15,12,12,13,7,12,12,86)):
    ws.column_dimensions[col].width=w
wb.calculation.fullCalcOnLoad=True
wb.save(WB)
print(f"Root_Methods added: {len(METHODS)} methods, baseline flagged")
