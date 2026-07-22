"""
Split the par rate into dynamic and hard-coded, and compare the two.

Bootstrap col E already feeds off SOFR_OIS_Quotes!H, which is the live BDP mid
falling back to the frozen S490 mid. That is the DYNAMIC input and the curve is
built from it.

Adds:
  Bootstrap!X   Par rate S hard-coded - read straight from S490_Snapshot, never
                touches BDP
  Bootstrap!Y   dynamic minus hard-coded, bp
  Bloomberg_S490_Validation!P:R   the same comparison per pillar, with max/mean

Off-terminal the two are identical by construction (H falls through to the same
frozen mids), so the column reads 0.00 and that is the wiring check. On a live
terminal it shows how far the pull has moved from the frozen capture.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
SF = PatternFill("solid", fgColor="D9E1F2")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

SNAP_D = "S490_Snapshot!$A$7:$A$38"
SNAP_M = "S490_Snapshot!$C$7:$C$38"


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


wb = load_workbook(WB)

# ---- Bootstrap: label E as the dynamic feed, add the hard-coded twin
b = wb["Bootstrap"]
put(b, "E7", "Par rate S (%) dynamic", BOLD, None, SF, True)
put(b, "X7", "Par rate S hard-coded (%)", BOLD, None, SF, True)
put(b, "Y7", "d dynamic - hard (bp)", BOLD, None, SF, True)
put(b, "X6", "from S490_Snapshot, never BDP", NOTE)
n = 0
for r in range(8, 73):
    m = f"MATCH(B{r},{SNAP_D},0)"
    put(b, f"X{r}", f'=IFERROR(INDEX({SNAP_M},{m}),"")', BLACK, "0.00000", None, True)
    put(b, f"Y{r}", f'=IF(OR(X{r}="",NOT(ISNUMBER(E{r}))),"",(E{r}-X{r})*100)',
        BLACK, "0.00", None, True)
    n += 1
put(b, "X75", "Pillars", BOLD)
put(b, "Y75", '=COUNT(X8:X72)&" / 32"', BLACK, None, OF, True)
put(b, "X76", "Max |d| (bp)", BOLD)
put(b, "Y76", '=IF(COUNT(Y8:Y72)=0,"",MAX(MAX(Y8:Y72),-MIN(Y8:Y72)))', BLACK, "0.000", OF, True)
put(b, "X78", "E is the live BDP mid via SOFR_OIS_Quotes!H, falling back to the frozen mids. "
              "X reads S490_Snapshot directly. Off-terminal both resolve to the same numbers "
              "so Y is 0.00; on a terminal Y is how far the pull has moved.", NOTE)

# ---- validation sheet: same comparison
v = wb["Bloomberg_S490_Validation"]
put(v, "P7", "Par S dynamic (%)", BOLD, None, SF, True)
put(v, "Q7", "Par S hard-coded (%)", BOLD, None, SF, True)
put(v, "R7", "d (bp)", BOLD, None, SF, True)
for r in range(8, 73):
    put(v, f"P{r}", f'=IF(ISNUMBER(Bootstrap!E{r}),Bootstrap!E{r},"")', BLACK, "0.00000", None, True)
    put(v, f"Q{r}", f'=IF(ISNUMBER(Bootstrap!X{r}),Bootstrap!X{r},"")', BLACK, "0.00000", None, True)
    put(v, f"R{r}", f'=IF(OR(P{r}="",Q{r}=""),"",(P{r}-Q{r})*100)', BLACK, "0.00", None, True)
put(v, "A80", "Input check: par rates", BOLD)
put(v, "C80", '=COUNT(R8:R72)&" pillars"', BLACK, None, OF, True)
put(v, "A81", "Max |d| dynamic vs hard-coded (bp)", BOLD)
put(v, "C81", '=IF(COUNT(R8:R72)=0,"",MAX(MAX(R8:R72),-MIN(R8:R72)))', BLACK, "0.000", OF, True)
put(v, "A82", "Mean d (bp)", BOLD)
put(v, "C82", '=IF(COUNT(R8:R72)=0,"",AVERAGE(R8:R72))', BLACK, "0.000", OF, True)
put(v, "A84", "P/Q/R compare the INPUT. L/E compare the OUTPUT. A non-zero mean in C82 means "
              "the live pull has drifted from the frozen capture, so the output diff in E is "
              "measuring against a different market.", NOTE)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"Bootstrap X/Y added ({n} rows); validation P:R comparison added")
