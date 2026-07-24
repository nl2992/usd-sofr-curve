"""
Add a single self-contained 'KRD reformed' sheet: the delta in Kevin's tenor
format (2D..30Y | dv01), on his standard grid (8Y/9Y dropped), with the working
that proves it right below the output table. Links to Quotes and Amort-Set-up.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter as CL

WB="bloomberg/KRW_IRS_Bootstrap_Book1.xlsx"
AMT="'Amort-Set-up'"
KEVROWS=[5,6,7,8,9,10,11,12,13,16,17,18,19]      # Quotes rows for Kevin's 13 nodes (skip 8Y=14,9Y=15)
NQ=len(KEVROWS)                                   # 13 nodes
BUCK=list(range(10))                              # first 10 nodes (3M..10Y) get bumped
NS=1+len(BUCK)                                    # base + 10 = 11 scenarios

F="Calibri"
H1=Font(name=F,size=14,bold=True); B=Font(name=F,size=11,bold=True); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=9,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
RES=PatternFill("solid",fgColor="FCE4D6"); GRN=Font(name=F,size=10,color="008000")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4); MONEY='#,##0'; PCT="0.000%"

wb=load_workbook(WB)
if "KRD reformed" in wb.sheetnames: del wb["KRD reformed"]
ws=wb.create_sheet("KRD reformed", wb.sheetnames.index("KRD")+1)

ws["A1"]="KRD reformed — delta in Kevin's format"; ws["A1"].font=H1
ws["A2"]=("Bump each tenor 1bp, re-bootstrap, re-price, read the NPV change. Kevin's grid "
          "(8Y/9Y dropped). Self-contained: links only to Quotes and Amort-Set-up.")
ws["A2"].font=SM

# ---------- working section (built first so the output can reference it) ----------
QM0=30                                            # quote matrix top row
QR0=47; QR1=QR0+39                                # quarter matrix rows
ANCH=QR0-1                                        # DF anchor row (=1)
RC0=7; DC0=19                                     # rate cols G.., DF cols S..
TAU="Quotes!$B$2"

ws.cell(28,1,"WORKING (proof)").font=B
# quote matrix
for i,h in enumerate(["Tenor","Year","Base"],1):
    c=ws.cell(QM0-1,i,h); c.font=HD; c.fill=HF
for t in range(NQ):
    r=QM0+t; qr=KEVROWS[t]
    ws.cell(r,1,f"=Quotes!A{qr}")
    ws.cell(r,2,f"=Quotes!C{qr}").number_format="0.00"
    ws.cell(r,3,f"=Quotes!B{qr}").number_format=PCT
    for j in range(NS):
        col=4+j
        if j==0: ws.cell(r,col,f"=$C{r}")
        else:
            tgt=QM0+BUCK[j-1]
            ws.cell(r,col,(f"=$C{r}+0.0001" if r==tgt else f"=$C{r}"))
        ws.cell(r,col).number_format=PCT
QMAT=f"$D${QM0}:${CL(3+NS)}${QM0+NQ-1}"; QYR=f"$B${QM0}:$B${QM0+NQ-1}"

# quarter matrix
for i,h in enumerate(["Qtr","Year","m","w","Notional"],1):
    c=ws.cell(QR0-1,i,h); c.font=HD; c.fill=HF
for j in range(NS):
    ws.cell(QR0-1,RC0+j,j+1)                        # rate scenario index header
    ws.cell(ANCH,DC0+j,1).number_format="0.00000000"   # DF anchor = 1
for qq in range(1,41):
    r=QR0+qq-1
    ws.cell(r,1,qq)
    ws.cell(r,2,f"=A{r}*0.25").number_format="0.00"
    ws.cell(r,3,f"=MIN(MATCH(B{r},{QYR},1),{NQ-1})")
    ws.cell(r,4,f"=(B{r}-INDEX({QYR},C{r}))/(INDEX({QYR},C{r}+1)-INDEX({QYR},C{r}))")
    ws.cell(r,5,f"={AMT}!E{19+qq}").number_format=MONEY     # notional from Amort-Set-up
    for j in range(NS):
        rc=RC0+j; dc=DC0+j; RL=CL(rc); DLx=CL(dc); hh=f"{RL}${QR0-1}"
        ws.cell(r,rc,f"=INDEX({QMAT},$C{r},{hh})+$D{r}*(INDEX({QMAT},$C{r}+1,{hh})-INDEX({QMAT},$C{r},{hh}))").number_format=PCT
        if qq==1: ws.cell(r,dc,f"=1/(1+{RL}{r}*{TAU})")
        else: ws.cell(r,dc,f"=(1-{RL}{r}*{TAU}*SUM({DLx}${QR0}:{DLx}{r-1}))/(1+{RL}{r}*{TAU})")
        ws.cell(r,dc).number_format="0.00000000"

# frozen par = base FloatPV / base annuity  (so base NPV = 0)
Nr=f"$E${QR0}:$E${QR1}"
S=CL(DC0)                                          # base DF column
FP,AN,PAR=QR1+2,QR1+3,QR1+4
ws.cell(FP,1,"base Float PV").font=B
ws.cell(FP,3,f"=SUMPRODUCT({Nr},{S}${ANCH}:{S}${QR1-1}-{S}${QR0}:{S}${QR1})").number_format=MONEY
ws.cell(AN,1,"base annuity").font=B
ws.cell(AN,3,f"={TAU}*SUMPRODUCT({Nr},{S}${QR0}:{S}${QR1})").number_format=MONEY
ws.cell(PAR,1,"frozen par rate").font=B
ws.cell(PAR,3,f"=C{FP}/C{AN}").number_format=PCT; ws.cell(PAR,3).fill=RES; ws.cell(PAR,3).font=B
PARC=f"$C${PAR}"

# NPV + delta per scenario
FR,FL,NP,DL=PAR+2,PAR+3,PAR+4,PAR+5
for a,t in ((FR,"Fixed PV"),(FL,"Float PV"),(NP,"NPV"),(DL,"DELTA")): ws.cell(a,1,t).font=B
for j in range(NS):
    dc=DC0+j; DLc=CL(dc)
    ws.cell(FR,dc,f"={PARC}*{TAU}*SUMPRODUCT({Nr},{DLc}${QR0}:{DLc}${QR1})").number_format=MONEY
    ws.cell(FL,dc,f"=SUMPRODUCT({Nr},{DLc}${ANCH}:{DLc}${QR1-1}-{DLc}${QR0}:{DLc}${QR1})").number_format=MONEY
    ws.cell(NP,dc,f"={DLc}{FR}-{DLc}{FL}").number_format=MONEY
    ws.cell(DL,dc,f"={DLc}{NP}-${CL(DC0)}{NP}").number_format=MONEY
# label delta cols with bumped tenor
for j in range(1,NS):
    dc=DC0+j; tgt=QM0+BUCK[j-1]
    ws.cell(DL+1,dc,f"=$A{tgt}").font=SM

# ---------- OUTPUT table (top) referencing the delta row ----------
ws.cell(4,1,"OUTPUT  (send to Kevin)").font=B
for i,h in enumerate(["Tenor","dv01"],1):
    c=ws.cell(5,i,h); c.font=HD; c.fill=HF; c.border=BOX
KEVIN=["2D","1W","1M","2M","3M","6M","9M","1Y","2Y","3Y","4Y","5Y","7Y","10Y","12Y","15Y","20Y","25Y","30Y"]
# map Kevin label -> the delta column that bumped it
lbl2node={"3M":0,"6M":1,"9M":2,"1Y":3,"2Y":4,"3Y":5,"4Y":6,"5Y":7,"7Y":8,"10Y":9}
r=6
for tn in KEVIN:
    ws.cell(r,1,tn).border=BOX
    if tn in lbl2node:
        dc=DC0+1+lbl2node[tn]   # delta col for that bump
        ws.cell(r,2,f"={CL(dc)}${DL}").number_format=MONEY
    else:
        ws.cell(r,2,0).number_format=MONEY
    ws.cell(r,2).border=BOX; ws.cell(r,2).font=GRN
    r+=1
ws.cell(r,1,"TOTAL").font=B; ws.cell(r,1).border=BOX
ws.cell(r,2,f"=SUM(B6:B{r-1})").number_format=MONEY; ws.cell(r,2).fill=RES; ws.cell(r,2).font=B; ws.cell(r,2).border=BOX
ws.cell(r+2,1,"Proof: TOTAL = sum of buckets = the parallel DV01. Base NPV (working) = 0.").font=SM

for col,w in zip("ABCDE",(10,16,12,10,16)): ws.column_dimensions[col].width=w
wb.save(WB)
print("KRD reformed added at position", wb.sheetnames.index("KRD reformed"), "of", wb.sheetnames)
