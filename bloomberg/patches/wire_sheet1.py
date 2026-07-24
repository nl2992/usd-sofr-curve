"""
(1) Create 'Sheet1' mirroring the master's folded Kevin-format table (Tenor A,
    dv01 B, rows 2-20, TOTAL 21) — values pulled from KRD-v3 so it's live here.
(2) Repoint the Hedge blotter to reference Sheet1, so it drops into the master
    (which already has Sheet1) with no repointing.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
WB="bloomberg/KRW_IRS_Bootstrap_Book1.xlsx"
F="Calibri"; HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="C00000")
B=Font(name=F,size=10,bold=True); RES=PatternFill("solid",fgColor="FCE4D6"); MONEY='#,##0'
wb=load_workbook(WB)

# ---------- (1) Sheet1 mirror ----------
if "Sheet1" in wb.sheetnames: del wb["Sheet1"]
s=wb.create_sheet("Sheet1")                       # matches master's tab
s["A1"]="Tenor"; s["B1"]="dv01"
for cc in ("A1","B1"): s[cc].font=HD; s[cc].fill=HF
# KRD-v3 output rows 8..26 -> Sheet1 rows 2..20 ; TOTAL 27 -> 21
for r in range(2,21):
    s.cell(r,1,f"='KRD-v3'!A{r+6}")
    s.cell(r,2,f"='KRD-v3'!B{r+6}").number_format=MONEY
s.cell(21,1,"TOTAL").font=B
s.cell(21,2,"='KRD-v3'!B27").number_format=MONEY; s.cell(21,2).fill=RES; s.cell(21,2).font=B
s.column_dimensions["A"].width=12; s.column_dimensions["B"].width=16

# ---------- (2) repoint Hedge formulas KRD-v3 -> Sheet1 ----------
# Sheet1 rows: 3M=6,6M=7,9M=8,1Y=9,2Y=10,3Y=11,4Y=12,5Y=13,7Y=14,10Y=15 ; TOTAL=21
MAP={"$B$12":"$B$6","$B$13":"$B$7","$B$14":"$B$8","$B$15":"$B$9","$B$16":"$B$10",
     "$B$17":"$B$11","$B$18":"$B$12","$B$19":"$B$13","$B$20":"$B$14","$B$21":"$B$15","$B$27":"$B$21"}
h=wb["Hedge"]
n=0
for row in h.iter_rows():
    for cell in row:
        if isinstance(cell.value,str) and "KRD-v3" in cell.value:
            v=cell.value
            for a,b in MAP.items(): v=v.replace("'KRD-v3'!"+a,"Sheet1!"+b)
            cell.value=v; n+=1
# fix the footer note to reflect it's now wired
for row in h.iter_rows(min_row=30,max_row=40,min_col=1,max_col=1):
    for cell in row:
        if isinstance(cell.value,str) and cell.value.startswith("In the master"):
            cell.value="Wired to Sheet1 (1Y=B9 … 10Y=B15). Drops into the master as-is; confirm Bootstrap!B3:B122/D3:D122 match."
print("repointed",n,"Hedge cells to Sheet1")
wb.save(WB)
print("sheets:",wb.sheetnames)
