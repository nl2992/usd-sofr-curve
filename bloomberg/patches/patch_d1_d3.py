"""
D1-D3 from CLAUDE.md: settlement date, par spread (3.3), upfront (3.2).

D1  Settlement T_s = T + 3 business days (p.6). Closed form T+3+IF(WEEKDAY(T,2)>=3,2,0),
    checked against a business-day loop for every weekday start over a year.
    07/21/2026 (Tue) -> 07/24/2026, matching the reference screen's Cash Settled On.
    D(T_s) read off Curve_Interface, so the DF component drives it.

D2  Par spread (3.3):
        S = Protection Leg / (Premium Leg - Accrued Interest * D(T_s)) | C=1bp
    PV01 is the premium leg NET of accrued interest, discounted from settlement. We had
    Protection/RPV01 with no netting. Applied in three places that must agree:
      - CDS_Pricer!B15
      - Hazard_Bootstrap!I (model spread)
      - Hazard_Solver objective f, i.e. the stripper itself (SS4 says (3.3) is what is solved)
    The netting term D0*D(Ts) is ~0.083 against an RPV01 of ~4.4, so it moves the stripped
    hazards by roughly 2%. It is not cosmetic.

D3  Market value (3.1) and upfront (3.2):
        Market Value = Protection Leg - Premium Leg          both discounted to T
        Upfront      = Market Value / D(T_s) + Accrued Interest
    Upfront is identically zero when C = S, which is the algebraic check that accrued
    interest enters as +A*C*D0.

NOT settled here: how (3.2)'s single Upfront maps onto CDSW's Principal / Accrued / Cash
three-way split. The paper gives the total, not the screen decomposition. Rows are labelled
and the split is left as before pending verification against a live screen at close.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLACK = Font(name="Calibri", size=11)
GREEN = Font(name="Calibri", size=11, color="008000")
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

BLOCKS = [6, 49, 92, 135, 178, 221]      # Hazard_Solver block start rows
NITER = 30
VAL = "CDS_Parameters!$B$4"


def mf(e):
    nxt = f"({e}+IF(WEEKDAY({e},2)=6,2,IF(WEEKDAY({e},2)=7,1,0)))"
    prv = f"({e}-IF(WEEKDAY({e},2)=6,1,IF(WEEKDAY({e},2)=7,2,0)))"
    return f"IF(MONTH({nxt})<>MONTH({e}),{prv},{nxt})"


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


wb = load_workbook(WB)

# ---------- D1: settlement block on CDS_Parameters ----------
p = wb["CDS_Parameters"]
put(p, "A20", "SETTLEMENT / ACCRUAL  (white paper p.6-7)", BOLD)
put(p, "A21", "Settlement date T_s", BLACK)
put(p, "B21", f'={VAL}+3+IF(WEEKDAY({VAL},2)>=3,2,0)', BLACK, "mm/dd/yyyy", OF, True)
put(p, "C21", "T + 3 business days (p.6). 07/21/26 -> 07/24/26, matching the screen.", NOTE)

K, L = "'Curve_Interface'!$K$8:$K$73", "'Curve_Interface'!$L$8:$L$73"
i = f"MATCH($B$21,{K},1)"
Li, Li1, Ki, Ki1 = f"INDEX({L},{i})", f"INDEX({L},{i}+1)", f"INDEX({K},{i})", f"INDEX({K},{i}+1)"
put(p, "A22", "D(T_s)", BLACK)
put(p, "B22", f'=IFERROR({Li}*({Li1}/{Li})^(($B$21-{Ki})/({Ki1}-{Ki})),{Li})',
    GREEN, "0.00000000", OF, True)
put(p, "C22", "Discount factor at settlement, off Curve_Interface. The DF component.", NOTE)

put(p, "A23", "Accrual start T_0", BLACK)
put(p, "B23", "=" + mf("EDATE($B$5,-3)"), BLACK, "mm/dd/yyyy", OF, True)
put(p, "C23", "Previous IMM roll, business-day adjusted.", NOTE)
put(p, "A24", "Accrued days", BLACK)
put(p, "B24", f'=MAX(0,{VAL}-$B$23+1)', BLACK, "0", OF, True)
put(p, "C24", "Includes the pricing date (p.7): one more than the daycount difference.", NOTE)
put(p, "A25", "Accrual fraction D0 (ACT/360)", BLACK)
put(p, "B25", "=$B$24/360", BLACK, "0.00000000", OF, True)
put(p, "A26", "PV01 netting term  D0 * D(T_s)", BLACK)
put(p, "B26", "=$B$25*$B$22", BLACK, "0.00000000", OF, True)
put(p, "C26", "Subtracted from the premium leg in (3.3). ~0.083 against RPV01 ~4.4, so it "
              "moves the stripped hazards by roughly 2%.", NOTE)

NET = "CDS_Parameters!$B$26"

# ---------- D2: (3.3) in all three places ----------
hs = wb["Hazard_Solver"]
n = 0
for b in BLOCKS:
    h1, h2 = b + 1, b + 2
    it0 = b + 10
    for r in range(it0, it0 + NITER):
        put(hs, f"T{r}", f"=(1-$F${h1})*($D${h2}+R{r})-$D${h1}*(($F${h2}+S{r})-{NET})",
            BLACK, "0.00000000")
        n += 1
put(hs, "A5", f"Objective is (3.3): protection leg minus S x PV01, where PV01 is the premium "
              f"leg NET of accrued interest x D(T_s) ({NET}).", NOTE)

hb = wb["Hazard_Bootstrap"]
for r in range(7, 13):
    put(hb, f"I{r}", f"=(1-CDS_Parameters!$B$8)*H{r}/(G{r}-{NET})*10000", BLACK, "0.0000")
put(hb, "A5", "Model spread per (3.3): protection / (RPV01 - accrued interest x D(T_s)).", NOTE)

# ---------- D2/D3: CDS_Pricer ----------
c = wb["CDS_Pricer"]
prot = 'SUMIF(CDS_Schedule!$C$7:$C$46,"<="&B8,CDS_Schedule!$O$7:$O$46)'
put(c, "A15", "Par spread (bp)  (3.3)", BLACK)
put(c, "B15", f"=(1-B5)*{prot}/(B12-{NET})*10000", BLACK, "0.0000", OF, True)
put(c, "C15", "Protection / (RPV01 net of accrued interest x D(T_s)).", NOTE)

put(c, "A19", "PV01  (premium leg net of AI, per 1bp)", BLACK)
put(c, "B19", f"=B4*(B12-{NET})*0.0001", BLACK, "#,##0.00", None, True)
put(c, "A20", "Market Value  (3.1)", BLACK)
put(c, "B20", "=B13-B14", BLACK, "#,##0.00", OF, True)
put(c, "C20", "Protection leg minus premium leg, both to the pricing date.", NOTE)
put(c, "A26", "Accrued Interest  A*C*D0", BLACK)
put(c, "B26", f"=B4*(B6/10000)*CDS_Parameters!$B$25", BLACK, "#,##0.00", None, True)
put(c, "A27", "Upfront  (3.2)", BOLD)
put(c, "B27", f"=B20/CDS_Parameters!$B$22+B26", BLACK, "#,##0.00", OF, True)
put(c, "C27", "Market Value / D(T_s) + Accrued Interest. Identically zero when C = S.", NOTE)

put(c, "A38", "Upfront (3.2) above is the paper's total. How it splits into CDSW's "
              "Principal / Accrued / Cash is NOT given by the paper - that mapping is "
              "unverified and awaits a live screen check.", WARN)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"D1 settlement block; D2 applied to {n} solver cells + Hazard_Bootstrap!I + CDS_Pricer!B15; D3 (3.1)/(3.2)")
