"""
Lay the working out properly: freeze panes, column widths, header styling,
consistent number formats.

Cosmetic only - no formula is touched.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(bottom=THIN)

# sheet -> (header row, first data row, last data row, {col: (width, numfmt)})
D6, D8, PCT, CCY, DT = "0.000000", "0.00000000", "0.0000", "#,##0.00", "mm/dd/yyyy"
LAYOUT = {
 "SOFR_OIS_Quotes": (4, 5, 36, {"A":(12,None),"B":(8,None),"C":(13,DT),"D":(11,D6),
    "E":(10,PCT),"F":(10,PCT),"G":(10,PCT),"H":(11,PCT),"I":(9,None),"J":(12,PCT),
    "K":(15,None),"N":(12,PCT),"O":(12,PCT),"P":(12,PCT),"Q":(10,"0.00"),"R":(10,"0.00"),"S":(10,"0.00")}),
 "S490_Snapshot": (6, 7, 38, {"A":(13,DT),"B":(9,None),"C":(15,PCT),"D":(14,PCT),"E":(12,D6)}),
 "Bootstrap": (7, 8, 72, {"A":(8,None),"B":(13,DT),"C":(10,D6),"D":(11,D6),"E":(13,PCT),
    "F":(11,D6),"G":(13,D8),"H":(14,D8),"I":(13,D8),"J":(13,PCT),"K":(11,PCT),"L":(12,PCT),
    "M":(11,"0.00"),"N":(11,None),"R":(12,"0.00E+00"),"S":(13,D8),
    "T":(12,PCT),"U":(12,D6),"V":(11,"0.00"),"W":(11,"0.00E+00"),"X":(15,PCT),"Y":(13,"0.00")}),
 "Curve_Interface": (9, 10, 73, {"A":(9,None),"B":(13,DT),"C":(11,D6),"D":(13,DT),"E":(13,DT),
    "F":(11,PCT),"G":(14,D8),"H":(12,PCT),"K":(13,DT),"L":(14,D8)}),
 "Bloomberg_S490_Validation": (7, 8, 72, {"A":(8,None),"B":(13,DT),"C":(12,PCT),"D":(12,PCT),
    "E":(10,"0.00"),"F":(13,None),"G":(14,PCT),"L":(14,PCT),"M":(12,D6),"N":(12,D6),
    "O":(11,"0.00E+00"),"P":(13,PCT),"Q":(15,PCT),"R":(10,"0.00")}),
 "Hazard_Bootstrap": (6, 7, 12, {"A":(8,None),"B":(13,DT),"C":(13,PCT),"D":(14,D8),
    "E":(12,D6),"F":(12,D6),"G":(12,D6),"H":(13,D8),"I":(13,PCT),"J":(12,"0.00E+00"),"L":(13,"0.00E+00")}),
 "CDS_Schedule": (6, 7, 46, {"A":(8,None),"B":(13,DT),"C":(13,DT),"D":(11,D6),"E":(11,D6),
    "F":(10,D6),"G":(12,D8),"H":(12,D8),"I":(11,D8),"J":(12,D8),"K":(12,D8),"L":(11,D8),
    "M":(13,D8),"N":(13,D8),"O":(13,D8)}),
 "CDS_Quotes": (6, 7, 12, {"A":(8,None),"B":(11,None),"C":(13,DT),"D":(24,None),
    "E":(15,PCT),"F":(15,PCT),"G":(15,D8),"H":(17,None),"I":(40,None)}),
}
FREEZE_ONLY = {"Hazard_Solver": "B6", "Curve_Solver": "B6", "Swap_Pricer": "B35",
               "CDS_Pricer": "B11", "SOFR_Futures": "B5", "Fwd_Interp": "B9"}

wb = load_workbook(WB)
touched = []
for sheet, (hrow, r0, r1, cols) in LAYOUT.items():
    ws = wb[sheet]
    for col, (w, fmt) in cols.items():
        ws.column_dimensions[col].width = w
        if fmt:
            for r in range(r0, r1 + 1):
                cell = ws[f"{col}{r}"]
                if isinstance(cell.value, (str, type(None))) and not str(cell.value or "").startswith("="):
                    continue
                try: cell.number_format = fmt
                except AttributeError: pass
        c = ws[f"{col}{hrow}"]
        try:
            if c.value is not None:
                c.font = HDR; c.fill = HFILL
                c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        except AttributeError:
            pass
    ws.row_dimensions[hrow].height = 30
    ws.freeze_panes = f"B{r0}"
    touched.append(sheet)

for sheet, anchor in FREEZE_ONLY.items():
    if sheet in wb.sheetnames:
        wb[sheet].freeze_panes = anchor
        touched.append(sheet)

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("laid out:", len(touched), "sheets;", ", ".join(sorted(set(touched))))
