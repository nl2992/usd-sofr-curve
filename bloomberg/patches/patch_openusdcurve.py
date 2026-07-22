"""
Patch USD_SOFR_Curve_Bloomberg_Pricer.xlsx in place:

  1. SOFR_OIS_Quotes  — add a Manual quote column so the OIS strip (and therefore
     the whole discount curve) computes with no Bloomberg terminal attached.
     Mirrors the Manual-spread pattern already used on CDS_Quotes.
  2. SOFR_Fixings     — same fallback for the overnight fixing that sets DFspot.
  3. Hazard_Solver    — NEW sheet. Solves each piecewise-constant hazard by
     in-cell bisection, so Hazard_Bootstrap fits exactly with no Goal Seek.
  4. Hazard_Bootstrap — column D now reads the solver, via a mode switch.
  5. CDS_Parameters   — Hazard mode dropdown (bootstrap vs flat) + flat rate.

The solver reads only the hazard-INDEPENDENT columns of CDS_Schedule
(alpha, dt, DF_end, DF_mid), so there is no circular reference: it computes its
own survival for each trial hazard, and CDS_Schedule consumes the answer exactly
as it does today.
"""

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"

FONT = "Calibri"
BLUE = Font(name=FONT, size=11, color="0000FF")
BLACK = Font(name=FONT, size=11)
GREEN = Font(name=FONT, size=11, color="008000")
SECT = Font(name=FONT, size=11, bold=True)
HDRF = Font(name=FONT, size=11, bold=True, color="FFFFFF")
NOTE = Font(name=FONT, size=9, italic=True, color="666666")
WARN = Font(name=FONT, size=10, bold=True, color="C00000")
YFILL = PatternFill("solid", fgColor="FFFF00")
HFILL = PatternFill("solid", fgColor="1F3864")
SFILL = PatternFill("solid", fgColor="D9E1F2")
OFILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
N8, N6 = "0.00000000", "0.000000"

# --- placeholder USD SOFR OIS strip -----------------------------------------
# NOT market data. A plausible Jul-2026 shape so the workbook computes offline.
# Overwrite with real quotes, or let BDP take over on a terminal.
OIS = {
    "1W": 4.05, "2W": 4.05, "3W": 4.04, "1M": 4.03, "2M": 4.00, "3M": 3.97,
    "4M": 3.94, "5M": 3.91, "6M": 3.88, "7M": 3.86, "8M": 3.84, "9M": 3.82,
    "10M": 3.80, "11M": 3.79, "1Y": 3.78, "18M": 3.72, "2Y": 3.70, "3Y": 3.71,
    "4Y": 3.74, "5Y": 3.78, "6Y": 3.82, "7Y": 3.86, "8Y": 3.90, "9Y": 3.93,
    "10Y": 3.96, "12Y": 4.01, "15Y": 4.05, "20Y": 4.06, "25Y": 4.02,
    "30Y": 3.97, "40Y": 3.85, "50Y": 3.75,
}
SOFR_ON = 4.06          # overnight SOFR fixing placeholder

# Hazard_Bootstrap tenors and the CDS_Schedule period each maturity lands on.
# Their convention: tenor n matures on the (4n)-th quarterly pay date.
TENORS = [1, 2, 3, 5, 7, 10]
NITER = 30
NSLOT = 12              # longest segment is 7Y->10Y = 12 quarters
SCH_ROW0 = 6            # CDS_Schedule period p sits on row SCH_ROW0 + p


def put(ws, cell, val, font=BLACK, fmt=None, fill=None, border=False, align=None):
    c = ws[cell]
    c.value = val
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = BOX
    if align:
        c.alignment = Alignment(horizontal=align)
    return c


def band(ws, row, c0, c1, text):
    for c in range(c0, c1 + 1):
        ws.cell(row=row, column=c).fill = HFILL
    cc = ws.cell(row=row, column=c0)
    cc.value = text
    cc.font = HDRF


# ============================================================ 1. OIS QUOTES ===
def patch_ois(wb):
    ws = wb["SOFR_OIS_Quotes"]
    put(ws, "J4", "Manual quote (%)", SECT, None, SFILL, True)
    put(ws, "K4", "Source", SECT, None, SFILL, True)
    n = 0
    for r in range(5, 40):
        tenor = ws[f"B{r}"].value
        if tenor not in OIS:
            continue
        put(ws, f"J{r}", OIS[tenor], BLUE, "0.0000", YFILL, True)
        put(ws, f"K{r}", "PLACEHOLDER", NOTE)
        # BDP mid -> BDP last -> manual
        put(ws, f"H{r}", f"=IFERROR((E{r}+F{r})/2,IFERROR(G{r}+0,J{r}))",
            BLACK, "0.0000", None, True)
        n += 1
    put(ws, "J1", "MANUAL FALLBACK — PLACEHOLDER RATES, NOT MARKET DATA", WARN)
    put(ws, "J2", "Column J is used only when BDP fails. These are a plausible "
                  "Jul-2026 SOFR OIS shape so the curve computes off-terminal — "
                  "they are invented, not observed. Overwrite with real quotes "
                  "before using any output for anything real. On a terminal, BDP "
                  "takes precedence automatically and column J is ignored.", NOTE)
    return n


def patch_fixing(wb):
    ws = wb["SOFR_Fixings"]
    cur = ws["C6"].value
    if "BDP" not in str(cur):
        return False
    put(ws, "E5", "Manual o/n (%)", SECT, None, SFILL, True)
    put(ws, "E6", SOFR_ON, BLUE, "0.0000", YFILL, True)
    put(ws, "F6", "PLACEHOLDER — used only if BDP fails. Sets DFspot.", NOTE)
    put(ws, "C6", '=IFERROR(BDP(SOFR_TKR,"PX_LAST")+0,E6)', BLACK, "0.0000",
        None, True)
    return True


# ========================================================= 3. HAZARD SOLVER ===
def build_solver(wb):
    if "Hazard_Solver" in wb.sheetnames:
        del wb["Hazard_Solver"]
    ws = wb.create_sheet("Hazard_Solver")
    ws.sheet_properties.tabColor = "1565C0"
    ws.column_dimensions["A"].width = 18
    for c in range(2, 21):
        ws.column_dimensions[CL(c)].width = 11

    c_q0 = 6                       # F
    c_qN = c_q0 + NSLOT - 1        # Q
    c_prot, c_rpv, c_f = c_qN + 1, c_qN + 2, c_qN + 3

    put(ws, "A1", "Hazard_Solver", SECT)
    put(ws, "A2", "Solves each piecewise-constant hazard by bisection, in cells. "
                  "Replaces the Goal Seek step: Hazard_Bootstrap fits exactly and "
                  "updates live when spreads, recovery or the SOFR curve move.", NOTE)
    put(ws, "A3", "Reads only the hazard-INDEPENDENT columns of CDS_Schedule "
                  "(D alpha, E dt, G DF_end, H DF_mid), so there is no circular "
                  "reference. Same leg conventions as CDS_Schedule M/N/O.", NOTE)
    put(ws, "A4", f"Bracket [0, 3], {NITER} bisection steps -> resolution ~3e-9.", NOTE)

    R = "CDS_Parameters!$B$8"
    row = 6
    prev = {"prot": "0", "rpv": "0", "q": "1"}
    sol_rows = []

    for k, T in enumerate(TENORS):
        p0 = 1 if k == 0 else 4 * TENORS[k - 1] + 1
        p1 = 4 * T
        hb = 7 + k                                  # Hazard_Bootstrap row

        band(ws, row, 1, c_f, f"TENOR {T}Y   —   periods {p0}-{p1}   —   solve lambda{k+1}")
        h1 = row + 1
        put(ws, f"A{h1}", "market S (bp)", SECT)
        put(ws, f"B{h1}", f"=Hazard_Bootstrap!$C${hb}", GREEN, "0.0000", None, True)
        put(ws, f"C{h1}", "S (dec)", SECT)
        put(ws, f"D{h1}", f"=B{h1}/10000", BLACK, N8, None, True)
        put(ws, f"E{h1}", "recovery R", SECT)
        put(ws, f"F{h1}", f"={R}", GREEN, "0.00%", None, True)
        rS, rR = f"$D${h1}", f"$F${h1}"

        h2 = row + 2
        put(ws, f"A{h2}", "Q(start)", SECT)
        put(ws, f"B{h2}", f"={prev['q']}", BLACK, N8, None, True)
        put(ws, f"C{h2}", "ProtFixed", SECT)
        put(ws, f"D{h2}", f"={prev['prot']}", BLACK, N8, None, True)
        put(ws, f"E{h2}", "RPV01Fixed", SECT)
        put(ws, f"F{h2}", f"={prev['rpv']}", BLACK, N8, None, True)
        rQ0, rPf, rRf = f"$B${h2}", f"$D${h2}", f"$F${h2}"

        # ---- slot headers, pulled straight off CDS_Schedule
        r_per, r_alp, r_dt, r_dfe, r_dfm, r_cum = (row + 3, row + 4, row + 5,
                                                   row + 6, row + 7, row + 8)
        for lr, lab in zip([r_per, r_alp, r_dt, r_dfe, r_dfm, r_cum],
                           ["period", "alpha 360", "dt 365", "DF(end)", "DF(mid)",
                            "cum dt in seg"]):
            put(ws, f"A{lr}", lab, SECT)

        for s in range(NSLOT):
            col = CL(c_q0 + s)
            p = p0 + s
            live = p <= p1
            srow = SCH_ROW0 + p
            put(ws, f"{col}{r_per}", p if live else "", BLACK, None, None, True, "center")
            for lr, sc in ((r_alp, "D"), (r_dt, "E"), (r_dfe, "G"), (r_dfm, "H")):
                put(ws, f"{col}{lr}",
                    f"=CDS_Schedule!${sc}${srow}" if live else 0,
                    GREEN if live else BLACK, N6, None, True)
            prevcum = "0" if s == 0 else f"{CL(c_q0+s-1)}{r_cum}"
            put(ws, f"{col}{r_cum}", f"={prevcum}+{col}{r_dt}", BLACK, N6, None, True)

        rng_alp = f"${CL(c_q0)}${r_alp}:${CL(c_qN)}${r_alp}"
        rng_dfe = f"${CL(c_q0)}${r_dfe}:${CL(c_qN)}${r_dfe}"
        rng_dfm = f"${CL(c_q0)}${r_dfm}:${CL(c_qN)}${r_dfm}"

        # ---- bisection
        ith = row + 9
        for c, lab in ((2, "lo"), (3, "hi"), (4, "lambda trial"), (5, "Q(start)"),
                       (c_prot, "ProtSeg"), (c_rpv, "RPV01Seg"), (c_f, "f(lambda)")):
            put(ws, f"{CL(c)}{ith}", lab, SECT, None, SFILL, True)
        put(ws, f"A{ith}", "bisection", SECT)

        it0 = ith + 1
        for it in range(NITER):
            r = it0 + it
            if it == 0:
                put(ws, f"B{r}", 0.0, BLACK, N8, None, True)
                put(ws, f"C{r}", 3.0, BLACK, N8, None, True)
            else:
                p = r - 1
                put(ws, f"B{r}", f"=IF({CL(c_f)}{p}<0,D{p},B{p})", BLACK, N8, None, True)
                put(ws, f"C{r}", f"=IF({CL(c_f)}{p}<0,C{p},D{p})", BLACK, N8, None, True)
            put(ws, f"D{r}", f"=(B{r}+C{r})/2", BLACK, N8, None, True)
            put(ws, f"E{r}", f"={rQ0}", BLACK, N8, None, True)
            for s in range(NSLOT):
                col = CL(c_q0 + s)
                p = p0 + s
                prevQ = "E" + str(r) if s == 0 else f"{CL(c_q0+s-1)}{r}"
                put(ws, f"{col}{r}",
                    f"={rQ0}*EXP(-$D{r}*{col}${r_cum})" if p <= p1 else f"={prevQ}",
                    BLACK, N8, None, True)
            rq = f"${CL(c_q0)}{r}:${CL(c_qN)}{r}"
            rqp = f"$E{r}:${CL(c_qN-1)}{r}"
            put(ws, f"{CL(c_prot)}{r}", f"=SUMPRODUCT({rng_dfm},{rqp}-{rq})",
                BLACK, N8, None, True)
            put(ws, f"{CL(c_rpv)}{r}",
                f"=SUMPRODUCT({rng_alp},{rng_dfe},{rq})"
                f"+0.5*SUMPRODUCT({rng_alp},{rng_dfm},{rqp}-{rq})",
                BLACK, N8, None, True)
            put(ws, f"{CL(c_f)}{r}",
                f"=(1-{rR})*({rPf}+{CL(c_prot)}{r})-{rS}*({rRf}+{CL(c_rpv)}{r})",
                BLACK, N8, None, True)

        rl = it0 + NITER - 1
        rs = rl + 1
        put(ws, f"A{rs}", f"lambda{k+1} SOLVED", SECT)
        put(ws, f"B{rs}", f"=D{rl}", BLACK, N8, OFILL, True)
        put(ws, f"C{rs}", "resid", SECT)
        put(ws, f"D{rs}", f"={CL(c_f)}{rl}", BLACK, "0.00E+00", None, True)
        put(ws, f"E{rs}", "Q(T_k)", SECT)
        put(ws, f"F{rs}", f"={CL(c_qN)}{rl}", BLACK, N8, None, True)
        put(ws, f"G{rs}", "ProtCum", SECT)
        put(ws, f"H{rs}", f"={rPf}+{CL(c_prot)}{rl}", BLACK, N8, None, True)
        put(ws, f"I{rs}", "RPV01Cum", SECT)
        put(ws, f"J{rs}", f"={rRf}+{CL(c_rpv)}{rl}", BLACK, N8, None, True)
        put(ws, f"K{rs}", "model S (bp)", SECT)
        put(ws, f"L{rs}", f"=IFERROR((1-{rR})*H{rs}/J{rs}*10000,0)",
            BLACK, "0.0000", None, True)

        sol_rows.append(rs)
        prev = {"prot": f"$H${rs}", "rpv": f"$J${rs}", "q": f"$F${rs}"}
        row = rs + 3

    return sol_rows


# ==================================================== 4/5. WIRE UP THE MODE ===
def patch_params(wb):
    ws = wb["CDS_Parameters"]
    put(ws, "A17", "Hazard mode", SECT)
    put(ws, "B17", "Bootstrap from CDS curve", BLUE, None, YFILL, True)
    put(ws, "C17", 'Dropdown. "Bootstrap from CDS curve" fits the term structure to '
                   'CDS_Quotes; "Flat hazard rate" overrides every tenor with B18.', NOTE)
    dv = DataValidation(type="list",
                        formula1='"Bootstrap from CDS curve,Flat hazard rate"',
                        allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B17"])
    put(ws, "A18", "Flat hazard rate", SECT)
    put(ws, "B18", 0.015, BLUE, N6, YFILL, True)
    put(ws, "C18", "Continuously-compounded intensity, used at every tenor when "
                   "B17 = Flat hazard rate. Credit-triangle guide: s/(1-R).", NOTE)


def patch_hazard(wb, sol_rows):
    ws = wb["Hazard_Bootstrap"]
    put(ws, "A2", "Piecewise-constant hazard curve. Column D is solved live by "
                  "Hazard_Solver (bisection) — no Goal Seek needed. Set "
                  "CDS_Parameters!B17 to switch to a flat hazard rate.", NOTE)
    put(ws, "A4", "Bootstrap maturity by maturity: hold earlier hazards fixed, solve "
                  "the current segment so model spread = market. Column D now reads "
                  "the in-cell bisection on Hazard_Solver, so column J (repricing "
                  "error) is ~0 automatically and stays there when inputs change.", NOTE)
    for k, rs in enumerate(sol_rows):
        r = 7 + k
        put(ws, f"D{r}",
            f'=IF(CDS_Parameters!$B$17="Flat hazard rate",CDS_Parameters!$B$18,'
            f"Hazard_Solver!$B${rs})", BLACK, N8, OFILL, True)
    put(ws, "L6", "Solver resid", SECT)
    for k, rs in enumerate(sol_rows):
        put(ws, f"L{7+k}", f"=Hazard_Solver!$D${rs}", BLACK, "0.00E+00", None, True)


def main():
    shutil.copy(WB, WB.replace(".xlsx", "_PRE_PATCH_BACKUP.xlsx"))
    wb = load_workbook(WB)
    n = patch_ois(wb)
    fx = patch_fixing(wb)
    sol = build_solver(wb)
    patch_params(wb)
    patch_hazard(wb, sol)
    wb.calculation.fullCalcOnLoad = True
    wb.save(WB)
    print(f"patched: {n} OIS manual quotes, fixing_fallback={fx}, "
          f"solver blocks at rows {sol}")


if __name__ == "__main__":
    main()
