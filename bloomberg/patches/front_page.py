"""
Steps becomes the front page: put a ticker in, read the CDSW screen out.

Everything was already connected, but the inputs were scattered - ticker on
Entities, notional and coupon on CDS_Parameters, maturity on CDS_Pricer. This
puts all of them in one block and points the downstream cells at it, so the
sheets below become working rather than places you have to visit.

The ticker drives the two-step pull directly (entity -> CDS ticker -> spread), so
on a terminal a ticker alone is enough. Blank falls back to the Entities table,
which is what you need off-terminal or for a name with manual spreads.

Solver is Brent by default, one call per tenor, with the in-cell ladder as
fallback.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

WB="bloomberg/CDS_Pricer.xlsx"
F="Calibri"
H1=Font(name=F,size=15,bold=True); H2=Font(name=F,size=11,bold=True)
N=Font(name=F,size=10); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
IN=PatternFill("solid",fgColor="FFFF00"); BLUE=Font(name=F,size=11,color="0000FF")
OUT=PatternFill("solid",fgColor="E2EFDA"); GRN=Font(name=F,size=11,color="008000")
RED=Font(name=F,size=10,bold=True,color="C00000"); BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(WB); st=wb["Steps"]
old=[(r,[st.cell(r,c).value for c in range(1,5)]) for r in range(1,60)]
for row in st.iter_rows(min_row=1,max_row=60,max_col=6):
    for c in row: c.value=None; c.fill=PatternFill(); c.border=Border()

st["A1"]="CDS Pricer"; st["A1"].font=H1
st["A2"]="Put a ticker in the yellow box. Everything below follows: spreads pull, hazard strips, CDSW figures come out."
st["A2"].font=SM

st["A4"]="INPUTS"; st["A4"].font=H2
INP=[("Reference entity ticker","HSBA LN Equity","BDP resolves the CDS tickers from this","@"),
     ("Notional",10000000,"","#,##0"),
     ("Standard coupon (bp)",100,"100 or 500","0"),
     ("Recovery R",0.4,"model parameter, conventionally 0.40","0.00"),
     ("Maturity (tenor yrs)",5,"","0"),
     ("Direction","Buy protection","","@"),
     ("Valuation date",None,"leave blank to keep the curve's date","mm/dd/yyyy")]
r=5
for lab,val,note,fmt in INP:
    st.cell(r,1,lab).font=N
    c=st.cell(r,2,val); c.fill=IN; c.font=BLUE; c.border=BOX; c.number_format=fmt
    st.cell(r,3,note).font=SM
    r+=1
dv=DataValidation(type="list",formula1='"Buy protection,Sell protection"',allow_blank=False)
st.add_data_validation(dv); dv.add(st["B10"])
st.cell(12,1,"Hazard solver").font=N
st.cell(12,2,"=CDS_Parameters!$B$30").font=GRN; st.cell(12,2).border=BOX
st.cell(12,3,"Brent, one call per tenor. Change at CDS_Parameters!B30.").font=SM

st["A14"]="STATUS"; st["A14"].font=H2
STAT=[("Curve link", '=IF(COUNT(Curve_Interface!$L$8:$L$73)=0,"curve workbook not open",'
                     'COUNT(Curve_Interface!$L$8:$L$73)&" pillars")'),
      ("Spreads",   '=IF(COUNTIF(CDS_Quotes!$H$7:$H$12,"BDP live")>0,'
                    'COUNTIF(CDS_Quotes!$H$7:$H$12,"BDP live")&" of 6 live","manual / demo")'),
      ("Strip",     '=IF(MAX(ABS(Hazard_Bootstrap!$J$7:$J$12))<0.01,"reprices, max err "&'
                    'TEXT(MAX(ABS(Hazard_Bootstrap!$J$7:$J$12)),"0.00E+00")&" bp",'
                    '"STRIP FAILED - check the quotes, see Model_Notes p.9")')]
r=15
for lab,f_ in STAT:
    st.cell(r,1,lab).font=N
    c=st.cell(r,2,f_); c.font=GRN; c.border=BOX
    r+=1

st["A19"]="CDSW SCREEN"; st["A19"].font=H2
st["C19"]="mirrors CDS_Pricer rows 15-37"; st["C19"].font=SM
OUTS=[("Par spread (bp)","=CDS_Pricer!$B$15","0.0000"),
      ("Points upfront","=CDS_Pricer!$B$31","0.0000"),
      ("Price","=CDS_Pricer!$B$32","0.0000"),
      ("Principal","=CDS_Pricer!$B$33","#,##0"),
      ("Accrued days","=CDS_Pricer!$B$34","0"),
      ("Accrued","=CDS_Pricer!$B$35","#,##0"),
      ("Cash amount","=CDS_Pricer!$B$36","#,##0"),
      ("Default exposure","=CDS_Pricer!$B$37","#,##0"),
      ("","",""),
      ("Spread DV01 / CS01","=CDS_Pricer!$B$22","#,##0"),
      ("IR DV01","=CDS_Pricer!$B$23","#,##0"),
      ("Rec risk 1%","=CDS_Pricer!$B$24","#,##0"),
      ("Jump to default","=CDS_Pricer!$B$25","#,##0")]
r=20
for lab,f_,fmt in OUTS:
    if lab:
        st.cell(r,1,lab).font=N
        c=st.cell(r,2,f_); c.fill=OUT; c.font=GRN; c.border=BOX; c.number_format=fmt
    r+=1

st.cell(34,1,"PIPELINE").font=H2
for i,h in enumerate(["Step","Sheet","What happens","What to check"],start=1):
    c=st.cell(35,i,h); c.font=HD; c.fill=HF; c.border=BOX
    c.alignment=Alignment(horizontal="left")
r=36
for s,sh,what,chk in [x[1] for x in old if x[1][0] and str(x[1][0]).isdigit()]:
    st.cell(r,1,s).font=H2; st.cell(r,2,sh).font=H2
    st.cell(r,3,what).font=N; st.cell(r,4,chk).font=N
    for k in range(1,5): st.cell(r,k).border=BOX
    r+=1
r+=1
st.cell(r,1,"Reference tabs").font=H2; r+=1
for sh,what in [("Model_Notes","B-Model equations (3.1)-(3.6), conventions, strippability check."),
                ("Brent_vs_Bisection","The two solvers side by side on the same objective."),
                ("Root_Methods","Eight root-finders from MATH5030 M2."),
                ("Hazard_Solver","In-cell bisection ladder, the fallback solver."),
                ("CDS_Entities","Data store behind Entities, plus the CDSW capture block.")]:
    st.cell(r,2,sh).font=H2; st.cell(r,3,what).font=N; r+=1
r+=1
st.cell(r,2,"Needs CDSBrent.bas and CDSRootFinders.bas imported, saved as .xlsm. Without them the strip falls back to the ladder and still prices.").font=SM
r+=1
st.cell(r,2,"Spreads are demo values. The SOFR curve is real market data; the credit curve is not.").font=RED

for col,w in zip("ABCD",(26,22,62,50)): st.column_dimensions[col].width=w
st.freeze_panes="A4"

# ---- point the downstream inputs at the front page --------------------------
p=wb["CDS_Parameters"]
p["B19"]='=IF(Steps!$B$5="",IF(Entities!$F$4="","",Entities!$F$4),Steps!$B$5)'
p["B9"] ='=IF(Steps!$B$6="",10000000,Steps!$B$6)'
p["B10"]='=IF(Steps!$B$7="",100,Steps!$B$7)'
p["B8"] ='=IF(Steps!$B$8="",0.4,Steps!$B$8)'
p["B15"]='=IF(Steps!$B$10="","Buy protection",Steps!$B$10)'
wb["CDS_Pricer"]["B7"]='=IF(Steps!$B$9="",5,Steps!$B$9)'
for a in ("B19","B9","B10","B8","B15"):
    p[a].font=GRN
print("front page wired: ticker, notional, coupon, recovery, tenor, direction")

wb.calculation.fullCalcOnLoad=True
wb.save(WB)
