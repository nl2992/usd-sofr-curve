"""
Repoint SOFR_OIS_Quotes!H at the new Testing layout.

The three-block rebuild moved every case: rates were Testing!C/G/K rows 7-38 in
the old single-block layout, and are now column B of each block at rows 8-47,
58-97 and 108-147. H was still reading the old cells, so a test case fed the
curve its BBG ZERO column instead of its swap rates.

Matching is now on the tenor label rather than row position, so a case with a
different pillar set cannot feed the wrong rate to a tenor.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
MODE = "Bootstrap!$G$4"
FIXED = "Fixed (S490 07/21/26)"
T = ["Test 1", "Test 2", "Test 3"]
BLK = [(8, 47), (58, 97), (108, 147)]

wb = load_workbook(WB)
q = wb["SOFR_OIS_Quotes"]
n = 0
for r in range(5, 40):
    if q[f"A{r}"].value is None or q[f"B{r}"].value is None:
        continue
    legs = []
    for i, (a, b) in enumerate(BLK):
        ten = f"Testing!$A${a}:$A${b}"
        rate = f"Testing!$B${a}:$B${b}"
        legs.append(f'IF({MODE}="{T[i]}",INDEX({rate},MATCH(B{r},{ten},0)),')
    live = f'IF(ISNUMBER(T{r}),T{r},J{r})'
    inner = f'IF({MODE}="{FIXED}",J{r},{live})'
    f = "=IFERROR(" + "".join(legs) + inner + ")))" + f",J{r})"
    c = q[f"H{r}"]
    c.value = f
    c.font = BLACK
    c.number_format = "0.00000"
    c.fill = OF
    c.border = BOX
    n += 1
q["A20"] = ("Col H is the mid in use. Test cases are matched on TENOR against the Testing "
            "blocks, not row position, so a case with a different pillar set cannot feed the "
            "wrong rate to a tenor.")
q["A20"].font = NOTE
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"repointed {n} quote rows at the three Testing blocks, matched on tenor")
