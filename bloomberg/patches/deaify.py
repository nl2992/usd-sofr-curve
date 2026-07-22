"""
Cut the prose back to desk-note register.

Tells being removed: em-dashes, "worth knowing / reading it honestly" framing,
paragraph-length cell notes, and headers that editorialise. Facts stay; the
commentary around them goes. A note earns its place by saying something you
cannot read off the formula.
"""
import re
from openpyxl import load_workbook

REPL = {
 "CDS_Pricer": {
  "A2": "Prices a CDS off the SOFR discount curve and the stripped hazard curve. Equations on Model_Notes.",
 },
 "Root_Methods": {
  "A2": "All methods solve the same f(h) = (1-R)*Prot(h) - S*(RPV01(h) - AI), B-Model (3.3). Same root; different cost.",
  "A3": "Baseline is the scheme, not the algorithm. Section 4 p.8 fixes piecewise-constant hazard solved maturity by maturity, each a 1-D root-find with earlier hazards fixed. It names no method. Hazard_Solver here is bisection.",
  "A4": "Needs CDSBrent.bas and CDSRootFinders.bas imported, saved as .xlsm.",
 },
}
NOTE_REPL = [
 ("All eight land on the same root to ~1e-16. On this problem the method is a cost question, not an accuracy one.",
  "All eight agree to ~1e-16. Method is a cost choice here, not an accuracy one."),
 ("Bisection needs 52 steps where Newton needs 6 - the price of never using the shape of f.",
  "Bisection: 52 steps. Newton: 6. That is the cost of ignoring the shape of f."),
 ("Halley and Householder d=3 do NOT beat Newton here. f is close to linear in h near the root, so the extra",
  "Halley and Householder d=3 do not beat Newton here: 6 iterations each. f is near-linear in h at the root,"),
 ("    order buys nothing before the tolerance is reached. Higher order is not free and not always faster.",
  "    so the extra order is spent before tolerance is reached."),
 ("Ridders is fastest at 5, but every open method (secant, Newton, Halley) can leave the domain on a bad start;",
  "Ridders is fastest at 5. But secant, Newton and Halley can all leave the domain on a bad start;"),
 ("    the bracketing ones cannot. On a strip that must never return h < 0, that guarantee is worth more than speed.",
  "    bracketing methods cannot. The strip must never return h < 0 (p.8), so the guarantee outranks the speed."),
 ("If no method converges, suspect the quote, not the solver - see the p.9 strippability note on Model_Notes.",
  "If nothing converges, suspect the quote. See the p.9 strippability check on Model_Notes."),
 ("Reading it honestly","Results"),
 ("Not covered by any of this","Limits"),
 ("The strip has no external validation — Bloomberg exports no hazard curve (Help Desk H#1330731572).",
  "No external validation of the strip: Bloomberg exports no hazard curve (Help Desk H#1330731572)."),
 ("Spreads in this workbook are demo values, not market. The SOFR curve is real; the credit curve is not.",
  "Spreads here are demo values. The SOFR curve is real market data; the credit curve is not."),
 ("A fail here is the p.9 condition, and it is a warning about the quotes, not about the solver.",
  "A fail is the p.9 condition: a warning about the quotes, not the solver."),
 ("The bound is approximate; it holds while cumulative default probability is low.",
  "Bound is approximate, valid while cumulative default probability is low."),
 ("Set On to 0 to take a name out. Overrides win over the Bloomberg pull; leave blank to use the pull.",
  "On = 0 takes a name out. Overrides beat the BDP pull; blank uses the pull."),
 ("Discount curve is the pasted snapshot. Hazard and survival come out of the strip.",
  "Discount curve is the pasted snapshot. Hazard and survival come from the strip."),
 ("Paste YELLOW A:E straight off the capture — tenor, date, swap mid, BBG zero, BBG discount. Set the curve date. Green is computed here; this sheet points at nothing outside itself.",
  "Paste A:E off the capture: tenor, date, swap mid, BBG zero, BBG discount. Set the curve date. Green is computed on this sheet."),
 ("Self-contained. Nothing on this sheet points at another sheet, so it survives Move or Copy into any workbook.",
  "Self-contained. No reference leaves this sheet, so Move or Copy cannot break it."),
]
LONG_CUT = {
 "CDS_Schedule": {"A4": "One row per quarterly coupon period. DF(end) and DF(mid) interpolate the SOFR curve; hazard is the piecewise-constant lambda for the segment, by pay date. Q(end) = Q(prev)*exp(-hazard*dt). See Model_Notes."},
 "CDS_Quotes":   {"A4": "Enter each tenor's CDS ticker, or let col D resolve it. Two-step pull per Help Desk H#1330731572, documented in K."},
}

n_em=n_txt=0
for wbn in ("bloomberg/CDS_Pricer.xlsx","bloomberg/Bootstrap_Check.xlsx"):
    wb=load_workbook(wbn)
    for sh,cells in list(REPL.items())+list(LONG_CUT.items()):
        if sh in wb.sheetnames:
            for a,v in cells.items():
                if wb[sh][a].value is not None:
                    wb[sh][a]=v; n_txt+=1
    for ws in wb:
        for row in ws.iter_rows():
            for c in row:
                v=c.value
                if isinstance(v,str) and not v.startswith("="):
                    new=v
                    for a,b in NOTE_REPL:
                        if a in new: new=new.replace(a,b); n_txt+=1
                    if "—" in new:
                        # em-dash as a clause break becomes a colon or full stop
                        new=re.sub(r"\s+—\s+", ": ", new, count=1)
                        new=new.replace("—","-")
                        n_em+=1
                    if new!=v: c.value=new
    wb.save(wbn)
print(f"{n_txt} notes rewritten, {n_em} cells cleared of em-dashes")

for wbn in ("bloomberg/CDS_Pricer.xlsx","bloomberg/Bootstrap_Check.xlsx"):
    wb=load_workbook(wbn); em=0; long=0
    for ws in wb:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value,str) and not c.value.startswith("="):
                    em+=c.value.count("—")
                    if len(c.value)>110: long+=1
    print(f"  {wbn.split('/')[-1]:<24} em-dashes {em}, cells>110ch {long}")
