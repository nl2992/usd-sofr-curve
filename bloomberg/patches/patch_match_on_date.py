"""
Match test-case rates on the DERIVED DATE, not the tenor label.

The capture writes 12M where SOFR_OIS_Quotes writes 1Y. Matching on the label
missed that pillar silently. Both derive to the same maturity, so matching on
date is label-agnostic and also survives 1YR / 12MO / 2YR and similar variants.

SOFR_OIS_Quotes!C is that row's maturity; Testing column F is the block's derived
maturity. Match one against the other.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
MODE = "Bootstrap!$G$4"
FIXED = "Fixed (S490 07/21/26)"
T = ["Test 1", "Test 2", "Test 3"]
BLK = [(8, 47), (58, 97), (108, 147)]

wb = load_workbook(WB)
q = wb["SOFR_OIS_Quotes"]
n = 0
for r in range(5, 40):
    if q[f"A{r}"].value is None or q[f"B{r}"].value is None:
        continue
    legs = []
    for i, (a, b) in enumerate(BLK):
        dates = f"Testing!$F${a}:$F${b}"
        rate = f"Testing!$B${a}:$B${b}"
        legs.append(f'IF({MODE}="{T[i]}",INDEX({rate},MATCH(C{r},{dates},0)),')
    live = f"IF(ISNUMBER(T{r}),T{r},J{r})"
    inner = f'IF({MODE}="{FIXED}",J{r},{live})'
    c = q[f"H{r}"]
    c.value = "=IFERROR(" + "".join(legs) + inner + ")))" + f",J{r})"
    c.font = BLACK; c.number_format = "0.00000"; c.fill = OF; c.border = BOX
    n += 1
q["A20"] = ("Col H is the mid in use. Test cases match on the derived MATURITY DATE, not the "
            "tenor label - a capture writing 12M where this sheet writes 1Y still lines up.")
q["A20"].font = NOTE
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"repointed {n} rows to match on date")
