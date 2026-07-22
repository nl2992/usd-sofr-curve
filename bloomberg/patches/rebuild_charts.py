"""
Rebuild the Curves charts: numeric axes, labelled, no crowding.

The SOFR zero and DF charts were LineCharts over 65 pillars, so the category axis
carried 65 date labels and was unreadable. Both are now scatter against T in
years - one axis, ticks every 5 years, no labels to crowd.

Hazard gets a genuine step series. It is piecewise constant by construction
(p.8), so a smooth line through six points misrepresents the model. The step
block below duplicates each segment's endpoints to draw flat-then-jump.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties

WB="bloomberg/CDS_Pricer.xlsx"
F="Calibri"
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4); SM=Font(name=F,size=9,color="808080")
B=Font(name=F,size=11,bold=True)

SOFR0,SOFR1 = 6,70        # SOFR block data rows (65 pillars)
HZ0 = 75                  # hazard block first data row
SP0 = 85                  # spread block first data row
STEP0 = 100               # new step block

wb=load_workbook(WB); cv=wb["Curves"]
cv._charts=[]

# ---- step coordinates for the piecewise-constant hazard ----------------------
cv.cell(STEP0-2,1,"Hazard step coordinates (drawing aid)").font=B
cv.cell(STEP0-2,3,"piecewise constant per p.8, so it is drawn flat-then-jump").font=SM
for i,h in enumerate(["t","hazard %"],start=1):
    c=cv.cell(STEP0-1,i,h); c.font=HD; c.fill=HF; c.border=BOX
r=STEP0
for j in range(6):
    hz=HZ0+j; yr=f"$E${hz}"; hv=f"$B${hz}"
    prev="0" if j==0 else f"$E${HZ0+j-1}"
    cv.cell(r,1,f"={prev}").number_format="0.00"
    cv.cell(r,2,f"={hv}").number_format="0.0000"
    cv.cell(r+1,1,f"={yr}").number_format="0.00"
    cv.cell(r+1,2,f"={hv}").number_format="0.0000"
    r+=2
STEP1=r-1

def axis(ch,xt,yt,xmin=None,xmax=None,xunit=None,ynum="General",xnum="General"):
    ch.x_axis.title=xt; ch.y_axis.title=yt
    ch.x_axis.delete=False; ch.y_axis.delete=False
    ch.x_axis.majorTickMark="out"; ch.y_axis.majorTickMark="out"
    ch.x_axis.tickLblPos="nextTo"; ch.y_axis.tickLblPos="nextTo"
    ch.y_axis.majorGridlines=ChartLines()
    ch.x_axis.majorGridlines=ChartLines()
    ch.x_axis.number_format=xnum; ch.y_axis.number_format=ynum
    if xmin is not None: ch.x_axis.scaling.min=xmin
    if xmax is not None: ch.x_axis.scaling.max=xmax
    if xunit is not None: ch.x_axis.majorUnit=xunit
    ch.height,ch.width=8.5,17
    ch.legend.position="b"

def sc(title,xcol,ycols,r0,r1,anchor,xt,yt,**kw):
    ch=ScatterChart(); ch.title=title; ch.style=13
    xref=Reference(cv,min_col=xcol,min_row=r0,max_row=r1)
    for yc,mk in ycols:
        s=Series(Reference(cv,min_col=yc,min_row=r0-1,max_row=r1),xref,title_from_data=True)
        s.smooth=False
        if mk: s.marker=Marker(symbol="circle",size=6)
        ch.series.append(s)
    axis(ch,xt,yt,**kw)
    cv.add_chart(ch,anchor)
    return ch

# ---- SOFR: numeric axis, ticks every 5y, no 65-label crowding ----------------
sc("SOFR zero curve",2,[(4,False)],SOFR0,SOFR1,"H4",
   "T (years)","zero rate (%)",xmin=0,xmax=50,xunit=5,ynum="0.00")
sc("SOFR discount factors",2,[(3,False)],SOFR0,SOFR1,"H22",
   "T (years)","D(0,t)",xmin=0,xmax=50,xunit=5,ynum="0.00")
sc("SOFR zero curve - first 10 years",2,[(4,False)],SOFR0,SOFR1,"H40",
   "T (years)","zero rate (%)",xmin=0,xmax=10,xunit=1,ynum="0.00")

# ---- credit ------------------------------------------------------------------
sc("CDS term structure - market vs model",6,[(2,True),(3,True)],SP0,SP0+5,"H58",
   "tenor (years)","spread (bp)",xmin=0,xmax=10,xunit=1,ynum="0")
sc("CDSW view - credit curve vs traded spread",6,[(2,True),(3,False),(5,False)],SP0,SP0+5,"H76",
   "tenor (years)","spread (bp)",xmin=0,xmax=10,xunit=1,ynum="0")

h=ScatterChart(); h.title="Hazard rate (piecewise constant)"; h.style=13
s=Series(Reference(cv,min_col=2,min_row=STEP0,max_row=STEP1),
         Reference(cv,min_col=1,min_row=STEP0,max_row=STEP1))
s.smooth=False; h.series.append(s)
axis(h,"t (years)","hazard (%)",xmin=0,xmax=10,xunit=1,ynum="0.00")
h.legend=None
cv.add_chart(h,"H94")

sc("Survival Q(t)",5,[(3,True)],HZ0,HZ0+5,"H112",
   "tenor (years)","Q(t)",xmin=0,xmax=10,xunit=1,ynum="0.00")
sc("Default probability 1 - Q(t)",5,[(4,True)],HZ0,HZ0+5,"H130",
   "tenor (years)","1 - Q(t)",xmin=0,xmax=10,xunit=1,ynum="0.0%")

cv["A2"]=("Discount curve from Curve_Interface. Hazard and survival from the strip. "
          "All charts use a numeric time axis, so pillar spacing is to scale.")
cv["A2"].font=SM
wb.calculation.fullCalcOnLoad=True
wb.save(WB)
print(f"rebuilt: {len(cv._charts)} charts, step block at rows {STEP0}-{STEP1}")
