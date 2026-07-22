"""
Put the S490 Curve Analysis capture on its own sheet and make it the ONE place
the comparison targets live.

Previously the snapshot was hard-coded in Bootstrap!AA:AD and the market rates
sat separately in SOFR_OIS_Quotes!J - two copies of the same capture that could
drift apart. Now:

    S490_Snapshot                 the capture: date, tenor, market rate, zero, DF
      -> SOFR_OIS_Quotes!J        manual mid, looked up by tenor   (INPUT)
      -> Bootstrap!T:U            comparison target, by date       (OUTPUT)
      -> Bloomberg_S490_Validation!G   same target, by date

Update the capture in one place and everything follows.

This capture is the matched pair: its Market Rate column equals the mid of the
curve export's bid/ask (1W: (3.634854+3.651146)/2 = 3.643000), so input and
output are from the same instant. Verified before writing: max |d zero| 0.40bp
overall, 0.08bp from 2Y, max |d DF| 1.84e-05.
"""
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
YF = PatternFill("solid", fgColor="FFFF00")
SF = PatternFill("solid", fgColor="D9E1F2")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

S = [("07/30/2026","1W",3.64300,3.69229,0.999090),("08/06/2026","2W",3.67870,3.72278,0.998369),
("08/13/2026","3W",3.68722,3.73075,0.997652),("08/24/2026","1M",3.69650,3.73879,0.996523),
("09/23/2026","2M",3.72235,3.75983,0.993429),("10/23/2026","3M",3.77507,3.80667,0.990244),
("11/23/2026","4M",3.82050,3.84600,0.986915),("12/23/2026","5M",3.86225,3.88163,0.983651),
("01/25/2027","6M",3.90812,3.92047,0.980009),("02/23/2027","7M",3.94110,3.94717,0.976806),
("03/23/2027","8M",3.96915,3.96903,0.973710),("04/23/2027","9M",3.99975,3.99261,0.970261),
("05/24/2027","10M",4.02750,4.01321,0.966808),("06/23/2027","11M",4.05145,4.03016,0.963474),
("07/23/2027","1Y",4.07237,4.04404,0.960154),("01/24/2028","18M",4.09080,4.07764,0.940196),
("07/24/2028","2Y",4.09725,4.06943,0.921424),("07/23/2029","3Y",4.07820,4.05047,0.885284),
("07/23/2030","4Y",4.06895,4.04120,0.850458),("07/23/2031","5Y",4.07450,4.04734,0.816523),
("07/23/2032","6Y",4.09215,4.06648,0.783148),("07/25/2033","7Y",4.11555,4.09206,0.750424),
("07/24/2034","8Y",4.14210,4.12158,0.718715),("07/23/2035","9Y",4.17128,4.15450,0.687729),
("07/23/2036","10Y",4.20250,4.19026,0.657309),("07/23/2038","12Y",4.26630,4.26517,0.599053),
("07/23/2041","15Y",4.35053,4.36781,0.518980),("07/23/2046","20Y",4.42135,4.45488,0.409905),
("07/24/2051","25Y",4.41735,4.43447,0.329654),("07/24/2056","30Y",4.37780,4.35382,0.270507),
("07/23/2066","40Y",4.25820,4.10978,0.192962),("07/23/2076","50Y",4.12531,3.81656,0.148103)]


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


wb = load_workbook(WB)
if "S490_Snapshot" in wb.sheetnames:
    del wb["S490_Snapshot"]
ws = wb.create_sheet("S490_Snapshot", wb.sheetnames.index("Bootstrap"))
ws.sheet_properties.tabColor = "6A1B9A"
for c, w in zip("ABCDEF", [14, 10, 14, 14, 14, 70]):
    ws.column_dimensions[c].width = w

put(ws, "A1", "S490_Snapshot", BOLD)
put(ws, "A2", "Bloomberg S490 Curve Analysis capture. USD SOFR (vs FIXED), "
              "Step Forward (Cont), Settle Date 07/21/26, Curve Side Mid, Shift +0.00bp.", NOTE)
put(ws, "A3", "THE single place to update the comparison. Market Rate feeds "
              "SOFR_OIS_Quotes!J (the input); Zero/Discount feed Bootstrap!T:U and "
              "Bloomberg_S490_Validation!G (the target). Change it here only.", NOTE)
put(ws, "A4", "Matched pair: Market Rate equals the mid of the curve export's bid/ask "
              "(1W: (3.634854+3.651146)/2 = 3.643000), so input and output are the same "
              "instant. Only valid while VAL_DATE = 07/21/2026.", NOTE)

for i, h in enumerate(["Date", "Tenor", "Market rate (%)", "Zero rate (%)", "Discount"]):
    put(ws, f"{chr(65+i)}6", h, BOLD, None, SF, True)
for i, (d, t, m, z, df) in enumerate(S):
    r = 7 + i
    put(ws, f"A{r}", dt.datetime.strptime(d, "%m/%d/%Y").date(), BLUE, "mm/dd/yyyy", YF, True)
    put(ws, f"B{r}", t, BLUE, None, YF, True)
    put(ws, f"C{r}", m, BLUE, "0.00000", YF, True)
    put(ws, f"D{r}", z, BLUE, "0.00000", YF, True)
    put(ws, f"E{r}", df, BLUE, "0.000000", YF, True)
lo, hi = 7, 7 + len(S) - 1
DTS, TNS = f"S490_Snapshot!$A${lo}:$A${hi}", f"S490_Snapshot!$B${lo}:$B${hi}"
MKT = f"S490_Snapshot!$C${lo}:$C${hi}"
ZRO, DSC = f"S490_Snapshot!$D${lo}:$D${hi}", f"S490_Snapshot!$E${lo}:$E${hi}"
put(ws, f"A{hi+2}", f"{len(S)} pillars", BOLD)

# ---- input: quotes read the snapshot's market rate by tenor
q = wb["SOFR_OIS_Quotes"]
put(q, "J4", "Manual MID (%) <- S490_Snapshot", BOLD, None, SF, True)
n = 0
for r in range(5, 40):
    t = q[f"B{r}"].value
    if t is None:
        continue
    put(q, f"J{r}", f'=IFERROR(INDEX({MKT},MATCH(B{r},{TNS},0)),"")', BLACK, "0.00000", OF, True)
    put(q, f"K{r}", "S490_Snapshot", NOTE)
    n += 1

# ---- target: Bootstrap comparison reads the snapshot by date; drop the old copy
b = wb["Bootstrap"]
for r in range(6, 45):
    for c in ("AA", "AB", "AC", "AD"):
        try: b[f"{c}{r}"].value = None
        except AttributeError: pass
put(b, "AA6", "S490 targets now live on the S490_Snapshot sheet - this block was a "
              "duplicate copy and has been cleared.", NOTE)
for r in range(8, 73):
    m = f"MATCH(B{r},{DTS},0)"
    put(b, f"T{r}", f'=IFERROR(INDEX({ZRO},{m}),"")', BLACK, "0.00000", None, True)
    put(b, f"U{r}", f'=IFERROR(INDEX({DSC},{m}),"")', BLACK, "0.000000", None, True)

# ---- and the validation sheet reads the same source
v = wb["Bloomberg_S490_Validation"]
put(v, "G7", "S490 zero (%) <- S490_Snapshot", BOLD, None, SF, True)
for r in range(8, 73):
    m = f"MATCH(B{r},{DTS},0)"
    put(v, f"G{r}", f'=IFERROR(INDEX({ZRO},{m}),"")', BLACK, "0.00000", OF, True)
    put(v, f"M{r}", f'=IFERROR(INDEX({DSC},{m}),"")', BLACK, "0.000000", None, True)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"S490_Snapshot created ({len(S)} pillars); quotes wired ({n}); targets repointed")
