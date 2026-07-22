"""
Testing sheet: three curve test cases and three CDS test cases.

CURVE (Testing!A:M)
  Three slots. Per case you paste the swap curve Bloomberg was built from and the
  zero + discount curve it generated. Bootstrap!G4 gains Test 1/2/3, which routes
  those market rates through SOFR_OIS_Quotes!H into the one bootstrap. The active
  case then shows delta-zero and delta-DF per pillar with a max.

  Only the selected case compares - there is one bootstrap, by design.

CDS (CDS_Entities rows 5-7, summarised on Testing)
  Each row gains the Help Desk two-step pull per tenor, so typing a ticker
  populates the term structure live:
      step 1  BDP(ticker, "CDS_SPREAD_TICKER_nY")     -> CDS ticker
      step 2  BDP(that & " BEST Curncy", "PX_LAST")   -> spread
  Manual columns H:M still win if BDP is unavailable, so a case can be driven
  either way.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
H1 = Font(name="Calibri", size=12, bold=True)
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
HF = PatternFill("solid", fgColor="1F3864")
CF = ["2E7D32", "C00000", "6A1B9A"]
YF = PatternFill("solid", fgColor="FFFF00")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
DTF, N5, N6, BP = "mm/dd/yyyy", "0.00000", "0.000000", "0.00"

TENORS = ["1W","2W","3W","1M","2M","3M","4M","5M","6M","7M","8M","9M","10M","11M","1Y",
          "18M","2Y","3Y","4Y","5Y","6Y","7Y","8Y","9Y","10Y","12Y","15Y","20Y","25Y",
          "30Y","40Y","50Y"]
R0 = 7
R1 = R0 + len(TENORS) - 1
CASE = {1: ("B","C","D","E"), 2: ("F","G","H","I"), 3: ("J","K","L","M")}
MODE = "Bootstrap!$G$4"
LIVE, FIXED = "Live (BDP)", "Fixed (S490 07/21/26)"
TESTS = ["Test 1", "Test 2", "Test 3"]


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False, al=None):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    if al: c.alignment = Alignment(horizontal=al, wrap_text=True, vertical="center")
    return c


wb = load_workbook(WB)
if "Testing" in wb.sheetnames:
    del wb["Testing"]
ws = wb.create_sheet("Testing", wb.sheetnames.index("Bootstrap") + 1)
ws.sheet_properties.tabColor = "B7950B"

put(ws, "A1", "Testing — curve and CDS test cases", H1)
put(ws, "A2", "Curve: paste the swap rates Bloomberg used and the zero + discount curve it "
              "produced. Select the case at Bootstrap!G4 to route it through the bootstrap.", NOTE)
put(ws, "A3", "Only the selected case compares — there is one bootstrap, deliberately. "
              "The other two hold their inputs and wait.", WARN)
put(ws, "A5", "Active source", BOLD)
put(ws, "B5", f"={MODE}", BLACK, None, OF, True)

ws.column_dimensions["A"].width = 9
put(ws, f"A{R0-1}", "Tenor", HDR, None, HF, True, "center")
for i, t in enumerate(TENORS):
    put(ws, f"A{R0+i}", t, BLACK, None, None, True, "center")

for k, (cd, cm, cz, cf) in CASE.items():
    band = PatternFill("solid", fgColor=CF[k-1])
    put(ws, f"{cd}{R0-3}", f"CURVE TEST CASE {k}", BOLD)
    put(ws, f"{cd}{R0-2}", "name / capture time:", NOTE)
    put(ws, f"{cm}{R0-2}", None, BLUE, None, YF, True)
    for c, h, w, fmt in ((cd,"Date",13,DTF), (cm,"Swap rate (mid)",14,N5),
                         (cz,"BBG zero %",12,N5), (cf,"BBG discount",13,N6)):
        ws.column_dimensions[c].width = w
        put(ws, f"{c}{R0-1}", h, HDR, None, band, True, "center")
        for i in range(len(TENORS)):
            put(ws, f"{c}{R0+i}", None, BLUE, fmt, YF, True)
ws.row_dimensions[R0-1].height = 30

# ---- active-case comparison
OUT = ["O","P","Q","R","S"]
for c, h, w in zip(OUT, ["Date (active)","Our zero %","Our DF","d zero (bp)","d DF"],
                   [13,12,13,12,12]):
    ws.column_dimensions[c].width = w
    put(ws, f"{c}{R0-1}", h, HDR, None, HF, True, "center")
put(ws, f"O{R0-3}", "ACTIVE CASE vs OUR BOOTSTRAP", BOLD)

def pick(col):
    """value of `col` for the active test case, blank if none selected"""
    a, b, c = CASE[1][col], CASE[2][col], CASE[3][col]
    return (f'IF({MODE}="{TESTS[0]}",{a}{{r}},IF({MODE}="{TESTS[1]}",{b}{{r}},'
            f'IF({MODE}="{TESTS[2]}",{c}{{r}},"")))')

BD, BJ, BH = "Bootstrap!$B$8:$B$72", "Bootstrap!$J$8:$J$72", "Bootstrap!$H$8:$H$72"
for i in range(len(TENORS)):
    r = R0 + i
    d = pick(0).format(r=r); z = pick(2).format(r=r); f_ = pick(3).format(r=r)
    put(ws, f"O{r}", f"={d}", BLACK, DTF, None, True)
    m = f"MATCH(O{r},{BD},0)"
    put(ws, f"P{r}", f'=IFERROR(INDEX({BJ},{m}),"")', BLACK, N5, None, True)
    put(ws, f"Q{r}", f'=IFERROR(INDEX({BH},{m}),"")', BLACK, "0.00000000", None, True)
    put(ws, f"R{r}", f'=IF(OR(P{r}="",NOT(ISNUMBER({z}))),"",(P{r}-({z}))*100)', BLACK, BP, None, True)
    put(ws, f"S{r}", f'=IF(OR(Q{r}="",NOT(ISNUMBER({f_}))),"",Q{r}-({f_}))', BLACK, "0.00E+00", None, True)

s = R1 + 2
put(ws, f"O{s}", "Pillars matched", BOLD)
put(ws, f"Q{s}", f'=COUNT(P{R0}:P{R1})&" / {len(TENORS)}"', BLACK, None, OF, True)
put(ws, f"O{s+1}", "Max |d zero| (bp)", BOLD)
put(ws, f"Q{s+1}", f'=IF(COUNT(R{R0}:R{R1})=0,"select a test case",MAX(MAX(R{R0}:R{R1}),-MIN(R{R0}:R{R1})))',
    BLACK, "0.000", OF, True)
put(ws, f"O{s+2}", "Max |d DF|", BOLD)
put(ws, f"Q{s+2}", f'=IF(COUNT(S{R0}:S{R1})=0,"",MAX(MAX(S{R0}:S{R1}),-MIN(S{R0}:S{R1})))',
    BLACK, "0.00E+00", OF, True)
put(ws, f"O{s+4}", "Benchmark: the 07/21/26 capture gives 0.397bp and 1.88e-05 through this "
                   "same path. A test case should land in that region.", NOTE)

# ---- extend the curve-source dropdown
b = wb["Bootstrap"]
dv = DataValidation(type="list", formula1=f'"{LIVE},{FIXED},{TESTS[0]},{TESTS[1]},{TESTS[2]}"',
                    allow_blank=False)
b.add_data_validation(dv); dv.add(b["G4"])
put(b, "I4", "Live / Fixed / Test 1-3. Test cases come from the Testing sheet and route "
             "through SOFR_OIS_Quotes!H, so the whole curve and the charts follow.", NOTE)

# ---- route test rates into the quotes sheet
q = wb["SOFR_OIS_Quotes"]
TA = f"Testing!$A${R0}:$A${R1}"
n = 0
for r in range(5, 40):
    t = q[f"B{r}"].value
    if t is None or q[f"A{r}"].value is None:
        continue
    m = f"MATCH(B{r},{TA},0)"
    t1 = f'INDEX(Testing!$C${R0}:$C${R1},{m})'
    t2 = f'INDEX(Testing!$G${R0}:$G${R1},{m})'
    t3 = f'INDEX(Testing!$K${R0}:$K${R1},{m})'
    live = f'IF(ISNUMBER(T{r}),T{r},J{r})'
    put(q, f"H{r}",
        f'=IFERROR(IF({MODE}="{TESTS[0]}",{t1},IF({MODE}="{TESTS[1]}",{t2},'
        f'IF({MODE}="{TESTS[2]}",{t3},IF({MODE}="{FIXED}",J{r},{live})))),J{r})',
        BLACK, N5, OF, True)
    n += 1

# ---- CDS: Help Desk two-step pull per entity row, so a ticker populates the row
e = wb["CDS_Entities"]
put(e, "AE3", "DYNAMIC PULL — Help Desk H#1330731572", BOLD)
put(e, "AE4", "Step 1 ticker", HDR, None, HF, True, "center")
TEN6 = ["1Y","2Y","3Y","5Y","7Y","10Y"]
SPCOL = {"1Y":"H","2Y":"I","3Y":"J","5Y":"K","7Y":"L","10Y":"M"}
from openpyxl.utils import get_column_letter as CL, column_index_from_string as CI
c0 = CI("AE")
for j, ten in enumerate(TEN6):
    c1 = CL(c0 + 1 + j*2)
    c2 = CL(c0 + 2 + j*2)
    put(e, f"{c1}4", f"{ten} ticker", HDR, None, HF, True, "center")
    put(e, f"{c2}4", f"{ten} bp live", HDR, None, HF, True, "center")
    e.column_dimensions[c1].width = 18
    e.column_dimensions[c2].width = 11
    for r in range(5, 15):
        put(e, f"{c1}{r}", f'=IFERROR(BDP($B{r},"CDS_SPREAD_TICKER_{ten}"),"")', BLACK, None, None, True)
        put(e, f"{c2}{r}", f'=IFERROR(BDP({c1}{r}&" BEST Curncy","PX_LAST")+0,"")',
            BLACK, "0.0000", None, True)
        # manual column wins only when the live pull is empty
        put(e, f"{SPCOL[ten]}{r}", f'=IF(ISNUMBER({c2}{r}),{c2}{r},"")', BLACK, "0.0000", OF, True)
put(e, "AE2", "Type a ticker in col B and the row populates. Cols H:M show what is in use: "
              "the live pull when it resolves. To hard-code instead, overwrite H:M directly.", NOTE)

# ---- point the Testing sheet at the three CDS cases
put(ws, "A" + str(s + 7), "CDS TEST CASES", H1)
put(ws, "A" + str(s + 8), "Rows 5, 6 and 7 of CDS_Entities. Type a ticker in col B and the term "
                          "structure pulls live; overwrite H:M to hard-code. Paste the CDSW "
                          "screen into O:AC. Select which one prices at CDS_Parameters!B27.", NOTE)
put(ws, "A" + str(s + 9), "The model-vs-capture comparison is on CDS_Pricer!A41 for whichever "
                          "name is active.", NOTE)
for i in range(3):
    rr = s + 11 + i
    put(ws, f"A{rr}", f"Case {i+1}", BOLD)
    put(ws, f"B{rr}", f'=IF(CDS_Entities!$A${5+i}="","(empty)",CDS_Entities!$A${5+i})', BLACK, None, None, True)
    put(ws, f"D{rr}", f'=IF(CDS_Entities!$B${5+i}="","no ticker","ticker: "&CDS_Entities!$B${5+i})', NOTE)
    put(ws, f"H{rr}", f'=IF(CDS_Parameters!$B$27=CDS_Entities!$A${5+i},"<< ACTIVE","")', BOLD)

ws.freeze_panes = f"B{R0}"
ws.sheet_view.showGridLines = False
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"Testing sheet: 3 curve cases x {len(TENORS)} pillars; quotes routed on {n} rows; "
      f"CDS dynamic pull on CDS_Entities")
