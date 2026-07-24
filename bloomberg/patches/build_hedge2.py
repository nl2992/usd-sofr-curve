"""
Hedge (developed): one-page blotter. Reads the folded Kevin-format deltas
(KRD-v3 in this book; Sheet1 in the master), sizes each benchmark swap live off
the curve annuities, states direction, and shows before/after residual.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
WB="bloomberg/KRW_IRS_Bootstrap_Book1.xlsx"; SRC="'KRD-v3'"
F="Calibri"
TITLE=Font(name=F,size=15,bold=True,color="1F3864")
SEC=Font(name=F,size=11,bold=True,color="FFFFFF"); SECF=PatternFill("solid",fgColor="1F3864")
B=Font(name=F,size=10,bold=True); SM=Font(name=F,size=9,color="808080"); N=Font(name=F,size=10)
HD=Font(name=F,size=9,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
GRN=Font(name=F,size=11,bold=True,color="006100"); REDF=Font(name=F,size=11,bold=True,color="9C0006")
KEY=PatternFill("solid",fgColor="FCE4D6"); BLU=PatternFill("solid",fgColor="DDEBF7")
thin=Side(style="thin",color="BFBFBF"); BOX=Border(thin,thin,thin,thin)
MONEY='#,##0'; BN='#,##0.0'; CTR=Alignment(horizontal="center")

# folded-table row map (KRD-v3): 1Y..10Y
ROW={"3M":12,"6M":13,"9M":14,"1Y":15,"2Y":16,"3Y":17,"4Y":18,"5Y":19,"7Y":20,"10Y":21}
def c(t): return f"{SRC}!$B${ROW[t]}"
TAU="Quotes!$B$2"; YR="Bootstrap!$B$3:$B$122"; DF="Bootstrap!$D$3:$D$122"

wb=load_workbook(WB)
if "Hedge" in wb.sheetnames: del wb["Hedge"]
ws=wb.create_sheet("Hedge", wb.sheetnames.index("KRD-v3")+1)

ws["A1"]="Hedge proposal — KRW amortising IRS (10Y, HSBC receives fixed)"; ws["A1"].font=TITLE
ws["A2"]=("Bucketed KRD (Sheet1) hedged at liquid benchmarks. Notional(T) = group DV01 / "
          "(annuity(T) x 1bp), sized live off the curve; direction from the bucket sign.")
ws["A2"].font=SM

def section(r,txt):
    ws.cell(r,1,txt).font=SEC
    for cc in range(1,7): ws.cell(r,cc).fill=SECF

# 1 · The risk
section(4,"1 · The risk")
ws.cell(5,1,"Net DV01").font=B
ws.cell(5,2,f"={SRC}!$B$27").number_format=MONEY; ws.cell(5,2).font=REDF; ws.cell(5,2).fill=KEY
ws.cell(5,3,"KRW per bp — net receiver").font=SM
ws.cell(6,1,"Front 1Y–3Y").font=B
ws.cell(6,2,f"={c('1Y')}+{c('2Y')}+{c('3Y')}").number_format=MONEY; ws.cell(6,2).font=GRN
ws.cell(6,3,"long (positive) — pays if rates rise").font=SM
ws.cell(7,1,"Back 4Y–10Y").font=B
ws.cell(7,2,f"={c('4Y')}+{c('5Y')}+{c('7Y')}+{c('10Y')}").number_format=MONEY; ws.cell(7,2).font=REDF
ws.cell(7,3,"short (negative) — concentrated at 10Y").font=SM
ws.cell(8,1,"Sub-1Y (3M/6M/9M)").font=B
ws.cell(8,2,f"={c('3M')}+{c('6M')}+{c('9M')}").number_format=MONEY; ws.cell(8,2).font=SM
ws.cell(8,3,"dust — leave unhedged").font=SM

# 2 · Strategy
section(10,"2 · Strategy")
for i,t in enumerate([
    "Hedge points, not buckets — trade the 4 liquid benchmarks (2Y/3Y/5Y/10Y), not all 10 nodes.",
    "Fold orphans into the nearest benchmark: 1Y→2Y, 4Y→5Y, 7Y→10Y. Size to null the group DV01.",
    "Skip the sub-1Y dust — its bid/offer costs more than the risk it removes."]):
    ws.cell(11+i,1,f"•  {t}").font=N

# annuity helper (right side)
ws.cell(15,8,"Benchmark annuity (curve)").font=B
for i,h in enumerate(["Tenor","T","Annuity"],8):
    cc=ws.cell(16,i,h); cc.font=HD; cc.fill=HF
for k,(tn,T) in enumerate([("2Y",2),("3Y",3),("5Y",5),("7Y",7),("10Y",10)]):
    r=17+k
    ws.cell(r,8,tn); ws.cell(r,9,T)
    ws.cell(r,10,f"={TAU}*SUMIF({YR},\"<=\"&I{r},{DF})").number_format="0.0000"
ANN={"2Y":"$J$17","3Y":"$J$18","5Y":"$J$19","7Y":"$J$20","10Y":"$J$21"}

# 3 · The trades (blotter)
section(15,"3 · The trades  (4 swaps)")
hdr=["Swap","Direction","Notional (bn)","Absorbs","Group DV01 (KRW/bp)","Annuity"]
for i,h in enumerate(hdr,1):
    cc=ws.cell(16,i,h); cc.font=HD; cc.fill=HF; cc.border=BOX
blot=[("2Y","1Y + 2Y",f"={c('1Y')}+{c('2Y')}","2Y"),
      ("3Y","3Y",f"={c('3Y')}","3Y"),
      ("5Y","4Y + 5Y",f"={c('4Y')}+{c('5Y')}","5Y"),
      ("10Y","7Y + 10Y",f"={c('7Y')}+{c('10Y')}","10Y")]
r0=17
for k,(tn,absorbs,dv,annk) in enumerate(blot):
    r=r0+k
    ws.cell(r,1,tn).font=B
    ws.cell(r,2,f'=IF(E{r}>0,"Receive fixed","Pay fixed")').font=B
    ws.cell(r,3,f"=ABS(E{r}/(F{r}*0.0001))/1000000000").number_format=BN; ws.cell(r,3).font=GRN; ws.cell(r,3).fill=BLU
    ws.cell(r,4,absorbs)
    ws.cell(r,5,dv).number_format=MONEY
    ws.cell(r,6,f"={ANN[annk]}").number_format="0.0000"
    for cc in range(1,7): ws.cell(r,cc).border=BOX

# 4 · Residual
rr=r0+len(blot)+1
section(rr,"4 · After the hedge")
ws.cell(rr+1,1,"Grouped buckets 2Y/3Y/5Y/10Y").font=B
ws.cell(rr+1,3,"neutralized to 0").font=N
ws.cell(rr+2,1,"Residual DV01").font=B
ws.cell(rr+2,3,f"={c('3M')}+{c('6M')}+{c('9M')}").number_format=MONEY; ws.cell(rr+2,3).fill=KEY; ws.cell(rr+2,3).font=B
ws.cell(rr+2,4,"sub-1Y dust only").font=SM
ws.cell(rr+3,1,"Curve residual").font=B
ws.cell(rr+3,3,"small 7s10s (7Y folded into 10Y) — see option below").font=SM

# 5 · Optional
ro=rr+5
section(ro,"5 · Optional — split 7Y out (kills the 7s10s residual, +1 ticket)")
for i,h in enumerate(hdr,1):
    cc=ws.cell(ro+1,i,h); cc.font=HD; cc.fill=HF; cc.border=BOX
opt=[("7Y","7Y",f"={c('7Y')}","7Y"),("10Y","10Y",f"={c('10Y')}","10Y")]
for k,(tn,absorbs,dv,annk) in enumerate(opt):
    r=ro+2+k
    ws.cell(r,1,tn).font=B
    ws.cell(r,2,f'=IF(E{r}>0,"Receive fixed","Pay fixed")').font=B
    ws.cell(r,3,f"=ABS(E{r}/(F{r}*0.0001))/1000000000").number_format=BN; ws.cell(r,3).font=GRN
    ws.cell(r,4,absorbs); ws.cell(r,5,dv).number_format=MONEY
    ws.cell(r,6,f"={ANN[annk]}").number_format="0.0000"
    for cc in range(1,7): ws.cell(r,cc).border=BOX
ws.cell(ro+2+len(opt),1,"In the master, repoint the KRD-v3 refs to Sheet1 (1Y=B9, 2Y=B10, 3Y=B11, 4Y=B12, 5Y=B13, 7Y=B14, 10Y=B15).").font=SM

for col,w in zip("ABCDEF",(11,15,14,14,22,10)): ws.column_dimensions[col].width=w
for col,w in zip("HIJ",(8,5,10)): ws.column_dimensions[col].width=w
wb.save(WB)
print("Hedge rebuilt at",wb.sheetnames.index("Hedge"))
