"""
Test 2 runs the bootstrap in VBA instead of in cells.

Test 1 and Test 3 stay as they are - INDEX/MATCH into the Bootstrap grid. Test 2
calls CurveVBA.SOFR_Curve() on the SAME quotes, so the two engines can be diffed
against each other as well as against Bloomberg. Column L is that head-to-head.

Two implementations agreeing is a real check on both. A third copy of the same
formulas would only have checked the copying.

Needs CurveVBA.bas imported and the workbook saved as .xlsm. Without it every
VBA cell is #NAME?, so the value columns are wrapped in IFERROR and H56 says
plainly what is missing.
"""
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
shutil.copy(WB, WB + ".bak")

R0, R1, SUM = 58, 97, 98
CD, TEN, RAT = "$D$55", "$A$58:$A$97", "$B$58:$B$97"
CALL = f'SOFR_Curve({CD},{TEN},{RAT},$A{{r}},"{{o}}")'
ACT = 'BOOT_MODE="Test 2"'

BLACK = Font(name="Calibri", size=11)
GREEN = Font(name="Calibri", size=11, color="008000")
HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="4472C4")
VFILL = PatternFill("solid", fgColor="E2EFDA")
BOX = Border(*[Side(style="thin", color="BFBFBF")] * 4)

wb = load_workbook(WB)
t = wb["Testing"]

HEADERS = {"G": "S used % (VBA)", "H": "Engine", "I": "VBA date", "J": "date ok?",
           "K": "cells DF", "L": "VBA - cells", "M": "our DF (VBA)",
           "N": "t ACT/365", "O": "our zero % (VBA)", "P": "d zero bp", "Q": "d DF"}
for col, txt in HEADERS.items():
    c = t[f"{col}57"]
    c.value = txt; c.font = HDR; c.fill = HFILL; c.border = BOX
    c.alignment = Alignment(horizontal="center", wrap_text=True)

def vba(r, o):
    return CALL.format(r=r, o=o)

n = 0
for r in range(R0, R1 + 1):
    g = lambda o: vba(r, o)
    t[f"G{r}"] = f'=IF(NOT({ACT}),"",IFERROR({g("PAR")},""))'
    t[f"H{r}"] = f'=IF(OR(NOT({ACT}),$A{r}=""),"","VBA SOFR_Curve()")'
    t[f"I{r}"] = f'=IF(NOT({ACT}),"",IFERROR({g("DATE")},""))'
    t[f"J{r}"] = (f'=IF(OR(NOT({ACT}),I{r}="",F{r}=""),"",'
                  f'IF(INT(I{r})=INT(F{r}),"ok","MISMATCH"))')
    t[f"K{r}"] = f'=IF(NOT({ACT}),"",IFERROR(INDEX(BOOT_DF,MATCH(F{r},BOOT_DATES,0)),""))'
    t[f"L{r}"] = f'=IF(OR(M{r}="",K{r}=""),"",M{r}-K{r})'
    t[f"M{r}"] = f'=IF(NOT({ACT}),"",IFERROR({g("DF")},""))'
    t[f"N{r}"] = f'=IF(NOT({ACT}),"",IFERROR({g("T")},""))'
    t[f"O{r}"] = f'=IF(NOT({ACT}),"",IFERROR({g("ZERO")},""))'
    for col, fmt in (("G", "0.00000"), ("I", "mm/dd/yyyy"), ("K", "0.00000000"),
                     ("L", "0.00E+00"), ("M", "0.00000000"), ("N", "0.0000"),
                     ("O", "0.00000")):
        t[f"{col}{r}"].number_format = fmt
    for col in "GHIJKLMNO":
        c = t[f"{col}{r}"]; c.border = BOX
        c.font = GREEN if col in "GIMNO" else BLACK
        if col in "GIMNO": c.fill = VFILL
    n += 1

# is the module actually there?
probe = f'SOFR_Curve({CD},{TEN},{RAT},"5Y","DF")'
t["H56"] = (f'=IF(NOT({ACT}),"",IF(ISERROR({probe}),'
            f'"VBA NOT AVAILABLE - import bloomberg/vba/CurveVBA.bas (Alt+F11 > File > '
            f'Import File) and save this workbook as .xlsm",'
            f'"VBA OK - CurveVBA.SOFR_Curve() is answering"))')
t["H56"].font = Font(name="Calibri", size=11, bold=True, color="C00000")

# summary: the two engines against each other, and the dates
t[f"G{SUM}"] = "max |d zero| bp vs BBG"
t[f"J{SUM}"] = "max |VBA - cells|"
t[f"L{SUM}"] = (f'=IF(NOT({ACT}),"",IF(COUNT(L{R0}:L{R1})=0,"",'
                f'MAX(MAX(L{R0}:L{R1}),-MIN(L{R0}:L{R1}))))')
t[f"L{SUM}"].number_format = "0.00E+00"
t[f"N{SUM}"] = "date mismatches"
t[f"O{SUM}"] = f'=IF(NOT({ACT}),"",COUNTIF(J{R0}:J{R1},"MISMATCH"))'
for a in (f"G{SUM}", f"J{SUM}", f"N{SUM}"):
    t[a].font = Font(name="Calibri", size=11, bold=True)
for a in (f"L{SUM}", f"O{SUM}"):
    t[a].font = BLACK; t[a].border = BOX

t["A56"] = ("Test 2 runs the SAME quotes through the VBA bootstrap (CurveVBA.SOFR_Curve) "
            "instead of the cell grid. Column L is VBA minus cells - two independent "
            "implementations, so agreement checks both. Needs .xlsm.")
t["A56"].font = Font(name="Calibri", size=9, italic=True, color="808080")

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"Test 2 rewired to VBA: {n} rows, 5 UDF columns + head-to-head in L")
