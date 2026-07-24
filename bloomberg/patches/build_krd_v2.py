"""
KRD-v2: delta in Kevin's tenor format, with the FRONT nodes (2D,1W,1M,2M) added
as LIVE bump columns so the sheet computes their zero itself (not hardcoded).
Front-node rate = the 3M quote (flat to 3M); bumping it moves only the sub-3M
segment, where the swap has no cashflow -> delta = 0, computed on the sheet.
Self-contained; links only to Quotes and Amort-Set-up.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as CL

WB="bloomberg/KRW_IRS_Bootstrap_Book1.xlsx"; AMT="'Amort-Set-up'"
FRONT=[("2D",2/365),("1W",7/365),("1M",1/12),("2M",2/12)]     # synthetic front nodes
KEVROWS=[5,6,7,8,9,10,11,12,13,16,17,18,19]                   # real nodes from Quotes (skip 8Y/9Y)
NQ=len(FRONT)+len(KEVROWS)                                    # 17 nodes
NB=len(FRONT)+10                                              # bump first 14 (2D..10Y)
NS=1+NB                                                       # base + 14 = 15 scenarios
F="Calibri"
H1=Font(name=F,size=14,bold=True); B=Font(name=F,size=11,bold=True); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=9,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
RES=PatternFill("solid",fgColor="FCE4D6"); GRN=Font(name=F,size=10,color="008000")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4); MONEY='#,##0'; PCT="0.000%"

wb=load_workbook(WB)
for nm in ("KRD reformed","KRD-v2"):
    if nm in wb.sheetnames: del wb[nm]
ws=wb.create_sheet("KRD-v2", wb.sheetnames.index("KRD")+1)

ws["A1"]="KRD-v2 — delta in Kevin's format (front nodes bumped live)"; ws["A1"].font=H1
ws["A2"]=("Bump each tenor 1bp, re-bootstrap, re-price, read the NPV change. 2D/1W/1M/2M are "
          "real bump columns here — they compute 0 because the first cashflow is at 3M. Kevin's "
          "grid (8Y/9Y dropped). Links only to Quotes and Amort-Set-up.")
ws["A2"].font=SM

QM0=30; QR0=51; QR1=QR0+39; ANCH=QR0-1
RC0=7; DC0=24; TAU="Quotes!$B$2"

ws.cell(28,1,"WORKING (proof)").font=B
for i,h in enumerate(["Tenor","Year","Base"],1):
    c=ws.cell(QM0-1,i,h); c.font=HD; c.fill=HF
labels=[f[0] for f in FRONT]+[None]*len(KEVROWS)
for t in range(NQ):
    r=QM0+t
    if t < len(FRONT):
        ws.cell(r,1,FRONT[t][0])
        ws.cell(r,2,round(FRONT[t][1],6)).number_format="0.000000"
        ws.cell(r,3,"=Quotes!$B$5").number_format=PCT          # front rate = 3M quote
    else:
        qr=KEVROWS[t-len(FRONT)]
        ws.cell(r,1,f"=Quotes!A{qr}")
        ws.cell(r,2,f"=Quotes!C{qr}").number_format="0.00"
        ws.cell(r,3,f"=Quotes!B{qr}").number_format=PCT
    for j in range(NS):
        col=4+j
        if j==0: ws.cell(r,col,f"=$C{r}")
        else:
            tgt=QM0+(j-1)                                       # bump nodes 0..13
            ws.cell(r,col,(f"=$C{r}+0.0001" if r==tgt else f"=$C{r}"))
        ws.cell(r,col).number_format=PCT
QMAT=f"$D${QM0}:${CL(3+NS)}${QM0+NQ-1}"; QYR=f"$B${QM0}:$B${QM0+NQ-1}"

for i,h in enumerate(["Qtr","Year","m","w","Notional"],1):
    c=ws.cell(QR0-1,i,h); c.font=HD; c.fill=HF
for j in range(NS):
    ws.cell(QR0-1,RC0+j,j+1)
    ws.cell(ANCH,DC0+j,1).number_format="0.00000000"
for qq in range(1,41):
    r=QR0+qq-1
    ws.cell(r,1,qq)
    ws.cell(r,2,f"=A{r}*0.25").number_format="0.00"
    ws.cell(r,3,f"=MIN(MATCH(B{r},{QYR},1),{NQ-1})")
    ws.cell(r,4,f"=(B{r}-INDEX({QYR},C{r}))/(INDEX({QYR},C{r}+1)-INDEX({QYR},C{r}))")
    ws.cell(r,5,f"={AMT}!E{19+qq}").number_format=MONEY
    for j in range(NS):
        rc=RC0+j; dc=DC0+j; RL=CL(rc); DLx=CL(dc); hh=f"{RL}${QR0-1}"
        ws.cell(r,rc,f"=INDEX({QMAT},$C{r},{hh})+$D{r}*(INDEX({QMAT},$C{r}+1,{hh})-INDEX({QMAT},$C{r},{hh}))").number_format=PCT
        if qq==1: ws.cell(r,dc,f"=1/(1+{RL}{r}*{TAU})")
        else: ws.cell(r,dc,f"=(1-{RL}{r}*{TAU}*SUM({DLx}${QR0}:{DLx}{r-1}))/(1+{RL}{r}*{TAU})")
        ws.cell(r,dc).number_format="0.00000000"

Nr=f"$E${QR0}:$E${QR1}"; S=CL(DC0)
FP,AN,PAR=QR1+2,QR1+3,QR1+4
ws.cell(FP,1,"base Float PV").font=B
ws.cell(FP,3,f"=SUMPRODUCT({Nr},{S}${ANCH}:{S}${QR1-1}-{S}${QR0}:{S}${QR1})").number_format=MONEY
ws.cell(AN,1,"base annuity").font=B
ws.cell(AN,3,f"={TAU}*SUMPRODUCT({Nr},{S}${QR0}:{S}${QR1})").number_format=MONEY
ws.cell(PAR,1,"frozen par rate").font=B
ws.cell(PAR,3,f"=C{FP}/C{AN}").number_format=PCT; ws.cell(PAR,3).fill=RES; ws.cell(PAR,3).font=B
PARC=f"$C${PAR}"
FR,FL,NP,DL=PAR+2,PAR+3,PAR+4,PAR+5
for a,t in ((FR,"Fixed PV"),(FL,"Float PV"),(NP,"NPV"),(DL,"DELTA")): ws.cell(a,1,t).font=B
for j in range(NS):
    dc=DC0+j; DLc=CL(dc)
    ws.cell(FR,dc,f"={PARC}*{TAU}*SUMPRODUCT({Nr},{DLc}${QR0}:{DLc}${QR1})").number_format=MONEY
    ws.cell(FL,dc,f"=SUMPRODUCT({Nr},{DLc}${ANCH}:{DLc}${QR1-1}-{DLc}${QR0}:{DLc}${QR1})").number_format=MONEY
    ws.cell(NP,dc,f"={DLc}{FR}-{DLc}{FL}").number_format=MONEY
    ws.cell(DL,dc,f"={DLc}{NP}-${CL(DC0)}{NP}").number_format=MONEY
for j in range(1,NS):
    dc=DC0+j; tgt=QM0+(j-1)
    ws.cell(DL+1,dc,f"=$A{tgt}").font=SM

ws.cell(4,1,"OUTPUT  (send to Kevin)").font=B
for i,h in enumerate(["Tenor","dv01"],1):
    c=ws.cell(5,i,h); c.font=HD; c.fill=HF; c.border=BOX
KEVIN=["2D","1W","1M","2M","3M","6M","9M","1Y","2Y","3Y","4Y","5Y","7Y","10Y","12Y","15Y","20Y","25Y","30Y"]
lbl2node={"2D":0,"1W":1,"1M":2,"2M":3,"3M":4,"6M":5,"9M":6,"1Y":7,"2Y":8,"3Y":9,"4Y":10,"5Y":11,"7Y":12,"10Y":13}
r=6
for tn in KEVIN:
    ws.cell(r,1,tn).border=BOX
    if tn in lbl2node:
        dc=DC0+1+lbl2node[tn]
        ws.cell(r,2,f"={CL(dc)}${DL}").number_format=MONEY
    else:
        ws.cell(r,2,0).number_format=MONEY
    ws.cell(r,2).border=BOX; ws.cell(r,2).font=GRN
    r+=1
ws.cell(r,1,"TOTAL").font=B; ws.cell(r,1).border=BOX
ws.cell(r,2,f"=SUM(B6:B{r-1})").number_format=MONEY; ws.cell(r,2).fill=RES; ws.cell(r,2).font=B; ws.cell(r,2).border=BOX
ws.cell(r+2,1,"2D/1W/1M/2M are live bumps and land on 0 — swap's first cashflow is at 3M, "
              "so the sub-3M curve is never touched. 12Y-30Y are 0: swap matures at 10Y.").font=SM
for col,w in zip("ABCDE",(10,16,12,10,16)): ws.column_dimensions[col].width=w
wb.save(WB)
print("KRD-v2 at pos", wb.sheetnames.index("KRD-v2"),"of",wb.sheetnames)
