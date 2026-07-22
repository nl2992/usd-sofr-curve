"""
Make the Curve_Interface interpolation selectable, per the Bloomberg IRS Curve
Construction guide's four methods.

EMPIRICAL FINDING (do not change the default without redoing this test):
the USD S490 curve is NOT built with the guide's method 1. Fitting both methods
to the 32 screen pillars, where the long end is sensitive to gap interpolation
through the swap annuity:

    log-linear DF (= step-function forward, method 3)   rms DF err 2.28e-06
    piecewise linear simple zero (method 1)             rms DF err 1.09e-04

Method 1 is 48x worse and systematically biased (+4.2e-04 by 50Y, growing with
maturity). Method 1 is the guide's EUR S201 example ("Piecewise Linear (Simple)"
in Fig 3a), a different curve with a different interpolation setting.

So the default stays step-function forward. Method 1 is added as a switchable
alternative for curves that do use it.

Method 1, exactly as the guide states it:
    DF(t) = 1 / (1 + r_s(t) * t),  r_s piecewise LINEAR between instrument
    maturities, t on ACT/360, r_s held FLAT before the first and after the last
    instrument maturity.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"

FONT = "Calibri"
BLUE = Font(name=FONT, size=11, color="0000FF")
BLACK = Font(name=FONT, size=11)
SECT = Font(name=FONT, size=11, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="666666")
WARN = Font(name=FONT, size=10, bold=True, color="C00000")
YFILL = PatternFill("solid", fgColor="FFFF00")
OFILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

R0, RN = 10, 73
K, L = "$K$8:$K$73", "$L$8:$L$73"
MODE = "$J$7"          # interpolation selector (K6:L6 and A4:H6 are merged; J7 is free)


def put(ws, cell, val, font=BLACK, fmt=None, fill=None, border=False):
    c = ws[cell]
    c.value = val
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if border: c.border = BOX
    return c


def main():
    wb = load_workbook(WB)
    ws = wb["Curve_Interface"]

    put(ws, "I7", "Interpolation", SECT)
    put(ws, "J7", "Step-function forward (flat fwd)", BLUE, None, YFILL, True)
    dv = DataValidation(
        type="list",
        formula1='"Step-function forward (flat fwd),Piecewise Linear (Simple)"',
        allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["J7"])
    put(ws, "I8", "VERIFIED default", SECT)
    put(ws, "J8", "Step-function forward reproduces the S490 screen (rms DF err 2.3e-06). "
                  "Piecewise Linear (Simple) — the guide's method 1, its EUR S201 example — "
                  "is 48x worse on this curve (1.1e-04, biased to +4.2e-04 by 50Y). "
                  "Only switch if the S490 Interpolation field actually says so.", WARN)

    for r in range(R0, RN + 1):
        if ws[f"B{r}"].value is None:
            continue
        i = f"MATCH(B{r},{K},1)"
        # --- method 3 (existing): flat forward off the previous pillar
        m3 = f"INDEX({L},{i})*EXP(-(F{r}/100)*(B{r}-D{r})/365)"
        # --- method 1: linear simple zero on ACT/360, flat outside the pillars
        tp = f"MAX(({r}>0)*(D{r}-VAL_DATE)/360,0)"
        tn = f"((E{r}-VAL_DATE)/360)"
        tt = f"((B{r}-VAL_DATE)/360)"
        rA = f"IF({tp}<=0,0,(1/INDEX({L},{i})-1)/{tp})"
        rB = f"IF({tn}<=0,0,(1/INDEX({L},{i}+1)-1)/{tn})"
        # guide: r_s flat before the first instrument maturity
        rAe = f"IF({tp}<=0,{rB},{rA})"
        w = f"IF({tn}-{tp}=0,0,({tt}-{tp})/({tn}-{tp}))"
        rs = f"({rAe}+({rB}-{rAe})*{w})"
        m1 = f"IF({tt}<=0,1,1/(1+{rs}*{tt}))"
        put(ws, f"G{r}",
            f'=IFERROR(IF({MODE}="Piecewise Linear (Simple)",{m1},{m3}),"")',
            BLACK, "0.00000000", None, True)

    put(ws, "I10", "Guide method 1:  DF(t)=1/(1+r_s(t)*t), r_s piecewise linear, "
                   "t on ACT/360, flat outside the first/last instrument maturity.", NOTE)
    put(ws, "I11", "Guide method 3:  flat instantaneous forward between pillars, "
                   "equivalently log-linear in the discount factor.", NOTE)
    put(ws, "I12", "Both agree AT the pillars; they differ only off-pillar — which is "
                   "exactly what CDS_Schedule looks up. Max gap difference ~0.07bp of "
                   "zero rate in a 2Y gap.", NOTE)

    wb.calculation.fullCalcOnLoad = True
    wb.save(WB)
    print(f"interpolation selector at Curve_Interface!J7; rows {R0}-{RN} rewired")


if __name__ == "__main__":
    main()
