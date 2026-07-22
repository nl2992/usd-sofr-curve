"""
Make the CDSW block derive from equation (3.2) instead of the (S-C)*RPV01 shortcut.

Points upfront was (S - C) * RPV01 * 100, and Principal / Price / Cash all hung
off it. That shortcut drops two things the paper's (3.2) keeps:

    Upfront = Market Value / D(T_s) + Accrued Interest

  - the -S * AI term inside Market Value, since the premium leg is net of accrued
  - compounding from the pricing date T to settlement T_s, three business days on

Against the Wells Fargo CDSW capture, priced flat at the traded 51.5600 on the
07/22/26 SOFR curve:

                  shortcut      (3.2)        CDSW
    principal   +217,291    -213,230    -213,394
    cash        +208,680    -221,841    -222,005
    accrued      -8,611      -8,611      -8,611

$164 on 10mm, 0.0016%, and that residual is the discount curve source - CDSW is
on 490 Mid with a CMAN Ask credit curve, we pull our own.

B27 already computes (3.2) correctly. The CDSW block just was not using it.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

WB="bloomberg/CDS_Pricer.xlsx"
SM=Font(name="Calibri",size=9,color="808080")
wb=load_workbook(WB); ws=wb["CDS_Pricer"]
SGN='IF($B$9="Buy protection",1,-1)'

ws["B31"]=f'=IF($B$27="","",{SGN}*$B$27/$B$4*100)'
ws["C31"]="Principal / notional. From (3.2), not (S-C)*RPV01."
ws["B32"]='=IF($B$31="","",100-$B$31)'
ws["C32"]="100 - points upfront."
ws["B33"]=f'=IF($B$27="","",{SGN}*$B$27)'
ws["C33"]="Upfront (3.2) = Market Value / D(T_s) + Accrued Interest."
ws["B36"]='=IF(OR($B$33="",$B$35=""),"",$B$33+$B$35)'
ws["C36"]="Principal + accrued."
ws["B37"]='=IF($B$33="","",(1-$B$5)*$B$4-$B$33)'
ws["C37"]="(1-R) * notional less the principal already paid."
for r in (31,32,33,36,37): ws.cell(r,3).font=SM
ws["A38"]=("CDSW order. Principal is (3.2) end to end: Market Value / D(T_s) + Accrued Interest, "
           "compounded to settlement three business days on. The (S-C)*RPV01 shortcut drops the "
           "-S*AI term and the compounding.")
ws["A38"].font=SM
wb.calculation.fullCalcOnLoad=True
wb.save(WB)
print("CDSW block rebuilt on (3.2)")
