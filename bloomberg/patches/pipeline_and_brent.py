"""
1. Hazard column can use Brent instead of the in-cell bisection ladder.
2. A Steps sheet laying out the pricing pipeline.
3. Tabs reordered to follow that pipeline.

On (1): Hazard_Solver is 3,712 cells of bisection ladder - 30 halvings per tenor,
each halving a row. It is correct but unreadable, which is the complaint. Brent
does the same job in one call. It is wired as the DEFAULT with the ladder as
fallback, so the workbook still prices as .xlsx if the VBA is not loaded, and
CDS_Parameters!B30 forces either one.

Nothing is deleted. Curves, Model_Notes, Root_Methods, CDS_Validation and
Brent_vs_Bisection are read by nobody because they are leaves - outputs and
reference, not dead weight.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

WB="bloomberg/CDS_Pricer.xlsx"
BLK=[(7,46,"1Y"),(50,89,"2Y"),(93,132,"3Y"),(136,175,"5Y"),(179,218,"7Y"),(222,261,"10Y")]

F="Calibri"
H1=Font(name=F,size=14,bold=True); B=Font(name=F,size=11,bold=True)
N=Font(name=F,size=10); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
IN=PatternFill("solid",fgColor="FFF2CC"); BLUE=Font(name=F,size=11,color="0000FF")
STEP=PatternFill("solid",fgColor="E2EFDA"); BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(WB)

# ---------- 1. Brent as the hazard solver -------------------------------------
p=wb["CDS_Parameters"]
p["A30"]="Hazard solver"; p["A30"].font=B
p["B30"]="Brent (VBA, falls back to bisection)"
p["B30"].fill=IN; p["B30"].font=BLUE; p["B30"].border=BOX
p["C30"]="Brent = one call per tenor. Bisection = the 3,712-cell ladder on Hazard_Solver."
p["C30"].font=SM
dv=DataValidation(type="list",
   formula1='"Brent (VBA, falls back to bisection),Bisection (in-cell ladder)"',allow_blank=False)
p.add_data_validation(dv); dv.add(p["B30"])

hb=wb["Hazard_Bootstrap"]
for i,(b,bis,ten) in enumerate(BLK):
    r=7+i
    a=(f"Hazard_Solver!$D${b},Hazard_Solver!$F${b},Hazard_Solver!$B${b+1},"
       f"Hazard_Solver!$D${b+1},Hazard_Solver!$F${b+1},CDS_Parameters!$B$26,"
       f"Hazard_Solver!$F${b+3}:$U${b+3},Hazard_Solver!$F${b+7}:$U${b+7},"
       f"Hazard_Solver!$F${b+5}:$U${b+5},Hazard_Solver!$F${b+6}:$U${b+6}")
    hb.cell(r,4,
      f'=IF(CDS_Parameters!$B$17="Flat hazard rate",CDS_Parameters!$B$18,'
      f'IF(LEFT(CDS_Parameters!$B$30,5)="Brent",IFERROR(CDS_Hazard({a}),Hazard_Solver!$B${bis}),'
      f'Hazard_Solver!$B${bis}))')
# row 5 is inside a merged note block; N5 is clear
hb["N5"]="Hazard: Brent when the VBA is loaded, else the in-cell ladder. Switch at CDS_Parameters!B30."
hb["N5"].font=SM
print("hazard column: Brent primary, bisection fallback, switchable at B30")

# ---------- 2. Steps sheet ----------------------------------------------------
if "Steps" in wb.sheetnames: del wb["Steps"]
st=wb.create_sheet("Steps",0)
st["A1"]="How this workbook prices a CDS"; st["A1"].font=H1
st["A2"]="Left to right along the tabs. Each step feeds the next."; st["A2"].font=SM

for i,h in enumerate(["Step","Sheet","What happens","What to check"],start=1):
    c=st.cell(4,i,h); c.font=HD; c.fill=HF; c.border=BOX
    c.alignment=Alignment(horizontal="left")

STEPS=[("1","Entities","Type the names and Bloomberg tickers. Pick the live one at B4.",
        "F4 shows the ticker the pull will use."),
       ("2","CDS_Parameters","Contract terms: recovery, notional, coupon, direction, maturity.",
        "B30 picks the hazard solver."),
       ("3","Curve_Interface","SOFR discount curve, linked from the curve workbook.",
        "LINK CHECK at N8. Open the curve workbook first."),
       ("4","CDS_Quotes","Two-step Bloomberg pull: entity -> CDS ticker -> spread.",
        "Col H says whether each tenor came live or manual."),
       ("5","Hazard_Bootstrap","Strips the hazard curve, one tenor at a time, earlier hazards held.",
        "Col J is the repricing error. It must go to zero."),
       ("6","CDS_Schedule","Builds the quarterly legs: DF, survival, accrual per period.",
        "One row per coupon period."),
       ("7","CDS_Pricer","Prices the trade and produces the CDSW settlement figures.",
        "Row 28 down is the CDSW screen block."),
       ("8","Curves","Charts: discount curve, hazard, survival, term structures.",
        "Hazard is drawn as a step because it is piecewise constant."),
       ("9","CDS_Validation","Internal checks on the strip.",
        "All should pass before the numbers are used.")]
r=5
for s,sh,what,chk in STEPS:
    st.cell(r,1,s).font=B; st.cell(r,1).fill=STEP
    st.cell(r,2,sh).font=B
    st.cell(r,3,what).font=N
    st.cell(r,4,chk).font=N
    for k in range(1,5): st.cell(r,k).border=BOX
    r+=1

r+=1
st.cell(r,1,"Reference tabs (not in the pricing path)").font=B; r+=1
for sh,what in [("Model_Notes","B-Model equations (3.1)-(3.6), conventions, strippability check."),
                ("Brent_vs_Bisection","The two solvers side by side on the same objective."),
                ("Root_Methods","Eight root-finders from MATH5030 M2 on the same objective."),
                ("Hazard_Solver","The in-cell bisection ladder. 30 halvings per tenor, one row each."),
                ("CDS_Entities","Data store behind Entities, plus the CDSW capture block.")]:
    st.cell(r,2,sh).font=B; st.cell(r,3,what).font=N; r+=1

r+=1
st.cell(r,1,"Needs VBA").font=B; r+=1
for t in ["CDSBrent.bas       hazard solver (Brent) and the objective",
          "CDSRootFinders.bas the eight methods on Root_Methods",
          "Without them the hazard column falls back to the in-cell ladder and still prices."]:
    st.cell(r,2,t).font=N if not t.startswith("Without") else SM; r+=1
r+=1
st.cell(r,2,"Spreads are demo values. The SOFR curve is real market data; the credit curve is not.").font=Font(name=F,size=10,bold=True,color="C00000")

for col,w in zip("ABCD",(7,20,66,52)): st.column_dimensions[col].width=w
print("Steps sheet added")

# ---------- 3. tab order ------------------------------------------------------
ORDER=["Steps","Entities","CDS_Parameters","Curve_Interface","CDS_Quotes",
       "Hazard_Bootstrap","CDS_Schedule","CDS_Pricer","Curves","CDS_Validation",
       "Model_Notes","Brent_vs_Bisection","Root_Methods","Hazard_Solver","CDS_Entities"]
wb._sheets=[wb[s] for s in ORDER if s in wb.sheetnames]+[ws for ws in wb._sheets if ws.title not in ORDER]
wb.active=0
wb.calculation.fullCalcOnLoad=True
wb.save(WB)
print("tabs ordered:", " > ".join(wb.sheetnames))
