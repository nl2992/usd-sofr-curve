"""
Clear the market data so the workbook ships as a blank template.

Removes: entity names, tickers, ccy, seniority, clause, all spreads (demo and
manual), every override, and the CDSW capture block. Keeps the contract
conventions - notional, coupon, recovery, tenor, direction - because those are
conventions from p.6 rather than data, and blanking them only produces division
by zero. They are inputs and can be overtyped.

Does not touch the curve link.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

WB="bloomberg/CDS_Pricer.xlsx"
NSLOT=10; R0=7; E0=5
F="Calibri"
SM=Font(name=F,size=9,color="808080")
RED=Font(name=F,size=10,bold=True,color="C00000")

wb=load_workbook(WB)
es=wb["Entities"]; ent=wb["CDS_Entities"]

n=0
# Entities: names, tickers, ccy, overrides, seniority, clause, recovery
for i in range(NSLOT):
    r=R0+i
    for col in list(range(1,15)):
        if es.cell(r,col).value is not None:
            es.cell(r,col).value=None; n+=1
es["B4"]=None
print(f"Entities cleared: {n} cells, all {NSLOT} slots empty")

# CDS_Entities: manual spreads, recovery literals, CDSW capture, captured date
m=0
for i in range(NSLOT):
    r=E0+i
    for col in list(range(7,14))+list(range(15,30)):   # G:M and O:AC
        if ent.cell(r,col).value is not None and not str(ent.cell(r,col).value).startswith("="):
            ent.cell(r,col).value=None; m+=1
print(f"CDS_Entities cleared: {m} literals (manual spreads + CDSW capture)")

# front page: ticker blank, conventions kept as defaults
st=wb["Steps"]
st["B5"]=None
st["C5"]="type a Bloomberg ticker here"
st["C5"].font=SM
st["A3"]=("Blank template. Enter a ticker, or add names and spreads on Entities. "
          "Notional, coupon, recovery, tenor and direction below are conventions, not data - overtype as needed.")
st["A3"].font=SM

# make the empty state say so rather than showing bare zeros
st["B17"]=('=IF(COUNT(CDS_Quotes!$F$7:$F$12)=0,"no spreads yet",'
           'IF(SUM(CDS_Quotes!$F$7:$F$12)=0,"spreads all zero - nothing to strip",'
           'IF(MAX(ABS(Hazard_Bootstrap!$J$7:$J$12))<0.01,"reprices, max err "&'
           'TEXT(MAX(ABS(Hazard_Bootstrap!$J$7:$J$12)),"0.00E+00")&" bp",'
           '"STRIP FAILED - check the quotes, see Model_Notes p.9")))')
st["B16"]=('=IF(COUNT(CDS_Quotes!$F$7:$F$12)=0,"none",'
           'IF(COUNTIF(CDS_Quotes!$H$7:$H$12,"BDP live")>0,'
           'COUNTIF(CDS_Quotes!$H$7:$H$12,"BDP live")&" of 6 live","manual"))')

pr=wb["CDS_Pricer"]
pr["A42"]=("Blank until a CDSW capture is typed into CDS_Entities O:AC for the live name. "
           "Notional, coupon, maturity and traded spread there are context, not compared.")
pr["A42"].font=SM

wb.calculation.fullCalcOnLoad=True
wb.save(WB)
print("front page ticker blank; conventions kept as defaults")
