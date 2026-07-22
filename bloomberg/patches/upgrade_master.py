"""
One-time upgrade of an existing master so Testing can be moved in cleanly.

    python3 upgrade_master.py "/path/to/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"

Applies the three things that live OUTSIDE the Testing sheet, so from then on a
Testing update is a plain Move or Copy:

  1. the BOOT_* defined names Testing binds to
  2. SOFR_OIS_Quotes!H  - reads the three Testing blocks, matched on maturity
                          DATE so a 12M label lines up with a 1Y row
  3. Instructions!B9    - VAL_DATE follows the active test case's curve date
  4. Bootstrap!G4       - dropdown gains Test 1/2/3

Takes a .bak first. Safe to re-run.
"""
import sys, shutil
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from openpyxl.styles import Font, PatternFill, Border, Side

if len(sys.argv) < 2:
    sys.exit("usage: upgrade_master.py <workbook.xlsx>")
PATH = sys.argv[1]
shutil.copy(PATH, PATH + ".bak")

NAMES = {"BOOT_MODE": "Bootstrap!$G$4", "BOOT_DFSPOT": "Bootstrap!$D$4",
         "BOOT_DATES": "Bootstrap!$B$8:$B$72", "BOOT_S": "Bootstrap!$E$8:$E$72",
         "BOOT_TAU0": "Bootstrap!$D$8:$D$72", "BOOT_TAUC": "Bootstrap!$F$8:$F$72",
         "BOOT_ANN": "Bootstrap!$G$8:$G$72", "BOOT_DF": "Bootstrap!$H$8:$H$72",
         "BOOT_T": "Bootstrap!$C$8:$C$72", "BOOT_ZERO": "Bootstrap!$J$8:$J$72"}
MODE, FIXED, LIVE = "Bootstrap!$G$4", "Fixed (S490 07/21/26)", "Live (BDP)"
T = ["Test 1", "Test 2", "Test 3"]
BLK = [(8, 47), (58, 97), (108, 147)]
CD = ["Testing!$D$5", "Testing!$D$55", "Testing!$D$105"]
BLACK = Font(name="Calibri", size=11)
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

wb = load_workbook(PATH)
missing = [s for s in ("Bootstrap", "SOFR_OIS_Quotes", "Instructions") if s not in wb.sheetnames]
if missing:
    sys.exit(f"ERROR: this workbook has no {missing} sheet - is it the right file?")

for n, ref in NAMES.items():
    if n in wb.defined_names:
        del wb.defined_names[n]
    wb.defined_names.add(DefinedName(n, attr_text=ref))
print(f"1. defined names: {len(NAMES)} added")

q = wb["SOFR_OIS_Quotes"]
n = 0
for r in range(5, 40):
    if q[f"A{r}"].value is None or q[f"B{r}"].value is None:
        continue
    legs = [f'IF({MODE}="{T[i]}",INDEX(Testing!$B${a}:$B${b},'
            f'MATCH(C{r},Testing!$F${a}:$F${b},0)),' for i, (a, b) in enumerate(BLK)]
    inner = f'IF({MODE}="{FIXED}",J{r},IF(ISNUMBER(T{r}),T{r},J{r}))'
    c = q[f"H{r}"]
    c.value = "=IFERROR(" + "".join(legs) + inner + ")))" + f",J{r})"
    c.font = BLACK; c.number_format = "0.00000"; c.fill = OF; c.border = BOX
    n += 1
print(f"2. SOFR_OIS_Quotes!H: {n} rows repointed, matched on date")

ins = wb["Instructions"]
c = ins["B9"]
c.value = (f'=IF({MODE}="{T[0]}",{CD[0]},IF({MODE}="{T[1]}",{CD[1]},'
           f'IF({MODE}="{T[2]}",{CD[2]},DATE(2026,7,21))))')
c.number_format = "mm/dd/yyyy"; c.fill = OF; c.border = BOX; c.font = BLACK
print("3. Instructions!B9: VAL_DATE follows the active test case")

b = wb["Bootstrap"]
b.data_validations = DataValidationList()
dv = DataValidation(type="list",
                    formula1=f'"{LIVE},{FIXED},{T[0]},{T[1]},{T[2]}"', allow_blank=False)
b.add_data_validation(dv); dv.add(b["G4"])
print("4. Bootstrap!G4: dropdown has Live / Fixed / Test 1-3")

wb.calculation.fullCalcOnLoad = True
wb.save(PATH)
print(f"\nupgraded {PATH}   (backup at {PATH}.bak)")
print("From now on: right-click Testing > Move or Copy > into this workbook,")
print("delete the old Testing, rename the new one to exactly 'Testing'.")
