"""
Stop the dead SOFR_Futures sheet from firing 301 Bloomberg requests.

SOFR_Futures holds 301 of the workbook's 434 BDP/BDS calls - roughly 70% - and
NOTHING reads it. Its only consumer was Bootstrap_Lehman, which was deleted when
we went to a single bootstrap. Every one of those requests hits the terminal,
can fail, and renders as an error, while feeding nothing.

Adds a master switch at SOFR_Futures!B3, default "No", gating every pull on the
sheet. The formulas are preserved verbatim inside the IF, so flipping it to
"Yes" restores the sheet exactly.

This does not touch the curve: Bootstrap reads SOFR_OIS_Quotes!H only.
"""
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
YF = PatternFill("solid", fgColor="FFFF00")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

wb = load_workbook(WB)
ws = wb["SOFR_Futures"]

def setv(cell, val, font=None, fill=None):
    try: ws[cell].value = val
    except AttributeError: return False
    if font: ws[cell].font = font
    if fill: ws[cell].fill = fill
    return True

setv("A3", "Enable Bloomberg pulls on this sheet", BOLD)
setv("B3", "No", BLUE, YF)
ws["B3"].border = BOX
dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws.add_data_validation(dv); dv.add(ws["B3"])
setv("C3", "This sheet holds 301 of the workbook's 434 Bloomberg requests and NOTHING "
           "reads it - its only consumer was the deleted Bootstrap_Lehman. Left on, it "
           "hammers the terminal and fills the sheet with errors for no benefit. Set to "
           "Yes to re-enable; the formulas are preserved unchanged.", NOTE)

n = 0
for row in ws.iter_rows():
    for c in row:
        v = c.value
        if not isinstance(v, str) or not v.startswith("="): continue
        if not re.search(r'\b(BDP|BDS|BDH)\s*\(', v): continue
        if "$B$3" in v: continue                      # already gated
        try: c.value = f'=IF($B$3<>"Yes","",{v[1:]})'
        except AttributeError: continue
        n += 1

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"gated {n} Bloomberg calls on SOFR_Futures behind B3 (default No)")
