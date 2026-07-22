"""
Reorganise into: A) data pulls  B) bootstrap working  C) curve interface + swaps
D) CDS. Also folds S490_Target into the existing Bloomberg_S490_Validation
(duplicate purpose) and hides the two bisection working sheets.

Result: 19 sheets, 17 visible - the original 18 minus the deleted Bootstrap_Lehman.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
SECT = Font(name="Calibri", size=11, bold=True)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
YF = PatternFill("solid", fgColor="FFFF00")
SF = PatternFill("solid", fgColor="D9E1F2")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

GROUPS = [
    ("META",      "808080", ["Instructions", "Conventions"]),
    ("A_DATA",    "C00000", ["SOFR_Fixings", "SOFR_OIS_Quotes", "SOFR_Futures"]),
    ("B_BOOTSTRAP","1F3864", ["Bootstrap", "Curve_Solver", "Fwd_Interp"]),
    ("C_CURVE",   "2E7D32", ["Curve_Interface", "Swap_Pricer",
                             "Bloomberg_S490_Validation", "Charts"]),
    ("D_CDS",     "6A1B9A", ["CDS_Parameters", "CDS_Quotes", "Hazard_Bootstrap",
                             "Hazard_Solver", "CDS_Schedule", "CDS_Pricer",
                             "CDS_Validation"]),
]
HIDE = ["Curve_Solver", "Hazard_Solver"]


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None      # merged
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


def fold_target(wb):
    """Move the S490 screen snapshot into Bloomberg_S490_Validation, then drop it."""
    src, dst = wb["S490_Target"], wb["Bloomberg_S490_Validation"]
    rows = [(src[f"A{r}"].value, src[f"B{r}"].value, src[f"C{r}"].value, src[f"D{r}"].value)
            for r in range(9, 41) if src[f"A{r}"].value is not None]

    put(dst, "V6", "S490 SCREEN SNAPSHOT (curve date 07/21/2026)", SECT)
    put(dst, "V7", "The Bloomberg Curve Analysis column, transcribed from the screen. "
                   "Columns G/M below look up this block by maturity date, so it stays "
                   "correct if rows move. Only valid while VAL_DATE = 07/21/2026.", NOTE)
    for i, h in enumerate(["Date", "Market rate", "Zero (screen)", "DF (screen)"]):
        put(dst, f"{chr(86+i)}8", h, SECT, None, SF, True)
    for i, (d, mkt, z, df) in enumerate(rows):
        r = 9 + i
        put(dst, f"V{r}", d, BLUE, "mm/dd/yy", YF, True)
        put(dst, f"W{r}", mkt, BLUE, "0.00000", YF, True)
        put(dst, f"X{r}", z, BLUE, "0.00000", YF, True)
        put(dst, f"Y{r}", df, BLUE, "0.000000", YF, True)
    lo, hi = 9, 9 + len(rows) - 1
    D, Z, F = f"$V${lo}:$V${hi}", f"$X${lo}:$X${hi}", f"$Y${lo}:$Y${hi}"

    put(dst, "M7", "DF (screen)", SECT, None, SF, True)
    put(dst, "N7", "DF (model)", SECT, None, SF, True)
    put(dst, "O7", "DF diff", SECT, None, SF, True)
    put(dst, "G7", "S490 zero (%) — screen snapshot", SECT, None, SF, True)
    for r in range(8, 73):
        m = f"MATCH(B{r},{D},0)"
        put(dst, f"G{r}", f'=IFERROR(INDEX({Z},{m}),"")', BLACK, "0.00000", None, True)
        put(dst, f"M{r}", f'=IFERROR(INDEX({F},{m}),"")', BLACK, "0.000000", None, True)
        put(dst, f"N{r}", f'=IF(M{r}="","",Bootstrap!H{r})', BLACK, "0.000000", None, True)
        put(dst, f"O{r}", f'=IF(M{r}="","",N{r}-M{r})', BLACK, "0.00E+00", None, True)

    put(dst, "A74", "Pillars matched to the screen", SECT)
    put(dst, "C74", f'=COUNT(M8:M72)&" / {len(rows)}"', BLACK, None, OF, True)
    put(dst, "A75", "Max |zero diff| vs screen (bp)", SECT)
    put(dst, "C75", '=MAX(MAX(E8:E72),-MIN(E8:E72))', BLACK, "0.000", OF, True)
    put(dst, "A76", "Max |DF diff| vs screen", SECT)
    put(dst, "C76", '=MAX(MAX(O8:O72),-MIN(O8:O72))', BLACK, "0.00E+00", OF, True)
    put(dst, "A78", "Column D = Bloomberg's own curve pulled live via BDS. Column G/M = "
                    "the 07/21/26 screen snapshot in V:Y. Column E compares your zero to "
                    "whichever is present.", NOTE)
    # column E: prefer the live BDS pull, fall back to the screen snapshot
    for r in range(8, 73):
        put(dst, f"E{r}",
            f'=IF(ISNUMBER(D{r}),(C{r}-D{r})*100,IF(ISNUMBER(G{r}),(C{r}-G{r})*100,""))',
            BLACK, "0.00", None, True)
    del wb["S490_Target"]


def main():
    wb = load_workbook(WB)
    fold_target(wb)
    order = []
    for _, colour, names in GROUPS:
        for n in names:
            if n in wb.sheetnames:
                wb[n].sheet_properties.tabColor = colour
                order.append(n)
    leftover = [s for s in wb.sheetnames if s not in order]
    wb._sheets = [wb[n] for n in order + leftover]
    for n in HIDE:
        if n in wb.sheetnames:
            wb[n].sheet_state = "hidden"
    wb.active = 0
    wb.calculation.fullCalcOnLoad = True
    wb.save(WB)
    vis = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
    print(f"{len(wb.sheetnames)} sheets, {len(vis)} visible, leftover={leftover}")
    for _, _, names in GROUPS:
        print("  ", " | ".join(n + ("(hidden)" if n in HIDE else "")
                               for n in names if n in wb.sheetnames))


if __name__ == "__main__":
    main()
