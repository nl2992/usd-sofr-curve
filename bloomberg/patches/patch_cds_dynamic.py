"""
Make the CDS spreads dynamic, completing the Bloomberg pickup chain.

The discount curve is already dynamic (BDP mid -> BDP last -> manual bid/ask).
The CDS side was not: CDS_Quotes column D (CDS ticker) was empty for every tenor,
so column F always fell through to the manual demo spreads.

Wires in the documented two-step pull:

  Step 1  ticker per tenor   =BDP(<entity ticker>,"CDS_SPREAD_TICKER_nY")
  Step 2  quote that ticker  =BDP(<returned ticker>&" BEST Curncy","PX_MID")

falling back to the manual spread when either leg fails, so the module still runs
off-terminal. Adds the entity-ticker input to CDS_Parameters.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"

FONT = "Calibri"
BLUE = Font(name=FONT, size=11, color="0000FF")
BLACK = Font(name=FONT, size=11)
SECT = Font(name=FONT, size=11, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="666666")
WARN = Font(name=FONT, size=10, bold=True, color="C00000")
YFILL = PatternFill("solid", fgColor="FFFF00")
SFILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TENORS = ["1Y", "2Y", "3Y", "5Y", "7Y", "10Y"]
TKR = "CDS_Parameters!$B$19"


def put(ws, cell, val, font=BLACK, fmt=None, fill=None, border=False):
    c = ws[cell]
    c.value = val
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if border: c.border = BOX
    return c


def main():
    wb = load_workbook(WB)

    # ---- entity ticker input
    p = wb["CDS_Parameters"]
    put(p, "A19", "Reference entity ticker", SECT)
    put(p, "B19", "HSBA LN Equity", BLUE, None, YFILL, True)
    put(p, "C19", 'Drives the CDS_SPREAD_TICKER_nY lookups on CDS_Quotes. Any '
                  'Bloomberg identifier whose CDS curve you want (equity ticker, '
                  'or the entity itself).', NOTE)

    # ---- two-step pull on CDS_Quotes
    q = wb["CDS_Quotes"]
    put(q, "D6", "CDS ticker (step 1, live)", SECT, None, SFILL, True)
    put(q, "H6", "Quote source in use", SECT, None, SFILL, True)
    for i, t in enumerate(TENORS):
        r = 7 + i
        # step 1: resolve the tenor's CDS ticker from the entity ticker
        put(q, f"D{r}", f'=IFERROR(BDP({TKR},"CDS_SPREAD_TICKER_{t}"),"")',
            BLACK, None, None, True)
        # step 2: quote it; fall back to the manual spread
        put(q, f"F{r}",
            f'=IFERROR(BDP(D{r}&" BEST Curncy","PX_MID")+0,'
            f'IFERROR(BDP(D{r},"PX_LAST")+0,E{r}))', BLACK, "0.0000", None, True)
        put(q, f"H{r}",
            f'=IF(ISNUMBER(IFERROR(BDP(D{r}&" BEST Curncy","PX_MID")+0,"")),'
            f'"BDP live","MANUAL (demo)")', BLACK, None, None, True)

    put(q, "A2", "Single-name CDS par spreads. Step 1 (col D) resolves each tenor's CDS "
                 "ticker from the entity ticker in CDS_Parameters!B19; step 2 (col F) "
                 "quotes it. Manual spreads in col E are used only when BDP fails.", NOTE)
    put(q, "A14", "WARNING: the manual spreads in column E (40/55/70/95/115/135) are "
                  "DEMO PLACEHOLDERS, not market data — unlike the SOFR OIS quotes, which "
                  "are the real 07/21/26 S490 levels. Column H tells you which source each "
                  "tenor is actually using. Replace column E, or connect a terminal, before "
                  "trusting any CDS output.", WARN)

    wb.calculation.fullCalcOnLoad = True
    wb.save(WB)
    print("CDS quotes wired to the two-step BDP pull; entity ticker at CDS_Parameters!B19")


if __name__ == "__main__":
    main()
