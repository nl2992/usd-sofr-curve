"""
Business-day roll the displayed maturity dates on SOFR_OIS_Quotes.

Column C parsed the tenor label into a raw date with no roll, so it showed
1M = 08/23/2026, 6M = 01/23/2027, 18M = 01/23/2028, 7Y = 07/23/2033 - all
weekend dates. S490 shows 08/24, 01/25, 01/24, 07/25.

This is DISPLAY ONLY: a reference scan shows only column H (the mid rate) is
read by other sheets (49 refs, all from Bootstrap), and Bootstrap computes its
own rolled dates. So the curve was never affected. But the sheet was showing
maturities that disagreed with both Bloomberg and the curve built from it, which
is a trap during reconciliation.

Same modified-following convention as everywhere else: weekends only, roll back
if the roll crosses a month end.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"


def mf(e):
    nxt = f"({e}+IF(WEEKDAY({e},2)=6,2,IF(WEEKDAY({e},2)=7,1,0)))"
    prv = f"({e}-IF(WEEKDAY({e},2)=6,1,IF(WEEKDAY({e},2)=7,2,0)))"
    return f"IF(MONTH({nxt})<>MONTH({e}),{prv},{nxt})"


wb = load_workbook(WB)
ws = wb["SOFR_OIS_Quotes"]
n = 0
for r in range(5, 40):
    v = ws[f"C{r}"].value
    if not (isinstance(v, str) and v.startswith("=") and "WEEKDAY" not in v):
        continue
    ws[f"C{r}"] = "=" + mf(v[1:])
    ws[f"C{r}"].number_format = "mm/dd/yyyy"
    n += 1
ws["M4"] = ("Maturity dates are modified-following business-day adjusted, matching S490 and "
            "the Bootstrap sheet. Display only - the curve reads column H (mid) alone.")
ws["M4"].font = Font(name="Calibri", size=9, italic=True, color="666666")
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"rolled {n} displayed maturity dates")
