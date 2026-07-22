"""
CDS_Entities: an intake sheet for the names being pulled at close.

One row per name. Holds the term structure (T3) and the CDSW screen outputs (T4)
side by side, plus a dropdown on CDS_Parameters selecting which row prices.
CDS_Quotes!E and the recovery rate then follow the selection, so swapping name is
one cell rather than retyping six spreads.

A comparison block diffs the model against the captured CDSW figures for whichever
name is selected - that is the acceptance test for the CDS module.

Row 5 is pre-filled with the CINDBK reference already on file, to show the
expected format. Only its 5Y spread and its CDSW outputs are known; the rest of
its term structure is blank because we never had it.
"""
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
H1 = Font(name="Calibri", size=12, bold=True)
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
HF = PatternFill("solid", fgColor="1F3864")
GF = PatternFill("solid", fgColor="2E7D32")
PF = PatternFill("solid", fgColor="6A1B9A")
YF = PatternFill("solid", fgColor="FFFF00")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
N4, CCY, DTF = "0.0000", "#,##0.00", "mm/dd/yyyy"

# (col, header, width, fmt, band)  band: I=identity  S=spreads  C=CDSW capture
COLS = [
 ("A","Entity name",26,None,"I"),("B","Ticker (BDP)",20,None,"I"),
 ("C","Ccy",7,None,"I"),("D","Seniority",13,None,"I"),("E","Clause",10,None,"I"),
 ("F","Recovery",10,"0.00%","I"),("G","Captured",13,DTF,"I"),
 ("H","1Y bp",10,N4,"S"),("I","2Y bp",10,N4,"S"),("J","3Y bp",10,N4,"S"),
 ("K","5Y bp",10,N4,"S"),("L","7Y bp",10,N4,"S"),("M","10Y bp",10,N4,"S"),
 ("O","Notional",14,CCY,"C"),("P","Coupon bp",11,N4,"C"),("Q","Maturity",13,DTF,"C"),
 ("R","Traded sprd bp",14,N4,"C"),("S","Pts upfront",13,"0.00000000","C"),
 ("T","Price",13,"0.00000000","C"),("U","Principal",14,CCY,"C"),
 ("V","Accr days",11,"0","C"),("W","Accrued",13,CCY,"C"),("X","Cash amount",14,CCY,"C"),
 ("Y","Spread DV01",13,CCY,"C"),("Z","IR DV01",12,CCY,"C"),("AA","Rec risk 1%",12,CCY,"C"),
 ("AB","Def exposure",14,CCY,"C"),("AC","Prob 5Y",10,"0.0000","C"),
]
R0, R1 = 5, 14


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False, al=None):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    if al: c.alignment = Alignment(horizontal=al, wrap_text=True, vertical="center")
    return c


wb = load_workbook(WB)
if "CDS_Entities" in wb.sheetnames:
    del wb["CDS_Entities"]
ws = wb.create_sheet("CDS_Entities", wb.sheetnames.index("CDS_Parameters"))
ws.sheet_properties.tabColor = "6A1B9A"

put(ws, "A1", "CDS_Entities — names pulled from Bloomberg", H1)
put(ws, "A2", "One row per name. Blue/yellow cells are typed in from the terminal. "
              "CDS_Parameters!B27 selects which row prices.", NOTE)
put(ws, "A3", "Spreads H:M are T3 (term structure). O:AC are T4 (the CDSW screen for that "
              "name) and are the acceptance test — see TEST_CASES.md.", NOTE)

for col, hdr, w, fmt, band in COLS:
    ws.column_dimensions[col].width = w
    fill = {"I": HF, "S": GF, "C": PF}[band]
    put(ws, f"{col}4", hdr, HDR, None, fill, True, "center")
    for r in range(R0, R1 + 1):
        put(ws, f"{col}{r}", None, BLUE, fmt, YF, True)
ws.row_dimensions[4].height = 30
put(ws, "H3", "TERM STRUCTURE (T3)", BOLD)
put(ws, "O3", "CDSW SCREEN CAPTURE (T4)", BOLD)

# worked example: the reference already on file
ex = {"A":"China CITIC Bank Intl","B":"","C":"USD","D":"Senior","E":"CR14","F":0.40,
      "G":dt.datetime(2026,7,21),"K":55.1514,"O":10_000_000,"P":100.0,
      "Q":dt.datetime(2031,6,20),"R":55.1514,"S":-1.97596265,"T":101.97596265,
      "U":-197597,"V":30,"W":-8333,"X":-205930,"Y":4482.41,"Z":47.71,"AA":72.00,
      "AB":6197596,"AC":0.0446}
for c, v in ex.items():
    put(ws, f"{c}{R0}", v, BLUE, None, YF, True)
put(ws, "A16", "Row 5 is the reference screen already on file. Only its 5Y spread is known — "
               "the rest of its term structure was never captured, which is exactly what T3 is "
               "for. Its CDSW outputs ARE known and are the target to hit.", WARN)

# ---- selector on CDS_Parameters, feeding the model
p = wb["CDS_Parameters"]
put(p, "A27", "Active entity", BOLD)
put(p, "B27", "China CITIC Bank Intl", BLUE, None, YF, True)
dv = DataValidation(type="list", formula1=f"=CDS_Entities!$A${R0}:$A${R1}", allow_blank=True)
p.add_data_validation(dv); dv.add(p["B27"])
put(p, "C27", "Picks a row on CDS_Entities. Spreads, recovery and the CDSW targets all follow.", NOTE)
M = f'MATCH($B$27,CDS_Entities!$A${R0}:$A${R1},0)'
put(p, "A28", "row on CDS_Entities", BLACK)
put(p, "B28", f'=IFERROR({M},0)', BLACK, "0", OF, True)
put(p, "A29", "Recovery from that row", BLACK)
put(p, "B29", f'=IFERROR(INDEX(CDS_Entities!$F${R0}:$F${R1},$B$28),"")', BLACK, "0.00%", OF, True)
put(p, "C29", "Type it into B8 to use it; left manual so a mis-selection cannot silently "
              "reprice the book.", NOTE)

# ---- CDS_Quotes reads the selected row's spreads
q = wb["CDS_Quotes"]
SPREAD_COL = {"1Y":"H","2Y":"I","3Y":"J","5Y":"K","7Y":"L","10Y":"M"}
for i, ten in enumerate(["1Y","2Y","3Y","5Y","7Y","10Y"]):
    r = 7 + i
    c = SPREAD_COL[ten]
    put(q, f"E{r}",
        f'=IFERROR(INDEX(CDS_Entities!${c}${R0}:${c}${R1},CDS_Parameters!$B$28),"")',
        BLUE, N4, YF, True)
put(q, "A19", "Col E now reads the active entity's row on CDS_Entities. BDP still takes "
              "precedence in col F when live.", NOTE)

# ---- model vs captured CDSW, for the selected name
c = wb["CDS_Pricer"]
put(c, "A41", "MODEL vs CAPTURED CDSW  (active entity)", BOLD)
put(c, "A42", "Blank until the CDSW capture is typed into CDS_Entities. This is the "
              "acceptance test for the module.", NOTE)
for i, h in enumerate(["Measure", "CDSW captured", "Model", "Diff"]):
    put(c, f"{chr(65+i)}43", h, HDR, None, HF, True, "center")
CHK = [("Points upfront","S","=B31"),("Price","T","=B32"),("Principal","U","=B33"),
       ("Accrued days","V","=B34"),("Accrued","W","=B35"),("Cash amount","X","=B36"),
       ("Def exposure","AB","=B37")]
for i,(lab,col,mdl) in enumerate(CHK):
    r = 44 + i
    put(c, f"A{r}", lab, BLACK)
    put(c, f"B{r}", f'=IFERROR(INDEX(CDS_Entities!${col}${R0}:${col}${R1},CDS_Parameters!$B$28),"")',
        BLACK, None, OF, True)
    put(c, f"C{r}", mdl, BLACK, None, None, True)
    put(c, f"D{r}", f'=IF(OR(B{r}="",NOT(ISNUMBER(B{r}))),"",C{r}-B{r})', BLACK, None, None, True)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"CDS_Entities created, {R1-R0+1} slots; selector at CDS_Parameters!B27; "
      f"acceptance block at CDS_Pricer!A41")
