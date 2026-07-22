"""
Point CDS_Pricer's Curve_Interface at the curve workbook instead of a snapshot.

Nigel wants the live BDP-driven curve rather than a pasted grid. Link form is
'[USD_SOFR_Curve_Bloomberg_Pricer.xlsx]Curve_Interface'!$K$n, which Excel
resolves by filename while both books are open and rewrites to a full path on
save.

Row mapping is NOT one-for-one. In the master, row 8 is the DF=1 anchor and the
first tenor sits at row 10; in this workbook the anchor is row 8 and the first
tenor is row 9. So:

    dest K8 /L8        <-  source K8 /L8      (anchor, DF = 1 at VAL_DATE)
    dest K9:L73        <-  source K10:L74     (the 65 pillars, 1W to 50Y)

Linking row-for-row instead would have dropped the 50Y pillar and left a blank
at row 9. The check block below reports the anchor DF, the first tenor date and
the pillar count, so a wrong offset shows up as a number rather than as a
pricing error.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB="bloomberg/CDS_Pricer.xlsx"
SRC="USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
LINK=f"'[{SRC}]Curve_Interface'"
R0,R1=8,73

F="Calibri"
SM=Font(name=F,size=9,color="808080"); B=Font(name=F,size=11,bold=True)
RED=Font(name=F,size=10,bold=True,color="C00000"); GRN=Font(name=F,size=11,color="008000")
LNK=Font(name=F,size=11,color="FF0000")          # red = link to another file
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(WB)
ci=wb["Curve_Interface"]

# anchor
ci.cell(8,11,f"={LINK}!$K$8").number_format="mm/dd/yyyy"
ci.cell(8,12,f"={LINK}!$L$8").number_format="0.00000000"
ci.cell(8,11).font=LNK; ci.cell(8,12).font=LNK
# 65 pillars, offset by one
n=0
for i in range(65):
    d=9+i; srcr=10+i
    ci.cell(d,11,f"={LINK}!$K${srcr}").number_format="mm/dd/yyyy"
    ci.cell(d,12,f"={LINK}!$L${srcr}").number_format="0.00000000"
    ci.cell(d,11).font=LNK; ci.cell(d,12).font=LNK
    n+=1
ci["D7"]=f"={LINK}!$D$7"; ci["D7"].number_format="mm/dd/yyyy"; ci["D7"].font=LNK
print(f"anchor + {n} pillars linked to {SRC} (source K8, then K10:K74)")

ci["K6"]="Curve grid - LIVE LINK to the curve workbook"
ci["M6"]=f"source: [{SRC}]Curve_Interface!K8 (anchor) + K10:L74 (65 pillars)"; ci["M6"].font=SM

# check block: a one-row slip or a stale link has to be visible, not inferred
ci["N8"]="LINK CHECK"; ci["N8"].font=B
rows=[("pillars linked",   f'=COUNT($L${R0}:$L${R1})'),
      ("first date",       f'=IF(COUNT($K${R0}:$K${R1})=0,"",MIN($K${R0}:$K${R1}))'),
      ("last date",        f'=IF(COUNT($K${R0}:$K${R1})=0,"",MAX($K${R0}:$K${R1}))'),
      ("anchor DF (want 1)",f'=IF($L$8="","",$L$8)'),
      ("first tenor date", f'=IF($K$9="","",$K$9)'),
      ("DF monotone?",     f'=IF(COUNT($L${R0}:$L${R1})<2,"",'
                           f'IF(SUMPRODUCT(--($L${R0+1}:$L${R1}>$L${R0}:$L${R1-1}))=0,"ok","NOT DECREASING"))')]
for i,(lab,f_) in enumerate(rows):
    r=9+i
    ci.cell(r,14,lab).font=Font(name=F,size=10)
    c=ci.cell(r,15,f_); c.border=BOX; c.font=GRN
    if "date" in lab: c.number_format="mm/dd/yyyy"
    if "DF (want" in lab: c.number_format="0.00000000"
ci["N15"]=(f'=IF(O9=0,"LINK NOT RESOLVING - open {SRC}",'
           f'IF(ABS(O12-1)>0.000001,"ROW 8 IS NOT THE ANCHOR - check the offset",'
           f'IF(O13-VAL_DATE>14,"FIRST TENOR LOOKS WRONG - expected about a week after the curve date",'
           f'IF(O9<60,"FEWER PILLARS THAN EXPECTED - check the source range","link ok"))))')
ci["N15"].font=RED
ci["N17"]=("Live only while the curve workbook is OPEN. Closed, Excel serves the last cached values and the "
           "BDP pulls in the source do not refresh.")
ci["N17"].font=SM
ci["N18"]="Data > Edit Links > Change Source if the file moves."
ci["N17"].font=SM

wb.calculation.fullCalcOnLoad=True
wb.save(WB)
print("check block added at N8:O13, status at N14")
