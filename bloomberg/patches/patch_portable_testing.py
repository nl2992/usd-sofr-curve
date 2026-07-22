"""
Make Testing portable, so it can be moved into the master with Move or Copy.

Copying a sheet between workbooks makes Excel rewrite every direct cross-sheet
reference as a link back to the source file:

    Bootstrap!$B$8:$B$72   ->   '[Update.xlsx]Bootstrap'!$B$8:$B$72

Defined names do not behave that way. If the destination already defines the
name, the copied sheet binds to the DESTINATION's definition. So Testing now
speaks only in names, and moving it is safe:

    BOOT_MODE    Bootstrap!$G$4       BOOT_DFSPOT  Bootstrap!$D$4
    BOOT_DATES   Bootstrap!$B$8:$B$72 BOOT_S       Bootstrap!$E$8:$E$72
    BOOT_TAU0    Bootstrap!$D$8:$D$72 BOOT_TAUC    Bootstrap!$F$8:$F$72
    BOOT_ANN     Bootstrap!$G$8:$G$72 BOOT_DF      Bootstrap!$H$8:$H$72
    BOOT_T       Bootstrap!$C$8:$C$72 BOOT_ZERO    Bootstrap!$J$8:$J$72

Non-volatile, unlike INDIRECT, which is the other way to do this and would put
~960 volatile calls on the sheet.

The destination must define these names once. patch_upgrade_master.py does that.
"""
import re
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
NAMES = {
    "BOOT_MODE":   "Bootstrap!$G$4",
    "BOOT_DFSPOT": "Bootstrap!$D$4",
    "BOOT_DATES":  "Bootstrap!$B$8:$B$72",
    "BOOT_S":      "Bootstrap!$E$8:$E$72",
    "BOOT_TAU0":   "Bootstrap!$D$8:$D$72",
    "BOOT_TAUC":   "Bootstrap!$F$8:$F$72",
    "BOOT_ANN":    "Bootstrap!$G$8:$G$72",
    "BOOT_DF":     "Bootstrap!$H$8:$H$72",
    "BOOT_T":      "Bootstrap!$C$8:$C$72",
    "BOOT_ZERO":   "Bootstrap!$J$8:$J$72",
}
# longest first so BOOT_DF does not eat BOOT_DFSPOT's target
SUBS = sorted(((v, k) for k, v in NAMES.items()), key=lambda x: -len(x[0]))

wb = load_workbook(WB)
for name, ref in NAMES.items():
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=ref))

ws = wb["Testing"]
n = 0
for row in ws.iter_rows():
    for c in row:
        v = c.value
        if not isinstance(v, str) or not v.startswith("="):
            continue
        new = v
        for ref, name in SUBS:
            new = new.replace(ref, name)
        if new != v:
            c.value = new
            n += 1

left = sum(1 for row in ws.iter_rows() for c in row
           if isinstance(c.value, str) and "Bootstrap!" in c.value)
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"Testing: {n} formulas moved onto defined names; direct Bootstrap! refs remaining: {left}")
