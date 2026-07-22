"""
Put the governing equations beside the working columns on each sheet.

Cites the Bloomberg B-Model white paper (3.x) and O'Kane-Turnbull (n) so anyone
reading a column can see which equation it implements without leaving the sheet.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as CL

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
H = Font(name="Calibri", size=11, bold=True, color="1F3864")
EQ = Font(name="Consolas", size=10)
TX = Font(name="Calibri", size=10)
SM = Font(name="Calibri", size=9, italic=True, color="666666")
BAND = PatternFill("solid", fgColor="D9E1F2")

CARDS = {
 "CDS_Schedule": ("Q", 3, 78, [
  ("H", "WHAT EACH COLUMN COMPUTES"),
  ("s", "Z(.) is the bootstrapped SOFR curve via Curve_Interface (B-Model p.6)."),
  ("", ""),
  ("t", "col I   hazard h_i, piecewise-constant, picked by pay date"),
  ("e", "col K   Q(t_i) = Q(t_i-1) * exp(-h_i * dt_i)                B-Model p.5"),
  ("e", "col L   dPD    = Q(t_i-1) - Q(t_i)"),
  ("", ""),
  ("t", "col M   premium factor                          (3.4) / O-T (5) 1st term"),
  ("e", "        alpha_i * Z(t_i) * Q(t_i)"),
  ("e", "        Coupons-in-Survival = A*C*SUM alpha_i Z(t_i) Q(t_i)"),
  ("", ""),
  ("t", "col N   accrual-on-default factor               (3.5) / O-T (5) 2nd term"),
  ("e", "        (alpha_i/2) * Z(mid_i) * dPD_i"),
  ("e", "        = A*C*INT Delta(T_n(t),t) (-dp_s/dt) D(t) dt"),
  ("s", "        Z at the MIDPOINT, not the pay date: (3.5) integrates D(t) at"),
  ("s", "        the default time. Z(mid) is 140x closer to it than O-T's Z(t_n)."),
  ("s", "        See CLAUDE.md D5."),
  ("", ""),
  ("t", "col O   protection factor                       (3.6) / O-T (7)"),
  ("e", "        Z(mid_i) * dPD_i"),
  ("e", "        Protection Leg = A(1-R) * INT (-dp_s/dt) D(t) dt"),
 ]),
 "Hazard_Solver": ("V", 3, 80, [
  ("H", "WHAT THIS SHEET SOLVES"),
  ("t", "For each tenor k, find h_k so the model par spread equals the quote."),
  ("", ""),
  ("e", "(3.3)  S = Protection / ( Premium - AccruedInterest*D(T_s) )  at C=1bp"),
  ("", ""),
  ("t", "as a root:"),
  ("e", "  f(h) = (1-R)*ProtCum(h) - S*( RPV01Cum(h) - D0*D(T_s) )"),
  ("e", "  ProtCum  = ProtFixed  + SUM Z(mid)*dPD          over this segment"),
  ("e", "  RPV01Cum = RPV01Fixed + SUM alpha*Z(end)*Q"),
  ("e", "                        + (1/2) SUM alpha*Z(mid)*dPD"),
  ("", ""),
  ("s", "ProtFixed / RPV01Fixed carry in from the shorter tenors, already solved."),
  ("s", "That carry IS the bootstrap: only h_k is unknown on each block."),
  ("", ""),
  ("e", "survival  Q(t) = Q(t_k-1) * exp(-h_k*(t - t_k-1))          B-Model p.5"),
  ("t", "method    1-D root search per maturity, earlier hazards held"),
  ("s", "          B-Model S4 p.8 ; O-T S9, footnote 8 names bisection"),
  ("t", "bracket   [0,3], 30 halvings -> ~3e-9.  h >= 0 is hard (p.8)"),
 ]),
 "CDS_Pricer": ("J", 3, 78, [
  ("H", "PRICING EQUATIONS"),
  ("e", "(3.1)  Market Value = Protection Leg - Premium Leg"),
  ("s", "       both discounted to the pricing date T"),
  ("e", "(3.2)  Upfront      = Market Value / D(T_s) + Accrued Interest"),
  ("e", "(3.3)  Par spread S = Protection / (Premium - AI*D(T_s))   at C=1bp"),
  ("", ""),
  ("e", "O-T (1b)  MTM = +/- [ S(t_V,t_N) - S(t_0,t_N) ] * RPV01"),
  ("s", "          equivalent form, + for long protection"),
  ("", ""),
  ("e", "Accrued Interest = A * C * D0,   D0 = accrued days / 360"),
  ("s", "accrued days include the pricing date (p.7)"),
  ("s", "T_s = T + 3 business days (p.6)"),
  ("", ""),
  ("s", "Upfront is identically 0 when C = S. That is the check that accrued"),
  ("s", "interest enters with a + sign, which the wording alone does not settle."),
 ]),
 "CDS_Quotes": ("K", 3, 82, [
  ("H", "BLOOMBERG HELP DESK H#1330731572"),
  ("t", "The documented two-step pull, implemented verbatim:"),
  ("e", "  step 1  col D = BDP( <entity>, \"CDS_SPREAD_TICKER_nY\" )   -> ticker"),
  ("e", "  step 2  col F = BDP( <ticker> & \" BEST Curncy\", \"PX_LAST\" )"),
  ("", ""),
  ("s", "Their tenors are 1Y 3Y 5Y 7Y 10Y. The 2Y row is ours -"),
  ("s", "CDS_SPREAD_TICKER_2Y may not be a valid field; that row is flagged."),
  ("", ""),
  ("t", "step 3  they suggest Hazard = S/(1-R), shown in col G."),
  ("s", "        Exact only for a flat curve. We bootstrap instead, which their"),
  ("s", "        own note says to do:"),
  ("s", "        \"Bloomberg does not provide a direct zero-coupon CDS spread"),
  ("s", "         output by maturity as an exportable field... you would"),
  ("s", "         bootstrap the hazard-rate curve using the par spreads and an"),
  ("s", "         assumed recovery rate.\""),
 ]),
}

wb = load_workbook(WB)
for sheet, (col, row0, width, lines) in CARDS.items():
    ws = wb[sheet]
    ws.column_dimensions[col].width = width
    r = row0
    for kind, text in lines:
        cell = f"{col}{r}"
        try:
            ws[cell].value = text or None
        except AttributeError:
            r += 1; continue
        if kind == "H":
            ws[cell].font = H
            for c in range(ws[cell].column, ws[cell].column + 1):
                ws.cell(row=r, column=c).fill = BAND
        elif kind == "e":
            ws[cell].font = EQ
        elif kind == "s":
            ws[cell].font = SM
        else:
            ws[cell].font = TX
        ws[cell].alignment = Alignment(horizontal="left")
        r += 1

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("formula cards written to", ", ".join(CARDS))
