"""
Curve_Check - a standalone sheet. Move it in, paste four columns, read the diffs.

The Testing sheet needed too much to work: the Bootstrap!G4 dropdown on the right
case, ten BOOT_* defined names present in the destination, and a macro run once.
Miss any one and it silently shows nothing - which is exactly what happened
(F48 read "not active" because the dropdown was on Fixed).

This sheet references NOTHING outside itself. Every formula points at its own
A/B columns and its own curve date. It does its own bootstrap through
CurveVBA.SOFR_Curve, so there is no dropdown, no gate, no defined names, and no
dependency on the Bootstrap grid. Move or Copy cannot break it, because there is
nothing for Excel to rewrite into an external link.

One dependency, stated on the sheet: CurveVBA.bas must be imported and the file
saved as .xlsm. C6 says so in red until it is.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUT = "/Users/nigelli/Desktop/openusdcurve/bloomberg/Curve_Check.xlsx"
R0, R1 = 9, 60
SUM = 62
TEN, RAT, CD = f"$A${R0}:$A${R1}", f"$B${R0}:$B${R1}", "$C$4"

F = "Calibri"
TITLE = Font(name=F, size=14, bold=True)
NOTE = Font(name=F, size=9, italic=True, color="808080")
HDR = Font(name=F, size=11, bold=True, color="FFFFFF")
BOLD = Font(name=F, size=11, bold=True)
BLACK = Font(name=F, size=11)
GREEN = Font(name=F, size=11, color="008000")
BLUE = Font(name=F, size=11, color="0000FF")
RED = Font(name=F, size=11, bold=True, color="C00000")
HF = PatternFill("solid", fgColor="4472C4")
PASTE = PatternFill("solid", fgColor="FFFF00")
CALC = PatternFill("solid", fgColor="E2EFDA")
INP = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")] * 4)

wb = Workbook()
ws = wb.active
ws.title = "Curve_Check"

ws["A1"] = "Curve_Check — paste a capture, read the diffs"
ws["A1"].font = TITLE
ws["A2"] = ("Self-contained. Nothing on this sheet points at another sheet, so it survives "
            "Move or Copy into any workbook.")
ws["A3"] = ("Paste the YELLOW columns off the S490 capture: tenor, swap rate (mid), BBG zero %, "
            "BBG discount. Set the curve date. Everything green is computed.")
for a in ("A2", "A3"):
    ws[a].font = NOTE

ws["B4"] = "Curve date →"
ws["B4"].font = BOLD
ws["C4"] = "=TODAY()"
ws["C4"].fill = INP; ws["C4"].font = BLUE; ws["C4"].border = BOX
ws["C4"].number_format = "mm/dd/yyyy"
ws["D4"] = "← type the capture date here"
ws["D4"].font = NOTE

ws["B5"] = "Spot (T+2bd) →"
ws["B5"].font = BOLD
# roll a weekend curve date forward first, then add two business days
BASE = f'({CD}+IF(WEEKDAY({CD},2)=6,2,IF(WEEKDAY({CD},2)=7,1,0)))'
ws["C5"] = (f'=IF({CD}="","",{BASE}+IF(WEEKDAY({BASE},2)>=4,4,2))')
ws["C5"].number_format = "mm/dd/yyyy"; ws["C5"].font = GREEN; ws["C5"].border = BOX

ws["B6"] = "VBA →"
ws["B6"].font = BOLD
ws["C6"] = (f'=IF(ISERROR(SOFR_Curve({CD},{TEN},{RAT},"5Y","DF")),'
            f'"NOT AVAILABLE — import CurveVBA.bas (Alt+F11 > File > Import File), save as .xlsm",'
            f'"OK — CurveVBA.SOFR_Curve() is answering")')
ws["C6"].font = RED

HEAD = ["Tenor", "Swap rate (mid) %", "BBG zero %", "BBG discount",
        "Date", "our DF", "our zero %", "t ACT/365", "d zero bp", "d DF", "rate used %"]
for i, h in enumerate(HEAD, start=1):
    c = ws.cell(row=8, column=i, value=h)
    c.font = HDR; c.fill = HF; c.border = BOX
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for r in range(R0, R1 + 1):
    g = lambda o: f'SOFR_Curve({CD},{TEN},{RAT},$A{r},"{o}")'
    ws[f"E{r}"] = f'=IF($A{r}="","",IFERROR({g("DATE")},""))'
    ws[f"F{r}"] = f'=IF($A{r}="","",IFERROR({g("DF")},""))'
    ws[f"G{r}"] = f'=IF($A{r}="","",IFERROR({g("ZERO")},""))'
    ws[f"H{r}"] = f'=IF($A{r}="","",IFERROR({g("T")},""))'
    ws[f"K{r}"] = f'=IF($A{r}="","",IFERROR({g("PAR")},""))'
    ws[f"I{r}"] = f'=IF(OR(G{r}="",NOT(ISNUMBER(C{r}))),"",(G{r}-C{r})*100)'
    ws[f"J{r}"] = f'=IF(OR(F{r}="",NOT(ISNUMBER(D{r}))),"",F{r}-D{r})'
    for col in "ABCD":
        c = ws[f"{col}{r}"]; c.fill = PASTE; c.border = BOX; c.font = BLACK
    for col in "EFGHK":
        c = ws[f"{col}{r}"]; c.fill = CALC; c.border = BOX; c.font = GREEN
    for col in "IJ":
        ws[f"{col}{r}"].border = BOX; ws[f"{col}{r}"].font = BLACK
    for col, fmt in (("B", "0.00000"), ("C", "0.00000"), ("D", "0.000000"),
                     ("E", "mm/dd/yyyy"), ("F", "0.00000000"), ("G", "0.00000"),
                     ("H", "0.0000"), ("I", "0.000"), ("J", "0.00E+00"),
                     ("K", "0.00000")):
        ws[f"{col}{r}"].number_format = fmt

ws[f"A{SUM}"] = "pillars pasted"
ws[f"C{SUM}"] = f"=COUNTA(A{R0}:A{R1})"
ws[f"D{SUM}"] = "computed"
ws[f"F{SUM}"] = f"=COUNT(G{R0}:G{R1})"
ws[f"G{SUM}"] = "max |d zero| bp"
ws[f"I{SUM}"] = (f'=IF(COUNT(I{R0}:I{R1})=0,"",MAX(MAX(I{R0}:I{R1}),-MIN(I{R0}:I{R1})))')
ws[f"I{SUM}"].number_format = "0.000"
ws[f"J{SUM}"] = "max |d DF|"
ws[f"K{SUM}"] = (f'=IF(COUNT(J{R0}:J{R1})=0,"",MAX(MAX(J{R0}:J{R1}),-MIN(J{R0}:J{R1})))')
ws[f"K{SUM}"].number_format = "0.00E+00"
for a in (f"A{SUM}", f"D{SUM}", f"G{SUM}", f"J{SUM}"):
    ws[a].font = BOLD
for a in (f"C{SUM}", f"F{SUM}", f"I{SUM}", f"K{SUM}"):
    ws[a].font = BLACK; ws[a].border = BOX; ws[a].fill = CALC

ws[f"A{SUM+2}"] = ("If 'computed' is less than 'pillars pasted', a tenor label was not "
                   "understood — use 1W/3M/18M/5Y form. 12M and 1Y both work.")
ws[f"A{SUM+2}"].font = NOTE

# a diff worth looking at twice goes red
ws.conditional_formatting.add(f"I{R0}:I{R1}",
    CellIsRule(operator="greaterThan", formula=["1"], font=RED))
ws.conditional_formatting.add(f"I{R0}:I{R1}",
    CellIsRule(operator="lessThan", formula=["-1"], font=RED))

for col, w in (("A", 10), ("B", 17), ("C", 12), ("D", 13), ("E", 12), ("F", 14),
               ("G", 12), ("H", 11), ("I", 11), ("J", 12), ("K", 12)):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A9"
wb.calculation = wb.calculation
wb.save(OUT)
print("built", OUT)
