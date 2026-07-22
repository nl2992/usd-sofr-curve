"""
A sheet that bootstraps by itself. Paste four columns, read the answer. No links.

Every previous version pointed at the Bootstrap grid, because that is where the
engine lived - which meant defined names, or a dropdown, or a find-and-replace,
and a sheet that showed nothing when any of it was missing. So the engine goes
on the sheet.

Layout per block:
  A:D  pasted        tenor / swap mid % / BBG zero % / BBG discount
  E:I  answer        date, our DF, our zero %, d zero bp, d DF
  K:AM working       the bootstrap itself (hidden by default, unhide to audit)

Conventions are the ones that match Bloomberg, same as the cell grid:
  DFspot = 1 (settle is the curve date on S490)   <=1Y money market, ACT/360 from
  the curve date   >1Y par, ACT/360 from spot   18M priced but NOT a coupon date,
  so 2Y accrues 1Y->2Y   gap years log-linear in DF (step forward)   dates
  modified following, weekends only.

The seven long-end pillars (12/15/20/25/30/40/50Y) have interpolated gap years
behind them, so their DF has to be solved. Done here by fixed-point iteration
across 16 columns: DF = (1 - S*(A_anchor + gaps(DF))) / (1 + S*tau). The map is a
contraction with factor ~S*tau ~ 0.04, so 16 passes is far past machine precision
and it needs no iterative-calculation setting and no circular reference.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter as CL
from openpyxl.formatting.rule import CellIsRule

OUT = "/Users/nigelli/Desktop/openusdcurve/bloomberg/Bootstrap_Check.xlsx"

SHORT = ["1W", "2W", "3W"] + [f"{m}M" for m in range(1, 12)]
QUOTED_Y = [2,3,4,5,6,7,8,9,10,12,15,20,25,30,40,50]
GRID = SHORT + ["12M", "18M"] + [f"{y}Y" for y in range(2, 51)]
SEGMENTS = [(10,12),(12,15),(15,20),(20,25),(25,30),(30,40),(40,50)]
NITER = 16

F="Calibri"
TITLE=Font(name=F,size=14,bold=True); B=Font(name=F,size=11,bold=True)
NOTE=Font(name=F,size=9,italic=True,color="808080")
HDR=Font(name=F,size=11,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
GRN=Font(name=F,size=11,color="008000"); BLK=Font(name=F,size=11)
BLU=Font(name=F,size=11,color="0000FF"); RED=Font(name=F,size=11,bold=True,color="C00000")
PASTE=PatternFill("solid",fgColor="FFFF00"); CALC=PatternFill("solid",fgColor="E2EFDA")
INP=PatternFill("solid",fgColor="FFF2CC"); WRK=PatternFill("solid",fgColor="F2F2F2")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

def modfol(raw):
    return (f'IF(MONTH(({raw})+IF(WEEKDAY(({raw}),2)=6,2,IF(WEEKDAY(({raw}),2)=7,1,0)))'
            f'<>MONTH({raw}),({raw})-IF(WEEKDAY(({raw}),2)=6,1,IF(WEEKDAY(({raw}),2)=7,2,0)),'
            f'({raw})+IF(WEEKDAY(({raw}),2)=6,2,IF(WEEKDAY(({raw}),2)=7,1,0)))')

def build_block(ws, top, label):
    """one self-contained block; returns the summary row"""
    CD  = f"$C${top+2}"          # curve date
    SP  = f"$C${top+3}"          # spot
    P0, P1 = top+7, top+7+51     # paste rows
    G0 = top+7                   # grid rows (same start, own columns)
    G1 = G0+len(GRID)-1
    grow = {t: G0+i for i,t in enumerate(GRID)}

    ws.cell(top,1,label).font=TITLE
    ws.cell(top+1,1,"Paste YELLOW A:E straight off the capture — tenor, date, swap mid, BBG zero, "
                    "BBG discount. Set the curve date. Green is computed here; this sheet points at "
                    "nothing outside itself.").font=NOTE
    ws.cell(top+2,2,"Curve date →").font=B
    c=ws.cell(top+2,3,None); c.fill=INP; c.font=BLU; c.border=BOX; c.number_format="mm/dd/yyyy"
    ws.cell(top+3,2,"Spot (T+2bd) →").font=B
    base=f"({CD}+IF(WEEKDAY({CD},2)=6,2,IF(WEEKDAY({CD},2)=7,1,0)))"
    c=ws.cell(top+3,3,f'=IF({CD}="","",{base}+IF(WEEKDAY({base},2)>=4,4,2))')
    c.font=GRN; c.border=BOX; c.number_format="mm/dd/yyyy"

    for i,h in enumerate(["Tenor","Date","Swap rate (mid) %","BBG zero %","BBG discount",
                          "Our zero %","Our discount","d zero bp","d DF",
                          "date check"],start=1):
        c=ws.cell(top+6,i,h); c.font=HDR; c.fill=HF; c.border=BOX
        c.alignment=Alignment(horizontal="center",wrap_text=True)

    # ---------------- working grid ----------------
    for i,h in enumerate(["tenor","yrs","date","T","tau0","tauC","sched","quoted",
                          "S %","w","DF","ann add"],start=11):
        c=ws.cell(top+6,i,h); c.font=Font(name=F,size=9,bold=True); c.fill=WRK; c.border=BOX

    for t in GRID:
        r=grow[t]; n=int(t[:-1]); u=t[-1]
        yrs = n/52 if u=="W" else (n/12 if u=="M" else n)
        raw = f"{SP}+7*{n}" if u=="W" else (f"EDATE({SP},{n})" if u=="M" else f"EDATE({SP},{12*n})")
        ws.cell(r,11,t).font=Font(name=F,size=9)
        ws.cell(r,12,yrs)
        ws.cell(r,13,f'=IF({CD}="","",{modfol(raw)})').number_format="mm/dd/yyyy"
        ws.cell(r,14,f'=IF(M{r}="","",(M{r}-{CD})/365)')
        ws.cell(r,15,f'=IF(M{r}="","",(M{r}-{CD})/360)' if yrs<=1 else f'=IF(M{r}="","",(M{r}-{SP})/360)')
        ws.cell(r,17, 1 if (t=="12M" or u=="Y") else 0)          # on the coupon schedule?
        ws.cell(r,18, 0)                                          # quoted? filled below

    # tauC: from the previous SCHEDULED pillar. 18M accrues from 12M but never advances it.
    prev_sched=None
    for t in GRID:
        r=grow[t]
        if prev_sched is None:
            ws.cell(r,16,f'=IF(M{r}="","",(M{r}-{CD})/360)')
        else:
            ws.cell(r,16,f'=IF(M{r}="","",(M{r}-M{prev_sched})/360)')
        if ws.cell(r,17).value==1: prev_sched=r

    # is this grid tenor present in the paste? 12M and 1Y are the same pillar
    for t in GRID:
        r=grow[t]
        # COUNTIF(range,"") counts BLANK cells, so an empty alt would flag every
        # gap year as quoted. Only add the second term when there really is one.
        alt = "1Y" if t=="12M" else ("12M" if t=="1Y" else None)
        terms=f'COUNTIF($A${P0}:$A${P1},"{t}")'
        if alt: terms+=f'+COUNTIF($A${P0}:$A${P1},"{alt}")'
        ws.cell(r,18,f'=IF({terms}>0,1,0)')

    # par rate: pasted where quoted, else linear in YEAR between bracketing quotes
    for t in GRID:
        r=grow[t]; u=t[-1]; n=int(t[:-1])
        alt = "1Y" if t=="12M" else ("12M" if t=="1Y" else t)
        direct=(f'IFERROR(INDEX($C${P0}:$C${P1},MATCH("{t}",$A${P0}:$A${P1},0)),'
                f'INDEX($C${P0}:$C${P1},MATCH("{alt}",$A${P0}:$A${P1},0)))')
        if u=="Y" and n not in QUOTED_Y:
            lo=max(y for y in QUOTED_Y if y<n); hi=min(y for y in QUOTED_Y if y>n)
            rl,rh=grow[f"{lo}Y"],grow[f"{hi}Y"]
            # gap year: linear in year between the bracketing quotes. Shown for
            # audit only - the gap DF comes from log-linear interpolation of the
            # DFs, not from this rate. IFERROR so an unquoted bracket cannot
            # put #N/A into a column nothing reads.
            ws.cell(r,19,f'=IFERROR(IF(R{r}=1,{direct},'
                         f'S{rl}+(S{rh}-S{rl})*{(n-lo)/(hi-lo):.10f}),"")')
        else:
            ws.cell(r,19,f'=IFERROR({direct},"")')
        ws.cell(r,19).number_format="0.00000"

    # w: log-linear weight for gap years, (T_g - T_A)/(T_P - T_A)
    gapmap={}
    for a,p in SEGMENTS:
        ra,rp=grow[f"{a}Y"],grow[f"{p}Y"]
        gaps=[grow[f"{y}Y"] for y in range(a+1,p)]
        gapmap[p]=(ra,rp,gaps)
        for g in gaps:
            ws.cell(g,20,f'=IF(OR(N{g}="",N{rp}=""),"",(N{g}-N{ra})/(N{rp}-N{ra}))')

    solved={p:v for p,v in gapmap.items()}
    solved_rows={grow[f"{p}Y"]:v for p,v in gapmap.items()}

    # ---------------- DF ----------------
    for t in GRID:
        r=grow[t]; yrs=ws.cell(r,12).value
        if yrs<=1:
            ws.cell(r,21,f'=IF(S{r}="","",1/(1+(S{r}/100)*O{r}))')
        elif r in solved_rows:
            ra,rp,gaps=solved_rows[r]
            last=CL(24+NITER-1)
            ws.cell(r,21,f'=IF(S{r}="","",{last}{r})')
        elif ws.cell(r,20).value is not None:      # gap year
            for a,p in SEGMENTS:
                if grow[f"{a}Y"]<r<grow[f"{p}Y"]:
                    ra,rp=grow[f"{a}Y"],grow[f"{p}Y"]
                    ws.cell(r,21,f'=IF(OR(T{r}="",U{rp}=""),"",U{ra}*(U{rp}/U{ra})^T{r})')
                    break
        else:
            prev=r-1
            ws.cell(r,21,f'=IF(S{r}="","",(1-(S{r}/100)*SUM($V${G0}:V{prev}))/(1+(S{r}/100)*P{r}))')
        ws.cell(r,21).number_format="0.00000000"
        ws.cell(r,22,f'=IF(OR(U{r}="",Q{r}=0),0,P{r}*U{r})')

    # ---------------- the seven solves ----------------
    for p,(ra,rp,gaps) in gapmap.items():
        g0,g1=gaps[0],gaps[-1]
        anchor=f"U{ra}"; Aanch=f"SUM($V${G0}:V{ra})"
        S=f"(S{rp}/100)"
        for j in range(NITER):
            col=24+j; c=ws.cell(rp,col)
            if j==0:
                c.value=f'=IF(S{rp}="","",{anchor}*0.9)'
            else:
                prev=f"{CL(col-1)}{rp}"
                gapsum=(f'SUMPRODUCT($P${g0}:$P${g1},$Q${g0}:$Q${g1},'
                        f'{anchor}*({prev}/{anchor})^$T${g0}:$T${g1})')
                c.value=(f'=IF(S{rp}="","",(1-{S}*({Aanch}+{gapsum}))/(1+{S}*P{rp}))')
            c.font=Font(name=F,size=8); c.number_format="0.00000000"
        ws.cell(top+6,24,"solve →").font=Font(name=F,size=9,bold=True)

    # ---------------- answer columns ----------------
    # A tenor | B date | C swap mid | D BBG zero | E BBG disc | F our zero |
    # G our disc | H d zero bp | I d DF        (paste A, then C:E)
    for r in range(P0,P1+1):
        m=f'MATCH($A{r},$K${G0}:$K${G1},0)'
        alt=f'MATCH(IF($A{r}="1Y","12M",IF($A{r}="12M","1Y",$A{r})),$K${G0}:$K${G1},0)'
        pick=f'IFERROR({m},{alt})'
        # B is PASTED. Left alone if the user pastes Bloomberg's date column;
        # falls back to the derived date only when B is empty.
        ws.cell(r,2,None).number_format="mm/dd/yyyy"
        ws.cell(r,7,f'=IF($A{r}="","",IFERROR(INDEX($U${G0}:$U${G1},{pick}),""))').number_format="0.00000000"
        ws.cell(r,6,f'=IF(OR($A{r}="",G{r}=""),"",-LN(G{r})/'
                    f'INDEX($N${G0}:$N${G1},{pick})*100)').number_format="0.00000"
        ws.cell(r,8,f'=IF(OR(F{r}="",NOT(ISNUMBER(D{r}))),"",(F{r}-D{r})*100)').number_format="0.000"
        ws.cell(r,9,f'=IF(OR(G{r}="",NOT(ISNUMBER(E{r}))),"",G{r}-E{r})').number_format="0.00E+00"
        for col in (1,2,3,4,5):
            c=ws.cell(r,col); c.fill=PASTE; c.border=BOX; c.font=BLK
        for col in (6,7):
            c=ws.cell(r,col); c.fill=CALC; c.border=BOX; c.font=GRN
        ws.cell(r,10,f'=IF(OR($A{r}="",NOT(ISNUMBER($B{r}))),"",'
                     f'IF(ABS($B{r}-INDEX($M${G0}:$M${G1},{pick}))<=3,"ok",'
                     f'"MISMATCH "&TEXT(INDEX($M${G0}:$M${G1},{pick}),"mm/dd/yy")))')
        for col in (8,9,10):
            ws.cell(r,col).border=BOX
        ws.cell(r,3).number_format="0.00000"; ws.cell(r,4).number_format="0.00000"
        ws.cell(r,5).number_format="0.000000"

    S=P1+2
    ws.cell(S,1,"pillars pasted").font=B
    ws.cell(S,3,f"=COUNTA($A${P0}:$A${P1})").border=BOX
    ws.cell(S,4,"priced").font=B
    ws.cell(S,6,f"=COUNT($F${P0}:$F${P1})").border=BOX
    ws.cell(S,7,"max |d zero| bp").font=B
    c=ws.cell(S,8,f'=IF(COUNT($H${P0}:$H${P1})=0,"",MAX(MAX($H${P0}:$H${P1}),-MIN($H${P0}:$H${P1})))')
    c.number_format="0.000"; c.border=BOX; c.fill=CALC
    ws.cell(S+1,7,"max |d DF|").font=B
    c=ws.cell(S+1,8,f'=IF(COUNT($I${P0}:$I${P1})=0,"",MAX(MAX($I${P0}:$I${P1}),-MIN($I${P0}:$I${P1})))')
    c.number_format="0.00E+00"; c.border=BOX; c.fill=CALC
    ws.cell(S,10,f'=IF(COUNTIF($J${P0}:$J${P1},"MISMATCH*")=0,"dates ok",'
                 f'COUNTIF($J${P0}:$J${P1},"MISMATCH*")&" DATE MISMATCHES — your pasted date column '
                 f'is out of step with the tenors; the model uses the derived date")')
    ws.cell(S,10).font=Font(name=F,size=11,bold=True,color="C00000")
    ws.cell(S+2,1,"If 'priced' is under 'pillars pasted', a tenor label is not on the internal grid "
                  "(1W-3W, 1M-12M, 18M, 1Y-50Y). 12M and 1Y are the same pillar.").font=NOTE
    ws.conditional_formatting.add(f"H{P0}:H{P1}",
        CellIsRule(operator="greaterThan",formula=["1"],font=RED))
    ws.conditional_formatting.add(f"H{P0}:H{P1}",
        CellIsRule(operator="lessThan",formula=["-1"],font=RED))
    return S

def harden(ws, top, G0, G1, NIT):
    """Wrap the working grid so an empty curve date blanks out instead of
    spraying #VALUE!. The old guards tested the swap rate, then divided by a tau
    that was still empty - which is exactly what Nigel hit."""
    n=0
    for r in range(G0,G1+1):
        for c in list(range(13,23))+list(range(24,24+NIT)):
            cell=ws.cell(r,c)
            v=cell.value
            if isinstance(v,str) and v.startswith("=") and not v.startswith("=IFERROR("):
                cell.value="=IFERROR("+v[1:]+',"")'; n+=1
            elif isinstance(v,str) and v.startswith("=IFERROR(") and not v.rstrip().endswith(',"")'):
                cell.value="=IFERROR("+v[1:]+',"")'; n+=1
    return n

wb=Workbook(); ws=wb.active; ws.title="Bootstrap_Check"
rows=[]
for i,top in enumerate([1,75,149],start=1):
    rows.append(build_block(ws,top,f"CURVE TEST CASE {i} — paste and read"))
    G0=top+7; G1=G0+len(GRID)-1
    harden(ws,top,G0,G1,NITER)
    # the one input that is easy to miss, made impossible to miss
    ws.cell(top+2,4,'=IF($C$%d="","<<<  ENTER THE CURVE DATE  —  nothing computes without it",'
                    '"")'%(top+2)).font=Font(name=F,size=11,bold=True,color="C00000")
for col,w in (("A",10),("B",17),("C",12),("D",13),("E",12),("F",14),("G",12),("H",11),("I",12)):
    ws.column_dimensions[col].width=w
ws.cell(1,11,"WORKING — the bootstrap itself, left visible so every step can be "
             "audited: dates, taus, annuity, then the 16 solver passes").font=NOTE
for j in range(11,24+NITER):
    ws.column_dimensions[CL(j)].width=9
wb.calculation.fullCalcOnLoad=True
wb.save(OUT)
print("built",OUT,"| blocks at rows 1 / 75 / 149 | working cols K:AM hidden")
