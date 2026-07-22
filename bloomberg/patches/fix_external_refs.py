"""
Strip external-workbook references from a workbook.

Copying a sheet between workbooks makes Excel rewrite its cross-sheet formulas as
links back to the source file:

    '[USD_SOFR_Curve_Bloomberg (26).xlsx]Bootstrap'!$G$4      ->      Bootstrap!$G$4

Run this on the file that has the copied sheet:

    python3 fix_external_refs.py "/path/to/your.xlsx"

It rewrites any [something.xlsx] prefix to a plain local sheet reference and
reports what it changed. Take a copy first.
"""
import re, sys, shutil
from openpyxl import load_workbook

if len(sys.argv) < 2:
    sys.exit("usage: fix_external_refs.py <workbook.xlsx>")
path = sys.argv[1]
shutil.copy(path, path + ".bak")

# '[Book.xlsx]Sheet'!  ->  Sheet!      and     [Book.xlsx]Sheet!  ->  Sheet!
PAT_Q = re.compile(r"'\[[^\]]+\]([^']+)'!")
PAT_P = re.compile(r"\[[^\]]+\]([A-Za-z0-9_]+)!")

wb = load_workbook(path)
n = 0
sheets = set(wb.sheetnames)
missing = set()
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, str) or "[" not in v:
                continue
            new = PAT_Q.sub(lambda m: (f"'{m.group(1)}'!" if " " in m.group(1)
                                       else f"{m.group(1)}!"), v)
            new = PAT_P.sub(r"\1!", new)
            if new != v:
                for m in PAT_Q.finditer(v):
                    if m.group(1) not in sheets: missing.add(m.group(1))
                for m in PAT_P.finditer(v):
                    if m.group(1) not in sheets: missing.add(m.group(1))
                c.value = new
                n += 1
wb.save(path)
print(f"rewrote {n} formulas in {path}   (backup at {path}.bak)")
if missing:
    print("WARNING - these sheets are referenced but not in this workbook:", sorted(missing))
else:
    print("all referenced sheets are present locally")
