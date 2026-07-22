"""
Align the CDS pull with Bloomberg Help Desk H#1330731572.

  1. Field is PX_LAST, not PX_MID. Their step 2 is:
       =BDP(A1 & " BEST Curncy", "PX_LAST")
     We had PX_MID primary with PX_LAST as fallback; swap to their documented
     method, keeping PX_MID as the fallback.

  2. Their tenor list is 1Y/3Y/5Y/7Y/10Y - no 2Y. The workbook has a 2Y row, so
     CDS_SPREAD_TICKER_2Y is requested and may not be a valid field. It fails
     safe (IFERROR to the manual spread) but is flagged.

  3. Recorded limitation, in their words: Bloomberg does not expose a
     zero-coupon CDS spread or bootstrapped hazard rates as an exportable field.
     So unlike the SOFR curve - which has S490 as a hard target, matched to
     0.36bp - the hazard curve has NO exportable ground truth. It can only be
     checked indirectly against CDSW screen outputs.

  4. Their suggested term structure is the credit triangle, Hazard = S/(1-R).
     That is the approximation the workbook seeds column D with; Hazard_Solver
     replaces it with a proper bootstrap. CDS_Validation already compares the
     two, so the difference stays visible.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
SECT = Font(name="Calibri", size=11, bold=True)
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

DOC = {"1Y", "3Y", "5Y", "7Y", "10Y"}      # tenors Bloomberg documented

wb = load_workbook(WB)
q = wb["CDS_Quotes"]
for r in range(7, 13):
    t = str(q[f"A{r}"].value)
    # Help Desk step 2: PX_LAST on "<ticker> BEST Curncy"
    q[f"F{r}"] = (f'=IFERROR(BDP(D{r}&" BEST Curncy","PX_LAST")+0,'
                  f'IFERROR(BDP(D{r}&" BEST Curncy","PX_MID")+0,E{r}))')
    q[f"F{r}"].number_format = "0.0000"
    q[f"F{r}"].border = BOX
    q[f"H{r}"] = (f'=IF(ISNUMBER(IFERROR(BDP(D{r}&" BEST Curncy","PX_LAST")+0,"")),'
                  f'"BDP live","MANUAL (demo)")')
    if t not in DOC:
        q[f"I{r}"] = "tenor not in the Help Desk list (1Y/3Y/5Y/7Y/10Y) - CDS_SPREAD_TICKER_2Y may not be a valid field"
        q[f"I{r}"].font = WARN

q["A16"] = ("SOURCE: Bloomberg Help Desk H#1330731572. Step 1 "
            "=BDP(<ticker>,\"CDS_SPREAD_TICKER_nY\") resolves each tenor's CDS ticker; "
            "step 2 =BDP(<returned>&\" BEST Curncy\",\"PX_LAST\") quotes it. "
            "Documented tenors: 1Y, 3Y, 5Y, 7Y, 10Y.")
q["A16"].font = NOTE
q["A17"] = ("LIMITATION (Bloomberg's words): they do not provide a zero-coupon CDS spread "
            "by maturity as an exportable field, and the bootstrapped hazard rates are not "
            "a standard export. So there is NO exportable ground truth for the hazard curve "
            "- unlike the SOFR curve, which is matched against S490 to 0.36bp. Validate the "
            "credit side indirectly against CDSW screen output.")
q["A17"].font = WARN
q["A18"] = ("Bloomberg's suggested term structure is the credit triangle, Hazard = S/(1-R). "
            "That is only exact for a flat curve; Hazard_Solver bootstraps properly instead. "
            "CDS_Validation rows 13-18 compare the two so the difference stays visible. "
            "Methodology: CDS Model White Paper {LPHP CDSW 0:1 2989941 <GO>}, {HELP CDSW <GO>}.")
q["A18"].font = NOTE

v = wb["CDS_Validation"]
v["A20"] = ("NOTE: there is no Bloomberg export of bootstrapped hazard rates or zero-coupon "
            "CDS spreads (Help Desk H#1330731572), so these checks are INTERNAL consistency "
            "only - they confirm the curve reprices its own inputs, not that it agrees with "
            "Bloomberg. For external validation compare CDSW screen output (points upfront, "
            "price, cash amount) for a known trade.")
v["A20"].font = WARN

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("CDS pull aligned to PX_LAST; limitation and tenor caveat recorded")
