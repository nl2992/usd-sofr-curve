"""
Put Test 2 back on the cell engine, matching Test 1 and Test 3.

Test 2 was briefly wired to CurveVBA.SOFR_Curve. Nigel wants all three blocks
non-VBA, as originally designed, so this restores the INDEX/MATCH-into-Bootstrap
form. CurveVBA.bas stays in the repo - it is still a validated second
implementation - it is just not what the Testing sheet uses.

Consequence of one bootstrap, restated on the sheet: only the block selected at
Bootstrap!G4 computes. The other two show their pasted data and dates and read
"not active" rather than stale numbers. That is by design, not a fault - it is
what the screenshot was showing.
"""
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
shutil.copy(WB, WB + ".bak")

R0, R1, SUM = 58, 97, 98
ACT = 'BOOT_MODE="Test 2"'
MATCH = "MATCH(F{r},BOOT_DATES,0)"

BLACK = Font(name="Calibri", size=11)
GREEN = Font(name="Calibri", size=11, color="008000")
HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="4472C4")
CALC = PatternFill("solid", fgColor="E2EFDA")
BOX = Border(*[Side(style="thin", color="BFBFBF")] * 4)

wb = load_workbook(WB)
t = wb["Testing"]

HEADERS = {"G": "S used %", "H": "Rule", "I": "tau", "J": "A(prior)", "K": "numerator",
           "L": "denominator", "M": "our DF", "N": "t ACT/365", "O": "our zero %",
           "P": "d zero bp", "Q": "d DF"}
for col, txt in HEADERS.items():
    c = t[f"{col}57"]
    c.value = txt; c.font = HDR; c.fill = HFILL; c.border = BOX
    c.alignment = Alignment(horizontal="center", wrap_text=True)

n = 0
for r in range(R0, R1 + 1):
    m = MATCH.format(r=r)
    t[f"G{r}"] = f'=IF(NOT({ACT}),"",IFERROR(INDEX(BOOT_S,{m}),""))'
    t[f"H{r}"] = (f'=IF(NOT({ACT}),"",IFERROR(IF({m}<=15,"short  DFspot/(1+S*tau)",'
                  f'"annual (DFspot-S*A)/(1+S*tau)"),""))')
    t[f"I{r}"] = f'=IF(NOT({ACT}),"",IFERROR(IF({m}<=15,INDEX(BOOT_TAU0,{m}),INDEX(BOOT_TAUC,{m})),""))'
    t[f"J{r}"] = f'=IF(NOT({ACT}),"",IFERROR(IF({m}<=15,0,INDEX(BOOT_ANN,{m})),""))'
    t[f"K{r}"] = f'=IF(OR(G{r}="",I{r}=""),"",BOOT_DFSPOT-(G{r}/100)*J{r})'
    t[f"L{r}"] = f'=IF(OR(G{r}="",I{r}=""),"",1+(G{r}/100)*I{r})'
    t[f"M{r}"] = f'=IF(NOT({ACT}),"",IFERROR(INDEX(BOOT_DF,{m}),""))'
    t[f"N{r}"] = f'=IF(NOT({ACT}),"",IFERROR(INDEX(BOOT_T,{m}),""))'
    t[f"O{r}"] = f'=IF(NOT({ACT}),"",IFERROR(INDEX(BOOT_ZERO,{m}),""))'
    t[f"P{r}"] = f'=IF(OR(O{r}="",NOT(ISNUMBER(C{r}))),"",(O{r}-C{r})*100)'
    t[f"Q{r}"] = f'=IF(OR(M{r}="",NOT(ISNUMBER(D{r}))),"",M{r}-D{r})'
    for col, fmt in (("G", "0.00000"), ("I", "0.000000"), ("J", "0.00000000"),
                     ("K", "0.00000000"), ("L", "0.00000000"), ("M", "0.00000000"),
                     ("N", "0.0000"), ("O", "0.00000"), ("P", "0.000"), ("Q", "0.00E+00")):
        t[f"{col}{r}"].number_format = fmt
    for col in "GHIJKLMNOPQ":
        c = t[f"{col}{r}"]; c.border = BOX
        c.font = GREEN if col in "GMNO" else BLACK
        if col in "GMNO": c.fill = CALC
    n += 1

# summary back to the two-column form the other blocks use
t[f"G{SUM}"] = "max |d zero| bp"
t[f"I{SUM}"] = (f'=IF(NOT({ACT}),"",IF(COUNT(P{R0}:P{R1})=0,"",'
                f'MAX(MAX(P{R0}:P{R1}),-MIN(P{R0}:P{R1}))))')
t[f"I{SUM}"].number_format = "0.000"
t[f"J{SUM}"] = "max |d DF|"
t[f"L{SUM}"] = (f'=IF(NOT({ACT}),"",IF(COUNT(Q{R0}:Q{R1})=0,"",'
                f'MAX(MAX(Q{R0}:Q{R1}),-MIN(Q{R0}:Q{R1}))))')
t[f"L{SUM}"].number_format = "0.00E+00"
t[f"N{SUM}"] = None
t[f"O{SUM}"] = None
t["H56"] = None

t["A56"] = ("Test 2 — paste a capture here, then select 'Test 2' at Bootstrap!G4 to make it "
            "the live case. There is one bootstrap, so only the selected block computes.")
t["A56"].font = Font(name="Calibri", size=9, italic=True, color="808080")

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"Test 2 back on the cell engine: {n} rows, matching Test 1 and Test 3")
