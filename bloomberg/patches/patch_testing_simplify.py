"""
Strip the per-case date columns off Testing.

Was: each case carried date | rate | zero | DF, so the same 32 dates were pasted
three times. Now the curve date sits at the top of each case and one derived
column gives the maturities for whichever case is active.

  A  tenor
  B  date, derived from the active case's curve date
  C:E  case 1   rate | BBG zero | BBG DF
  F:H  case 2
  I:K  case 3

Three columns to paste per case instead of four.

Column B uses a native formula rather than the new TenorDate() UDF so the sheet
keeps working in .xlsx without the macro. TenorDate() is the VBA equivalent and
gives identical dates - both reproduce all 32 on the S490 capture.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
HF = PatternFill("solid", fgColor="1F3864")
CF = ["2E7D32", "C00000", "6A1B9A"]
YF = PatternFill("solid", fgColor="FFFF00")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
DTF, N5, N6 = "mm/dd/yyyy", "0.00000", "0.000000"
R0, R1 = 7, 38
MODE = "Bootstrap!$G$4"
T = ["Test 1", "Test 2", "Test 3"]
CASE = {1: ("C", "D", "E"), 2: ("F", "G", "H"), 3: ("I", "J", "K")}
CDATE = {1: "$C$4", 2: "$F$4", 3: "$I$4"}


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False, al=None):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    if al: c.alignment = Alignment(horizontal=al, wrap_text=True, vertical="center")
    return c


wb = load_workbook(WB)
ws = wb["Testing"]

# clear the old four-column-per-case layout
for col in "BCDEFGHIJKLM":
    for r in range(4, R1 + 1):
        put(ws, f"{col}{r}", None)

put(ws, "A3", "Curve date per case at row 4. Maturities derive from the tenor, so paste three "
              "columns per case: rate, BBG zero, BBG discount.", NOTE)

# active curve date -> spot -> tenor date, native formulas
act_cd = (f'IF({MODE}="{T[0]}",{CDATE[1]},IF({MODE}="{T[1]}",{CDATE[2]},'
          f'IF({MODE}="{T[2]}",{CDATE[3]},"")))')
ws.column_dimensions["B"].width = 13
put(ws, "B6", "Date", HDR, None, HF, True, "center")
put(ws, "A6", "Tenor", HDR, None, HF, True, "center")
for k, (cm, cz, cf) in CASE.items():
    band = PatternFill("solid", fgColor=CF[k-1])
    put(ws, f"{cm}3", f"CURVE TEST CASE {k}", BOLD)
    put(ws, f"{cm}4", None, BLUE, DTF, YF, True)
    put(ws, f"{cz}4", "<- curve date", NOTE)
    for c, h, w, fmt in ((cm, "Swap rate (mid)", 14, N5), (cz, "BBG zero %", 12, N5),
                         (cf, "BBG discount", 13, N6)):
        ws.column_dimensions[c].width = w
        put(ws, f"{c}6", h, HDR, None, band, True, "center")
        for i in range(R1 - R0 + 1):
            put(ws, f"{c}{R0+i}", None, BLUE, fmt, YF, True)
ws.row_dimensions[6].height = 30

for i in range(R1 - R0 + 1):
    r = R0 + i
    cd = f"({act_cd})"
    spot = f"({cd}+2+IF(WEEKDAY({cd},2)>=4,2,0))"
    n = f"VALUE(LEFT(A{r},LEN(A{r})-1))"
    raw = (f'IF(RIGHT(A{r},1)="W",{spot}+7*{n},'
           f'IF(RIGHT(A{r},1)="M",EDATE({spot},{n}),EDATE({spot},12*{n})))')
    nxt = f"({raw}+IF(WEEKDAY({raw},2)=6,2,IF(WEEKDAY({raw},2)=7,1,0)))"
    prv = f"({raw}-IF(WEEKDAY({raw},2)=6,1,IF(WEEKDAY({raw},2)=7,2,0)))"
    put(ws, f"B{r}", f'=IF({cd}="","",IF(MONTH({nxt})<>MONTH({raw}),{prv},{nxt}))',
        BLACK, DTF, None, True)

# repoint the working block: date now comes from B, values from the 3-col cases
def pick(idx, r):
    a, b, c = CASE[1][idx], CASE[2][idx], CASE[3][idx]
    return (f'IF({MODE}="{T[0]}",{a}{r},IF({MODE}="{T[1]}",{b}{r},'
            f'IF({MODE}="{T[2]}",{c}{r},"")))')

B = lambda col: f"Bootstrap!${col}$8:${col}$72"
for i in range(R1 - R0 + 1):
    r = R0 + i
    m = f"MATCH(O{r},{B('B')},0)"
    put(ws, f"O{r}", f"=B{r}", BLACK, DTF, None, True)
    put(ws, f"Z{r}", f"={pick(1, r)}", BLACK, N5, None, True)
    put(ws, f"AA{r}", f"={pick(2, r)}", BLACK, N6, None, True)

put(ws, "A4", "Active source", BOLD)
put(ws, "B4", f"={MODE}", BLACK, None, OF, True)
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("Testing simplified: date columns removed, curve date at row 4, dates derived")
