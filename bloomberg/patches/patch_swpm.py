"""
Let Swap_Pricer handle sub-annual / off-grid swaps, and add the SWPM cross-check.

The coupon schedule was gated on A{r} <= $B$9 with the tenor in whole years, so
a 7-day swap (the SWPM screen's deal: effective 07/23/2026, maturity 07/30/2026)
generated NO coupons at all - zero annuity, undefined par coupon. The sheet
could not express the trade Bloomberg was showing.

  * New maturity override at B17. Blank = derive from tenor as before; type a
    date and it is used directly.
  * The schedule now walks from the effective date, capping each pay date at
    maturity and stopping once maturity is reached. A short swap collapses to a
    single stub period; long swaps are unchanged.
  * SWPM cross-check block with the screen's figures hard-coded.

SWPM screen (Fixed vs SOFR, 10mm, curve date 07/21/26, valuation 07/23/26):
  effective 07/23/2026, maturity 07/30/2026, coupon 3.63840, ACT/360, annual
  Par Cpn 3.638400 | NPV 0.00 | Accrued 0.00 | PV01 = DV01 = 19.42

Leg NPVs (9,995,885.46) are NOT reproduced here and are not expected to be:
SWPM discounts from curve date to a separate valuation date, as the existing
note on row 32 already records.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
YF = PatternFill("solid", fgColor="FFFF00")
SF = PatternFill("solid", fgColor="D9E1F2")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)


def mf(e):
    nxt = f"({e}+IF(WEEKDAY({e},2)=6,2,IF(WEEKDAY({e},2)=7,1,0)))"
    prv = f"({e}-IF(WEEKDAY({e},2)=6,1,IF(WEEKDAY({e},2)=7,2,0)))"
    return f"IF(MONTH({nxt})<>MONTH({e}),{prv},{nxt})"


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


wb = load_workbook(WB)
ws = wb["Swap_Pricer"]

# ---- maturity override
put(ws, "A17", "Maturity override (blank = from tenor)", BLACK)
put(ws, "B17", None, BLUE, "mm/dd/yyyy", YF, True)
put(ws, "C17", "Type a date to price an off-grid or sub-annual swap, e.g. 07/30/2026 for "
               "the SWPM 1W deal. Leave blank to use Tenor (years).", NOTE)
put(ws, "B10", f'=IF($B$17<>"",$B$17,{mf("EDATE($B$8,12*$B$9)")})', BLACK, "mm/dd/yyyy", None, True)

# ---- schedule: walk from effective, cap at maturity, stop when reached
for r in range(36, 86):
    k = r - 35
    prev = "$B$8" if r == 36 else f'IF(OR(B{r-1}="",B{r-1}>=$B$10),"",B{r-1})'
    put(ws, f"C{r}", f"={prev}" if r == 36 else f"={prev}", BLACK, "mm/dd/yyyy")
    put(ws, f"B{r}", f'=IF(C{r}="","",MIN($B$10,{mf(f"EDATE($B$8,12*{k})")}))',
        BLACK, "mm/dd/yyyy")
    put(ws, f"D{r}", f'=IF(C{r}="","",(B{r}-C{r})/360)', BLACK, "0.000000")
    for col, prevexpr in (("E", None),):
        pass
    # DF at the pay date, honouring the interpolation selector
    K, L = "$K$6:$K$71", "$L$6:$L$71"
    i = f"MATCH(B{r},{K},1)"
    Li, Li1, Ki, Ki1 = f"INDEX({L},{i})", f"INDEX({L},{i}+1)", f"INDEX({K},{i})", f"INDEX({K},{i}+1)"
    m3 = f"{Li}*({Li1}/{Li})^((B{r}-{Ki})/({Ki1}-{Ki}))"
    tp, tn, tt = f"(({Ki}-VAL_DATE)/360)", f"(({Ki1}-VAL_DATE)/360)", f"((B{r}-VAL_DATE)/360)"
    rB = f"IF({tn}<=0,0,(1/{Li1}-1)/{tn})"
    rA = f"IF({tp}<=0,{rB},(1/{Li}-1)/{tp})"
    w = f"IF({tn}-{tp}=0,0,({tt}-{tp})/({tn}-{tp}))"
    m1 = f"IF({tt}<=0,1,1/(1+({rA}+({rB}-{rA})*{w})*{tt}))"
    put(ws, f"E{r}",
        f'=IF(C{r}="","",IFERROR(IF(Curve_Interface!$J$7="Piecewise Linear (Simple)",{m1},{m3}),{Li}))',
        BLACK, "0.00000000")
    put(ws, f"F{r}", f'=IF(C{r}="","",$B$6*$B$11/100*D{r})', BLACK, "#,##0.00")
    put(ws, f"G{r}", f'=IF(C{r}="","",F{r}*E{r})', BLACK, "#,##0.00")
    put(ws, f"H{r}", f'=IF(C{r}="","",D{r}*E{r})', BLACK, "0.00000000")

# ---- SWPM cross-check
put(ws, "A88", "SWPM CROSS-CHECK", BOLD)
put(ws, "A89", "Bloomberg SWPM, Fixed vs SOFR, 10mm, curve date 07/21/26, valuation 07/23/26, "
               "OIS DC stripping. To reproduce: Effective 07/23/2026, Maturity override "
               "07/30/2026, Notional 10,000,000, Coupon 3.63840, ACT/360.", NOTE)
for i, h in enumerate(["Measure", "SWPM screen", "This sheet", "Diff"]):
    put(ws, f"{chr(65+i)}90", h, BOLD, None, SF, True)
chk = [("Par coupon (%)", 3.638400, "=B25", "0.000000"),
       ("Net swap NPV", 0.00, "=B28", "#,##0.00"),
       ("Accrued", 0.00, "=B30", "#,##0.00"),
       ("PV01 / DV01 (1bp)", 19.42, "=B29", "0.00")]
for i, (lab, tgt, f, fmt) in enumerate(chk):
    r = 91 + i
    put(ws, f"A{r}", lab, BLACK)
    put(ws, f"B{r}", tgt, BLUE, fmt, YF, True)
    put(ws, f"C{r}", f, BLACK, fmt, OF, True)
    put(ws, f"D{r}", f"=C{r}-B{r}", BLACK, fmt, None, True)
put(ws, "A96", "Leg NPVs (9,995,885.46) are deliberately NOT compared: SWPM discounts from "
               "the curve date to a separate valuation date, so its leg PVs are not on the "
               "same basis as this sheet's. See the note on row 32.", NOTE)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("maturity override at B17; schedule handles sub-annual; SWPM cross-check at A88")
