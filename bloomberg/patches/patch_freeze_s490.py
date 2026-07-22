"""
Freeze on the final S490 capture and hard-code the target into column L.

  * S490_Snapshot updated to the final capture (market rate / zero / discount).
  * Bloomberg_S490_Validation!L holds the S490 zero rates as LITERAL numbers,
    written onto the matching pillar rows - no lookup, no live pull.
  * Column E (d zero) now compares our bootstrap against L, the frozen target.
    D (live BDS) and G (snapshot lookup) stay as secondary references.

Verified before writing, input and output from this same capture:
  max |d zero| 0.40bp overall, 0.08bp from 2Y, max |d DF| 1.88e-05
"""
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
YF = PatternFill("solid", fgColor="FFFF00")
SF = PatternFill("solid", fgColor="D9E1F2")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

S = [("07/30/2026","1W",3.63760,3.68682,0.999091),("08/06/2026","2W",3.67330,3.71731,0.998372),
("08/13/2026","3W",3.68739,3.73043,0.997652),("08/24/2026","1M",3.69655,3.73852,0.996524),
("09/23/2026","2M",3.72255,3.75985,0.993429),("10/23/2026","3M",3.77525,3.80673,0.990244),
("11/23/2026","4M",3.82085,3.84626,0.986914),("12/23/2026","5M",3.86220,3.88151,0.983652),
("01/25/2027","6M",3.90800,3.92030,0.980010),("02/23/2027","7M",3.94150,3.94751,0.976805),
("03/23/2027","8M",3.96975,3.96958,0.973707),("04/23/2027","9M",4.00070,3.99350,0.970254),
("05/24/2027","10M",4.02865,4.01430,0.966800),("06/23/2027","11M",4.05230,4.03096,0.963467),
("07/23/2027","1Y",4.07315,4.04476,0.960147),("01/24/2028","18M",4.09090,4.07772,0.940194),
("07/24/2028","2Y",4.09656,4.06872,0.921438),("07/23/2029","3Y",4.07655,4.04880,0.885328),
("07/23/2030","4Y",4.06655,4.03876,0.850541),("07/23/2031","5Y",4.07150,4.04430,0.816648),
("07/23/2032","6Y",4.08910,4.06340,0.783293),("07/25/2033","7Y",4.11260,4.08910,0.750580),
("07/24/2034","8Y",4.13925,4.11874,0.718878),("07/23/2035","9Y",4.16870,4.15198,0.687885),
("07/23/2036","10Y",4.19995,4.18778,0.657473),("07/23/2038","12Y",4.26400,4.26298,0.599211),
("07/23/2041","15Y",4.34865,4.36615,0.519110),("07/23/2046","20Y",4.41995,4.45385,0.409989),
("07/24/2051","25Y",4.41615,4.43370,0.329718),("07/24/2056","30Y",4.37685,4.35343,0.270538),
("07/23/2066","40Y",4.25755,4.10984,0.192957),("07/23/2076","50Y",4.12427,3.81556,0.148177)]
BYDATE = {dt.datetime.strptime(d, "%m/%d/%Y").date(): (t, m, z, df) for d, t, m, z, df in S}


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


wb = load_workbook(WB)

# ---- 1. refresh the snapshot sheet to the final capture
ws = wb["S490_Snapshot"]
for i, (d, t, m, z, df) in enumerate(S):
    r = 7 + i
    put(ws, f"A{r}", dt.datetime.strptime(d, "%m/%d/%Y").date(), BLUE, "mm/dd/yyyy", YF, True)
    put(ws, f"B{r}", t, BLUE, None, YF, True)
    put(ws, f"C{r}", m, BLUE, "0.00000", YF, True)
    put(ws, f"D{r}", z, BLUE, "0.00000", YF, True)
    put(ws, f"E{r}", df, BLUE, "0.000000", YF, True)
put(ws, "A2", "Bloomberg S490 Curve Analysis capture - FROZEN. USD SOFR (vs FIXED), "
              "Step Forward (Cont), Settle 07/21/26, Curve Side Mid, Shift +0.00bp.", NOTE)

# ---- 2. hard-code the zero target into Bloomberg_S490_Validation!L
# the sheet mirrors Bootstrap rows 8..72; work out which row holds which pillar date
v = wb["Bloomberg_S490_Validation"]
put(v, "L7", "S490 zero (%) FROZEN", BOLD, None, SF, True)
# Bootstrap!B dates are formulas, so derive the row order from the snapshot's own
# ordering: pillars appear in date order, and the annual grid rows are contiguous.
# Match on the tenor label already present in column A of this sheet.
import json
ROWMAP = json.load(open("/private/tmp/claude-501/-Users-nigelli-Desktop-Curve-Bootstrapping/"
                        "a03716b7-925b-41c8-abbb-c9250247d791/scratchpad/rowmap.json"))
BYSTR = {d: (z, df) for d, t, m, z, df in S}
n = 0
for r in range(8, 73):
    ent = ROWMAP.get(str(r))
    hit = BYSTR.get(ent[1]) if ent else None
    if hit:
        put(v, f"L{r}", hit[0], BLUE, "0.00000", YF, True)
        n += 1
    else:
        put(v, f"L{r}", None)
put(v, "L6", f"{n} pillars hard-coded", NOTE)

# ---- 3. compare against the frozen column
for r in range(8, 73):
    put(v, f"E{r}", f'=IF(ISNUMBER(L{r}),(C{r}-L{r})*100,"")', BLACK, "0.00", None, True)
put(v, "A74", "Pillars matched to the frozen target", BOLD)
put(v, "C74", '=COUNT(L8:L72)&" / 32"', BLACK, None, OF, True)
put(v, "A75", "Max |d zero| vs frozen S490 (bp)", BOLD)
put(v, "C75", '=MAX(MAX(E8:E72),-MIN(E8:E72))', BLACK, "0.000", OF, True)
put(v, "A78", "Column L is the FROZEN S490 zero curve, hard-coded from the 07/21/26 capture "
              "- no lookup, no live pull. Column E compares our bootstrap against it. "
              "D (live BDS) and G (snapshot lookup) remain as secondary references.", NOTE)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"frozen: snapshot refreshed, {n} zero rates hard-coded into column L")
