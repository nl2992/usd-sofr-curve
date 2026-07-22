"""
Delete Hazard_Solver. Everything that used it now calls CDSStrip.CDS_StripHazard.

Consequence, stated once: with the ladder gone there is no in-cell fallback, so
the workbook needs the VBA. Hazards resolve to #N/A without it rather than to a
plausible wrong number, and the front page says so.

Brent_vs_Bisection and Root_Methods are kept, as asked, but now compare methods
through the same direct call instead of against a sheet of halvings.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList

WB="bloomberg/CDS_Pricer.xlsx"
ARG=("Hazard_Bootstrap!$B$7:$B$12,Hazard_Bootstrap!$C$7:$C$12,CDS_Parameters!$B$8,"
     "CDS_Parameters!$B$26,CDS_Schedule!$C$7:$C$46,CDS_Schedule!$D$7:$D$46,"
     "CDS_Schedule!$E$7:$E$46,CDS_Schedule!$G$7:$G$46,CDS_Schedule!$H$7:$H$46")
METHODS=["BRENT","BISECTION","FALSE POSITION","SECANT","NEWTON","HALLEY","HOUSEHOLDER","RIDDERS"]
TEN=["1Y","2Y","3Y","5Y","7Y","10Y"]

F="Calibri"
H1=Font(name=F,size=13,bold=True); B=Font(name=F,size=11,bold=True)
N=Font(name=F,size=10); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
IN=PatternFill("solid",fgColor="FFF2CC"); BLUE=Font(name=F,size=11,color="0000FF")
BASE=PatternFill("solid",fgColor="FFF2CC"); RED=Font(name=F,size=10,bold=True,color="C00000")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(WB)

# ---- 1. hazard column: direct, no ladder ------------------------------------
hb=wb["Hazard_Bootstrap"]
for i in range(6):
    hb.cell(7+i,4,
      f'=IF(CDS_Parameters!$B$17="Flat hazard rate",CDS_Parameters!$B$18,'
      f'IFERROR(CDS_StripHazard({i+1},CDS_Parameters!$B$30,{ARG}),NA()))')
hb["N5"]=("Stripped by CDSStrip.CDS_StripHazard straight off CDS_Schedule. "
          "Method at CDS_Parameters!B30. #N/A here means the VBA is not loaded.")
hb["N5"].font=SM

p=wb["CDS_Parameters"]
p["B30"]="BRENT"
p.data_validations=DataValidationList()
dv=DataValidation(type="list",formula1='"'+",".join(METHODS)+'"',allow_blank=False)
p.add_data_validation(dv); dv.add(p["B30"])
p["A30"]="Root-finding method"; p["A30"].font=B
p["C30"]="All eight solve the same segment objective. Brent unless you are testing."
p["C30"].font=SM

# ---- 2. Brent_vs_Bisection, rebuilt on the direct call ----------------------
bv=wb["Brent_vs_Bisection"]
for row in bv.iter_rows(min_row=1,max_row=40,max_col=12):
    for c in row: c.value=None
bv["A1"]="Brent vs bisection"; bv["A1"].font=H1
bv["A2"]="Same segment objective (3.3), same inputs, same sequential scheme. Only the root-finder differs."
bv["A2"].font=SM
bv["A3"]="Hazard_Solver has been removed; both now run through CDSStrip.CDS_StripHazard."
bv["A3"].font=SM
for i,h in enumerate(["Tenor","Market bp","Brent","Bisection","diff","Brent its"],start=1):
    c=bv.cell(5,i,h); c.font=HD; c.fill=HF; c.border=BOX
for i,t in enumerate(TEN):
    r=6+i
    bv.cell(r,1,t)
    bv.cell(r,2,f"=Hazard_Bootstrap!$C${7+i}").number_format="0.00"
    bv.cell(r,3,f'=IFERROR(CDS_StripHazard({i+1},"BRENT",{ARG}),"module not loaded")').number_format="0.000000000"
    bv.cell(r,4,f'=IFERROR(CDS_StripHazard({i+1},"BISECTION",{ARG}),"")').number_format="0.000000000"
    bv.cell(r,5,f'=IF(OR(NOT(ISNUMBER(C{r})),NOT(ISNUMBER(D{r}))),"",C{r}-D{r})').number_format="0.00E+00"
    bv.cell(r,6,'=IFERROR(CDS_StripIterations(),"")')
    for k in range(1,7): bv.cell(r,k).border=BOX
bv.cell(13,1,"max |diff|").font=B
bv.cell(13,5,'=IF(COUNT(E6:E11)=0,"module not loaded",MAX(MAX(E6:E11),-MIN(E6:E11)))').number_format="0.00E+00"
bv.cell(15,1,"Both find the same root. Bisection needs ~52 evaluations per segment, Brent ~9.").font=N
bv.cell(16,1,"The 3,712-cell ladder that used to sit behind bisection is gone; the cost is now one call either way.").font=N

# ---- 3. Root_Methods on the direct call -------------------------------------
rm=wb["Root_Methods"]
for row in rm.iter_rows(min_row=6,max_row=40,max_col=12):
    for c in row: c.value=None
rm["A3"]=("Baseline is the scheme, not the algorithm. Section 4 p.8 fixes piecewise-constant hazard "
          "solved maturity by maturity, each a 1-D root-find with earlier hazards fixed. It names no method.")
rm["A4"]="Needs CDSStrip.bas imported, saved as .xlsm."
for i,h in enumerate(["Method","Family","Order","5Y hazard","its","vs Brent"],start=1):
    c=rm.cell(6,i,h); c.font=HD; c.fill=HF; c.border=BOX
    c.alignment=Alignment(horizontal="center",wrap_text=True)
FAM={"BRENT":("bracketing hybrid","hybrid"),"BISECTION":("bracketing","linear"),
     "FALSE POSITION":("bracketing","super-linear"),"SECANT":("open","~1.618"),
     "NEWTON":("open","2"),"HALLEY":("open","3"),"HOUSEHOLDER":("open","d+1"),
     "RIDDERS":("bracketing hybrid","hybrid")}
for i,m in enumerate(METHODS):
    r=7+i
    rm.cell(r,1,m).font=N
    rm.cell(r,2,FAM[m][0]).font=N; rm.cell(r,3,FAM[m][1]).font=N
    rm.cell(r,4,f'=IFERROR(CDS_StripHazard(4,"{m}",{ARG}),"module not loaded")').number_format="0.000000000"
    rm.cell(r,5,'=IFERROR(CDS_StripIterations(),"")')
    rm.cell(r,6,f'=IF(NOT(ISNUMBER(D{r})),"",D{r}-$D$7)').number_format="0.00E+00"
    for k in range(1,7): rm.cell(r,k).border=BOX
rm.cell(7,1).fill=BASE
rm.cell(16,1,"Spread of the eight roots").font=B
rm.cell(16,4,'=IF(COUNT(D7:D14)=0,"module not loaded",MAX(D7:D14)-MIN(D7:D14))').number_format="0.00E+00"
rm.cell(16,5,"they solve the same equation, so this is the real check").font=SM

# ---- 4. delete the sheet -----------------------------------------------------
del wb["Hazard_Solver"]

st=wb["Steps"]
st.cell(12,2,"=CDS_Parameters!$B$30")
st.cell(12,3,"Brent by default. No in-cell fallback: without the VBA the hazards read #N/A.").font=RED
wb.calculation.fullCalcOnLoad=True
wb.save(WB)

import re
wb2=load_workbook(WB)
bad=[(ws.title,c.coordinate) for ws in wb2 for row in ws.iter_rows() for c in row
     if isinstance(c.value,str) and "Hazard_Solver" in c.value]
print("sheets:",len(wb2.sheetnames))
print("remaining references to Hazard_Solver:", bad or "NONE")
