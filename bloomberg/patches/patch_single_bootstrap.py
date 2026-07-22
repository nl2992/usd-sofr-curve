"""
Reduce to a single bootstrap: delete Bootstrap_Lehman, the futures-based second
copy. `Bootstrap` (purely-OIS) is the one verified against the S490 screen.

Bootstrap_Lehman has NO formula dependencies - only three prose mentions
(Instructions!A41, Bootstrap!A5, Bootstrap!A74), which are rewritten here. It was
also the source of 160 of the workbook's 164 off-terminal errors.

SOFR_Futures is left in place: it is a quote sheet, not a bootstrap, and deleting
market data nobody asked me to delete is the riskier move. It now has no formula
consumers, so it is flagged as unused rather than removed.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")

def safe_set(ws, cell, text):
    try:
        ws[cell].value = text
        ws[cell].font = NOTE
        return True
    except AttributeError:
        return False   # merged cell

wb = load_workbook(WB)
before = list(wb.sheetnames)

# rewrite the prose that referenced the deleted sheet
b = wb["Bootstrap"]
safe_set(b, "A5",
    "Purely-OIS USD SOFR curve: short single-payment OIS (<=1Y) + annual OIS swaps "
    "(18M-50Y). THE single production curve - used by Swap_Pricer, Curve_Interface "
    "and the CDS module. Verified against the Bloomberg S490 screen (see S490_Target): "
    "32/32 pillars, max zero diff 0.26bp, max DF diff 1.2e-05.")
safe_set(b, "A74",
    "Long-end quoted pillars (12Y+) are solved on Curve_Solver to break the gap-row "
    "annuity circularity. Gap rows keep log-linear DF interpolation (= step-function "
    "forward), the method verified against S490.")
safe_set(wb["Instructions"], "A41",
    "Bootstrap - the single production SOFR curve, purely from OIS quotes. "
    "(The former Bootstrap_Lehman futures variant has been removed.)")
safe_set(wb["SOFR_Futures"], "A2",
    "UNUSED: no formula on any sheet reads this tab since the futures-based bootstrap "
    "was removed. Kept as reference data only - the production curve is purely OIS.")

del wb["Bootstrap_Lehman"]
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("removed Bootstrap_Lehman")
print("sheets:", len(before), "->", len(wb.sheetnames))
