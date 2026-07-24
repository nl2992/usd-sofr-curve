"""
Rebuild the Hedge sheet to MIRROR the master's layout exactly:
  delta table B4:B23 | 1 risk 27-31 | 2 strategy 33-36 | 3 trades 38-43 | 4 cost 44-51
Blotter DV01 reads the on-sheet delta table; cost table reads the blotter (E/F/C 40:43),
all ABSOLUTE so the block can't misalign. Delta values sourced from KRD-v3 (master uses KRD-v2
at the same cells, so refs are identical).
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
WB="bloomberg/KRW_IRS_Bootstrap_Book1.xlsx"; V="'KRD-v3'"
TAU="Quotes!$B$2"; YR="Bootstrap!$B$3:$B$122"; DF="Bootstrap!$D$3:$D$122"
F="Calibri"
TITLE=Font(name=F,size=14,bold=True); SEC=Font(name=F,size=11,bold=True,color="FFFFFF"); SECF=PatternFill("solid",fgColor="C00000")
HD=Font(name=F,size=9,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
B=Font(name=F,size=10,bold=True); SM=Font(name=F,size=9,color="808080"); N=Font(name=F,size=10)
GRN=Font(name=F,size=10,bold=True,color="006100"); RED=Font(name=F,size=10,bold=True,color="9C0006")
INP=Font(name=F,size=10,color="0000FF"); MONO=Font(name="Consolas",size=9)
KEY=PatternFill("solid",fgColor="FCE4D6"); BLU=PatternFill("solid",fgColor="DDEBF7"); INPF=PatternFill("solid",fgColor="FFF2CC"); SKIPF=PatternFill("solid",fgColor="FCE4D6")
thin=Side(style="thin",color="BFBFBF"); BOX=Border(thin,thin,thin,thin)
MONEY='#,##0'; BN='#,##0.0'

wb=load_workbook(WB)
if "Hedge" in wb.sheetnames: del wb["Hedge"]
ws=wb.create_sheet("Hedge", wb.sheetnames.index("KRD-v3")+1)
def sec(r,txt,ncol=6):
    ws.cell(r,1,txt).font=SEC
    for cc in range(1,ncol+1): ws.cell(r,cc).fill=SECF

ws["A1"]="Hedging the trade"; ws["A1"].font=TITLE
# --- folded delta table B4:B23 (mirror master rows) ---
ws["A3"]="Tenor"; ws["B3"]="dv01"
for cc in ("A3","B3"): ws[cc].font=HD; ws[cc].fill=HF
labels=["2D","1W","1M","2M","3M","6M","9M","1Y","2Y","3Y","4Y","5Y","7Y","10Y","12Y","15Y","20Y","25Y","30Y"]
for i,lab in enumerate(labels):
    r=4+i
    ws.cell(r,1,lab)
    ws.cell(r,2,f"={V}!B{r+4}").number_format=MONEY     # KRD-v3 folded rows 8..26
ws["A23"]="TOTAL"; ws["A23"].font=B
ws["B23"]=f"={V}!B27"; ws["B23"].number_format=MONEY; ws["B23"].fill=KEY; ws["B23"].font=B
# side annotations
ws["E4"]="y1 exposure"; ws["F4"]="=$B$8+$B$9+$B$10+$B$11"; ws["F4"].number_format=MONEY
ws["E5"]="y4-y10 exposure"; ws["F5"]="=$B$14+$B$15+$B$16+$B$17"; ws["F5"].number_format=MONEY
ws["E8"]="Original Trade"; ws["F8"]="HSBC pay Float 3m CD, Q/A365"; ws["F8"].font=B
ws["F9"]="fixed leg -> back-end risk"; ws["F9"].font=SM

# --- 1 risk ---
sec(27,"1 · The risk")
ws["A28"]="Net DV01"; ws["A28"].font=B
ws["B28"]="=$B$23"; ws["B28"].number_format=MONEY; ws["B28"].font=RED; ws["B28"].fill=KEY
ws["C28"]="KRW per bp, net receiver"; ws["C28"].font=SM
ws["A29"]="Front 1Y–3Y"; ws["A29"].font=B
ws["B29"]="=$B$11+$B$12+$B$13"; ws["B29"].number_format=MONEY; ws["B29"].font=GRN
ws["C29"]="long (positive), pays if rates rise"; ws["C29"].font=SM
ws["A30"]="Back 4Y–10Y"; ws["A30"].font=B
ws["B30"]="=$B$14+$B$15+$B$16+$B$17"; ws["B30"].number_format=MONEY; ws["B30"].font=RED
ws["C30"]="short (negative), concentrated at 10Y"; ws["C30"].font=SM
ws["A31"]="Sub-1Y (3M/6M/9M)"; ws["A31"].font=B
ws["B31"]="=$B$8+$B$9+$B$10"; ws["B31"].number_format=MONEY; ws["B31"].font=SM
ws["C31"]="miniscule"; ws["C31"].font=SM

# --- 2 strategy ---
sec(33,"2 · Strategy")
ws["A36"]="Receive fixed front-end; Pay Fixed for long end; Short-end not worth hedging — doesn't even cover b/a"; ws["A36"].font=N

# --- 3 trades (blotter 40-43) ---
sec(38,"3 · The trades  -->  4 swaps?")
for i,h in enumerate(["Swap","Direction","Notional (bn)","Absorbs","Group DV01 (KRW)","Annuity SUM(DF*Tau)"],1):
    c=ws.cell(39,i,h); c.font=HD; c.fill=HF; c.border=BOX
blot=[("2Y","1Y + 2Y","=$B$11+$B$12",2),("3Y","3Y","=$B$13",3),
      ("5Y","4Y + 5Y","=$B$14+$B$15",5),("10Y","7Y + 10Y","=$B$16+$B$17",10)]
for k,(tn,ab,dv,T) in enumerate(blot):
    r=40+k
    ws.cell(r,1,tn).font=B
    ws.cell(r,2,f'=IF(E{r}>0,"Receive fixed","Pay fixed")').font=B
    ws.cell(r,3,f"=ABS(E{r}/(F{r}*0.0001))/1000000000").number_format=BN; ws.cell(r,3).font=GRN; ws.cell(r,3).fill=BLU
    ws.cell(r,4,ab)
    ws.cell(r,5,dv).number_format=MONEY
    ws.cell(r,6,f'={TAU}*SUMIF({YR},"<="&{T},{DF})').number_format="0.0000"
    for cc in range(1,7): ws.cell(r,cc).border=BOX
# benchmark annuity display (right)
ws.cell(38,8,"Benchmark annuity (curve)").font=B
for i,h in enumerate(["Tenor","T","Annuity"],8):
    c=ws.cell(39,i,h); c.font=HD; c.fill=HF
for k,(tn,T) in enumerate([("2Y",2),("3Y",3),("5Y",5),("7Y",7),("10Y",10)]):
    r=40+k; ws.cell(r,8,tn); ws.cell(r,9,T)
    ws.cell(r,10,f'={TAU}*SUMIF({YR},"<="&I{r},{DF})').number_format="0.0000"

# --- 4 cost analysis (44-51) ---
sec(44,"4 · Hedge Cost Analysis",13)
hdr=["Tenor","Ticker (edit)","Bid","Ask","Spread (bp)","Avg spr 1M (bp)","DV01 (KRW)","Annuity","Notl (bn)","Cost (KRW)","Break-even (bp)","Vol USDmm","Verdict"]
for i,h in enumerate(hdr,1):
    c=ws.cell(45,i,h); c.font=HD; c.fill=HF; c.border=BOX
# 4 trades reference the blotter (absolute); skips self-source from delta table
crows=[
 ("2Y","KWSWNI2 BGN Curncy","=$E$40","=$F$40","=$C$40",False),
 ("3Y","KWSWNI3 BGN Curncy","=$E$41","=$F$41","=$C$41",False),
 ("5Y","KWSWNI5 BGN Curncy","=$E$42","=$F$42","=$C$42",False),
 ("10Y","KWSWNI10 BGN Curncy","=$E$43","=$F$43","=$C$43",False),
 ("7Y split","KWSWNI7 BGN Curncy","=$B$16",f'={TAU}*SUMIF({YR},"<="&7,{DF})',None,True),
 ("Dust 3-9M","(front=CD/FRA, GC S205)","=$B$8+$B$9+$B$10",f'={TAU}*SUMIF({YR},"<="&1,{DF})',None,True),
]
for k,(tn,tk,dv,ann,notl,skip) in enumerate(crows):
    r=46+k
    ws.cell(r,1,tn).font=B
    ws.cell(r,2,tk).font=INP; ws.cell(r,2).fill=INPF
    ws.cell(r,3,f'=IFERROR(BDP(B{r},"PX_BID"),"n/a")').number_format="0.000"
    ws.cell(r,4,f'=IFERROR(BDP(B{r},"PX_ASK"),"n/a")').number_format="0.000"
    ws.cell(r,5,f'=IFERROR((D{r}-C{r})*100,"n/a")').number_format="0.00"
    ws.cell(r,6,f'=IFERROR((AVERAGE(BDH(B{r},"PX_ASK",TEXT(TODAY()-30,"yyyymmdd"),TEXT(TODAY(),"yyyymmdd"),"Dts=H"))'
                f'-AVERAGE(BDH(B{r},"PX_BID",TEXT(TODAY()-30,"yyyymmdd"),TEXT(TODAY(),"yyyymmdd"),"Dts=H")))*100,"n/a")').number_format="0.00"
    ws.cell(r,7,dv).number_format=MONEY
    ws.cell(r,8,ann).number_format="0.0000"
    ws.cell(r,9,(notl if notl else f"=ABS(G{r}/(H{r}*0.0001))/1000000000")).number_format=BN
    ws.cell(r,10,f'=IFERROR(I{r}*1000000000*H{r}*(E{r}/2)*0.0001,"n/a")').number_format=MONEY
    ws.cell(r,11,f'=IFERROR(J{r}/ABS(G{r}),"n/a")').number_format="0.00"
    ws.cell(r,12,0).font=INP; ws.cell(r,12).fill=INPF
    ws.cell(r,13,("SKIP" if tn.startswith("Dust") else f'=IFERROR(IF(K{r}<1,"TRADE","SKIP"),"SKIP")'))
    ws.cell(r,13).font=RED if skip else GRN
    for cc in range(1,14):
        ws.cell(r,cc).border=BOX
        if skip and cc not in (2,12): ws.cell(r,cc).fill=SKIPF
ws.cell(52,1,"Break-even (bp) = cost / DV01 ≈ half the bid/offer. <1bp -> TRADE. Cross-check SDRV volume.").font=SM
ws.cell(53,1,'Commands: GC S205 <GO> (tickers) · SDRV <GO> Rates>IRS/FRA>KRW (volume) · SDR <GO> (trades).').font=MONO

for col,w in zip("ABCDEFGHIJKLM",(12,22,8,8,11,14,16,10,10,15,13,11,10)): ws.column_dimensions[col].width=w
wb.save(WB)
print("Hedge rebuilt to mirror master.")
