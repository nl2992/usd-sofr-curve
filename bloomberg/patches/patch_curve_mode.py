"""
Dropdown to run the swap curve off LIVE BDP or the FIXED S490 capture.

The switch goes into SOFR_OIS_Quotes!H, the single cell every consumer reads:
the 32 quoted Bootstrap!E cells, the 17-row P/Q helper that interpolates the gap
years, and therefore the bootstrap, the zero curve, the DF grid, the CDS module
and the charts. One formula per row rather than patching 49 downstream cells.

  Live (BDP)              H = BDP mid, falling back to the fixed capture if the
                              terminal is unavailable, so the curve always builds
  Fixed (S490 07/21/26)   H = the frozen capture, BDP ignored entirely

Column T keeps the pure BDP mid regardless of mode, so the live-vs-fixed
difference stays measurable from either setting.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
YF = PatternFill("solid", fgColor="FFFF00")
HF = PatternFill("solid", fgColor="1F3864")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
LIVE, FIXED = "Live (BDP)", "Fixed (S490 07/21/26)"
MODE = "Bootstrap!$G$4"


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False, al=None):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    if al: c.alignment = Alignment(horizontal=al, wrap_text=True, vertical="center")
    return c


wb = load_workbook(WB)

# ---- the control, on the curve sheet where it is visible
b = wb["Bootstrap"]
put(b, "F4", "Curve source", BOLD)
put(b, "G4", FIXED, BLUE, None, YF, True)
dv = DataValidation(type="list", formula1=f'"{LIVE},{FIXED}"', allow_blank=False)
b.add_data_validation(dv); dv.add(b["G4"])
put(b, "I4", "Drives every quote via SOFR_OIS_Quotes!H, so the whole curve, the CDS "
             "module and the charts follow. Live falls back to the fixed capture if "
             "BDP is down.", NOTE)

# ---- the switch itself
q = wb["SOFR_OIS_Quotes"]
put(q, "T4", "Live mid (BDP only)", HDR, None, HF, True, "center")
put(q, "U4", "d live - fixed (bp)", HDR, None, HF, True, "center")
n = 0
for r in range(5, 40):
    if q[f"A{r}"].value is None or q[f"B{r}"].value is None:
        continue
    put(q, f"T{r}", f'=IFERROR((E{r}+F{r})/2,IFERROR(G{r}+0,""))', BLACK, "0.00000", None, True)
    put(q, f"U{r}", f'=IF(OR(T{r}="",NOT(ISNUMBER(J{r}))),"",(T{r}-J{r})*100)',
        BLACK, "0.00", None, True)
    put(q, f"H{r}",
        f'=IF({MODE}="{FIXED}",J{r},IF(ISNUMBER(T{r}),T{r},J{r}))',
        BLACK, "0.00000", OF, True)
    n += 1
put(q, "T2", f'=IF({MODE}="{FIXED}","FIXED - BDP ignored, curve on the 07/21/26 capture",'
             f'"LIVE - BDP mid, falling back to the capture if unavailable")', BOLD)
put(q, "T3", "Column H is the mid actually used. T is always the raw BDP mid, so the "
             "live-vs-fixed gap stays visible in either mode.", NOTE)
put(q, "T41", "Max |d live-fixed| (bp)", BOLD)
put(q, "U41", '=IF(COUNT(U5:U36)=0,"BDP unavailable",MAX(MAX(U5:U36),-MIN(U5:U36)))',
    BLACK, "0.000", OF, True)

# ---- relabel the Bootstrap comparison to say what it now means
put(b, "Y7", "d in-use vs fixed (bp)", HDR, None, HF, True, "center")
put(b, "X78", "X is the fixed 07/21/26 capture, always. Y is what the curve is actually "
              "using minus that: 0.00 in Fixed mode, the live drift in Live mode. "
              "Raw BDP mid and the live-vs-fixed gap are on SOFR_OIS_Quotes T:U.", NOTE)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"curve-source dropdown at Bootstrap!G4; switch applied to {n} quote rows")
