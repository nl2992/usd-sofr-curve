"""
'Hedge' sheet: live hedge sizing off the KRD buckets + Bootstrap curve.
notional(T) = group DV01 / (annuity(T) * 1bp); sign -> receive/pay fixed.
Groups: 1Y+2Y->2Y, 3Y->3Y, 4Y+5Y->5Y, 7Y+8Y+9Y+10Y->10Y. 3M/6M/9M left as dust.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
WB="bloomberg/KRW_IRS_Bootstrap_Book1.xlsx"
F="Calibri"
H1=Font(name=F,size=14,bold=True); B=Font(name=F,size=11,bold=True); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=9,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
RES=PatternFill("solid",fgColor="FCE4D6"); GRN=Font(name=F,size=10,color="008000")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4); MONEY='#,##0'; BN='#,##0.0'
wb=load_workbook(WB)
if "Hedge" in wb.sheetnames: del wb["Hedge"]
ws=wb.create_sheet("Hedge", wb.sheetnames.index("KRD-v3")+1)
TAU="Quotes!$B$2"; YR="Bootstrap!$B$3:$B$122"; DF="Bootstrap!$D$3:$D$122"

ws["A1"]="Hedge — live notional sizing off the KRD buckets"; ws["A1"].font=H1
ws["A2"]=("notional(T) = group DV01 / (annuity(T) x 1bp). Positive DV01 -> receive fixed, "
          "negative -> pay fixed. Buckets grouped to liquid benchmarks; 3M/6M/9M left unhedged "
          "(dust). Re-sizes automatically as the trade or curve moves.")
ws["A2"].font=SM

# --- benchmark annuities from the curve ---
ws.cell(4,1,"Benchmark PV01 (from Bootstrap curve)").font=B
for i,h in enumerate(["Tenor","T (yrs)","Annuity","DV01 / 100bn"],1):
    c=ws.cell(5,i,h); c.font=HD; c.fill=HF
for k,(tn,T) in enumerate([("2Y",2),("3Y",3),("5Y",5),("7Y",7),("10Y",10)]):
    r=6+k
    ws.cell(r,1,tn)
    ws.cell(r,2,T)
    ws.cell(r,3,f"={TAU}*SUMIF({YR},\"<=\"&B{r},{DF})").number_format="0.0000"
    ws.cell(r,4,f"=100000000000*C{r}*0.0001").number_format=MONEY
ANN={"2Y":"$C$6","3Y":"$C$7","5Y":"$C$8","7Y":"$C$9","10Y":"$C$10"}

# --- hedge proposal (4 swaps) ---
ws.cell(12,1,"Hedge proposal — 4 swaps").font=B
for i,h in enumerate(["Hedge swap","Absorbs buckets","Group DV01 (KRW/bp)","Annuity","Notional (bn)","Direction"],1):
    c=ws.cell(13,i,h); c.font=HD; c.fill=HF
rows=[
 ("2Y","1Y + 2Y","=KRD!X69+KRD!Y69","2Y"),
 ("3Y","3Y","=KRD!Z69","3Y"),
 ("5Y","4Y + 5Y","=KRD!AA69+KRD!AB69","5Y"),
 ("10Y","7Y + 8Y + 9Y + 10Y","=KRD!AC69+KRD!AD69+KRD!AE69+KRD!AF69","10Y"),
]
r0=14
for k,(tn,absorbs,dv,annk) in enumerate(rows):
    r=r0+k
    ws.cell(r,1,tn).font=B
    ws.cell(r,2,absorbs)
    ws.cell(r,3,dv).number_format=MONEY
    ws.cell(r,4,f"={ANN[annk]}").number_format="0.0000"
    ws.cell(r,5,f"=ABS(C{r}/(D{r}*0.0001))/1000000000").number_format=BN; ws.cell(r,5).font=GRN
    ws.cell(r,6,f'=IF(C{r}>0,"Receive fixed","Pay fixed")').font=B
rlast=r0+len(rows)-1
# dust row
rd=rlast+1
ws.cell(rd,1,"(unhedged)").font=SM
ws.cell(rd,2,"3M + 6M + 9M (dust)").font=SM
ws.cell(rd,3,"=KRD!U69+KRD!V69+KRD!W69").number_format=MONEY; ws.cell(rd,3).font=SM
ws.cell(rd,6,"leave — bid/offer > risk").font=SM

# --- verification ---
rv=rd+2
ws.cell(rv,1,"Net DV01 before hedge:").font=B
ws.cell(rv,3,f"=SUM(C14:C17)+C{rd}").number_format=MONEY; ws.cell(rv,3).fill=RES; ws.cell(rv,3).font=B
ws.cell(rv+1,1,"Hedged buckets (2Y/3Y/5Y/10Y groups):").font=B
ws.cell(rv+1,3,"neutralized to 0 by construction").font=SM
ws.cell(rv+2,1,"Residual DV01 after hedge:").font=B
ws.cell(rv+2,3,f"=C{rd}").number_format=MONEY; ws.cell(rv+2,3).fill=RES; ws.cell(rv+2,3).font=B
ws.cell(rv+2,4,"= sub-1Y dust only").font=SM

# --- optional 5th trade ---
ro=rv+4
ws.cell(ro,1,"Optional — split 7Y out to kill the 7s10s residual").font=B
for i,h in enumerate(["Hedge swap","Absorbs buckets","Group DV01 (KRW/bp)","Annuity","Notional (bn)","Direction"],1):
    c=ws.cell(ro+1,i,h); c.font=HD; c.fill=HF
opt=[
 ("7Y","7Y + 8Y (part)","=KRD!AC69+KRD!AD69","7Y"),
 ("10Y","9Y + 10Y (part)","=KRD!AE69+KRD!AF69","10Y"),
]
for k,(tn,absorbs,dv,annk) in enumerate(opt):
    r=ro+2+k
    ws.cell(r,1,tn).font=B; ws.cell(r,2,absorbs)
    ws.cell(r,3,dv).number_format=MONEY
    ws.cell(r,4,f"={ANN[annk]}").number_format="0.0000"
    ws.cell(r,5,f"=ABS(C{r}/(D{r}*0.0001))/1000000000").number_format=BN; ws.cell(r,5).font=GRN
    ws.cell(r,6,f'=IF(C{r}>0,"Receive fixed","Pay fixed")').font=B
ws.cell(ro+2+len(opt)+1,1,"Replaces the single 10Y line above with these two; +1 bid/offer.").font=SM

for col,w in zip("ABCDEF",(12,22,20,10,13,15)): ws.column_dimensions[col].width=w
wb.save(WB)
print("Hedge sheet at",wb.sheetnames.index("Hedge"),"of",wb.sheetnames)
