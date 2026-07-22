"""
Make CDS_Entities earn its place.

Two real problems, found by tracing which columns anything reads:

1. Four captured CDSW fields were typed in and never compared. Spread DV01,
   IR DV01, Rec risk and Prob 5Y sat on the sheet as dead literals while the
   model computed all four. The comparison block now uses them.

2. Seniority, clause and recovery were typed on CDS_Entities while every other
   input had moved to Entities, so adding a name meant visiting two sheets again.
   They are inputs on Entities now and CDS_Entities reads them.

Ticker (B) and ccy (C) stay as pass-through display, but they are read by nothing
and are labelled as such rather than looking load-bearing.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

WB="bloomberg/CDS_Pricer.xlsx"
NSLOT=10; R0=7; E0=5
F="Calibri"
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
IN=PatternFill("solid",fgColor="FFF2CC"); BLUE=Font(name=F,size=11,color="0000FF")
SM=Font(name=F,size=9,color="808080"); N=Font(name=F,size=10); B=Font(name=F,size=11,bold=True)
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(WB); es=wb["Entities"]; ent=wb["CDS_Entities"]; pr=wb["CDS_Pricer"]

# ---- 1. compare the four captured fields that were dead --------------------
EXTRA=[("Spread DV01","Y","=B22"),("IR DV01","Z","=B23"),
       ("Rec risk 1%","AA","=B24"),
       ("Prob 5Y (1-Q)","AC","=IFERROR(INDEX(Hazard_Bootstrap!$F$7:$F$12,MATCH(5,CDS_Quotes!$B$7:$B$12,0)),\"\")")]
r=51
for lab,col,model in EXTRA:
    pr.cell(r,1,lab).font=N
    pr.cell(r,2,f'=IFERROR(INDEX(CDS_Entities!${col}$5:${col}$14,CDS_Parameters!$B$28),"")')
    pr.cell(r,3,model)
    pr.cell(r,4,f'=IF(OR(B{r}="",NOT(ISNUMBER(B{r}))),"",C{r}-B{r})')
    for k in range(1,5): pr.cell(r,k).border=BOX
    r+=1
pr["A42"]=("Blank until a CDSW capture is typed into CDS_Entities O:AC. "
           "Notional, coupon, maturity and traded spread there are context for the capture, not compared.")
pr["A42"].font=SM
print(f"comparison block extended: {len(EXTRA)} previously dead capture fields now compared")

# ---- 2. seniority / clause / recovery become Entities inputs ----------------
for i,h in enumerate(["Seniority","Clause","Recovery"],start=12):
    c=es.cell(6,i,h); c.font=HD; c.fill=HF; c.border=BOX
    c.alignment=Alignment(horizontal="center")
es.cell(5,12,"identity - picks the ticker").font=SM
for i in range(NSLOT):
    r=R0+i; src=E0+i
    es.cell(r,12, ent.cell(src,4).value if not isinstance(ent.cell(src,4).value,str) or not ent.cell(src,4).value.startswith("=") else None)
    es.cell(r,13, ent.cell(src,5).value)
    es.cell(r,14, ent.cell(src,6).value if isinstance(ent.cell(src,6).value,(int,float)) else 0.4)
    for k in (12,13,14):
        c=es.cell(r,k); c.fill=IN; c.font=BLUE; c.border=BOX
    es.cell(r,14).number_format="0.00"
    ent.cell(src,4,f'=IF(Entities!$L{r}="","",Entities!$L{r})')
    ent.cell(src,5,f'=IF(Entities!$M{r}="","",Entities!$M{r})')
    ent.cell(src,6,f'=IF(Entities!$N{r}="",0.4,Entities!$N{r})')
es["E5"]="override"; es["E5"].font=SM
print("seniority, clause and base recovery moved onto Entities")

# ---- 3. label the pass-through columns so they do not look load-bearing -----
ent["B3"]="display only - the ticker in use comes from Entities!F4"; ent["B3"].font=SM
ent["O3"]="CDSW capture - context"; ent["O3"].font=SM
ent["S3"]="CDSW capture - compared on CDS_Pricer rows 44-54"; ent["S3"].font=SM

wb.calculation.fullCalcOnLoad=True
wb.save(WB)
