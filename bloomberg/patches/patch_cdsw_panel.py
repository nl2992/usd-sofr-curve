"""
Fold the CDSW screen output panel into CDS_Pricer, so the workbook is the single
deliverable and the standalone CDSW_CDS_Pricer.xlsx can be retired.

CDS_Pricer had the analytics (RPV01, legs, par spread, PV, CS01/IR01/Rec01) but
not the CDSW screen's own settlement figures. Adds, in the screen's order:

    Points upfront, Price, Principal, Accrued days, Accrued, Cash Amount,
    Default Exposure

Conventions, matched against the reference CDSW screen (CINDBK 5Y, 07/21/26):
  Points upfront   = (S - C) * RPV01 * 100          protection buyer
  Price            = 100 - points upfront
  Principal        = direction * (S - C) * RPV01 * notional
  Accrual start    = previous IMM roll, business-day adjusted (20 Jun 2026 is a
                     Saturday, so the screen shows 1st Accr Start 06/22/26)
  Accrued days     = valuation - accrual start + 1   (ISDA counts both ends;
                     the screen shows 30 days, giving 100bp*10mm*30/360 = 8,333)
  Cash Amount      = Principal + Accrued
  Default Exposure = (1-R)*notional - Principal
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
CCY = '$#,##0;($#,##0);-'

N, R = "CDS_Parameters!$B$9", "CDS_Parameters!$B$8"
CPN = "CDS_Parameters!$B$10/10000"
DIR = 'IF(CDS_Parameters!$B$15="Buy protection",1,-1)'
ROLL = "CDS_Parameters!$B$5"


def mf(e):
    nxt = f"({e}+IF(WEEKDAY({e},2)=6,2,IF(WEEKDAY({e},2)=7,1,0)))"
    prv = f"({e}-IF(WEEKDAY({e},2)=6,1,IF(WEEKDAY({e},2)=7,2,0)))"
    return f"IF(MONTH({nxt})<>MONTH({e}),{prv},{nxt})"


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


wb = load_workbook(WB)
ws = wb["CDS_Pricer"]
put(ws, "A28", "CDSW SCREEN OUTPUT", BOLD)
put(ws, "A29", "Settlement figures in the order the CDSW screen shows them. "
               "S = par spread implied by the curve (B15); C = contractual coupon (B6).", NOTE)

# accrual start = previous IMM roll, business-day adjusted
accr = mf(f"EDATE({ROLL},-3)")
S = "($B$15/10000)"
rows = [
    ("1st accrual start", f"={accr}", "mm/dd/yy",
     "Previous IMM roll, business-day adjusted. 20 Jun 2026 is a Saturday, so the "
     "screen shows 06/22/26."),
    ("Points upfront", f"=({S}-{CPN})*$B$12*100", "0.00000000",
     "(S - C) * RPV01 * 100, protection buyer."),
    ("Price", "=100-B31", "0.00000000", "100 - points upfront."),
    ("Principal", f"={DIR}*({S}-{CPN})*$B$12*{N}", CCY, "Clean upfront in cash."),
    ("Accrued days", "=MAX(0,CDS_Parameters!$B$4-B30+1)", "0",
     "ISDA counts both endpoints."),
    ("Accrued", f"=-{DIR}*{CPN}*{N}*B34/360", CCY, "Coupon * notional * days/360."),
    ("Cash Amount", "=B33+B35", CCY, "Principal + accrued."),
    ("Default Exposure", f"=(1-{R})*{N}-B33", CCY,
     "(1-R) * notional less the principal already paid."),
]
for i, (lab, f, fmt, note) in enumerate(rows):
    r = 30 + i
    hl = lab in ("Points upfront", "Price", "Cash Amount")
    put(ws, f"A{r}", lab, BOLD if hl else BLACK)
    put(ws, f"B{r}", f, BLACK, fmt, OF if hl else None, True)
    put(ws, f"C{r}", note, NOTE)

put(ws, "A39", "Reference CDSW screen (CINDBK 5Y, 07/21/26) for validation: points upfront "
               "-1.97596265, price 101.97596265, principal -197,597, accrued -8,333 (30 days), "
               "cash -205,930, default exposure 6,197,596. Those are TARGETS for a like-for-like "
               "trade, not model output - the demo spreads here are not that entity's curve.", NOTE)
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("CDSW panel added to CDS_Pricer")
