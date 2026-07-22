"""
CDS_Pricer.xlsx - the credit half, split out of the curve workbook.

The dependency runs one way: CDS -> Curve_Interface -> Bootstrap. Nothing on the
curve side reads a CDS sheet, so the cut is clean. Curve_Interface comes across
as a COPY and becomes an input sheet: its 131 ties to Bootstrap (D7 and K9:L73)
are replaced with values. Every other CDS formula is untouched, because the sheet
keeps its name.

Adds:
  Entities  - pick the live name, override any pulled input, switch a name off with 0
  Curves    - the three curves that matter here, charted

Discount curve is a dated snapshot, not a live link. Two workbooks in one folder
is exactly when a stray external link looks fine and then breaks later.
"""
import sys, datetime as dt
sys.path.insert(0,"/private/tmp/claude-501/-Users-nigelli-Desktop-Curve-Bootstrapping/a03716b7-925b-41c8-abbb-c9250247d791/scratchpad")
import vbacurve as V
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, ScatterChart, Reference, Series

SRC="bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
OUT="bloomberg/CDS_Pricer.xlsx"
CDS=["CDS_Entities","CDS_Parameters","CDS_Quotes","CDS_Schedule","Hazard_Bootstrap",
     "Hazard_Solver","CDS_Pricer","CDS_Validation","Brent_vs_Bisection"]
KEEP=CDS+["Curve_Interface"]
VALD=dt.date(2026,7,21)
CAP=[('1W',3.63760),('2W',3.67330),('3W',3.68739),('1M',3.69655),('2M',3.72255),('3M',3.77525),
('4M',3.82085),('5M',3.86220),('6M',3.90800),('7M',3.94150),('8M',3.96975),('9M',4.00070),
('10M',4.02865),('11M',4.05230),('1Y',4.07315),('18M',4.09090),('2Y',4.09656),('3Y',4.07655),
('4Y',4.06655),('5Y',4.07150),('6Y',4.08910),('7Y',4.11260),('8Y',4.13925),('9Y',4.16870),
('10Y',4.19995),('12Y',4.26400),('15Y',4.34865),('20Y',4.41995),('25Y',4.41615),('30Y',4.37685),
('40Y',4.25755),('50Y',4.12427)]

F="Calibri"
H1=Font(name=F,size=13,bold=True); B=Font(name=F,size=11,bold=True)
N=Font(name=F,size=11); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
IN=PatternFill("solid",fgColor="FFF2CC"); BLUE=Font(name=F,size=11,color="0000FF")
GRN=Font(name=F,size=11,color="008000"); RED=Font(name=F,size=11,bold=True,color="C00000")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(SRC)
for s in [s for s in wb.sheetnames if s not in KEEP]:
    del wb[s]
for n in list(wb.defined_names):
    del wb.defined_names[n]

# ---- Curve_Interface becomes an input sheet -----------------------------------
ci=wb["Curve_Interface"]
ten,dates,T,par,DF=V.build(VALD,CAP)
spot=V.addbus(VALD,2)
# A4:H6 is a merged note block, so the input goes right of the grid
ci["N5"]="Valuation date"; ci["N5"].font=B
ci["O5"]=dt.datetime(*VALD.timetuple()[:3]); ci["O5"].fill=IN; ci["O5"].font=BLUE
ci["O5"].number_format="mm/dd/yyyy"; ci["O5"].border=BOX
ci["N6"]="drives every date on the CDS sheets"; ci["N6"].font=SM
wb.defined_names.add(DefinedName("VAL_DATE",attr_text="Curve_Interface!$O$5"))

ci["D7"]=dt.datetime(*spot.timetuple()[:3]); ci["D7"].number_format="mm/dd/yyyy"
ci["K6"]="Curve grid - snapshot from the SOFR bootstrap"
ci["M6"]=f"pasted {VALD:%m/%d/%Y}, not linked"; ci["M6"].font=SM
# paste the FULL 65-pillar grid, matching what Bootstrap fed it. The gap years
# lie on the log-linear line between quoted pillars, so the curve is the same
# either way - but keeping all 65 means the interpolation grid is untouched.
r=9
for t,d,df in zip(ten,dates,DF):
    ci.cell(r,11,dt.datetime(*d.timetuple()[:3])).number_format="mm/dd/yyyy"
    ci.cell(r,12,round(df,12)).number_format="0.00000000"
    r+=1
for rr in range(r,74):
    ci.cell(rr,11,None); ci.cell(rr,12,None)
print(f"Curve_Interface: {r-9} pillars pasted as values (rows 9-{r-1})")

# ---- Entities: select, override, switch off -----------------------------------
if "Entities" in wb.sheetnames: del wb["Entities"]
es=wb.create_sheet("Entities",0)
ent=wb["CDS_Entities"]
names=[ent.cell(rr,1).value for rr in range(5,15) if ent.cell(rr,1).value]

es["A1"]="Entities"; es["A1"].font=H1
es["A2"]="Set On to 0 to take a name out. Overrides win over the Bloomberg pull; leave blank to use the pull."
es["A2"].font=SM
es["A4"]="Live name"; es["A4"].font=B
es["B4"]=names[0] if names else ""; es["B4"].fill=IN; es["B4"].font=BLUE; es["B4"].border=BOX
if names:
    dv=DataValidation(type="list",formula1='"'+",".join(names)+'"',allow_blank=False)
    es.add_data_validation(dv); dv.add(es["B4"])
es["C4"]="only this name is stripped"; es["C4"].font=SM

hdr=["On","Entity","Ticker","Ccy","Recovery","1Y","2Y","3Y","5Y","7Y","10Y"]
for i,h in enumerate(hdr,start=1):
    c=es.cell(6,i,h); c.font=HD; c.fill=HF; c.border=BOX
    c.alignment=Alignment(horizontal="center")
es["F5"]="override spreads (bp)"; es["F5"].font=SM
for j,nm in enumerate(names):
    rr=7+j; src=5+j
    es.cell(rr,1,1).fill=IN; es.cell(rr,1).font=BLUE
    es.cell(rr,2,f"=CDS_Entities!A{src}")
    es.cell(rr,3,f"=CDS_Entities!B{src}")
    es.cell(rr,4,f"=CDS_Entities!C{src}")
    es.cell(rr,5,None).fill=IN; es.cell(rr,5).font=BLUE; es.cell(rr,5).number_format="0.00"
    for k in range(6):
        c=es.cell(rr,6+k,None); c.fill=IN; c.font=BLUE; c.number_format="0.0"
    for k in range(1,12): es.cell(rr,k).border=BOX

last=6+len(names)
es.cell(last+2,1,"Effective (what the strip uses)").font=B
for i,h in enumerate(["","Entity","","","Recovery","1Y","2Y","3Y","5Y","7Y","10Y"],start=1):
    if h: c=es.cell(last+3,i,h); c.font=HD; c.fill=HF; c.border=BOX
for j,nm in enumerate(names):
    rr=last+4+j; ov=7+j; src=5+j
    es.cell(rr,2,f"=B{ov}")
    es.cell(rr,5,f'=IF($A{ov}=0,"off",IF(ISNUMBER(E{ov}),E{ov},CDS_Entities!F{src}))')
    for k,col in enumerate("HIJKLM"):
        # override cells sit in F:K (1Y 2Y 3Y 5Y 7Y 10Y) - chr(70+k), not 70+k+1,
        # or a 5Y override lands in the 3Y slot
        oc=chr(70+k)
        es.cell(rr,6+k,f'=IF($A{ov}=0,0,IF(ISNUMBER({oc}{ov}),{oc}{ov},'
                       f'IF(ISNUMBER(CDS_Entities!{col}{src}),CDS_Entities!{col}{src},0)))')
        es.cell(rr,6+k).number_format="0.0"
    for k in range(1,12): es.cell(rr,k).border=BOX; es.cell(rr,k).font=GRN
for col,w in zip("ABCDEFGHIJK",(6,26,18,7,10,9,9,9,9,9,9)):
    es.column_dimensions[col].width=w
print(f"Entities: {len(names)} names, On/Off + overrides")

# ---- Curves --------------------------------------------------------------------
if "Curves" in wb.sheetnames: del wb["Curves"]
cv=wb.create_sheet("Curves",1)
cv["A1"]="Curves"; cv["A1"].font=H1
cv["A2"]="Discount curve is the pasted snapshot. Hazard and survival come out of the strip."
cv["A2"].font=SM

cv["A4"]="SOFR discount curve"; cv["A4"].font=B
for i,h in enumerate(["Date","T","DF","Zero %"],start=1):
    c=cv.cell(5,i,h); c.font=HD; c.fill=HF; c.border=BOX
NPIL=65
for j in range(NPIL):
    rr=6+j; s=9+j
    cv.cell(rr,1,f"='Curve_Interface'!K{s}").number_format="mm/dd/yyyy"
    cv.cell(rr,2,f"=('Curve_Interface'!K{s}-VAL_DATE)/365").number_format="0.00"
    cv.cell(rr,3,f"='Curve_Interface'!L{s}").number_format="0.000000"
    cv.cell(rr,4,f'=IF(B{rr}<=0,"",-LN(C{rr})/B{rr}*100)').number_format="0.0000"

r0=6+NPIL+2
cv.cell(r0,1,"Hazard and survival").font=B
for i,h in enumerate(["Tenor","Hazard %","Survival Q"],start=1):
    c=cv.cell(r0+1,i,h); c.font=HD; c.fill=HF; c.border=BOX
for j in range(6):
    rr=r0+2+j; s=7+j
    cv.cell(rr,1,f"=Hazard_Bootstrap!A{s}")
    cv.cell(rr,2,f"=Hazard_Bootstrap!D{s}*100").number_format="0.0000"
    cv.cell(rr,3,f"=Hazard_Bootstrap!E{s}").number_format="0.0000"

r1=r0+2+6+2
cv.cell(r1,1,"Spread term structure").font=B
for i,h in enumerate(["Tenor","Market bp","Model bp","Diff bp"],start=1):
    c=cv.cell(r1+1,i,h); c.font=HD; c.fill=HF; c.border=BOX
for j in range(6):
    rr=r1+2+j; s=7+j
    cv.cell(rr,1,f"=Hazard_Bootstrap!A{s}")
    cv.cell(rr,2,f"=Hazard_Bootstrap!C{s}").number_format="0.00"
    cv.cell(rr,3,f"=Hazard_Bootstrap!I{s}").number_format="0.00"
    cv.cell(rr,4,f"=Hazard_Bootstrap!J{s}").number_format="0.0000"

def line(title,ytitle,minr,maxr,catcol,valcols,anchor,step=False):
    ch=LineChart(); ch.title=title; ch.y_axis.title=ytitle; ch.x_axis.title=None
    ch.height, ch.width = 7.5, 15
    for vc in valcols:
        ref=Reference(cv,min_col=vc,min_row=minr-1,max_row=maxr)
        ch.add_data(ref,titles_from_data=True)
    ch.set_categories(Reference(cv,min_col=catcol,min_row=minr,max_row=maxr))
    if step:
        for s in ch.series: s.smooth=False
    for s in ch.series: s.smooth=False
    cv.add_chart(ch,anchor)

line("SOFR zero curve","zero %",6,6+NPIL-1,1,[4],"F4")
line("Discount factors","DF",6,6+NPIL-1,1,[3],"F20")
line("Hazard (piecewise constant) and survival","%",r0+2,r0+7,1,[2,3],"F36")
line("Spread term structure - market vs model","bp",r1+2,r1+7,1,[2,3],"F52")
for col,w in zip("ABCD",(12,8,12,10)): cv.column_dimensions[col].width=w
print("Curves: 4 charts")

wb.active=0
wb.calculation.fullCalcOnLoad=True
wb.save(OUT)
print("built",OUT)
