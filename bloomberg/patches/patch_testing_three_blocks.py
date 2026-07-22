"""
Three working blocks on Testing, one per curve test case.

Was: one shared tenor column, one shared date column, and a single working block
for whichever case is active. Cases 2 and 3 had no visible working at all - you
could not check their dates or their pasted data without selecting them.

Now each case is a self-contained block with its own curve date, its own spot,
its own derived dates and its own comparison. Each block references only its own
curve-date cell, so the nested three-case IF is gone entirely - that nesting is
what produced the 8,729-character formulas that corrupted the file.

Model columns (our zero, our DF, tau, annuity, numerator, denominator, deltas)
populate only for the ACTIVE case, because there is one bootstrap and only one
case can be running through it. The other two blocks show their dates and pasted
data and say "not active" rather than showing stale numbers that look live.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidationList

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
H1 = Font(name="Calibri", size=12, bold=True)
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
MONO = Font(name="Consolas", size=9)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
CF = ["2E7D32", "C00000", "6A1B9A"]
YF = PatternFill("solid", fgColor="FFFF00")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
DTF, N5, N6, N8, BP = "mm/dd/yyyy", "0.00000", "0.000000", "0.00000000", "0.00"

MODE = "Bootstrap!$G$4"
TEST = ["Test 1", "Test 2", "Test 3"]
NROW = 40
BLOCKS = [8, 58, 108]          # first data row of each block
COLS = [("A","Tenor",10,None),("B","Swap rate (mid)",14,N5),("C","BBG zero %",12,N5),
        ("D","BBG discount",13,N6),("E","raw date",12,DTF),("F","Date",13,DTF),
        ("G","S used %",10,N5),("H","Rule",30,None),("I","tau",10,N6),
        ("J","A(prior)",12,N8),("K","numerator",12,N8),("L","denominator",12,N8),
        ("M","our DF",13,N8),("N","t ACT/365",10,N6),("O","our zero %",11,N5),
        ("P","d zero bp",10,BP),("Q","d DF",11,"0.00E+00")]
B = lambda c: f"Bootstrap!${c}$8:${c}$72"


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False, al=None):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    if al: c.alignment = Alignment(horizontal=al, wrap_text=True, vertical="center")
    return c


wb = load_workbook(WB)
del wb["Testing"]
ws = wb.create_sheet("Testing", wb.sheetnames.index("Bootstrap") + 1)
ws.sheet_properties.tabColor = "B7950B"
ws.data_validations = DataValidationList()
for c, _, w, _ in COLS:
    ws.column_dimensions[c].width = w

put(ws, "A1", "Testing — three curve test cases", H1)
put(ws, "A2", "Per case paste four columns off the capture: tenor, swap rate (mid), BBG zero, "
              "BBG discount. Curve date once in the block header.", NOTE)
put(ws, "A3", "The tenor is pasted with each case and dates derive from tenor + that case's "
              "curve date, so cases may differ in pillar set or length without shifting.", NOTE)
put(ws, "A4", "Model columns G:Q are live only for the case selected at Bootstrap!G4 — there "
              "is one bootstrap. The other two blocks show dates and pasted data and read "
              "'not active' rather than stale numbers.", WARN)
put(ws, "A5", "Active source", BOLD)
put(ws, "B5", f"={MODE}", BLACK, None, OF, True)
put(ws, "D5", "Benchmark: the 07/21/26 capture gives 0.397bp and 1.88e-05.", NOTE)

for bi, r0 in enumerate(BLOCKS):
    k = bi + 1
    band = PatternFill("solid", fgColor=CF[bi])
    hdr, cd, sp = r0 - 1, r0 - 3, r0 - 2
    r1 = r0 + NROW - 1
    live = f'$B${cd+1}=""'                       # placeholder, replaced below
    put(ws, f"A{cd}", f"CURVE TEST CASE {k}", H1)
    put(ws, f"C{cd}", "curve date ->", NOTE)
    put(ws, f"D{cd}", None, BLUE, DTF, YF, True)
    put(ws, f"E{cd}", "name / time ->", NOTE)
    put(ws, f"F{cd}", None, BLUE, None, YF, True)
    CDATE, SPOT = f"$D${cd}", f"$D${sp}"
    put(ws, f"C{sp}", "spot (T+2bd) ->", NOTE)
    put(ws, f"D{sp}", f'=IF({CDATE}="","",{CDATE}+2+IF(WEEKDAY({CDATE},2)>=4,2,0))',
        BLACK, DTF, OF, True)
    ACT = f'{MODE}="{TEST[bi]}"'
    put(ws, f"F{sp}", f'=IF({ACT},"ACTIVE — model columns live","not active")', BOLD)

    for c, h, w, fmt in COLS:
        put(ws, f"{c}{hdr}", h, HDR, None, band, True, "center")
    ws.row_dimensions[hdr].height = 30

    for r in range(r0, r1 + 1):
        for c, fmt in (("A", None), ("B", N5), ("C", N5), ("D", N6)):
            put(ws, f"{c}{r}", None, BLUE, fmt, YF, True)
        n = f"VALUE(LEFT(A{r},LEN(A{r})-1))"
        raw = (f'IF(RIGHT(A{r},1)="W",{SPOT}+7*{n},IF(RIGHT(A{r},1)="M",'
               f'EDATE({SPOT},{n}),EDATE({SPOT},12*{n})))')
        put(ws, f"E{r}", f'=IF(OR(A{r}="",{SPOT}=""),"",{raw})', BLACK, DTF, None, True)
        nxt = f'(E{r}+IF(WEEKDAY(E{r},2)=6,2,IF(WEEKDAY(E{r},2)=7,1,0)))'
        prv = f'(E{r}-IF(WEEKDAY(E{r},2)=6,1,IF(WEEKDAY(E{r},2)=7,2,0)))'
        put(ws, f"F{r}", f'=IF(E{r}="","",IF(MONTH({nxt})<>MONTH(E{r}),{prv},{nxt}))',
            BLACK, DTF, None, True)
        m = f"MATCH(F{r},{B('B')},0)"
        gate = f'=IF(NOT({ACT}),"",IFERROR('
        put(ws, f"G{r}", f'{gate}INDEX({B("E")},{m}),""))', BLACK, N5, None, True)
        put(ws, f"H{r}", f'{gate}IF({m}<=15,"short  DFspot/(1+S*tau)",'
                         f'"annual (DFspot-S*A)/(1+S*tau)"),""))', MONO, None, None, True)
        put(ws, f"I{r}", f'{gate}IF({m}<=15,INDEX({B("D")},{m}),INDEX({B("F")},{m})),""))',
            BLACK, N6, None, True)
        put(ws, f"J{r}", f'{gate}IF({m}<=15,0,INDEX({B("G")},{m})),""))', BLACK, N8, None, True)
        put(ws, f"K{r}", f'=IF(OR(G{r}="",I{r}=""),"",Bootstrap!$D$4-(G{r}/100)*J{r})',
            BLACK, N8, None, True)
        put(ws, f"L{r}", f'=IF(OR(G{r}="",I{r}=""),"",1+(G{r}/100)*I{r})', BLACK, N8, None, True)
        put(ws, f"M{r}", f'{gate}INDEX({B("H")},{m}),""))', BLACK, N8, None, True)
        put(ws, f"N{r}", f'{gate}INDEX({B("C")},{m}),""))', BLACK, N6, None, True)
        put(ws, f"O{r}", f'{gate}INDEX({B("J")},{m}),""))', BLACK, N5, None, True)
        put(ws, f"P{r}", f'=IF(OR(O{r}="",NOT(ISNUMBER(C{r}))),"",(O{r}-C{r})*100)',
            BLACK, BP, None, True)
        put(ws, f"Q{r}", f'=IF(OR(M{r}="",NOT(ISNUMBER(D{r}))),"",M{r}-D{r})',
            BLACK, "0.00E+00", None, True)

    s = r1 + 1
    put(ws, f"A{s}", "pillars pasted", BOLD)
    put(ws, f"C{s}", f'=COUNTA(A{r0}:A{r1})', BLACK, "0", OF, True)
    put(ws, f"D{s}", "matched", BOLD)
    put(ws, f"F{s}", f'=IF(NOT({ACT}),"not active",COUNT(O{r0}:O{r1}))', BLACK, None, OF, True)
    put(ws, f"G{s}", "max |d zero| bp", BOLD)
    put(ws, f"I{s}", f'=IF(NOT({ACT}),"",IF(COUNT(P{r0}:P{r1})=0,"",'
                     f'MAX(MAX(P{r0}:P{r1}),-MIN(P{r0}:P{r1}))))', BLACK, "0.000", OF, True)
    put(ws, f"J{s}", "max |d DF|", BOLD)
    put(ws, f"L{s}", f'=IF(NOT({ACT}),"",IF(COUNT(Q{r0}:Q{r1})=0,"",'
                     f'MAX(MAX(Q{r0}:Q{r1}),-MIN(Q{r0}:Q{r1}))))', BLACK, "0.00E+00", OF, True)
    put(ws, f"M{s}", f'=IF(AND({ACT},C{s}>0,F{s}<C{s}),"CHECK: matched < pasted, dates are '
                     f'not hitting Bootstrap!B — check the curve date","")', WARN)

ws.freeze_panes = "A8"
ws.sheet_view.showGridLines = False
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"Testing rebuilt: {len(BLOCKS)} self-contained blocks of {NROW} rows")
