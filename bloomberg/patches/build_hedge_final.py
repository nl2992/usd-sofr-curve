"""
Rebuild the Hedge sheet with the cost analysis placed RIGHT UNDER the trades:
  1 risk | 2 strategy | 3 trades | 4 cost & liquidity (KWSWNI) | 5 residual | 6 optional
Then delete the standalone 'Hedge Cost Data Pull' sheet.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
WB="bloomberg/KRW_IRS_Bootstrap_Book1.xlsx"; SRC="Sheet1"
F="Calibri"
TITLE=Font(name=F,size=15,bold=True,color="1F3864")
SEC=Font(name=F,size=11,bold=True,color="FFFFFF"); SECF=PatternFill("solid",fgColor="1F3864")
B=Font(name=F,size=10,bold=True); SM=Font(name=F,size=9,color="808080"); N=Font(name=F,size=10)
HD=Font(name=F,size=9,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
GRN=Font(name=F,size=10,bold=True,color="006100"); RED=Font(name=F,size=10,bold=True,color="9C0006")
INP=Font(name=F,size=10,color="0000FF"); MONO=Font(name="Consolas",size=9)
KEY=PatternFill("solid",fgColor="FCE4D6"); BLU=PatternFill("solid",fgColor="DDEBF7"); INPF=PatternFill("solid",fgColor="FFF2CC")
SKIPF=PatternFill("solid",fgColor="FCE4D6")
thin=Side(style="thin",color="BFBFBF"); BOX=Border(thin,thin,thin,thin)
MONEY='#,##0'; BN='#,##0.0'
# Sheet1 folded-table rows
S={"3M":6,"6M":7,"9M":8,"1Y":9,"2Y":10,"3Y":11,"4Y":12,"5Y":13,"7Y":14,"10Y":15}
def c(t): return f"{SRC}!$B${S[t]}"
TAU="Quotes!$B$2"; YR="Bootstrap!$B$3:$B$122"; DF="Bootstrap!$D$3:$D$122"

wb=load_workbook(WB)
if "Hedge Cost Data Pull" in wb.sheetnames: del wb["Hedge Cost Data Pull"]
if "Hedge" in wb.sheetnames: del wb["Hedge"]
ws=wb.create_sheet("Hedge", wb.sheetnames.index("KRD-v3")+1)

ws["A1"]="Hedge proposal — KRW amortising IRS (10Y, HSBC receives fixed)"; ws["A1"].font=TITLE
ws["A2"]="Bucketed KRD hedged at liquid benchmarks; cost analysis (live bid/offer + SDRV volume) sits under the trades."
ws["A2"].font=SM
def section(r,txt,ncol=10):
    ws.cell(r,1,txt).font=SEC
    for cc in range(1,ncol+1): ws.cell(r,cc).fill=SECF

# annuity helper (right)
ws.cell(4,12,"Benchmark annuity").font=B
for i,h in enumerate(["Tenor","T","Annuity"],12):
    cc=ws.cell(5,i,h); cc.font=HD; cc.fill=HF
ANN={}
for k,(tn,T) in enumerate([("2Y",2),("3Y",3),("5Y",5),("7Y",7),("10Y",10)]):
    r=6+k; ws.cell(r,12,tn); ws.cell(r,13,T)
    ws.cell(r,14,f'={TAU}*SUMIF({YR},"<="&M{r},{DF})').number_format="0.0000"
    ANN[tn]=f"$N${r}"

# 1 · risk
section(4,"1 · The risk")
ws.cell(5,1,"Net DV01").font=B
ws.cell(5,2,f"={SRC}!$B$21").number_format=MONEY; ws.cell(5,2).font=RED; ws.cell(5,2).fill=KEY
ws.cell(5,3,"KRW per bp — net receiver").font=SM
ws.cell(6,1,"Front 1Y–3Y").font=B
ws.cell(6,2,f"={c('1Y')}+{c('2Y')}+{c('3Y')}").number_format=MONEY; ws.cell(6,2).font=GRN
ws.cell(6,3,"long (positive)").font=SM
ws.cell(7,1,"Back 4Y–10Y").font=B
ws.cell(7,2,f"={c('4Y')}+{c('5Y')}+{c('7Y')}+{c('10Y')}").number_format=MONEY; ws.cell(7,2).font=RED
ws.cell(7,3,"short (negative), concentrated at 10Y").font=SM
ws.cell(8,1,"Sub-1Y (3M/6M/9M)").font=B
ws.cell(8,2,f"={c('3M')}+{c('6M')}+{c('9M')}").number_format=MONEY; ws.cell(8,2).font=SM
ws.cell(8,3,"dust — leave").font=SM

# 2 · strategy
section(10,"2 · Strategy")
for i,t in enumerate([
    "Hedge liquid benchmarks (2Y/3Y/5Y/10Y), not every bucket.",
    "Fold orphans into the nearest benchmark: 1Y→2Y, 4Y→5Y, 7Y→10Y.",
    "Skip the sub-1Y dust — bid/offer > risk."]):
    ws.cell(11+i,1,f"•  {t}").font=N

# 3 · trades
section(14,"3 · The trades  (4 swaps)")
for i,h in enumerate(["Swap","Direction","Notional (bn)","Absorbs","Group DV01 (KRW/bp)","Annuity"],1):
    cc=ws.cell(15,i,h); cc.font=HD; cc.fill=HF; cc.border=BOX
blot=[("2Y","1Y + 2Y",f"={c('1Y')}+{c('2Y')}","2Y"),("3Y","3Y",f"={c('3Y')}","3Y"),
      ("5Y","4Y + 5Y",f"={c('4Y')}+{c('5Y')}","5Y"),("10Y","7Y + 10Y",f"={c('7Y')}+{c('10Y')}","10Y")]
for k,(tn,ab,dv,ak) in enumerate(blot):
    r=16+k
    ws.cell(r,1,tn).font=B
    ws.cell(r,2,f'=IF(E{r}>0,"Receive fixed","Pay fixed")').font=B
    ws.cell(r,3,f"=ABS(E{r}/(F{r}*0.0001))/1000000000").number_format=BN; ws.cell(r,3).font=GRN; ws.cell(r,3).fill=BLU
    ws.cell(r,4,ab); ws.cell(r,5,dv).number_format=MONEY
    ws.cell(r,6,f"={ANN[ak]}").number_format="0.0000"
    for cc in range(1,7): ws.cell(r,cc).border=BOX

# 4 · cost & liquidity  (right under trades)
section(21,"4 · Cost to hedge & liquidity   (KWSWNI tickers — confirm on GC S205 > Table)",14)
hdr=["Tenor","Ticker (edit)","Bid","Ask","Spread (bp)","Avg spr 1M (bp)","DV01 (KRW/bp)","Annuity",
     "Notl (bn)","Cost (KRW)","Break-even (bp)","Vol USDmm","Verdict"]
for i,h in enumerate(hdr,1):
    cc=ws.cell(22,i,h); cc.font=HD; cc.fill=HF; cc.border=BOX
# 4 trades reference the blotter rows 16-19 ; skip rows self-source from Sheet1
crows=[
 ("2Y","KWSWNI2 BGN Curncy","2Y",16,False),
 ("3Y","KWSWNI3 BGN Curncy","3Y",17,False),
 ("5Y","KWSWNI5 BGN Curncy","5Y",18,False),
 ("10Y","KWSWNI10 BGN Curncy","10Y",19,False),
 ("7Y split","KWSWNI7 BGN Curncy","7Y",None,True),
 ("Dust 3-9M","(front=CD/FRA, GC S205)","1Y",None,True),
]
r0=23
for k,(tn,tk,ak,br,skip) in enumerate(crows):
    r=r0+k
    ws.cell(r,1,tn).font=B
    ws.cell(r,2,tk).font=INP; ws.cell(r,2).fill=INPF
    ws.cell(r,3,f'=BDP(B{r},"PX_BID")').number_format="0.000"
    ws.cell(r,4,f'=BDP(B{r},"PX_ASK")').number_format="0.000"
    ws.cell(r,5,f"=(D{r}-C{r})*100").number_format="0.00"
    ws.cell(r,6,f'=(AVERAGE(BDH(B{r},"PX_ASK",TEXT(TODAY()-30,"yyyymmdd"),TEXT(TODAY(),"yyyymmdd")))'
                f'-AVERAGE(BDH(B{r},"PX_BID",TEXT(TODAY()-30,"yyyymmdd"),TEXT(TODAY(),"yyyymmdd"))))*100').number_format="0.00"
    if br:  # 4 trades: pull DV01/annuity/notional from the blotter line above
        ws.cell(r,7,f"=E{br}").number_format=MONEY
        ws.cell(r,8,f"=F{br}").number_format="0.0000"
        ws.cell(r,9,f"=C{br}").number_format=BN
    else:   # skip candidates: self-source
        dv = f"={c('7Y')}" if tn.startswith("7Y") else f"={c('3M')}+{c('6M')}+{c('9M')}"
        ws.cell(r,7,dv).number_format=MONEY
        ws.cell(r,8,f"={ANN['7Y']}" if tn.startswith("7Y") else f'={TAU}*SUMIF({YR},"<="&1,{DF})').number_format="0.0000"
        ws.cell(r,9,f"=ABS(G{r}/(H{r}*0.0001))/1000000000").number_format=BN
    ws.cell(r,10,f"=I{r}*1000000000*H{r}*(E{r}/2)*0.0001").number_format=MONEY
    ws.cell(r,11,f"=J{r}/ABS(G{r})").number_format="0.00"
    ws.cell(r,12,0).font=INP; ws.cell(r,12).fill=INPF
    ws.cell(r,13,f'=IF(K{r}<1,"TRADE","SKIP")'); ws.cell(r,13).font=RED if skip else GRN
    for cc in range(1,14):
        ws.cell(r,cc).border=BOX
        if skip and cc not in (2,12): ws.cell(r,cc).fill=SKIPF
rt=r0+len(crows)
ws.cell(rt,1,"TOTAL cost (4 trades)").font=B
ws.cell(rt,10,f"=SUM(J{r0}:J{r0+3})").number_format=MONEY; ws.cell(rt,10).font=B
ws.cell(rt+1,1,"Yellow = inputs: exact ticker from GC S205 > Table, and SDRV volume (USDmm) per tenor.").font=SM
ws.cell(rt+2,1,"Break-even (bp) = cost / DV01 ≈ half the bid/offer. <1bp -> TRADE. Cross-check with SDRV volume.").font=SM
ws.cell(rt+3,1,'Commands:  GC S205 <GO> (tickers/curve) · SDRV <GO> Rates>IRS/FRA>KRW (volume) · SDR <GO> (trade drill).').font=MONO

# 5 · residual
rr=rt+5
section(rr,"5 · After the hedge")
ws.cell(rr+1,1,"Grouped buckets 2Y/3Y/5Y/10Y").font=B; ws.cell(rr+1,3,"neutralized to 0").font=N
ws.cell(rr+2,1,"Residual DV01").font=B
ws.cell(rr+2,3,f"={c('3M')}+{c('6M')}+{c('9M')}").number_format=MONEY; ws.cell(rr+2,3).fill=KEY; ws.cell(rr+2,3).font=B
ws.cell(rr+2,4,"sub-1Y dust only").font=SM

# 6 · optional
ro=rr+4
section(ro,"6 · Optional — split 7Y out (kills 7s10s residual, +1 ticket)")
for i,h in enumerate(["Swap","Direction","Notional (bn)","Absorbs","Group DV01 (KRW/bp)","Annuity"],1):
    cc=ws.cell(ro+1,i,h); cc.font=HD; cc.fill=HF; cc.border=BOX
opt=[("7Y","7Y",f"={c('7Y')}","7Y"),("10Y","10Y",f"={c('10Y')}","10Y")]
for k,(tn,ab,dv,ak) in enumerate(opt):
    r=ro+2+k
    ws.cell(r,1,tn).font=B
    ws.cell(r,2,f'=IF(E{r}>0,"Receive fixed","Pay fixed")').font=B
    ws.cell(r,3,f"=ABS(E{r}/(F{r}*0.0001))/1000000000").number_format=BN; ws.cell(r,3).font=GRN
    ws.cell(r,4,ab); ws.cell(r,5,dv).number_format=MONEY; ws.cell(r,6,f"={ANN[ak]}").number_format="0.0000"
    for cc in range(1,7): ws.cell(r,cc).border=BOX

for col,w in zip("ABCDEFGHIJKLMN",(11,22,8,8,11,14,18,9,10,15,13,11,10,10)): ws.column_dimensions[col].width=w
wb.save(WB)
print("Hedge rebuilt with cost under trades. sheets:",wb.sheetnames)
