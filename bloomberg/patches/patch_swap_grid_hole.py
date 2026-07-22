"""
Fix a hole in the Swap_Pricer curve grid.

A prose note had been written INTO the grid and MERGED across K39:L40,
destroying all four of those data cells. The grid maps K(m) -> Bootstrap!B(m+1), so
K39 and K40 should hold the 18Y and 19Y nodes (Bootstrap rows 40 and 41). The
Bootstrap row sequence jumped 39 -> 42.

MATCH(date, $K$6:$K$71, 1) requires ascending order. A text value and a blank
break the sequence, so every coupon date past 17Y matched at K38 and the
interpolation fell back (via its own IFERROR) to a flat DF(17Y).

Symptom: swaps up to 15Y repriced their input to within 0.03bp, while 20Y/30Y/50Y
were 1.3-2.1bp off - a discontinuity, not a drift, because only tenors with
coupons beyond 17Y touched the hole.

Curve_Interface's grid was checked and is contiguous, so the CDS module was never
affected by this.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
wb = load_workbook(WB)
sp = wb["Swap_Pricer"]
# the note had been MERGED across K39:L40, destroying all four grid cells
if "K39:L40" in [str(m) for m in sp.merged_cells.ranges]:
    sp.unmerge_cells("K39:L40")
sp["K39"] = "=Bootstrap!B40"      # 18Y
sp["L39"] = "=Bootstrap!H40"
sp["K40"] = "=Bootstrap!B41"      # 19Y
sp["L40"] = "=Bootstrap!H41"
for c in ("K39", "K40"):
    sp[c].number_format = "mm/dd/yy"
for c in ("L39", "L40"):
    sp[c].number_format = "0.00000000"
# the displaced note goes somewhere that is not a data cell
sp["N4"] = ("Log-linear DF interpolation on this grid (dates $K$6:$K$71, DFs $L$6:$L$71). "
            "Do NOT write text into K or L - MATCH needs a strictly ascending date column, "
            "and a stray label here silently flattened every DF past 17Y.")
sp["N4"].font = Font(name="Calibri", size=9, italic=True, color="C00000")
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("grid hole closed: K39/K40 now carry the 18Y and 19Y nodes")
