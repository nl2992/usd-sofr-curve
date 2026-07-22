"""
Repair the two faults that made Excel report damaged content.

1. Bootstrap!G4 carried TWO data validations - patch_curve_mode added the
   Live/Fixed list and patch_testing added the Live/Fixed/Test1-3 list without
   clearing the first. Two validations on one cell is invalid. Rebuilt with one.

2. Testing!B7:B38 were 8,729 characters, over Excel's 8,192 formula limit. The
   date formula nested the three-case IF inside spot inside the raw date inside
   both business-day-roll branches, so it expanded combinatorially. Broken into
   helpers: M4 active curve date, M5 spot, L the unrolled date, B the roll -
   each a few hundred characters.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
BOLD = Font(name="Calibri", size=11, bold=True)
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
DTF = "mm/dd/yyyy"
MODE = "Bootstrap!$G$4"
T = ["Test 1", "Test 2", "Test 3"]
LIVE, FIXED = "Live (BDP)", "Fixed (S490 07/21/26)"
R0, R1 = 7, 38


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


wb = load_workbook(WB)

# ---- 1. one validation on G4
b = wb["Bootstrap"]
b.data_validations = DataValidationList()
dv = DataValidation(type="list",
                    formula1=f'"{LIVE},{FIXED},{T[0]},{T[1]},{T[2]}"', allow_blank=False)
b.add_data_validation(dv)
dv.add(b["G4"])

# ---- 2. break the date formula into helpers
ws = wb["Testing"]
cd = (f'IF({MODE}="{T[0]}",$C$4,IF({MODE}="{T[1]}",$F$4,'
      f'IF({MODE}="{T[2]}",$I$4,"")))')
put(ws, "L4", "curve date ->", NOTE)
put(ws, "M4", f"={cd}", BLACK, DTF, OF, True)
put(ws, "L5", "spot (T+2bd) ->", NOTE)
put(ws, "M5", '=IF($M$4="","",$M$4+2+IF(WEEKDAY($M$4,2)>=4,2,0))', BLACK, DTF, OF, True)
put(ws, "L6", "raw date", BOLD)
ws.column_dimensions["L"].width = 14
ws.column_dimensions["M"].width = 14

n = f"VALUE(LEFT(A{{r}},LEN(A{{r}})-1))"
for r in range(R0, R1 + 1):
    nn = n.format(r=r)
    put(ws, f"L{r}",
        f'=IF($M$5="","",IF(RIGHT(A{r},1)="W",$M$5+7*{nn},'
        f'IF(RIGHT(A{r},1)="M",EDATE($M$5,{nn}),EDATE($M$5,12*{nn}))))',
        BLACK, DTF, None, True)
    nxt = f'(L{r}+IF(WEEKDAY(L{r},2)=6,2,IF(WEEKDAY(L{r},2)=7,1,0)))'
    prv = f'(L{r}-IF(WEEKDAY(L{r},2)=6,1,IF(WEEKDAY(L{r},2)=7,2,0)))'
    put(ws, f"B{r}", f'=IF(L{r}="","",IF(MONTH({nxt})<>MONTH(L{r}),{prv},{nxt}))',
        BLACK, DTF, None, True)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("repaired: single validation on Bootstrap!G4; Testing dates broken into helpers")
