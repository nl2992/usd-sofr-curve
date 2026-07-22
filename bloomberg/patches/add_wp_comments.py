"""
Tie the CDS workbook back to the Bloomberg B-Model white paper.

Source: bloomberg/reference/cds_white_paper/ (photographed pages). Page numbers
below are the paper's own, read off those pages - 6 (CDS Pricer static data and
conventions), 7 (3.1-3.4), 8 (3.5, 3.6 and the Curve Stripper), 9 (strippability,
credit triangle, DV01s).

Adds cell comments on the formulas that implement an equation, plus a Model_Notes
sheet holding the equations themselves and a strippability test taken from p.9.
"""
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

WB="bloomberg/CDS_Pricer.xlsx"
F="Calibri"
H1=Font(name=F,size=13,bold=True); B=Font(name=F,size=11,bold=True)
N=Font(name=F,size=10); SM=Font(name=F,size=9,color="808080")
MONO=Font(name="Consolas",size=10)
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)
RED=Font(name=F,size=10,bold=True,color="C00000")

wb=load_workbook(WB)

C={
 ("CDS_Pricer","B8"):  "p.6 — accrual dates are IMM: 3/20, 6/20, 9/20, 12/20. Payment dates roll to the\nnext business day; the last accrual end date T_M does not roll.",
 ("CDS_Pricer","B12"): "(3.4) coupons-in-survival + (3.5) accrued-on-default.\nRPV01 is the premium leg per 1bp of coupon.",
 ("CDS_Pricer","B13"): "(3.6)  Protection Leg = A(1-R) * INT_T^TM (-dp_s/dt) D(t) dt\np.8: the timeline is cut into segments short enough to treat the forward\ndiscount rate and the hazard as constant, and no longer than a premium\naccrual period. The integral is then analytic on each segment.",
 ("CDS_Pricer","B14"): "(3.4)  Coupons-in-Survival = A * C * SUM delta(T_i-1,T_i) p_s(T_i) D(T_i)",
 ("CDS_Pricer","B15"): "(3.3)  S = Protection Leg / (Premium Leg - Accrued Interest * D(T_s)) at C=1bp.\np.7: the par spread is the replacement spread, quoted in bp, expressed as the\nprotection leg over the premium leg net of accrued interest.",
 ("CDS_Pricer","B18"): "(3.2)  Upfront = Market Value / D(T_s) + Accrued Interest.\np.7: quoted at the settlement date T_s, typically 3 business days after pricing,\nso the market value is compounded from T to T_s. Identically 0 when C = S.",
 ("CDS_Pricer","B20"): "(3.1)  Market Value = Protection Leg - Premium Leg, both discounted to the\npricing date T.",
 ("CDS_Pricer","B5"):  "p.6 — the recovery rate is a model parameter, conventionally 40%, not an\nobservable. It sets the payout (1-R)A and feeds the credit triangle.",
 ("CDS_Pricer","B6"):  "p.6 — coupons are standardised to 100bp / 500bp for most names, ACT/360.\nOvertype to price an off-market coupon.",
 ("Hazard_Bootstrap","D7"): "Section 4, p.8 — sequential 1-D root-finding. h_1 is solved on [0, T_M1] to\nmatch the first quote; each later h_i+1 is solved on (T_Mi, T_Mi+1] holding all\nearlier hazards fixed. That is why this column is solved top-down, not jointly.",
 ("Hazard_Bootstrap","E7"): "p.8 — the hazard is the intensity of decline of survival. Default is a\nterminating event, so p_s cannot increase and h >= 0 is a hard constraint the\nmodel cannot violate.",
 ("Hazard_Bootstrap","J7"): "Repricing error: model spread (3.3) minus the market quote. This is the\nclosure check on the strip - it must go to zero at every quoted maturity.",
 ("CDS_Quotes","G7"): "p.9 credit triangle — (1-R) p_d(T) ~= S*T for small p_d, giving h ~= S/(1-R).\nExact only for a flat curve; used here as a starting guess and a sanity check,\nnot as the answer.",
 ("CDS_Schedule","A6"): "p.8 — one row per segment. Segments are short enough to justify constant\nforward discounting and constant hazard, and never longer than a premium\naccrual period, so (3.5) and (3.6) are analytic on each.",
 ("Curve_Interface","J7"): "p.6 — since the IBOR to RFR transition the standard ISDA model, and the\nB-Model in CDSW, discount USD off the SOFR RFR curve. That is what this grid is.",
 ("CDS_Validation","B7"): "p.8 — h(t) >= 0 is a hard model constraint, not a fitted outcome. A negative\nhazard means the strip has failed, not that the name is unusual.",
}
n=0
for (sh,cell),txt in C.items():
    if sh in wb.sheetnames:
        c=wb[sh][cell]
        c.comment=Comment(txt,"B-Model white paper"); c.comment.width=460; c.comment.height=150
        n+=1
print(f"{n} cell comments")

if "Model_Notes" in wb.sheetnames: del wb["Model_Notes"]
ws=wb.create_sheet("Model_Notes",2)
r=1
def put(txt,font=N,col=1):
    global r
    ws.cell(r,col,txt).font=font; r+=1
put("Model notes — Bloomberg B-Model (CDSW)",H1)
put("Equation numbers and page cites are the white paper's own. Scans in bloomberg/reference/cds_white_paper/.",SM)
r+=1
put("Pricing",B)
for eq,txt in [
 ("(3.1)","Market Value = Protection Leg - Premium Leg,  both discounted to pricing date T"),
 ("(3.2)","Upfront = Market Value / D(T_s) + Accrued Interest"),
 ("(3.3)","S = Protection Leg / (Premium Leg - Accrued Interest * D(T_s))  at C = 1bp"),
 ("(3.4)","Coupons-in-Survival = A * C * SUM delta(T_i-1, T_i) p_s(T_i) D(T_i)"),
 ("(3.5)","Accrued-on-Default = A * C * INT_T^TM delta(T_n(t), t) (-dp_s/dt) D(t) dt"),
 ("(3.6)","Protection Leg = A (1-R) * INT_T^TM (-dp_s/dt) D(t) dt")]:
    ws.cell(r,1,eq).font=MONO; ws.cell(r,2,txt).font=MONO; r+=1
r+=1
put("Conventions that the sheets depend on",B)
for p,txt in [
 ("p.6","Accrual dates are IMM (3/20, 6/20, 9/20, 12/20). Payment dates roll to the next business day; the last accrual end date T_M does not roll."),
 ("p.6","Coupon standardised to 100bp / 500bp for most names. ACT/360."),
 ("p.6","Protection starts immediately, so the protection period is T_M - T + 1 days."),
 ("p.7","Accrued interest includes the pricing date."),
 ("p.7","T_s = settlement, typically 3 business days after pricing. The upfront is quoted there, so market value is compounded T to T_s."),
 ("p.6","USD discounting is the SOFR RFR curve — ISDA moved to RFR curves and the B-Model follows."),
 ("p.6","Recovery R is a model parameter, conventionally 40%, not an observable.")]:
    ws.cell(r,1,p).font=SM; ws.cell(r,2,txt).font=N; r+=1
r+=1
put("Curve stripper (section 4, p.8)",B)
for txt in [
 "Sequential one-dimensional root-finding, not a joint fit. h_1 is solved on [0, T_M1] to match the first quote;",
 "each later h_i+1 is solved on (T_Mi, T_Mi+1] with all earlier hazards held fixed.",
 "Hazard is piecewise constant. h(t) >= 0 is a hard constraint - default is terminating, so survival cannot rise.",
 "The timeline is discretised finely enough that the forward discount rate and hazard are constant on a segment,",
 "and no segment is longer than a premium accrual period; (3.5) and (3.6) are then analytic on each segment."]:
    put(txt)
r+=1
put("Strippability (p.9)",B)
for txt in [
 "With h_i+1 running 0 to infinity the achievable par spread at the next maturity is a bounded interval, and it",
 "narrows sharply once survival is already low. A quote outside that interval makes the calibration fail.",
 "The paper's reading: usually inconsistent data sources rather than genuine arbitrage. Common with inverted",
 "curves (S_i+1 < S_i, though mildly inverted curves are often still strippable) and distressed non-flat curves.",
 "Approximate lower bound, valid while cumulative default probability is low:   S_i+1  >=  S_i * T_i / T_i+1"]:
    put(txt)
r+=1
put("Strippability check on the live quotes",B)
hdr=r
for i,h in enumerate(["Tenor","T","S (bp)","bound S*T_prev/T","ok?"],start=1):
    c=ws.cell(hdr,i,h); c.font=HD; c.fill=HF; c.border=BOX
for j in range(6):
    rr=hdr+1+j; s=7+j
    ws.cell(rr,1,f"=Hazard_Bootstrap!A{s}")
    ws.cell(rr,2,f"=CDS_Quotes!B{s}")
    ws.cell(rr,3,f"=Hazard_Bootstrap!C{s}").number_format="0.00"
    if j==0:
        ws.cell(rr,4,"—").font=SM
        ws.cell(rr,5,"first pillar, unconstrained").font=SM
    else:
        ws.cell(rr,4,f"=C{rr-1}*B{rr-1}/B{rr}").number_format="0.00"
        ws.cell(rr,5,f'=IF(C{rr}>=D{rr},"ok","below bound — may not strip")')
    for k in range(1,6): ws.cell(rr,k).border=BOX
r=hdr+8
put("A fail here is the p.9 condition, and it is a warning about the quotes, not about the solver.",SM)
put("The bound is approximate; it holds while cumulative default probability is low.",SM)
r+=1
put("Not covered by any of this",RED)
put("The strip has no external validation — Bloomberg exports no hazard curve (Help Desk H#1330731572).",N)
put("Spreads in this workbook are demo values, not market. The SOFR curve is real; the credit curve is not.",N)
ws.column_dimensions["A"].width=13
ws.column_dimensions["B"].width=104
for col in "CDE": ws.column_dimensions[col].width=15
print("Model_Notes sheet added")

wb.calculation.fullCalcOnLoad=True
wb.save(WB)
print("saved",WB)
