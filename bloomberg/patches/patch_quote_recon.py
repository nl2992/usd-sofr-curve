"""
Add a reconciliation block to SOFR_OIS_Quotes: paste the Bloomberg curve export's
Bid/Ask next to the BDP pull and diff them.

Settles in one glance whether a BDP-vs-screen difference is timing (small, mixed
sign, moves between refreshes) or systematic (consistent shift across all 32,
i.e. wrong pricing source, field or ticker).

Columns
  N  export Bid   (paste)      Q  d Bid  (bp)
  O  export Ask   (paste)      R  d Ask  (bp)
  P  export Mid   (derived)    S  d Mid  (bp)   <- the one that matters, H feeds the curve
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
SECT = Font(name="Calibri", size=11, bold=True)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
YF = PatternFill("solid", fgColor="FFFF00")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

wb = load_workbook(WB)
ws = wb["SOFR_OIS_Quotes"]
for c, h in [("N", "Export Bid (paste)"), ("O", "Export Ask (paste)"), ("P", "Export Mid"),
             ("Q", "d Bid (bp)"), ("R", "d Ask (bp)"), ("S", "d Mid (bp)")]:
    ws[f"{c}4"] = h
    ws[f"{c}4"].font = SECT
    ws[f"{c}4"].border = BOX
for r in range(5, 37):
    if ws[f"A{r}"].value is None:
        continue
    for c in ("N", "O"):
        ws[f"{c}{r}"].fill = YF
        ws[f"{c}{r}"].font = BLUE
        ws[f"{c}{r}"].number_format = "0.00000"
        ws[f"{c}{r}"].border = BOX
    ws[f"P{r}"] = f'=IF(COUNT(N{r}:O{r})=2,(N{r}+O{r})/2,"")'
    ws[f"Q{r}"] = f'=IF(AND(ISNUMBER(N{r}),ISNUMBER(E{r})),(E{r}-N{r})*100,"")'
    ws[f"R{r}"] = f'=IF(AND(ISNUMBER(O{r}),ISNUMBER(F{r})),(F{r}-O{r})*100,"")'
    ws[f"S{r}"] = f'=IF(AND(ISNUMBER(P{r}),ISNUMBER(H{r})),(H{r}-P{r})*100,"")'
    for c in ("P",):
        ws[f"{c}{r}"].number_format = "0.00000"
    for c in ("Q", "R", "S"):
        ws[f"{c}{r}"].number_format = "0.00"
        ws[f"{c}{r}"].border = BOX

ws["N1"] = "RECONCILIATION vs the Bloomberg curve export"
ws["N1"].font = Font(name="Calibri", size=10, bold=True, color="C00000")
ws["N2"] = ("Paste the export's Bid/Ask into N:O. Q/R/S show BDP minus export in bp. "
            "Mixed signs of a few tenths = the two were captured at different moments and "
            "the rates ticked. A consistent shift across all 32 = wrong pricing source, "
            "field or ticker, and needs fixing. Only column S matters for the curve: H is "
            "(PX_BID+PX_ASK)/2 and is the sole input the bootstrap reads.")
ws["N2"].font = NOTE
ws["N40"] = "Max |d Mid| (bp)"
ws["N40"].font = SECT
ws["P40"] = "=IF(COUNT(S5:S36)=0,\"paste the export first\",MAX(MAX(S5:S36),-MIN(S5:S36)))"
ws["P40"].fill = OF
ws["P40"].border = BOX
ws["N41"] = "Mean d Mid (bp) - a non-zero mean is the tell for a systematic offset"
ws["N41"].font = SECT
ws["P41"] = '=IF(COUNT(S5:S36)=0,"",AVERAGE(S5:S36))'
ws["P41"].number_format = "0.000"
ws["P41"].fill = OF
ws["P41"].border = BOX

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("reconciliation block added at SOFR_OIS_Quotes!N:S")
