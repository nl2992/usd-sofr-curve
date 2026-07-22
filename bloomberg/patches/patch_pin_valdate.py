"""
Pin VAL_DATE to the capture date.

Instructions!B9 was =TODAY(). The moment the clock passed midnight the curve rebuilt on new
pillar dates while the frozen S490 targets stayed on 07/21/26, so the comparison collapsed
from 32/32 to 9/32 and settlement moved to 07/27 with 31 accrued days.

Freezing the market data without freezing the date it was captured on is only half a freeze.
Hard-code the date; it is a yellow input, so roll it forward deliberately when re-capturing.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import datetime as dt

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
wb = load_workbook(WB)
ws = wb["Instructions"]
ws["B9"] = dt.datetime(2026, 7, 21)
ws["B9"].number_format = "mm/dd/yyyy"
ws["B9"].font = Font(name="Calibri", size=11, color="0000FF")
ws["B9"].fill = PatternFill("solid", fgColor="FFFF00")
ws["B9"].border = Border(*[Side(style="thin", color="BFBFBF")]*4)
try:
    ws["D9"] = ("Pinned to the S490 capture date, not =TODAY(). The frozen targets, the "
                "settlement date and the accrued count are all anchored to it. Change it "
                "only when re-capturing market data.")
    ws["D9"].font = Font(name="Calibri", size=9, italic=True, color="C00000")
except AttributeError:
    pass
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("VAL_DATE pinned to 07/21/2026")
