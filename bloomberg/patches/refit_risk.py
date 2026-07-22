"""
Risk measures by full revalue, matching how CDSW computes them.

The old block held hazards fixed and multiplied RPV01 by a bump. That answers a
different question from Bloomberg's, which is why Rec01 came out -4,792 against
their +72.64 - opposite sign, not a calibration difference. Each measure is now
the difference of two full valuations, with the hazard curve RE-STRIPPED under
the bumped inputs.

Row numbers below are unchanged on purpose: Steps and the CDS_Entities
comparison block reference them by row. The four valuation cells go in F, out of
the way, so the working is visible without moving anything.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB="bloomberg/CDS_Pricer.xlsx"
ARG=("Hazard_Bootstrap!$B$7:$B$12,Hazard_Bootstrap!$C$7:$C$12,CDS_Parameters!$B$8,"
     "CDS_Parameters!$B$26,CDS_Schedule!$C$7:$C$46,CDS_Schedule!$D$7:$D$46,"
     "CDS_Schedule!$E$7:$E$46,CDS_Schedule!$F$7:$F$46,CDS_Schedule!$G$7:$G$46,"
     "CDS_Schedule!$H$7:$H$46,$B$4,$B$6,$B$8")
def mv(bs,bd,br):
    return f'=IFERROR(CDS_MarketValue({bs},{bd},{br},{ARG}),"")'

F="Calibri"
B=Font(name=F,size=11,bold=True); N=Font(name=F,size=10); SM=Font(name=F,size=9,color="808080")
GRN=Font(name=F,size=11,color="008000"); WRK=PatternFill("solid",fgColor="F2F2F2")
BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(WB); ws=wb["CDS_Pricer"]

ws["A21"]="RISK MEASURES (full revalue, curve re-stripped on each bump)"
ws["A21"].font=B
ws["N21"]="full-revalue working"; ws["N21"].font=B
for r,(lab,f_) in enumerate([("MV base",mv(0,0,0)),
                             ("MV  spreads +1bp",mv(1,0,0)),
                             ("MV  rates +1bp",mv(0,1,0)),
                             ("MV  recovery +1%",mv(0,0,0.01))],start=22):
    ws.cell(r,14,lab).font=N
    c=ws.cell(r,15,f_); c.number_format="#,##0.00"; c.font=GRN; c.fill=WRK; c.border=BOX

ws["B22"]='=IF(OR($O$22="",$O$23=""),"",$O$23-$O$22)'
ws["C22"]="Spread DV01: revalue at +1bp on every quote, curve re-stripped (p.9)."
ws["B23"]='=IF(OR($O$22="",$O$24=""),"",$O$24-$O$22)'
ws["C23"]="IR DV01: parallel +1bp on the discount curve, curve re-stripped."
ws["B24"]='=IF(OR($O$22="",$O$25=""),"",$O$25-$O$22)'
ws["C24"]="Rec01: recovery +1%, hazards re-fitted so the quotes still reprice."
for r in (22,23,24):
    ws.cell(r,2).number_format="#,##0.00"; ws.cell(r,3).font=SM
ws["A22"]="CS01 / Spread DV01  (+1bp)"
ws["A23"]="IR DV01  (+1bp rates)"
ws["A24"]="Rec01  (+1% recovery)"
ws["N26"]=("Each measure is a difference of two full valuations. Holding hazards fixed and "
           "multiplying RPV01 by a bump answers a different question - that is what made Rec01 "
           "come out with the wrong sign.")
ws["N26"].font=SM
wb.calculation.fullCalcOnLoad=True
wb.save(WB)
print("CDS_Pricer risk block rebuilt on CDS_MarketValue; row numbers unchanged")
