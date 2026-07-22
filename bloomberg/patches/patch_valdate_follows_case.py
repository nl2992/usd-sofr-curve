"""
Make VAL_DATE follow the active test case's curve date.

The bug: Testing derived its maturities from the case's own curve date, while
Bootstrap built the curve from VAL_DATE, pinned at 07/21/26. A case captured on
07/22 derived 1W = 07/31 while Bootstrap held 1W = 07/30, so the dates never met
and only the handful of pillars that coincidentally aligned matched - 9 of 32.

Selecting a test case must move the whole workbook to that snapshot, otherwise
the comparison measures a curve built on one date against targets captured on
another. That is the same error as fitting the o/n stub to the answer, in a
different disguise.

  Live / Fixed   VAL_DATE = 07/21/2026, the pinned capture date
  Test 1/2/3     VAL_DATE = that block's curve date

No circularity: VAL_DATE reads the case's curve-date INPUT cell; the case's
derived dates read its own curve date, not VAL_DATE.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
MODE = "Bootstrap!$G$4"
T = ["Test 1", "Test 2", "Test 3"]
CD = ["Testing!$D$5", "Testing!$D$55", "Testing!$D$105"]
PINNED = "DATE(2026,7,21)"

wb = load_workbook(WB)
ins = wb["Instructions"]
f = (f'=IF({MODE}="{T[0]}",{CD[0]},IF({MODE}="{T[1]}",{CD[1]},'
     f'IF({MODE}="{T[2]}",{CD[2]},{PINNED})))')
c = ins["B9"]
c.value = f
c.number_format = "mm/dd/yyyy"
c.fill = OF
c.border = BOX
c.font = BLACK
try:
    ins["D9"] = ("Follows the active test case's curve date; otherwise the pinned capture date "
                 "07/21/2026. A test case must move the WHOLE workbook to its snapshot, or the "
                 "bootstrap builds on one date and the targets come from another.")
    ins["D9"].font = WARN
except AttributeError:
    pass

# make the mismatch impossible to miss on the Testing blocks
ts = wb["Testing"]
for bi, (cd_row, sum_row) in enumerate([(5, 48), (55, 98), (105, 148)]):
    act = f'{MODE}="{T[bi]}"'
    try:
        ts[f"N{cd_row}"] = (f'=IF(NOT({act}),"",IF($D${cd_row}="","set the curve date",'
                            f'IF($D${cd_row}<>VAL_DATE,"curve date not driving VAL_DATE",'
                            f'"VAL_DATE follows this case")))')
        ts[f"N{cd_row}"].font = WARN
    except AttributeError:
        pass
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("VAL_DATE now follows the active test case")
