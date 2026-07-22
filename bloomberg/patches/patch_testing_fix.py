"""
Fix two faults introduced by patch_testing.py.

1. CDS_Entities!H:M were overwritten with =IF(ISNUMBER(live),live,"") which
   returns an empty STRING when BDP is unavailable. CDS_Quotes!E read that and
   the empty string propagated into arithmetic through the whole hazard chain -
   4,131 errors across Hazard_Solver, CDS_Schedule, CDS_Pricer.

   H:M go back to being plain manual inputs. A separate in-use block resolves
   live-over-manual and is guaranteed numeric, and CDS_Quotes!E reads that.

2. Testing!B5 held the active-source readout but the case-1 header block wrote
   over it. Moved to B4.
"""
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter as CL, column_index_from_string as CI

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
HF = PatternFill("solid", fgColor="1F3864")
GF = PatternFill("solid", fgColor="2E7D32")
YF = PatternFill("solid", fgColor="FFFF00")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
TEN6 = ["1Y", "2Y", "3Y", "5Y", "7Y", "10Y"]
MAN = {"1Y": "H", "2Y": "I", "3Y": "J", "5Y": "K", "7Y": "L", "10Y": "M"}


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False, al=None):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    if al: c.alignment = Alignment(horizontal=al, wrap_text=True, vertical="center")
    return c


wb = load_workbook(WB)
e = wb["CDS_Entities"]

# 1a. H:M back to manual inputs
for ten, col in MAN.items():
    put(e, f"{col}4", f"{ten} bp (manual)", HDR, None, GF, True, "center")
    for r in range(5, 15):
        put(e, f"{col}{r}", None, BLUE, "0.0000", YF, True)
put(e, f"K5", 55.1514, BLUE, "0.0000", YF, True)      # the one spread we actually have

# 1b. in-use block, always numeric
live0 = CI("AE") + 2          # AG, AI, ... hold the live bp from the earlier patch
c_use = CI("AR")
put(e, "AR3", "IN USE  =  live if it resolved, else manual, else 0", BOLD)
for j, ten in enumerate(TEN6):
    cu = CL(c_use + j)
    cl = CL(live0 + j * 2)
    e.column_dimensions[cu].width = 11
    put(e, f"{cu}4", f"{ten} used", HDR, None, HF, True, "center")
    for r in range(5, 15):
        put(e, f"{cu}{r}",
            f'=IF(ISNUMBER({cl}{r}),{cl}{r},IF(ISNUMBER({MAN[ten]}{r}),{MAN[ten]}{r},0))',
            BLACK, "0.0000", OF, True)
put(e, "AR2", "Guaranteed numeric. CDS_Quotes reads THIS, never H:M directly - an empty "
              "string here propagates into the hazard chain and takes out the module.", NOTE)

# 1c. point CDS_Quotes at the in-use block
q = wb["CDS_Quotes"]
for i, ten in enumerate(TEN6):
    r = 7 + i
    cu = CL(c_use + i)
    put(q, f"E{r}",
        f'=IFERROR(INDEX(CDS_Entities!${cu}$5:${cu}$14,CDS_Parameters!$B$28),0)',
        BLACK, "0.0000", OF, True)
put(q, "A19", "Col E reads CDS_Entities in-use (AR:AW) for the active row: live pull when it "
              "resolves, else the manual spread, else 0. Always numeric.", NOTE)

# 2. move the active-source readout clear of the case-1 header
t = wb["Testing"]
put(t, "A4", "Active source", BOLD)
put(t, "B4", "=Bootstrap!$G$4", BLACK, None, OF, True)
put(t, "A5", None)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("H:M restored as manual; in-use block at AR:AW; CDS_Quotes repointed; readout moved to B4")
