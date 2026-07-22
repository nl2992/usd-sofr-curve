"""
Entities becomes the input surface for up to 10 names; Curves gains the CDSW view.

Before, names lived as literals on CDS_Entities and the ticker for the BDP pull
was hardcoded on CDS_Parameters!B19. Adding a name meant editing two sheets and
remembering the second. Now Entities!B:D are the inputs, CDS_Entities reads them,
and picking a name at B4 drives both the active entity and the ticker the pull
uses.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from openpyxl.chart import LineChart, Reference

WB="bloomberg/CDS_Pricer.xlsx"
NSLOT=10
R0=7                      # first entity row on Entities
E0=5                      # matching first row on CDS_Entities
TEN=["1Y","2Y","3Y","5Y","7Y","10Y"]
SRCCOL=["H","I","J","K","L","M"]      # manual spread cols on CDS_Entities

F="Calibri"
H1=Font(name=F,size=13,bold=True); B=Font(name=F,size=11,bold=True)
SM=Font(name=F,size=9,color="808080"); N=Font(name=F,size=10)
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
IN=PatternFill("solid",fgColor="FFF2CC"); BLUE=Font(name=F,size=11,color="0000FF")
GRN=Font(name=F,size=10,color="008000"); BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=load_workbook(WB)
es=wb["Entities"]; ent=wb["CDS_Entities"]

seed=[(ent.cell(E0+i,1).value, ent.cell(E0+i,2).value, ent.cell(E0+i,3).value,
       ent.cell(E0+i,6).value) for i in range(NSLOT)]

for row in es.iter_rows(min_row=1,max_row=60,max_col=14):
    for c in row: c.value=None
es.data_validations=DataValidationList()

es["A1"]="Entities"; es["A1"].font=H1
es["A2"]="Type a name and its Bloomberg ticker on any free row. On = 0 removes a name. Overrides beat the BDP pull; blank uses the pull."
es["A2"].font=SM
es["A4"]="Live name"; es["A4"].font=B
es["B4"]=seed[0][0] or ""; es["B4"].fill=IN; es["B4"].font=BLUE; es["B4"].border=BOX
es["C4"]="only this name is stripped and priced"; es["C4"].font=SM
es["E4"]="ticker in use"; es["E4"].font=B
es["F4"]=f'=IFERROR(INDEX($C${R0}:$C${R0+NSLOT-1},MATCH($B$4,$B${R0}:$B${R0+NSLOT-1},0)),"")'
es["F4"].font=GRN; es["F4"].border=BOX

es["F6"]="override spreads (bp)"; es["F6"].font=SM
hdr=["On","Entity name","Bloomberg ticker","Ccy","Recovery"]+TEN
for i,h in enumerate(hdr,start=1):
    c=es.cell(6,i,h); c.font=HD; c.fill=HF; c.border=BOX
    c.alignment=Alignment(horizontal="center")

for i in range(NSLOT):
    r=R0+i
    es.cell(r,1, 1 if seed[i][0] else None)
    es.cell(r,2, seed[i][0]); es.cell(r,3, seed[i][1]); es.cell(r,4, seed[i][2])
    es.cell(r,5, None).number_format="0.00"
    for k in range(6): es.cell(r,6+k,None).number_format="0.0"
    for k in range(1,12):
        c=es.cell(r,k); c.fill=IN; c.font=BLUE; c.border=BOX

dv=DataValidation(type="list",formula1=f"=$B${R0}:$B${R0+NSLOT-1}",allow_blank=False)
es.add_data_validation(dv); dv.add(es["B4"])

L0=R0+NSLOT+2
es.cell(L0,1,"Effective (what the strip uses)").font=B
es.cell(L0,6,"blank override falls back to the pulled or manual spread").font=SM
for i,h in enumerate(["","Entity","","","Recovery"]+TEN,start=1):
    if h:
        c=es.cell(L0+1,i,h); c.font=HD; c.fill=HF; c.border=BOX
for i in range(NSLOT):
    r=L0+2+i; ov=R0+i; src=E0+i
    es.cell(r,2,f'=IF($B{ov}="","",$B{ov})')
    es.cell(r,5,f'=IF($B{ov}="","",IF($A{ov}=0,"off",IF(ISNUMBER($E{ov}),$E{ov},CDS_Entities!$F${src})))')
    for k,col in enumerate(SRCCOL):
        oc=chr(70+k)
        es.cell(r,6+k,f'=IF($B{ov}="","",IF($A{ov}=0,0,IF(ISNUMBER(${oc}{ov}),${oc}{ov},'
                      f'IF(ISNUMBER(CDS_Entities!${col}${src}),CDS_Entities!${col}${src},0))))').number_format="0.0"
    for k in range(1,12): es.cell(r,k).border=BOX; es.cell(r,k).font=GRN

# CDS_Entities now reads its identity columns from Entities
for i in range(NSLOT):
    src=E0+i; r=R0+i
    ent.cell(src,1,f"=IF(Entities!$B{r}=\"\",\"\",Entities!$B{r})")
    ent.cell(src,2,f"=IF(Entities!$C{r}=\"\",\"\",Entities!$C{r})")
    ent.cell(src,3,f"=IF(Entities!$D{r}=\"\",\"\",Entities!$D{r})")
    if ent.cell(src,6).value is None: ent.cell(src,6, 0.4)

p=wb["CDS_Parameters"]
p["B27"]="=Entities!$B$4"
p["B19"]='=IF(Entities!$F$4="","",Entities!$F$4)'
p["C19"]="from Entities!F4 - the selected name's ticker"; p["C19"].font=SM
print(f"Entities: {NSLOT} slots, dropdown on B4, ticker feeds CDS_Parameters!B19")

# ---------------- Curves: default probability + the CDSW view ----------------
cv=wb["Curves"]
r0=73          # hazard/survival header
cv.cell(r0+1,4,"Default prob 1-Q").font=Font(name=F,size=10,bold=True,color="FFFFFF")
cv.cell(r0+1,4).fill=HF; cv.cell(r0+1,4).border=BOX
for j in range(6):
    rr=r0+2+j; s=7+j
    cv.cell(rr,4,f"=Hazard_Bootstrap!F{s}").number_format="0.0000"

r1=83
cv.cell(r1+1,5,"Traded spread").font=Font(name=F,size=10,bold=True,color="FFFFFF")
cv.cell(r1+1,5).fill=HF; cv.cell(r1+1,5).border=BOX
for j in range(6):
    rr=r1+2+j
    cv.cell(rr,5,"=CDS_Pricer!$B$15").number_format="0.00"

def line(title,ytitle,minr,maxr,valcols,anchor):
    ch=LineChart(); ch.title=title; ch.y_axis.title=ytitle
    ch.height,ch.width=7.5,15
    for vc in valcols:
        ch.add_data(Reference(cv,min_col=vc,min_row=minr-1,max_row=maxr),titles_from_data=True)
    ch.set_categories(Reference(cv,min_col=1,min_row=minr,max_row=maxr))
    for s in ch.series: s.smooth=False
    cv.add_chart(ch,anchor)

line("Default probability 1-Q(t)","probability",r0+2,r0+7,[4],"F68")
line("CDSW view - credit curve vs traded spread","bp",r1+2,r1+7,[2,3,5],"F84")
print("Curves: default-probability chart + CDSW credit-curve view")

wb.calculation.fullCalcOnLoad=True
wb.save(WB)
