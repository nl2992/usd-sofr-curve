"""
KRW workbook rebuilt on Kevin's standard tenor grid (8Y/9Y dropped), plus a
DV01 sheet in his exact reporting format.

Kevin's grid: 2D 1W 1M 2M 3M 6M 9M 1Y 2Y 3Y 4Y 5Y 7Y 10Y 12Y 15Y 20Y 25Y 30Y.
Our data starts at 3M, and the swap is 10Y, so:
  - 2D..2M  : no quote / no sensitivity for a 10Y swap  -> 0
  - 3M..10Y : the live buckets (bump each quote 1bp, re-bootstrap, reprice)
  - 12Y..30Y: past the 10Y maturity                     -> 0
Dropping 8Y/9Y as nodes lets the 7Y and 10Y bumps span the whole 7-10Y region,
so the buckets sum exactly to the parallel DV01.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter as CL

OUT="bloomberg/KRW_KevinGrid.xlsx"
F="Calibri"
H1=Font(name=F,size=14,bold=True); B=Font(name=F,size=11,bold=True); SM=Font(name=F,size=9,color="808080")
HD=Font(name=F,size=10,bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="4472C4")
IN=PatternFill("solid",fgColor="FFF2CC"); BLUE=Font(name=F,size=11,color="0000FF")
CALC=PatternFill("solid",fgColor="E2EFDA"); GRN=Font(name=F,size=10,color="008000")
RES=PatternFill("solid",fgColor="FCE4D6"); BOX=Border(*[Side(style="thin",color="BFBFBF")]*4)
MONEY='#,##0'; PCT="0.000%"

# Kevin's curve nodes we actually have (3M..10Y + 15/20/30)
QT=[("3Mo",2.91,0.25),("6Mo",3.125,0.5),("9Mo",3.315,0.75),("1Yr",3.4675,1),("2Yr",3.825,2),
    ("3Yr",3.965,3),("4Yr",4.029,4),("5Yr",4.0775,5),("7Yr",4.1425,7),("10Yr",4.20,10),
    ("15Yr",4.1775,15),("20Yr",4.075,20),("30Yr",3.845,30)]
NQ=len(QT); QR0=5; QR1=QR0+NQ-1     # Quotes data rows

wb=Workbook()
# ---------------- Quotes ----------------
q=wb.active; q.title="Quotes"
q["A1"]="KRW IRS  (Kevin grid, 8Y/9Y dropped)"; q["A1"].font=H1
q["A2"]="tau"; q["B2"]=0.25; q["B2"].fill=IN; q["B2"].font=BLUE
for i,h in enumerate(["Tenor","Coupon","Year"],1): q.cell(4,i,h).font=B
for r,(l,c,y) in enumerate(QT,QR0):
    q.cell(r,1,l); q.cell(r,2,c/100).number_format=PCT; q.cell(r,3,y).number_format="0.00"
    q.cell(r,2).fill=IN; q.cell(r,2).font=BLUE
QY=f"$C${QR0}:$C${QR1}"; QC=f"$B${QR0}:$B${QR1}"

# ---------------- Interpolation (quarterly) ----------------
it=wb.create_sheet("Interpolation")
for i,h in enumerate(["Year","Swap %","Helper"],1): it.cell(1,i,h).font=B
for i in range(120):
    r=2+i; T=round(0.25*(i+1),2)
    it.cell(r,1,T).number_format="0.00"
    it.cell(r,3,f"=MIN(MATCH(A{r},Quotes!{QY},1),{NQ-1})")
    it.cell(r,2,f"=INDEX(Quotes!{QC},C{r})+(A{r}-INDEX(Quotes!{QY},C{r}))/"
                f"(INDEX(Quotes!{QY},C{r}+1)-INDEX(Quotes!{QY},C{r}))*"
                f"(INDEX(Quotes!{QC},C{r}+1)-INDEX(Quotes!{QC},C{r}))").number_format=PCT

# ---------------- Bootstrap ----------------
bs=wb.create_sheet("Bootstrap")
for i,h in enumerate(["Year","Swap %","DF","Zero %","Future %"],1): bs.cell(2,i,h).font=B
for i in range(120):
    r=3+i
    bs.cell(r,1,f"=Interpolation!A{2+i}").number_format="0.00"
    bs.cell(r,2,f"=Interpolation!B{2+i}").number_format=PCT
    if i==0: bs.cell(r,3,f"=1/(1+B{r}*Quotes!$B$2)"); bs.cell(r,5,f"=(1/C{r}-1)/Quotes!$B$2")
    else:
        bs.cell(r,3,f"=(1-B{r}*Quotes!$B$2*SUM($C$3:C{r-1}))/(1+B{r}*Quotes!$B$2)")
        bs.cell(r,5,f"=(C{r-1}/C{r}-1)/Quotes!$B$2")
    bs.cell(r,3).number_format="0.00000000"; bs.cell(r,4,f"=(1/C{r}-1)/A{r}").number_format=PCT
    bs.cell(r,5).number_format=PCT

# ---------------- Amort-Set-up ----------------
a=wb.create_sheet("Amort-Set-up")
a["A1"]="Amortising IRS"; a["A1"].font=H1
a["A5"]="Start"; a["B5"]="=DATE(2026,7,27)"; a["B5"].number_format="mm/dd/yyyy"
S0,S1=20,59
a["A11"]="Solved par"; a["B11"]=f"=SUMPRODUCT($E${S0}:$E${S1},$F${S0}:$F${S1},$D${S0}:$D${S1})/SUMPRODUCT($E${S0}:$E${S1},$D${S0}:$D${S1})"
a["B11"].number_format=PCT; a["B11"].fill=RES; a["B11"].font=B
a["A12"]="Trade fixed rate"; a["B12"]="=B11"; a["B12"].number_format=PCT; a["B12"].fill=IN; a["B12"].font=BLUE
a["A15"]="Fixed leg PV"; a["B15"]=f"=SUM($J${S0}:$J${S1})"; a["B15"].number_format=MONEY
a["A16"]="Float leg PV"; a["B16"]=f"=SUM($K${S0}:$K${S1})"; a["B16"].number_format=MONEY
a["A17"]="NPV"; a["B17"]="=B15-B16"; a["B17"].number_format=MONEY; a["B17"].fill=RES; a["B17"].font=B
for i,h in enumerate(["Period","Start","End","DF","Notional","Forward %","Fixed CF","Float CF","","PV fixed","PV float"],1):
    if h: a.cell(19,i,h).font=HD; a.cell(19,i).fill=HF
prof=[10]*4+[30]*4+[60]*4+[100]*4+[80]*4+[60]*4+[50]*4+[50]*4+[40]*4+[40]*4
for p in range(1,41):
    r=19+p; br=2+p
    a.cell(r,1,p); a.cell(r,2,f"=EDATE($B$5,3*{p-1})").number_format="mm/dd/yyyy"
    a.cell(r,3,f"=EDATE($B$5,3*{p})").number_format="mm/dd/yyyy"
    a.cell(r,4,f"=Bootstrap!C{br}").number_format="0.00000000"
    a.cell(r,5,prof[p-1]*1_000_000_000).number_format=MONEY; a.cell(r,5).fill=IN; a.cell(r,5).font=BLUE
    a.cell(r,6,f"=Bootstrap!E{br}").number_format=PCT
    a.cell(r,7,f"=E{r}*$B$12*Quotes!$B$2").number_format=MONEY
    a.cell(r,8,f"=E{r}*F{r}*Quotes!$B$2").number_format=MONEY
    a.cell(r,10,f"=G{r}*D{r}").number_format=MONEY; a.cell(r,11,f"=H{r}*D{r}").number_format=MONEY
a.cell(S1+2,5,"PLACEHOLDER notionals - paste Kevin's schedule into E20:E59").font=Font(name=F,size=9,italic=True,color="C00000")

# ---------------- KRD (bucket the 10 tenors <=10Y) ----------------
BUCK=[i for i in range(NQ) if QT[i][2]<=10]   # tenor indices <=10Y = 0..9 (3M..10Y)
NS=1+len(BUCK)                                  # base + 10 bumps
k=wb.create_sheet("KRD")
k["A1"]="Bucketed key-rate delta (Kevin grid)"; k["A1"].font=H1
k["D3"]="='Amort-Set-up'!$B$12"; k["D3"].number_format=PCT
k["E4"]="=Quotes!$B$2"
qr0=7
for i,h in enumerate(["Tenor","Year","Base"],1): k.cell(6,i,h).font=HD; k.cell(6,i).fill=HF
for t in range(NQ):
    r=qr0+t
    k.cell(r,1,f"=Quotes!A{QR0+t}"); k.cell(r,2,f"=Quotes!C{QR0+t}").number_format="0.00"
    k.cell(r,3,f"=Quotes!B{QR0+t}").number_format=PCT
    for j in range(NS):
        col=4+j
        if j==0: k.cell(r,col,f"=$C{r}")
        else:
            tgt=qr0+BUCK[j-1]
            k.cell(r,col,(f"=$C{r}+0.0001" if r==tgt else f"=$C{r}"))
        k.cell(r,col).number_format=PCT
QMAT=f"$D$7:${CL(3+NS)}${qr0+NQ-1}"; TYR=f"$B$7:$B${qr0+NQ-1}"
QR_=25; QE=QR_+39; RC0=6; DC0=RC0+NS+1
for i,h in enumerate(["Qtr","Year","m","w","Notional"],1): k.cell(24,i,h).font=HD; k.cell(24,i).fill=HF
for j in range(NS):
    k.cell(24,RC0+j,j+1); k.cell(24,DC0+j,1)
for qq in range(1,41):
    r=QR_+qq-1
    k.cell(r,1,qq); k.cell(r,2,f"=A{r}*0.25").number_format="0.00"
    k.cell(r,3,f"=MIN(MATCH(B{r},{TYR},1),{NQ-1})")
    k.cell(r,4,f"=(B{r}-INDEX({TYR},C{r}))/(INDEX({TYR},C{r}+1)-INDEX({TYR},C{r}))")
    k.cell(r,5,f"='Amort-Set-up'!E{19+qq}").number_format=MONEY
    for j in range(NS):
        rc=RC0+j; dc=DC0+j; RL=CL(rc); DLx=CL(dc); hh=f"{RL}$24"
        k.cell(r,rc,f"=INDEX({QMAT},$C{r},{hh})+$D{r}*(INDEX({QMAT},$C{r}+1,{hh})-INDEX({QMAT},$C{r},{hh}))").number_format=PCT
        if qq==1: k.cell(r,dc,f"=1/(1+{RL}{r}*$E$4)")
        else: k.cell(r,dc,f"=(1-{RL}{r}*$E$4*SUM({DLx}${QR_}:{DLx}{r-1}))/(1+{RL}{r}*$E$4)")
        k.cell(r,dc).number_format="0.00000000"
FR,FL,NP,DL_=QE+2,QE+3,QE+4,QE+5
for a2,t in ((FR,"Fixed PV"),(FL,"Float PV"),(NP,"NPV"),(DL_,"DELTA")): k.cell(a2,1,t).font=B
Nr=f"$E${QR_}:$E${QE}"
for j in range(NS):
    dc=DC0+j; DLc=CL(dc)
    k.cell(FR,dc,f"=$D$3*$E$4*SUMPRODUCT({Nr},{DLc}${QR_}:{DLc}${QE})").number_format=MONEY
    k.cell(FL,dc,f"=SUMPRODUCT({Nr},{DLc}$24:{DLc}${QE-1}-{DLc}${QR_}:{DLc}${QE})").number_format=MONEY
    k.cell(NP,dc,f"={DLc}{FR}-{DLc}{FL}").number_format=MONEY
    k.cell(DL_,dc,f"={DLc}{NP}-${CL(DC0)}{NP}").number_format=MONEY
# label each delta col with the bumped tenor
for j in range(1,NS):
    dc=DC0+j; tgt=qr0+BUCK[j-1]
    k.cell(DL_+1,dc,f"=$A{tgt}").font=SM

# ---------------- DV01 output (Kevin's exact format) ----------------
d=wb.create_sheet("DV01",0)
d["A1"]="Delta - bump each tenor 1bp"; d["A1"].font=H1
d["A2"]="KRW IRS loan hedge, 10Y. Each row = NPV change for +1bp on that tenor."; d["A2"].font=SM
for i,h in enumerate(["Tenor","dv01"],1): c=d.cell(4,i,h); c.font=HD; c.fill=HF; c.border=BOX
# map Kevin's full tenor list to the KRD delta columns (by tenor label)
KEVIN=["2D","1W","1M","2M","3M","6M","9M","1Y","2Y","3Y","4Y","5Y","7Y","10Y","12Y","15Y","20Y","25Y","30Y"]
lblmap={"3M":"3Mo","6M":"6Mo","9M":"9Mo","1Y":"1Yr","2Y":"2Yr","3Y":"3Yr","4Y":"4Yr","5Y":"5Yr","7Y":"7Yr","10Y":"10Yr"}
# KRD delta columns DC0+1 .. DC0+len(BUCK), labelled at row DL_+1
delta_cols={}
for j in range(1,NS):
    delta_cols[QT[BUCK[j-1]][0]]=CL(DC0+j)
r=5
for tn in KEVIN:
    d.cell(r,1,tn)
    src=lblmap.get(tn)
    if src and src in delta_cols:
        d.cell(r,2,f"=KRD!{delta_cols[src]}${DL_}").number_format=MONEY
    else:
        d.cell(r,2,0).number_format=MONEY
    d.cell(r,1).border=BOX; d.cell(r,2).border=BOX
    r+=1
d.cell(r,1,"TOTAL").font=B
d.cell(r,2,f"=SUM(B5:B{r-1})").number_format=MONEY; d.cell(r,2).fill=RES; d.cell(r,2).font=B
d.cell(r,1).border=BOX; d.cell(r,2).border=BOX
for col,w in zip("AB",(10,18)): d.column_dimensions[col].width=w

wb.save(OUT)
print("built",OUT,"| sheets:",wb.sheetnames,"| buckets:",[QT[i][0] for i in BUCK])
