"""
Wire Bloomberg_S490_Validation to the live BDS curve dump, so the benchmark
comparison runs automatically on a terminal instead of needing a manual paste.

BDS(S490_TKR,"CURVE_TENOR_RATES") spills a block whose column ordering I cannot
verify without a terminal. So rather than hard-coding an assumed schema, the
lookup is parametrised:

    I5  = which column of the dump holds the MATURITY DATE
    I6  = which column of the dump holds the ZERO RATE

with a preview of the dump's first rows right beside them. If the schema differs
from the default guess, you change two numbers and everything realigns — and a
"matched N / M" counter makes a wrong guess obvious instead of silently
comparing against the wrong tenors.

Match is by maturity DATE, not tenor label: dates are unambiguous, whereas the
label set ("1W" vs "1 WK") varies between Bloomberg fields.

Column D now reads live, falling back to a manual paste column G, so the sheet
still works exactly as before off-terminal.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"

FONT = "Calibri"
BLUE = Font(name=FONT, size=11, color="0000FF")
BLACK = Font(name=FONT, size=11)
SECT = Font(name=FONT, size=11, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="666666")
WARN = Font(name=FONT, size=10, bold=True, color="C00000")
YFILL = PatternFill("solid", fgColor="FFFF00")
SFILL = PatternFill("solid", fgColor="D9E1F2")
OFILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

R0, RN = 8, 72                      # Bootstrap pillar rows mirrored on this sheet
DUMP = "$J$12:$T$82"                # generous block for the BDS spill
# merged ranges on this sheet are H7:L7, A6:E6, B4:E4 - controls placed clear of them


def put(ws, cell, val, font=BLACK, fmt=None, fill=None, border=False):
    c = ws[cell]
    c.value = val
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if border: c.border = BOX
    return c


def main():
    wb = load_workbook(WB)
    ws = wb["Bloomberg_S490_Validation"]

    put(ws, "A6", "Column D pulls Bloomberg's own S490 zero curve live from the BDS dump "
                  "and falls back to the manual paste in column G. Column E is your "
                  "bootstrap minus Bloomberg, in bp — single digits = agreement.", NOTE)

    # ---- live-lookup controls (rows 3-6, clear of the H7:L7 / A6:E6 / B4:E4 merges)
    put(ws, "H2", "LIVE BENCHMARK — BDS dump controls", SECT)
    put(ws, "H3", "Dump col holding MATURITY DATE", BLACK)
    put(ws, "I3", 1, BLUE, "0", YFILL, True)
    put(ws, "H4", "Dump col holding ZERO RATE", BLACK)
    put(ws, "I4", 5, BLUE, "0", YFILL, True)
    put(ws, "H5", "Zero units (1 = pct, 100 = decimal)", BLACK)
    put(ws, "I5", 1, BLUE, "0.###", YFILL, True)
    put(ws, "H6", "VERIFY I3/I4/I5 against the dump at J12 before trusting column E — "
                  "the defaults are a GUESS at the CURVE_TENOR_RATES schema, not confirmed.",
        WARN)
    put(ws, "H7", "Raw CURVE_TENOR_RATES dump (spills right and down from J12)", SECT)
    put(ws, "J12", '=BDS(S490_TKR,"CURVE_TENOR_RATES")', BLACK, None, None, True)

    # ---- per-pillar lookup
    put(ws, "D7", "S490 zero (%) — live", SECT, None, SFILL, True)
    put(ws, "F7", "Source", SECT, None, SFILL, True)
    put(ws, "G7", "S490 zero (%) [manual paste]", SECT, None, SFILL, True)

    datecol = f"INDEX({DUMP},0,$I$3)"
    for r in range(R0, RN + 1):
        live = (f'INDEX({DUMP},MATCH(B{r},{datecol},0),$I$4)*$I$5')
        put(ws, f"D{r}", f'=IFERROR({live},IF(ISNUMBER(G{r}),G{r},""))',
            BLACK, "0.00000", None, True)
        put(ws, f"F{r}",
            f'=IF(ISNUMBER(IFERROR({live},"")),"BDS live",'
            f'IF(ISNUMBER(G{r}),"manual paste","-"))', BLACK, None, None, True)
        put(ws, f"G{r}", None, BLUE, "0.00000", YFILL, True)
        put(ws, f"E{r}", f'=IF(OR(D{r}="",NOT(ISNUMBER(D{r}))),"",(C{r}-D{r})*100)',
            BLACK, "0.00", None, True)

    # ---- summary
    s = RN + 2
    put(ws, f"A{s}", "Pillars matched to the dump", SECT)
    put(ws, f"C{s}", f'=COUNTIF(F{R0}:F{RN},"BDS live")&" / "&'
                     f'COUNTA(B{R0}:B{RN})&"  (only ~32 of the 65 rows are quoted '
                     f'Bloomberg pillars; interpolated gap rows will not match)"',
        BLACK, None, OFILL, True)
    put(ws, f"A{s+1}", "Max |Δz| over matched pillars (bp)", SECT)
    put(ws, f"C{s+1}", f'=IFERROR(MAX(ABS(IF(ISNUMBER(E{R0}:E{RN}),E{R0}:E{RN}))),"")',
        BLACK, "0.000", OFILL, True)
    put(ws, f"D{s+1}", "Array formula — Ctrl+Shift+Enter in older Excel.", NOTE)
    put(ws, f"A{s+3}", "If 'matched' reads 0 / 65 the column offsets in I3/I4 are wrong "
                       "for this dump's schema — read the preview at J10 and correct them. "
                       "A wrong offset produces plausible but meaningless diffs.", WARN)

    wb.calculation.fullCalcOnLoad = True
    wb.save(WB)
    print(f"wired rows {R0}-{RN}; offsets at I3/I4, units at I5, dump at J12")


if __name__ == "__main__":
    main()
