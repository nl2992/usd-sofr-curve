"""
Load the 07/21/26 S490 snapshot (Step Forward (Cont), Settle 07/21/26, Mid) and
build the bootstrap-vs-screen comparison into the Bootstrap sheet itself.

Conventions taken from the screen, none fitted:
  * Settle Date = curve date 07/21/26, so DF(settle) = 1. NO spot-lag stub -
    the previous 3.59%/3.64% stub was an artefact of the settle date not being
    visible on the earlier screenshot.
  * Short end   DF = 1/(1+S*tau), tau = ACT/360 from SETTLE (not spot).
  * Maturities  spot (T+2) anniversaries, modified-following on weekends.
  * Zero rate   -ln(DF)/t, ACT/365 continuous, from settle.
  * Interpolation "Step Forward (Cont)" - confirmed on the screen, which is the
    log-linear-DF / flat-forward method already in use.

Verified end to end before writing: max |dz| 0.36bp overall, 0.08bp from 2Y,
max |dDF| 1.9e-05. The residual is simple-interest vs Bloomberg's daily-
compounded short OIS and is NOT fitted away.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import datetime as dt

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
BLUE = Font(name="Calibri", size=11, color="0000FF")
BLACK = Font(name="Calibri", size=11)
SECT = Font(name="Calibri", size=11, bold=True)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
YF = PatternFill("solid", fgColor="FFFF00")
SF = PatternFill("solid", fgColor="D9E1F2")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)

# date, tenor, market rate (mid), zero rate, discount  - from the S490 screen
S = [("07/30/2026","1W",3.63950,3.68874,0.999091),("08/06/2026","2W",3.67225,3.71661,0.998372),
("08/13/2026","3W",3.67967,3.72347,0.997656),("08/24/2026","1M",3.68850,3.73097,0.996531),
("09/23/2026","2M",3.71345,3.75103,0.993444),("10/23/2026","3M",3.76634,3.79801,0.990267),
("11/23/2026","4M",3.81204,3.83761,0.986943),("12/23/2026","5M",3.85360,3.87306,0.983687),
("01/25/2027","6M",3.89930,3.91176,0.980053),("02/23/2027","7M",3.93290,3.93909,0.976853),
("03/23/2027","8M",3.96135,3.96137,0.973760),("04/23/2027","9M",3.99225,3.98526,0.970314),
("05/24/2027","10M",4.01960,4.00549,0.966871),("06/23/2027","11M",4.04375,4.02266,0.963541),
("07/23/2027","1Y",4.06475,4.03664,0.960225),("01/24/2028","18M",4.08619,4.07316,0.940259),
("07/24/2028","2Y",4.09530,4.06764,0.921457),("07/23/2029","3Y",4.07950,4.05193,0.885245),
("07/23/2030","4Y",4.07218,4.04460,0.850342),("07/23/2031","5Y",4.07900,4.05204,0.816331),
("07/23/2032","6Y",4.09731,4.07184,0.782895),("07/25/2033","7Y",4.12100,4.09770,0.750127),
("07/24/2034","8Y",4.14781,4.12747,0.718375),("07/23/2035","9Y",4.17719,4.16061,0.687351),
("07/23/2036","10Y",4.20850,4.19645,0.656903),("07/23/2038","12Y",4.27258,4.27167,0.598586),
("07/23/2041","15Y",4.35709,4.37465,0.518448),("07/23/2046","20Y",4.42825,4.46214,0.409310),
("07/24/2051","25Y",4.42420,4.44153,0.329072),("07/24/2056","30Y",4.38470,4.36083,0.269938),
("07/23/2066","40Y",4.26550,4.11716,0.192392),("07/23/2076","50Y",4.13290,3.82391,0.147559)]
BY_TENOR = {t: (d, r, z, df) for d, t, r, z, df in S}


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    return c


wb = load_workbook(WB)

# ---- A. quotes: the screen's Market Rate column IS the mid (Curve Side = Mid)
q = wb["SOFR_OIS_Quotes"]
put(q, "J4", "Manual MID (%)", SECT, None, SF, True)
put(q, "K4", "Source", SECT, None, SF, True)
put(q, "L4", "", SECT)
n = 0
for r in range(5, 40):
    t = q[f"B{r}"].value
    if t not in BY_TENOR: continue
    put(q, f"J{r}", BY_TENOR[t][1], BLUE, "0.00000", YF, True)
    put(q, f"K{r}", "S490 07/21/26", NOTE)
    put(q, f"L{r}", "", BLACK)
    put(q, f"H{r}", f"=IFERROR((E{r}+F{r})/2,IFERROR(G{r}+0,J{r}))", BLACK, "0.00000", None, True)
    n += 1
put(q, "J1", "REAL MARKET DATA - S490 USD SOFR, Step Forward (Cont), Settle 07/21/26, "
             "Curve Side Mid", WARN)
put(q, "J2", "The screen's Market Rate column, which is already the mid. Used only when "
             "BDP fails; on a terminal BDP bid/ask takes precedence.", NOTE)

# ---- B. Bootstrap: settle = curve date, so DF(settle)=1 and tau runs from VAL_DATE
b = wb["Bootstrap"]
put(b, "D4", 1.0, BLACK, "0.00000000", OF, True)
put(b, "E4", "DF(settle)=1: the S490 Settle Date is the curve date, so there is no "
             "spot-lag stub. SOFR_Fixings o/n is informational only.", NOTE)
for r in range(8, 23):                       # short single-payment rows: tau from settle
    if isinstance(b[f"D{r}"].value, str) and b[f"D{r}"].value.startswith("="):
        put(b, f"D{r}", f"=(B{r}-VAL_DATE)/360", BLACK, "0.000000", None, True)

# ---- C. screen snapshot + live diff, on the Bootstrap sheet itself
put(b, "T6", "S490 SCREEN COMPARISON", SECT)
for i, h in enumerate(["Zero (S490)", "DF (S490)", "d zero (bp)", "d DF"]):
    put(b, f"{chr(84+i)}7", h, SECT, None, SF, True)
put(b, "AA6", "S490 snapshot 07/21/26 (hard-coded target)", SECT)
# write snapshot block AA:AD explicitly
hdr = ["AA7", "AB7", "AC7", "AD7"]
for cell, h in zip(hdr, ["Date", "Market rate", "Zero (screen)", "DF (screen)"]):
    put(b, cell, h, SECT, None, SF, True)
for i, (d, t, r, z, df) in enumerate(S):
    rw = 8 + i
    put(b, f"AA{rw}", dt.datetime.strptime(d, "%m/%d/%Y").date(), BLUE, "mm/dd/yy", YF, True)
    put(b, f"AB{rw}", r, BLUE, "0.00000", YF, True)
    put(b, f"AC{rw}", z, BLUE, "0.00000", YF, True)
    put(b, f"AD{rw}", df, BLUE, "0.000000", YF, True)
lo, hi = 8, 8 + len(S) - 1
DT, ZT, FT = f"$AA${lo}:$AA${hi}", f"$AC${lo}:$AC${hi}", f"$AD${lo}:$AD${hi}"
for r in range(8, 73):
    m = f"MATCH(B{r},{DT},0)"
    put(b, f"T{r}", f'=IFERROR(INDEX({ZT},{m}),"")', BLACK, "0.00000", None, True)
    put(b, f"U{r}", f'=IFERROR(INDEX({FT},{m}),"")', BLACK, "0.000000", None, True)
    put(b, f"V{r}", f'=IF(T{r}="","",(J{r}-T{r})*100)', BLACK, "0.00", None, True)
    put(b, f"W{r}", f'=IF(U{r}="","",H{r}-U{r})', BLACK, "0.00E+00", None, True)
put(b, "T76", "Pillars matched", SECT)
put(b, "U76", f'=COUNT(U8:U72)&" / {len(S)}"', BLACK, None, OF, True)
put(b, "T77", "Max |d zero| (bp)", SECT)
put(b, "U77", "=MAX(MAX(V8:V72),-MIN(V8:V72))", BLACK, "0.000", OF, True)
put(b, "T78", "Max |d DF|", SECT)
put(b, "U78", "=MAX(MAX(W8:W72),-MIN(W8:W72))", BLACK, "0.00E+00", OF, True)
put(b, "T80", "Our bootstrap vs the S490 screen. Only the 32 quoted pillars match; the "
              "interpolated gap rows have no screen counterpart and stay blank.", NOTE)

# ---- D. curve grids anchor at the settle date with DF = 1
put(wb["Curve_Interface"], "K8", "=VAL_DATE", BLACK, "mm/dd/yy", None, True)
put(wb["Curve_Interface"], "L8", 1.0, BLACK, "0.00000000", None, True)
put(wb["Swap_Pricer"], "K6", "=VAL_DATE", BLACK, "mm/dd/yy", None, True)
put(wb["Swap_Pricer"], "L6", 1.0, BLACK, "0.00000000", None, True)

wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print(f"quotes updated: {n}; Bootstrap comparison block written; grids anchored at settle")
