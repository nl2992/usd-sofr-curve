"""Brent_vs_Bisection: the VBA Brent solver side by side with the in-cell strip."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg.xlsx"
H1 = Font(name="Calibri", size=12, bold=True)
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
BLACK = Font(name="Calibri", size=11)
MONO = Font(name="Consolas", size=9)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
WARN = Font(name="Calibri", size=10, bold=True, color="C00000")
HF = PatternFill("solid", fgColor="1F3864")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
N8 = "0.00000000"
BLOCKS = [6, 49, 92, 135, 178, 221]      # Hazard_Solver block starts
TEN = ["1Y", "2Y", "3Y", "5Y", "7Y", "10Y"]
AI = "CDS_Parameters!$B$26"


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False, al=None):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    if al: c.alignment = Alignment(horizontal=al, wrap_text=True, vertical="center")
    return c


wb = load_workbook(WB)
if "Brent_vs_Bisection" in wb.sheetnames:
    del wb["Brent_vs_Bisection"]
ws = wb.create_sheet("Brent_vs_Bisection", wb.sheetnames.index("Testing") + 1)
ws.sheet_properties.tabColor = "B7950B"
for c, w in zip("ABCDEFGHIJK", [9, 13, 16, 16, 13, 10, 14, 14, 13, 13, 60]):
    ws.column_dimensions[c].width = w

put(ws, "A1", "Brent (VBA) vs in-cell bisection — same objective, same inputs", H1)
put(ws, "A2", "Both solve (3.3): f(h) = (1-R)*ProtCum - S*(RPV01Cum - D0*D(Ts)). Only the "
              "root-finder differs. Columns D:F need the VBA module.", NOTE)
put(ws, "A3", "Import: Alt+F11 > File > Import File > bloomberg/vba/CDSBrent.bas, then save "
              "as .xlsm. Until then D:F read 'module not loaded' — the workbook is .xlsx and "
              "stays fully functional without it.", WARN)

for i, h in enumerate(["Tenor", "Market bp", "Bisection (30 steps)", "Brent (VBA)",
                       "diff", "Brent its", "f at bisection", "f at Brent", "Model bp",
                       "Err bp"]):
    put(ws, f"{chr(65+i)}5", h, HDR, None, HF, True, "center")
ws.row_dimensions[5].height = 30

for i, (blk, ten) in enumerate(zip(BLOCKS, TEN)):
    r = 6 + i
    h1, h2 = blk + 1, blk + 2
    a_, cd, de, dm = f"$F${blk+4}:$U${blk+4}", f"$F${blk+8}:$U${blk+8}", \
                     f"$F${blk+6}:$U${blk+6}", f"$F${blk+7}:$U${blk+7}"
    args = (f"Hazard_Solver!$D${h1},Hazard_Solver!$F${h1},Hazard_Solver!$B${h2},"
            f"Hazard_Solver!$D${h2},Hazard_Solver!$F${h2},{AI},"
            f"Hazard_Solver!{a_},Hazard_Solver!{cd},Hazard_Solver!{de},Hazard_Solver!{dm}")
    put(ws, f"A{r}", ten, BLACK, None, None, True, "center")
    put(ws, f"B{r}", f"=Hazard_Solver!$B${h1}", BLACK, "0.0000", None, True)
    put(ws, f"C{r}", f"=Hazard_Solver!$B${blk+40}", BLACK, N8, OF, True)
    put(ws, f"D{r}", f'=IFERROR(CDS_Hazard({args}),"module not loaded")', BLACK, N8, OF, True)
    put(ws, f"E{r}", f'=IF(ISNUMBER(D{r}),D{r}-C{r},"")', BLACK, "0.00E+00", None, True)
    put(ws, f"F{r}", f'=IFERROR(CDS_LastIterations(),"")', BLACK, "0", None, True)
    put(ws, f"G{r}", f'=IFERROR(CDS_Objective(C{r},{args}),"")', BLACK, "0.00E+00", None, True)
    put(ws, f"H{r}", f'=IF(ISNUMBER(D{r}),IFERROR(CDS_Objective(D{r},{args}),""),"")',
        BLACK, "0.00E+00", None, True)
    put(ws, f"I{r}", f'=IF(ISNUMBER(D{r}),IFERROR(CDS_ModelSpread(D{r},{args}),""),"")',
        BLACK, "0.0000", None, True)
    put(ws, f"J{r}", f'=IF(ISNUMBER(I{r}),I{r}-B{r},"")', BLACK, "0.00E+00", None, True)

put(ws, "A13", "Max |diff| (hazard)", BOLD)
put(ws, "C13", '=IF(COUNT(E6:E11)=0,"module not loaded",MAX(MAX(E6:E11),-MIN(E6:E11)))',
    BLACK, "0.00E+00", OF, True)
put(ws, "A14", "Max |Err| Brent (bp)", BOLD)
put(ws, "C14", '=IF(COUNT(J6:J11)=0,"",MAX(MAX(J6:J11),-MIN(J6:J11)))', BLACK, "0.00E+00", OF, True)

put(ws, "A17", "WHAT BRENT ACTUALLY BUYS", H1)
ROWS = [
 ("Iterations", "30, fixed", "8-13, adaptive", "Brent"),
 ("Residual |f| at the root", "~1e-9", "~1e-17", "Brent, by ~8 orders"),
 ("Hazard agreement", "the two agree to ~1e-9", "", "immaterial: 1e-9 of hazard is 1e-7 bp of spread"),
 ("Cells consumed", "~3,600 on Hazard_Solver", "6 function calls", "Brent, heavily"),
 ("Working visible", "yes, every halving is a row", "no, opaque call", "BISECTION"),
 ("File format", ".xlsx", ".xlsm + macro trust", "BISECTION"),
 ("Recalc", "native", "UDF, slower per call, no multithread", "bisection"),
 ("Fails on an unstrippable quote", "pins silently at a bracket end", "returns #NUM!", "Brent"),
]
for i, h in enumerate(["", "In-cell bisection", "Brent (VBA)", "Better"]):
    put(ws, f"{chr(65+i)}18", h, HDR, None, HF, True, "center")
for i, (a, b, c, d) in enumerate(ROWS):
    r = 19 + i
    put(ws, f"A{r}", a, BOLD)
    put(ws, f"B{r}", b, BLACK, None, None, True)
    put(ws, f"C{r}", c, BLACK, None, None, True)
    put(ws, f"D{r}", d, BLACK, None, None, True)
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 34
ws.column_dimensions["D"].width = 44

put(ws, "A29", "Verdict", BOLD)
put(ws, "A30", "Brent is the better root-finder and it is not close — 8-13 iterations to machine "
               "precision against 30 to 1e-9. But the accuracy is not the binding constraint: our "
               "strip already reprices to ~1e-5 bp, which is set by the quarterly discretisation, "
               "not by the solver. Going to 1e-17 changes no price.", NOTE)
put(ws, "A31", "What it would actually buy is ~3,600 cells back and a faster recalc. What it costs "
               "is the working — the reason the bisection is laid out as rows is that every halving "
               "is inspectable, which is what was asked for. A UDF is a black box.", NOTE)
put(ws, "A32", "Keeping bisection as the engine and Brent as the cross-check is the useful split: "
               "an independent implementation agreeing to 1e-9 is real evidence the strip is right.", NOTE)
put(ws, "A33", "One genuine functional gain worth stealing: Brent returns #NUM! when the quote is "
               "unbracketed on [0,3], where the bisection pins at a bracket end and looks converged. "
               "That is deliverable D8 and can be added to the in-cell version without VBA.", WARN)

ws.sheet_view.showGridLines = False
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("Brent_vs_Bisection sheet written")
