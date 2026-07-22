"""
Show the working on the Testing sheet, not just the deltas.

CURVE (Testing!O:AC) - per pillar of the active case, every intermediate the
bootstrap used, surfaced from Bootstrap by date lookup so it is the real working
and not a re-derivation:

    date | tenor | S used | which rule | tau | A(prior) | numerator |
    denominator | our DF | t | our zero | BBG zero | BBG DF | d zero | d DF

so DF can be checked by hand:
    short  1W-1Y   DF = DFspot / (1 + S*tau)          tau ACT/360 from settle
    annual 18M-50Y DF = (DFspot - S*A) / (1 + S*tau)  A = SUM tau_i*DF_i prior
    zero           z  = -ln(DF)/t                     t ACT/365 from settle

CDS (Testing, lower block) - per tenor of the active entity: market spread, the
solved hazard, survival, the two legs, the model spread and the repricing error.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

WB = "/Users/nigelli/Desktop/openusdcurve/bloomberg/USD_SOFR_Curve_Bloomberg_Pricer.xlsx"
H1 = Font(name="Calibri", size=12, bold=True)
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
BLACK = Font(name="Calibri", size=11)
MONO = Font(name="Consolas", size=10)
NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
HF = PatternFill("solid", fgColor="1F3864")
WF = PatternFill("solid", fgColor="7B1FA2")
OF = PatternFill("solid", fgColor="FFF2CC")
BOX = Border(*[Side(style="thin", color="BFBFBF")]*4)
DTF, N5, N6, N8, BP = "mm/dd/yyyy", "0.00000", "0.000000", "0.00000000", "0.00"
R0, R1 = 7, 38
MODE = "Bootstrap!$G$4"
T = ["Test 1", "Test 2", "Test 3"]


def put(ws, cell, v, f=BLACK, fmt=None, fill=None, b=False, al=None):
    c = ws[cell]
    try: c.value = v
    except AttributeError: return None
    c.font = f
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if b: c.border = BOX
    if al: c.alignment = Alignment(horizontal=al, wrap_text=True, vertical="center")
    return c


wb = load_workbook(WB)
ws = wb["Testing"]

COLS = [("O","Date",13,DTF),("P","Tenor",8,None),("Q","S used %",10,N5),
        ("R","Rule",26,None),("S","tau",10,N6),("T","A(prior)",12,N8),
        ("U","numerator",12,N8),("V","denominator",12,N8),("W","our DF",13,N8),
        ("X","t ACT/365",10,N6),("Y","our zero %",11,N5),("Z","BBG zero %",11,N5),
        ("AA","BBG DF",12,N6),("AB","d zero bp",10,BP),("AC","d DF",11,"0.00E+00")]
for c, h, w, fmt in COLS:
    ws.column_dimensions[c].width = w
    put(ws, f"{c}{R0-1}", h, HDR, None, WF, True, "center")
put(ws, f"O{R0-3}", "WORKING — active case, every intermediate the bootstrap used", H1)
put(ws, f"O{R0-2}", "Pulled from Bootstrap by date, so this is the actual working, not a "
                    "re-derivation. Check any row by hand from the Rule column.", NOTE)

def pick(col):
    return (f'IF({MODE}="{T[0]}",{col}1{{r}},IF({MODE}="{T[1]}",{col}2{{r}},'
            f'IF({MODE}="{T[2]}",{col}3{{r}},"")))')

# case input columns: date / rate / bbg zero / bbg df
CD = {"1": ("B","C","D","E"), "2": ("F","G","H","I"), "3": ("J","K","L","M")}
def case(idx, r):
    a, b, c = CD["1"][idx], CD["2"][idx], CD["3"][idx]
    return (f'IF({MODE}="{T[0]}",{a}{r},IF({MODE}="{T[1]}",{b}{r},'
            f'IF({MODE}="{T[2]}",{c}{r},"")))')

B = lambda col: f"Bootstrap!${col}$8:${col}$72"
for i in range(R1 - R0 + 1):
    r = R0 + i
    d = case(0, r); z = case(2, r); f_ = case(3, r)
    m = f"MATCH(O{r},{B('B')},0)"
    put(ws, f"O{r}", f"={d}", BLACK, DTF, None, True)
    put(ws, f"P{r}", f"=A{r}", BLACK, None, None, True)
    put(ws, f"Q{r}", f'=IFERROR(INDEX({B("E")},{m}),"")', BLACK, N5, None, True)
    put(ws, f"R{r}",
        f'=IFERROR(IF({m}<=15,"short  DFspot/(1+S*tau)","annual (DFspot-S*A)/(1+S*tau)"),"")',
        MONO, None, None, True)
    # short rows use tau0 (col D), annual rows use the coupon tau (col F)
    put(ws, f"S{r}", f'=IFERROR(IF({m}<=15,INDEX({B("D")},{m}),INDEX({B("F")},{m})),"")',
        BLACK, N6, None, True)
    put(ws, f"T{r}", f'=IFERROR(IF({m}<=15,0,INDEX({B("G")},{m})),"")', BLACK, N8, None, True)
    put(ws, f"U{r}", f'=IF(OR(Q{r}="",S{r}=""),"",Bootstrap!$D$4-(Q{r}/100)*T{r})',
        BLACK, N8, None, True)
    put(ws, f"V{r}", f'=IF(OR(Q{r}="",S{r}=""),"",1+(Q{r}/100)*S{r})', BLACK, N8, None, True)
    put(ws, f"W{r}", f'=IFERROR(INDEX({B("H")},{m}),"")', BLACK, N8, None, True)
    put(ws, f"X{r}", f'=IFERROR(INDEX({B("C")},{m}),"")', BLACK, N6, None, True)
    put(ws, f"Y{r}", f'=IFERROR(INDEX({B("J")},{m}),"")', BLACK, N5, None, True)
    put(ws, f"Z{r}", f"={z}", BLACK, N5, None, True)
    put(ws, f"AA{r}", f"={f_}", BLACK, N6, None, True)
    put(ws, f"AB{r}", f'=IF(OR(Y{r}="",NOT(ISNUMBER(Z{r}))),"",(Y{r}-Z{r})*100)', BLACK, BP, None, True)
    put(ws, f"AC{r}", f'=IF(OR(W{r}="",NOT(ISNUMBER(AA{r}))),"",W{r}-AA{r})', BLACK, "0.00E+00", None, True)

s = R1 + 2
for lab, col, cell, fmt in [("Pillars matched","AB",f'=COUNT(Y{R0}:Y{R1})&" / 32"',None),
                            ("Max |d zero| (bp)","AB",
                             f'=IF(COUNT(AB{R0}:AB{R1})=0,"select a test case",MAX(MAX(AB{R0}:AB{R1}),-MIN(AB{R0}:AB{R1})))',"0.000"),
                            ("Max |d DF|","AC",
                             f'=IF(COUNT(AC{R0}:AC{R1})=0,"",MAX(MAX(AC{R0}:AC{R1}),-MIN(AC{R0}:AC{R1})))',"0.00E+00")]:
    put(ws, f"O{s}", lab, BOLD)
    put(ws, f"Q{s}", cell, BLACK, fmt, OF, True)
    s += 1
put(ws, f"O{s+1}", "Benchmark through this same path: 0.397bp and 1.88e-05 on the 07/21/26 "
                   "capture. A test case should land in that region.", NOTE)

# ---------- CDS working ----------
c0 = R1 + 14
put(ws, f"A{c0}", "CDS WORKING — active entity, per tenor", H1)
put(ws, f"A{c0+1}", "The strip for whichever name is selected at CDS_Parameters!B27. "
                    "Objective is (3.3): protection = S x (RPV01 net of accrued interest).", NOTE)
CH = [("A","Tenor",9,None),("B","Maturity",13,DTF),("C","Market bp",11,"0.0000"),
      ("D","lambda",13,N8),("E","Q(T)",11,N6),("F","1-Q",11,N6),
      ("G","RPV01",12,N6),("H","Protection",13,N8),("I","Model bp",11,"0.0000"),
      ("J","Err bp",12,"0.00E+00"),("K","s/(1-R)",12,N8),("L","lambda - s/(1-R)",15,N8)]
for c, h, w, fmt in CH:
    put(ws, f"{c}{c0+2}", h, HDR, None, WF, True, "center")
for i in range(6):
    r, hb = c0 + 3 + i, 7 + i
    for c, src in [("A","A"),("B","B"),("C","C"),("D","D"),("E","E"),("F","F"),
                   ("G","G"),("H","H"),("I","I"),("J","J")]:
        put(ws, f"{c}{r}", f"=Hazard_Bootstrap!{src}{hb}", BLACK,
            dict(CH)[c] if False else None, None, True)
    put(ws, f"K{r}", f"=CDS_Quotes!G{hb}", BLACK, N8, None, True)
    put(ws, f"L{r}", f'=IF(OR(NOT(ISNUMBER(D{r})),NOT(ISNUMBER(K{r}))),"",D{r}-K{r})',
        BLACK, N8, None, True)
for c, h, w, fmt in CH:
    for i in range(6):
        if fmt: ws[f"{c}{c0+3+i}"].number_format = fmt
put(ws, f"A{c0+10}", "K/L compare the strip against the credit triangle s/(1-R). They are not "
                     "meant to agree: lambda is the FORWARD hazard over each segment, s/(1-R) "
                     "the flat-equivalent to that tenor. The gap widens with curve slope.", NOTE)
put(ws, f"A{c0+12}", "Full bisection for every tenor is on Hazard_Solver (hidden): 30 halvings "
                     "of [0,3] per segment, with the carried-forward legs at the top of each "
                     "block. That sheet is the complete working.", NOTE)

ws.sheet_view.showGridLines = False
wb.calculation.fullCalcOnLoad = True
wb.save(WB)
print("working blocks written: curve O:AC, CDS at row", c0)
